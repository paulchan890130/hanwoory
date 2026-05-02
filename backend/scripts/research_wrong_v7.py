"""
Research the 13 WRONG NEEDS_MANUAL_CHECK rows from Task 1 v5_patch_research.

READ-ONLY. No DB modification. No LLM call.
Output: backend/data/manuals/v7_wrong_research.json + .xlsx
"""
import json
import re
import sys
from pathlib import Path

import fitz

ROOT = Path(__file__).parent.parent.parent
PDF_RES = ROOT / "backend" / "data" / "manuals" / "unlocked_체류민원.pdf"
PDF_VIS = ROOT / "backend" / "data" / "manuals" / "unlocked_사증민원.pdf"
DB_PATH = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
OUT_JSON = ROOT / "backend" / "data" / "manuals" / "v7_wrong_research.json"
OUT_XLSX = ROOT / "backend" / "data" / "manuals" / "v7_wrong_research.xlsx"

# Per-row research strategies — for each WRONG row, define search keywords + likely page hints
TARGETS = [
    # (row_id, detailed_code, action_type, title, search_keywords, manual_hint)
    ("M1-0086", "F-4-R", "CHANGE",
     "지역특화형 재외동포",
     ["F-4-R", "지역특화 동포", "지역특화동포", "지역특화형 재외동포"],
     "체류민원"),
    ("M1-0091", "F-2-71", "GRANT",
     "K-STAR 거주자의 국내출생 미성년자녀",
     ["F-2-71", "K-STAR", "국내출생", "체류자격 부여", "체류자격부여"],
     "체류민원"),
    ("M1-0112", "D-8-2", "EXTEND",
     "벤처투자",
     ["D-8-2", "벤처투자", "벤처기업", "기업투자(D-8)", "체류기간 연장"],
     "체류민원"),
    ("M1-0113", "D-8-3", "EXTEND",
     "개인기업투자",
     ["D-8-3", "개인기업투자", "개인기업", "체류기간 연장"],
     "체류민원"),
    ("M1-0124", "D-9-4", "CHANGE",
     "개인사업자 무역경영",
     ["D-9-4", "개인사업자", "무역경영(D-9)", "체류자격 변경"],
     "체류민원"),
    ("M1-0175", "E-7-4", "CHANGE",
     "숙련기능인력",
     ["E-7-4", "숙련기능인력", "체류자격 변경", "변경허가"],
     "체류민원"),
    ("M1-0198", "E-10", "CHANGE",
     "선원취업",
     ["E-10", "선원취업", "체류자격 변경"],
     "체류민원"),
    ("M1-0256", "F-2-71", "GRANT",
     "점수제 우수인재 국내출생 자녀",
     ["F-2-71", "K-STAR", "점수제 우수인재", "국내출생", "체류자격 부여"],
     "체류민원"),
    ("M1-0320", "E-7-S", "CHANGE",
     "네거티브방식 전문인력",
     ["E-7-S", "네거티브", "네거티브방식", "특정활동(E-7)"],
     "체류민원"),
    ("M1-0364", "F-5-1", "CHANGE",
     "영주 변경 — 국민의 배우자·자녀",
     ["F-5-1", "국민의 배우자", "국민의배우자·자녀", "영주(F-5) 자격변경"],
     "체류민원"),
    ("M1-0365", "F-5-2", "CHANGE",
     "영주 변경 — 영주권자의 미성년 자녀",
     ["F-5-2", "영주권자의 미성년", "영주(F-5) 자격변경"],
     "체류민원"),
    ("M1-0367", "F-5-10", "CHANGE",
     "영주 변경 — 재외동포(F-4) 2년",
     ["F-5-10", "재외동포(F-4) 자격으로 2년", "재외동포(F-4) 자격으로  2년"],
     "체류민원"),
    ("M1-0369", "F-5-14", "CHANGE",
     "영주 변경 — H-2 제조업 4년",
     ["F-5-14", "방문취업(H-2)", "제조업", "F-5-14"],
     "체류민원"),
]


def page_heading_lines(page_text: str, n: int = 3) -> str:
    lines = [l.strip() for l in page_text.splitlines() if l.strip()]
    top = []
    for l in lines:
        if re.match(r"^- \d+ -$", l) or l == "목차":
            continue
        top.append(l)
        if len(top) >= n:
            break
    return " / ".join(top)[:200]


def find_pages_with_terms(doc, terms: list[str], max_hits: int = 4) -> list[tuple[int, str, str]]:
    """Returns list of (page_no_1based, matched_term, heading_snippet)."""
    out = []
    for p in range(doc.page_count):
        txt = doc.load_page(p).get_text() or ""
        for term in terms:
            if term in txt:
                heading = page_heading_lines(txt)
                out.append((p + 1, term, heading))
                break
        if len(out) >= max_hits:
            break
    return out


def has_application_artifacts(page_text: str) -> dict:
    """Check if page contains 신청요건 or 제출서류 markers."""
    return {
        "has_요건":      bool(re.search(r"신청요건|허가요건|허가\s*요건|허용\s*대상|허용대상|허가\s*기준", page_text)),
        "has_제출서류":  bool(re.search(r"제출서류|첨부서류|신청서류", page_text)),
        "has_action_change": bool(re.search(r"체류자격\s*변경|변경허가", page_text)),
        "has_action_extend": bool(re.search(r"체류기간\s*연장|연장허가", page_text)),
        "has_action_grant":  bool(re.search(r"체류자격\s*부여|자격부여", page_text)),
        "has_action_reentry": bool(re.search(r"재입국허가", page_text)),
        "has_action_register": bool(re.search(r"외국인등록", page_text)),
    }


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    docs = {
        "체류민원": fitz.open(str(PDF_RES)),
        "사증민원": fitz.open(str(PDF_VIS)),
    }

    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    idx = {r["row_id"]: r for r in db["master_rows"]}

    report = []
    for rid, code, action, title, terms, manual_hint in TARGETS:
        cur_row = idx.get(rid, {})
        cur_refs = cur_row.get("manual_ref") or []
        cur_summary = ", ".join(
            f"{e.get('manual')}/p.{e.get('page_from')}-{e.get('page_to')}" for e in cur_refs[:2]
        ) or "(empty)"

        # Search both manuals
        candidates = []
        for manual_name in ["체류민원", "사증민원"]:
            doc = docs[manual_name]
            hits = find_pages_with_terms(doc, terms, max_hits=8)
            for page, term, heading in hits:
                page_text = doc.load_page(page - 1).get_text() or ""
                arts = has_application_artifacts(page_text)
                # Score: heading strong + application artifacts present
                score = 0
                if term in heading:
                    score += 50
                if arts["has_요건"]:
                    score += 20
                if arts["has_제출서류"]:
                    score += 30
                # action-specific bonuses
                if action == "CHANGE" and arts["has_action_change"]:
                    score += 15
                if action == "EXTEND" and arts["has_action_extend"]:
                    score += 15
                if action == "GRANT" and arts["has_action_grant"]:
                    score += 15
                candidates.append({
                    "manual": manual_name,
                    "page": page,
                    "matched_term": term,
                    "heading": heading,
                    "has_요건": arts["has_요건"],
                    "has_제출서류": arts["has_제출서류"],
                    "score": score,
                })

        candidates.sort(key=lambda x: -x["score"])
        top3 = candidates[:5]

        # Pick best
        if top3 and top3[0]["score"] > 0:
            best = top3[0]
            best_page = best["page"]
            best_manual = best["manual"]
            best_heading = best["heading"]
            confidence = "HIGH" if best["score"] >= 60 else ("MEDIUM" if best["score"] >= 30 else "LOW")
        else:
            best = None
            best_page = None
            best_manual = manual_hint
            best_heading = ""
            confidence = "NONE"

        cur_page = (cur_refs[0].get("page_from") if cur_refs else None)
        cur_manual = (cur_refs[0].get("manual") if cur_refs else None)

        report.append({
            "row_id":          rid,
            "detailed_code":   code,
            "action_type":     action,
            "title":           title,
            "current_manual":  cur_manual,
            "current_page":    cur_page,
            "current_summary": cur_summary,
            "found_manual":    best_manual,
            "found_page":      best_page,
            "heading_snippet": best_heading,
            "confidence":      confidence,
            "best_score":      best["score"] if best else 0,
            "top_candidates":  top3,
        })

    OUT_JSON.write_text(json.dumps({
        "meta": {
            "total_targets": len(TARGETS),
            "scoring": "heading_match=50 / 신청요건=20 / 제출서류=30 / action_kw=15"
        },
        "rows": report,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("research")
        headers = ["row_id", "code", "action", "title",
                   "current_manual/page", "found_manual", "found_page",
                   "heading_snippet", "confidence", "best_score"]
        widths = [10, 12, 14, 32, 24, 14, 12, 80, 12, 12]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for x in report:
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                f"{x['current_manual']} p.{x['current_page']}",
                x["found_manual"], x["found_page"],
                x["heading_snippet"], x["confidence"], x["best_score"],
            ])
            color = {"HIGH": "C6F6D5", "MEDIUM": "FFE599",
                     "LOW": "FFCC99", "NONE": "FFCCCB"}.get(x["confidence"], "FFFFFF")
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    print()
    print("=== Per-row research result ===")
    print(f"{'row_id':<8} {'code':<8} {'action':<14} {'cur':<22} {'→':<3} {'found':<22} {'conf':<8}")
    for x in report:
        cur = f"{x['current_manual']} p.{x['current_page']}"
        found = (f"{x['found_manual']} p.{x['found_page']}" if x['found_page'] else "(no match)")
        print(f"{x['row_id']:<8} {x['detailed_code']:<8} {x['action_type']:<14} {cur:<22} → {found:<22} {x['confidence']:<8}")
        if x["heading_snippet"]:
            print(f"         heading: {x['heading_snippet'][:130]}")


if __name__ == "__main__":
    main()
