"""
Patch v9 — preferred_manual_missing HIGH-confidence fixes (6 rows).

DEFAULT: dry-run. `--apply` required for DB write.

All 6 rows ADD the missing preferred_manual page (사증민원 or 체류민원)
while preserving the existing entry for the other manual.

Targets:
  Group A — add 사증민원 page to VISA_CONFIRM rows:
    M1-0116 D-8-3   VISA_CONFIRM → 사증 p.111 (D-8-3 명시, 사증발급인정서 발급대상)
    M1-0151 E-3     VISA_CONFIRM → 사증 p.160 (E-3 연구 사증발급인정서 발급대상)
    M1-0161 E-5     VISA_CONFIRM → 사증 p.160 (전문직업 독립전문가 사증발급인정서, E-5 포함)

  Group B — add 체류민원 page to residence-action rows:
    M1-0291 H-1     EXTEND       → 체류 p.517 (관광취업 체류기간연장 + 제출서류)
    M1-0292 H-1     REGISTRATION → 체류 p.517 (관광취업 외국인등록 포함)
    M1-0293 H-1     REENTRY      → 체류 p.517 (관광취업 재입국허가 포함)

Note: M1-0290 H-1 CHANGE is NOT patched here — 체류민원 p.517 explicitly states
"다른 체류자격에서 관광취업자격으로 자격변경 불가". Added to disposition file instead.

IMPORTANT: For rows that already have an entry for the non-preferred manual,
we ADD the new preferred-manual entry rather than replacing the full ref list.
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

ROOT       = Path(__file__).parent.parent.parent
DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"
DRY_JSON   = ROOT / "backend" / "data" / "manuals" / "v9_patch_dryrun.json"
DRY_XLSX   = ROOT / "backend" / "data" / "manuals" / "v9_patch_dryrun.xlsx"

# For each row: new_manual_ref replaces the entire list.
# We explicitly keep the old entry + add the new preferred-manual entry.
PATCHES: dict[str, dict] = {
    # Group A — add 사증민원 to VISA_CONFIRM rows
    "M1-0116": {
        "detailed_code": "D-8-3",
        "action_type":   "VISA_CONFIRM",
        "title":         "개인기업투자",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.111 '사증발급인정서 / 발급대상 / 3. 국민이 경영하는 개인기업에 투자한 외국인 "
            "기업투자(D-8-3) 자격' — D-8-3 명시적 등장 + req=True + doc=True"
        ),
        "non_contiguous": False,
        "old_ref_keep": True,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  111,
                "page_to":    111,
                "match_type": "manual_override",
                "match_text": "D-8-3 개인기업투자 VISA_CONFIRM 사증발급인정서 발급대상 (사증민원 p.111) — 수동 검증",
            },
        ],
    },
    "M1-0151": {
        "detailed_code": "E-3",
        "action_type":   "VISA_CONFIRM",
        "title":         "연구",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.160 'E-3 연구' 코드 확인 + 사증발급인정서 / 발급대상 / 독립전문가 항목 포함 "
            "+ req=True doc=True"
        ),
        "non_contiguous": False,
        "old_ref_keep": True,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  160,
                "page_to":    160,
                "match_type": "manual_override",
                "match_text": "E-3 연구 VISA_CONFIRM 사증발급인정서 발급대상 (사증민원 p.160) — 수동 검증",
            },
        ],
    },
    "M1-0161": {
        "detailed_code": "E-5",
        "action_type":   "VISA_CONFIRM",
        "title":         "전문직업",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.160 'E-5' 코드 확인 + 사증발급인정서 / 발급대상 + 독립전문가 항목 "
            "(전문직업(E-5) 포함) + doc=True"
        ),
        "non_contiguous": False,
        "old_ref_keep": True,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  160,
                "page_to":    160,
                "match_type": "manual_override",
                "match_text": "E-5 전문직업 VISA_CONFIRM 사증발급인정서 발급대상 (사증민원 p.160) — 수동 검증",
            },
        ],
    },
    # Group B — add 체류민원 to H-1 residence rows
    "M1-0291": {
        "detailed_code": "H-1",
        "action_type":   "EXTEND",
        "title":         "관광취업",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.517: ext=True, reg=True, ree=True, doc=True — "
            "'체류기간 / 연장허가' 섹션 + 관광취업(H-1) 외국인등록·재입국허가 공통 페이지"
        ),
        "non_contiguous": False,
        "old_ref_keep": True,
        "add_entries": [
            {
                "manual":     "체류민원",
                "page_from":  517,
                "page_to":    517,
                "match_type": "manual_override",
                "match_text": "H-1 관광취업 EXTEND 체류기간연장허가 (체류민원 p.517) — 수동 검증",
            },
        ],
    },
    "M1-0292": {
        "detailed_code": "H-1",
        "action_type":   "REGISTRATION",
        "title":         "관광취업",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.517: reg=True, doc=True — 관광취업(H-1) 외국인등록 포함 (ext·ree도 동페이지)"
        ),
        "non_contiguous": False,
        "old_ref_keep": True,
        "add_entries": [
            {
                "manual":     "체류민원",
                "page_from":  517,
                "page_to":    517,
                "match_type": "manual_override",
                "match_text": "H-1 관광취업 REGISTRATION 외국인등록 (체류민원 p.517) — 수동 검증",
            },
        ],
    },
    "M1-0293": {
        "detailed_code": "H-1",
        "action_type":   "REENTRY",
        "title":         "관광취업",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.517: ree=True, doc=True — 관광취업(H-1) 재입국허가 포함 (ext·reg도 동페이지)"
        ),
        "non_contiguous": False,
        "old_ref_keep": True,
        "add_entries": [
            {
                "manual":     "체류민원",
                "page_from":  517,
                "page_to":    517,
                "match_type": "manual_override",
                "match_text": "H-1 관광취업 REENTRY 재입국허가 (체류민원 p.517) — 수동 검증",
            },
        ],
    },
}


def _abort(msg: str) -> None:
    print(f"\n[ABORT] {msg}", file=sys.stderr)
    sys.exit(1)


def make_backup(db_bytes: bytes) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"immigration_guidelines_db_v2.v9_patch_backup_{ts}.json"
    bk.write_bytes(db_bytes)
    if not bk.exists() or bk.stat().st_size != len(db_bytes):
        _abort(f"백업 작성 검증 실패: {bk}")
    try:
        json.loads(bk.read_text(encoding="utf-8"))
    except Exception as e:
        _abort(f"백업 JSON readback 실패: {e}")
    return bk


def diff_rows(before: list, after: list) -> list:
    bi = {r.get("row_id"): r for r in before}
    ai = {r.get("row_id"): r for r in after}
    out = []
    for rid in set(bi) | set(ai):
        b = bi.get(rid); a = ai.get(rid)
        if b is None or a is None:
            out.append({"row_id": rid, "added_or_removed": True}); continue
        diff = [k for k in set(b) | set(a) if b.get(k) != a.get(k)]
        if diff:
            out.append({"row_id": rid, "fields": sorted(diff)})
    return out


def load_db() -> tuple:
    raw = DB_PATH.read_bytes()
    return json.loads(raw.decode("utf-8")), raw


def build_patched_db(db: dict) -> tuple:
    new_db = json.loads(json.dumps(db, ensure_ascii=False))
    rows = new_db.get("master_rows") or []
    found = {r.get("row_id"): r for r in rows if r.get("row_id") in PATCHES}
    missing = [rid for rid in PATCHES if rid not in found]
    if missing:
        _abort(f"target row_ids 누락: {missing}")
    patch_report = []
    for rid, spec in PATCHES.items():
        row = found[rid]
        if row.get("detailed_code") != spec["detailed_code"]:
            _abort(f"{rid} detailed_code 불일치")
        if row.get("action_type") != spec["action_type"]:
            _abort(f"{rid} action_type 불일치")
        old_ref = row.get("manual_ref") or []
        # Build new_manual_ref: keep existing entries + add new preferred-manual entries
        kept = list(old_ref)
        new_entries = spec.get("add_entries") or []
        # Avoid duplicates by page signature
        existing_sigs = {(e.get("manual"), e.get("page_from"), e.get("page_to")) for e in kept}
        for ne in new_entries:
            sig = (ne["manual"], ne["page_from"], ne["page_to"])
            if sig not in existing_sigs:
                kept.append(ne)
        row["manual_ref"] = kept
        patch_report.append({
            "row_id":           rid,
            "detailed_code":    spec["detailed_code"],
            "action_type":      spec["action_type"],
            "title":            spec["title"],
            "preferred_manual": spec["preferred_manual"],
            "old_manual_ref":   old_ref,
            "new_manual_ref":   kept,
            "added_entries":    new_entries,
            "page_evidence":    spec["page_evidence"],
            "non_contiguous":   spec["non_contiguous"],
        })
    return new_db, patch_report


def verify_patch_safety(before_db: dict, after_db: dict) -> list:
    errors = []
    bm = before_db.get("master_rows") or []; am = after_db.get("master_rows") or []
    if len(bm) != 369: errors.append(f"master_rows 개수 이상: {len(bm)}")
    if len(am) != len(bm): errors.append(f"패치 후 개수 불일치: {len(am)}")
    bi = {r.get("row_id"): r for r in bm}; ai = {r.get("row_id"): r for r in am}
    if set(bi) != set(ai): errors.append("row_id 집합 변경됨")
    for rid in bi:
        b = bi[rid]; a = ai.get(rid, {})
        ck = set(b) | set(a)
        if rid in PATCHES:
            for k in ck:
                if k == "manual_ref": continue
                if b.get(k) != a.get(k):
                    errors.append(f"{rid} non-manual_ref 필드 변경됨: {k}")
        else:
            for k in ck:
                if b.get(k) != a.get(k):
                    errors.append(f"비-target row {rid} 필드 변경됨: {k}")
    for k in set(before_db) | set(after_db):
        if k == "master_rows": continue
        if before_db.get(k) != after_db.get(k):
            errors.append(f"top-level 키 변경됨: {k}")
    return errors


def write_dryrun_reports(patch_report: list, summary: dict) -> None:
    DRY_JSON.write_text(json.dumps({
        "patch_meta": {"version": "v9", "mode": "DRY_RUN",
                       "scope": "6 rows: D-8-3/E-3/E-5 + 사증민원 / H-1 EXTEND/REG/REENTRY + 체류민원"},
        "summary": summary, "rows": patch_report,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {DRY_JSON.relative_to(ROOT)} ({DRY_JSON.stat().st_size:,} bytes)")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("patches")
        headers = ["row_id","code","action","title","preferred","old_ref","added_entries","evidence"]
        widths  = [10,12,14,36,14,80,100,80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for x in patch_report:
            ws.append([x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                       x["preferred_manual"],
                       json.dumps(x["old_manual_ref"], ensure_ascii=False)[:600],
                       json.dumps(x["added_entries"], ensure_ascii=False)[:800],
                       x["page_evidence"][:400]])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCB")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(DRY_XLSX)
        print(f"[OUT] {DRY_XLSX.relative_to(ROOT)} ({DRY_XLSX.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[skip xlsx] {e}")


def main() -> int:
    try: sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    is_apply = args.apply
    if not DB_PATH.exists(): _abort(f"DB 없음: {DB_PATH}")
    db, raw = load_db()
    bytes_before = len(raw)
    bm = db.get("master_rows") or []
    if len(bm) != 369: _abort("master_rows 개수 비정상")
    print(f"[patch] mode={'APPLY' if is_apply else 'DRY-RUN'}  DB={bytes_before:,}  master_rows={len(bm)}")
    print(f"[patch] targets ({len(PATCHES)}): {sorted(PATCHES.keys())}")
    new_db, patch_report = build_patched_db(db)
    errors = verify_patch_safety(db, new_db)
    if errors:
        for e in errors: print(f"  [error] {e}")
        _abort("safety check 실패")
    diffs = diff_rows(bm, new_db.get("master_rows") or [])
    diff_ids = {d["row_id"] for d in diffs}
    if diff_ids != set(PATCHES.keys()):
        _abort(f"diff 불일치: diff={sorted(diff_ids)} expected={sorted(PATCHES.keys())}")
    for d in diffs:
        if d.get("fields") != ["manual_ref"]:
            _abort(f"{d['row_id']} manual_ref 외 필드 변경: {d.get('fields')}")
    summary = {
        "mode": "APPLY" if is_apply else "DRY_RUN",
        "production_db_modified": False,
        "patch_candidate_count": len(patch_report),
        "diff_row_ids": sorted(diff_ids),
        "diff_count": len(diff_ids),
        "non_target_row_diffs": 0,
        "non_manual_ref_field_diffs": 0,
        "db_byte_size_before": bytes_before,
    }
    write_dryrun_reports(patch_report, summary)
    if not is_apply:
        print(f"\n{'='*70}\nPATCH v9 — DRY-RUN ({len(PATCHES)} rows)\n{'='*70}")
        print(f"  targets: {sorted(diff_ids)}")
        print(f"  non-target diffs: 0  /  non-manual_ref field diffs: 0")
        print(f"  DB: {bytes_before:,} → {bytes_before:,} (unchanged)")
        print()
        for x in patch_report:
            print(f"  {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}")
            print(f"    OLD refs: {len(x['old_manual_ref'])} entries  →  NEW refs: {len(x['new_manual_ref'])} entries")
            for e in x["added_entries"]:
                print(f"    + {e['manual']} p.{e['page_from']}-{e['page_to']}  [{e['match_text'][:80]}]")
        print()
        print("  Apply command (NOT EXECUTED):\n    python backend/scripts/patch_v9.py --apply")
        return 0
    # APPLY
    print("\n[apply] 백업 생성...")
    bk = make_backup(raw)
    print(f"  backup: {bk.relative_to(ROOT)}")
    DB_PATH.write_text(json.dumps(new_db, ensure_ascii=False, indent=2), encoding="utf-8")
    readback = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rb_rows = readback.get("master_rows") or []
    if len(rb_rows) != 369:
        _abort(f"readback 개수 이상: {len(rb_rows)}\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    errs2 = verify_patch_safety(json.loads(bk.read_text(encoding="utf-8")), readback)
    if errs2:
        for e in errs2: print(f"  [error] {e}")
        _abort(f"사후 verify 실패\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    bytes_after = DB_PATH.stat().st_size
    print(f"\n{'='*70}\nPATCH v9 — APPLIED\n{'='*70}")
    print(f"  backup: {bk.relative_to(ROOT)}")
    print(f"  DB: {bytes_before:,} → {bytes_after:,}  rows updated: {len(patch_report)}")
    print(f"  rollback: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
