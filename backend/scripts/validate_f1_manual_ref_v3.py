"""
F-1 manual_ref 정밀 검증 (READ-ONLY, REPORT-ONLY)

검증 대상 (4건, USER_SPECIFIED_HIGH):
  M1-0209  F-1-21  VISA_CONFIRM  주한 외국공관원 가사보조인
  M1-0211  F-1-15  EXTEND        우수인재·투자자·유학생 부모 방문동거
  M1-0227  F-1-22  EXTEND        고액투자가 가사보조인 방문동거
  M1-0230  F-1-24  EXTEND        해외우수인재 가사보조인 방문동거

현재 DB 매핑: 체류민원 p.543 + 사증민원 p.404 (4건 동일)
검증 결과 (PDF 본문 텍스트 직접 인용):
  체류 p.543 / 사증 p.404 = "외국국적동포가족 동반(F-3) 및 방문동거(F-1) 사증·체류 세부절차"
                            (재외동포(F-4) 자격 취득자의 부모/배우자/자녀 = 폐지된 F-1-72)
  → 4건 모두 "가사보조인" 또는 "비-동포 부모" 카테고리이므로 543/404 와 무관

쓰기 보장: DB / blocklist / triage / verify / 매뉴얼 모든 파일 무수정.
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

ROOT     = Path(__file__).parent.parent.parent
DB_PATH  = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUPS  = ROOT / "backend" / "data" / "backups"
ANOMALY  = ROOT / "backend" / "data" / "manuals" / "manual_ref_post_apply_anomalies_v3.json"
PDF_RES  = ROOT / "backend" / "data" / "manuals" / "unlocked_체류민원.pdf"
PDF_VIS  = ROOT / "backend" / "data" / "manuals" / "unlocked_사증민원.pdf"

OUT_JSON = ROOT / "backend" / "data" / "manuals" / "f1_manual_ref_validation_v3.json"
OUT_XLSX = ROOT / "backend" / "data" / "manuals" / "f1_manual_ref_validation_v3.xlsx"

TARGETS = ["M1-0209", "M1-0211", "M1-0227", "M1-0230"]

# 페이지 본문 발췌 (PDF에서 직접 추출 후 의미 단위로 잘라낸 인용)
RES_543_HEADING = "6. 외국국적동포가족 동반(F-3) 및 방문동거(F-1) 사증·체류 세부절차"
RES_543_PREVIEW = (
    "외국국적동포의 가족에 대한 동반(F-3) 및 방문동거(F-1)에 관한 사증발급 및 체류의 "
    "세부절차에 관한 사항 / 나. 외국국적동포가족 방문동거(F-1) 자격 기본 대상 = "
    "재외동포(F-4) 자격을 취득한 자의 가족(부모) / 폐지된 F-1-72 영주(F-5-7) 신청자 가족."
)
VIS_404_HEADING = "6. 외국국적동포가족 동반(F-3) 및 방문동거(F-1) 사증·체류 세부절차"
VIS_404_PREVIEW = (
    "재외동포(F-4) 가족에 대한 사증발급 및 체류 세부절차 — 동반(F-3)은 배우자·미성년자녀, "
    "방문동거(F-1)는 부모. 폐지된 F-1-72(영주 F-5-7 신청자 가족 방문동거) 후속 처리."
)

# 실제 정답 페이지 (PDF 본문에서 직접 확인)
CORRECT_PAGES = {
    "M1-0209": {  # F-1-21 VISA_CONFIRM 외국공관원 가사보조인
        "preferred_manual": "사증민원",
        "correct_residence": [
            ("체류민원", 337, 337, "방문동거(F-1) 자격 개요 (해당자 = 주한외국공관원의 가사보조인)"),
            ("체류민원", 340, 341, "체류자격 변경허가 1. 주한 외국공관 등 비세대동거인 및 가사보조인"),
        ],
        "correct_visa": [
            ("사증민원", 298, 298, "2. 주한 외국공관원의 가사보조인(F-1-21) 단수사증 발급"),
            ("사증민원", 306, 307, "3. 주한 외국공관원의 가사보조인(F-1-21) 사증발급인정서 첨부서류"),
        ],
    },
    "M1-0211": {  # F-1-15 EXTEND 우수인재·투자자·유학생 부모
        "preferred_manual": "체류민원",
        "correct_residence": [
            ("체류민원", 349, 349, "체류기간 연장허가 8. 우수인재, 투자자 및 유학생 부모 — 제출서류"),
        ],
        "correct_visa": [
            ("사증민원", 305, 305, "사증발급인정서 1. 우수인재, 투자자 및 유학생 부모(F-1-15) — 초청자 요건/첨부서류"),
        ],
    },
    "M1-0227": {  # F-1-22 EXTEND 고액투자가 가사보조인
        "preferred_manual": "체류민원",
        "correct_residence": [
            ("체류민원", 348, 348, "체류기간 연장허가 (외국인투자자 / 우수전문인력의 가사보조인 절차)"),
        ],
        "correct_visa": [
            ("사증민원", 299, 299, "투자가·전문인력의 외국인 가사보조인(F-1-22, F-1-23, F-1-24) 사증발급 기본"),
            ("사증민원", 306, 306, "사증발급인정서 2. 고액투자가(F-1-22) 및 해외우수인재(F-1-24)의 가사보조인"),
        ],
    },
    "M1-0230": {  # F-1-24 EXTEND 해외우수인재 가사보조인
        "preferred_manual": "체류민원",
        "correct_residence": [
            ("체류민원", 348, 348, "체류기간 연장허가 (외국인투자자 / 우수전문인력의 가사보조인 절차)"),
            ("체류민원", 668, 669, "최우수인재 부모·가사보조인 절차 — 신청서류"),
        ],
        "correct_visa": [
            ("사증민원", 299, 299, "투자가·전문인력의 외국인 가사보조인(F-1-22, F-1-23, F-1-24) 사증발급 기본"),
            ("사증민원", 306, 306, "사증발급인정서 2. 고액투자가(F-1-22) 및 해외우수인재(F-1-24)의 가사보조인"),
        ],
    },
}


def find_latest_backup() -> Path | None:
    if not BACKUPS.exists():
        return None
    bs = sorted(BACKUPS.glob("immigration_guidelines_db_v2.manual_ref_backup_*.json"))
    return bs[-1] if bs else None


def build_recommended_ref(target_id: str) -> list[dict]:
    """preferred_manual 의 정답 페이지만 manual_ref 형태로."""
    spec = CORRECT_PAGES[target_id]
    pref = spec["preferred_manual"]
    src = spec["correct_residence"] if pref == "체류민원" else spec["correct_visa"]
    return [
        {
            "manual": m,
            "page_from": pf,
            "page_to": pt,
            "match_text": txt,
            "match_type": "user_validated_section",
        }
        for (m, pf, pt, txt) in src
    ]


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not DB_PATH.exists():
        print(f"[ABORT] DB 없음: {DB_PATH}"); sys.exit(1)
    backup = find_latest_backup()
    if not backup:
        print(f"[ABORT] backup 없음"); sys.exit(1)

    print("[validate] DB 로드...")
    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rows = db.get("master_rows") or []
    cur_index = {r.get("row_id"): r for r in rows}

    print(f"[validate] backup 로드: {backup.name}")
    backup_db = json.loads(backup.read_text(encoding="utf-8"))
    bk_index = {r.get("row_id"): r for r in (backup_db.get("master_rows") or [])}

    print("[validate] anomaly v3 로드...")
    if ANOMALY.exists():
        anomaly = json.loads(ANOMALY.read_text(encoding="utf-8"))
        anom_index = {r.get("row_id"): r for r in (anomaly.get("rows") or [])}
    else:
        anom_index = {}

    out_rows: list[dict] = []
    for rid in TARGETS:
        cur = cur_index.get(rid)
        if not cur:
            print(f"  [skip] {rid} not in DB")
            continue
        bk = bk_index.get(rid, {})
        spec = CORRECT_PAGES[rid]
        pref = spec["preferred_manual"]
        action = cur.get("action_type")
        title = cur.get("business_name", "")

        # residence p.543 / visa p.404 검증
        # 본문 = 외국국적동포가족 (F-3 동반 + F-1 방문동거 = 재외동포 부모, 폐지된 F-1-72 후속).
        # 4건 모두 가사보조인 또는 비-동포 부모 → 본문이 row 와 무관 → "wrong"
        residence_check = {
            "page": 543,
            "heading": RES_543_HEADING,
            "text_preview": RES_543_PREVIEW,
            "relevance": "wrong",
            "reason": (
                "본문은 외국국적동포가족 (재외동포 F-4 자격 취득자의 부모) 방문동거 절차 "
                f"(폐지된 F-1-72 후속). 본 row({title}) 와 무관."
            ),
        }
        visa_check = {
            "page": 404,
            "heading": VIS_404_HEADING,
            "text_preview": VIS_404_PREVIEW,
            "relevance": "wrong",
            "reason": (
                "본문은 체류 p.543 과 동일 — 외국국적동포가족 방문동거 절차. "
                f"본 row({title}) 와 무관."
            ),
        }

        # 1차 진단 단계: 정답 override 가 모듈에 등록되기 전이므로 MANUAL_OVERRIDE_REQUIRED.
        # (DB·indexer 둘 다 이번에는 변경하지 않음 — 패치 스크립트 별도 단계.)
        recommended_action = "MANUAL_OVERRIDE_REQUIRED"
        recommended_ref = build_recommended_ref(rid)
        reason_lines = [
            f"action_type={action} → preferred={pref}",
            f"current_manual_ref 의 체류 p.543 본문 = {RES_543_HEADING} (외국국적동포가족 방문동거).",
            f"current_manual_ref 의 사증 p.404 본문 = 동일 (외국국적동포가족 방문동거).",
            f"본 row 는 '{title}' 카테고리이므로 543/404 의 동포가족 절차와 무관.",
            f"정답 페이지 (PDF 본문 직접 확인): "
            + "; ".join(
                f"{m} p.{pf}" + (f"-{pt}" if pt != pf else "") + f" ({txt})"
                for (m, pf, pt, txt) in (
                    spec["correct_residence"] if pref == "체류민원" else spec["correct_visa"]
                )
            ),
            "MANUAL_PAGE_OVERRIDE 등록 + 재인덱싱 (또는 surgical patch) 필요.",
        ]

        out_rows.append({
            "row_id":              rid,
            "detailed_code":       cur.get("detailed_code"),
            "action_type":         action,
            "title":               title,
            "current_manual_ref":  cur.get("manual_ref") or [],
            "backup_manual_ref":   bk.get("manual_ref") or [],
            "preferred_manual":    pref,
            "residence_page_check": residence_check,
            "visa_page_check":      visa_check,
            "correct_residence_pages": [
                {"manual": m, "page_from": pf, "page_to": pt, "context": ctx}
                for (m, pf, pt, ctx) in spec["correct_residence"]
            ],
            "correct_visa_pages": [
                {"manual": m, "page_from": pf, "page_to": pt, "context": ctx}
                for (m, pf, pt, ctx) in spec["correct_visa"]
            ],
            "recommended_manual_ref": recommended_ref,
            "recommended_action":     recommended_action,
            "reason":                 " // ".join(reason_lines),
            "anomaly_v3_found":       rid in anom_index,
        })

    summary = {
        "targets":              TARGETS,
        "rows_validated":       len(out_rows),
        "all_current_543_404":  all(
            any(r.get("page_from") == 543 for r in (x["current_manual_ref"] or []))
            and any(r.get("page_from") == 404 for r in (x["current_manual_ref"] or []))
            for x in out_rows
        ),
        "recommended_actions":  {
            x["row_id"]: x["recommended_action"] for x in out_rows
        },
        "preferred_manuals":    {
            x["row_id"]: x["preferred_manual"] for x in out_rows
        },
        "production_db_modified": False,
    }

    out = {
        "validation_meta": {
            "version":     "v3",
            "scope":       "F-1 USER_SPECIFIED_HIGH 4 rows",
            "db_path":     str(DB_PATH.relative_to(ROOT)),
            "backup_path": str(backup.relative_to(ROOT)),
            "rules": {
                "EXTEND":       "preferred 체류민원, 사증민원은 명시적 사증 참조 시에만",
                "VISA_CONFIRM": "preferred 사증민원, 체류민원은 명시적 체류 참조 시에만",
                "broad_section_correct": "KEEP_PRIMARY_ONLY",
                "wrong_section":         "MANUAL_REVIEW",
                "subtype_specific_known": "MANUAL_OVERRIDE_REQUIRED",
            },
            "page_evidence": {
                "체류민원_p543": RES_543_HEADING + " — " + RES_543_PREVIEW,
                "사증민원_p404": VIS_404_HEADING + " — " + VIS_404_PREVIEW,
            },
        },
        "summary": summary,
        "rows":    out_rows,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 80

        ws = wb.create_sheet("validation")
        headers = [
            "row_id", "detailed_code", "action_type", "title",
            "preferred_manual", "recommended_action",
            "current_manual_ref",
            "residence_p543_relevance", "residence_p543_reason",
            "visa_p404_relevance", "visa_p404_reason",
            "correct_residence_pages", "correct_visa_pages",
            "recommended_manual_ref", "anomaly_v3_found", "reason",
        ]
        widths = [10, 12, 14, 32, 12, 26, 50, 14, 60, 14, 60, 60, 60, 60, 14, 100]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for x in out_rows:
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                x["preferred_manual"], x["recommended_action"],
                json.dumps(x["current_manual_ref"], ensure_ascii=False)[:300],
                x["residence_page_check"]["relevance"],
                x["residence_page_check"]["reason"][:500],
                x["visa_page_check"]["relevance"],
                x["visa_page_check"]["reason"][:500],
                json.dumps(x["correct_residence_pages"], ensure_ascii=False)[:500],
                json.dumps(x["correct_visa_pages"], ensure_ascii=False)[:500],
                json.dumps(x["recommended_manual_ref"], ensure_ascii=False)[:500],
                x["anomaly_v3_found"],
                x["reason"][:1500],
            ])
            color = {
                "MANUAL_OVERRIDE_REQUIRED": "FFCCCB",
                "MANUAL_REVIEW":            "FFE599",
                "KEEP_PRIMARY_ONLY":        "C6F6D5",
                "KEEP_BOTH":                "C6F6D5",
            }.get(x["recommended_action"])
            if color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=color)

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # ── 콘솔 ──
    print("\n" + "=" * 78)
    print("F-1 MANUAL_REF VALIDATION REPORT")
    print("=" * 78)
    print("\n1. Files created:")
    print(f"   - {OUT_JSON.relative_to(ROOT)}")
    print(f"   - {OUT_XLSX.relative_to(ROOT)}")

    print("\n2. Production DB modified or not:")
    print("   No — DB / blocklist / triage / verify / 매뉴얼 모든 파일 무수정")

    print("\n3. Per-row decision:")
    for x in out_rows:
        print(
            f"\n   {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}  {x['title']}"
        )
        print(f"     preferred_manual:        {x['preferred_manual']}")
        print(f"     recommended_action:      {x['recommended_action']}")
        print(f"     current_manual_ref:      "
              + json.dumps(x['current_manual_ref'], ensure_ascii=False))
        print(f"     recommended_manual_ref:  "
              + json.dumps(x['recommended_manual_ref'], ensure_ascii=False))

    print("\n4. Should a surgical patch script be created next?")
    print("   YES — 4 rows have MANUAL_OVERRIDE_REQUIRED.")
    print("   권장 절차:")
    print("     (a) backend/services/manual_indexer_v6.MANUAL_PAGE_OVERRIDE 에 4 키 추가")
    print("         예: \"F-1-15|EXTEND\": [{체류, 349, 349}, {사증, 305, 305}]")
    print("     (b) 또는 surgical patch script `patch_f1_manual_ref_v3.py`")
    print("         - DB read-merge-write, 4 row 의 manual_ref 만 직접 교체")
    print("         - 자동 백업 + 사전·사후 검증 + 다른 138 row 무변경 보장")
    print("     (c) 패치 후 audit_post_apply_manual_ref_anomalies_v3.py 재실행으로 ")
    print("         USER_SPECIFIED_HIGH/SAME_PAGE_CLUSTER 23 row 클러스터 축소 확인")


if __name__ == "__main__":
    main()
