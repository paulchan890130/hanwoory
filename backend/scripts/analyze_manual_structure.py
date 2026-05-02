"""
매뉴얼 전체 구조 분석 (Phase 1)

매뉴얼 PDF 한 권을 통째로 LLM에 전달하여 구조 분석.
청크 분할 없이 1M context 활용 → 매뉴얼 전체 맥락 이해.

산출물 (영구 저장 — 재추출 안 함):
  backend/data/manuals/structure/
    체류민원_structure.json
    체류민원_structure.xlsx
    사증민원_structure.json
    사증민원_structure.xlsx

결과 활용:
  - 자격별 시작·끝 페이지 (정확)
  - sub-category 모든 정식 등장 페이지
  - 자격별 변경/연장/등록 action 페이지
  - PDF 페이지 ↔ 매뉴얼 인쇄 페이지 오프셋
"""
from __future__ import annotations
import json, os, sys, time, re
from pathlib import Path
from typing import Optional

import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

MANUALS_DIR = ROOT / "backend" / "data" / "manuals"
OUT_DIR     = MANUALS_DIR / "structure"
OUT_DIR.mkdir(parents=True, exist_ok=True)

MANUAL_PATHS = {
    "체류민원": MANUALS_DIR / "unlocked_체류민원.pdf",
    "사증민원": MANUALS_DIR / "unlocked_사증민원.pdf",
}


def extract_full_text(pdf_path: Path) -> str:
    """매뉴얼 전체 텍스트 추출 (PDF 페이지 마커 포함)."""
    doc = fitz.open(pdf_path)
    parts = []
    for p in doc:
        parts.append(f"\n----- p.{p.number + 1} -----\n")
        parts.append(p.get_text())
    doc.close()
    return "".join(parts)


SYSTEM_PROMPT = """\
당신은 한국 출입국·외국인정책본부 매뉴얼을 정독하고 구조를 정밀하게 추출하는 분석 전문가입니다.
매뉴얼 전체 텍스트를 받아 다음을 추출합니다:

1. 챕터/섹션 구조 (Ⅰ, Ⅱ, Ⅲ ... 또는 1, 2, 3 ...)
2. 각 체류자격(A-1, F-4, F-5, H-2 등)의 시작·끝 페이지 (해당 자격이 본문에서 다뤄지는 영역)
3. sub-category(예: F-5-1, F-5-4) 정식 등장 페이지 (헤더에 코드와 한글 명칭 함께 표기되는 곳)
4. 각 자격의 action_type(변경/연장/등록/재입국/사증발급/거소신고/근무처/체류자격외활동/자격부여) 페이지
5. PDF 페이지 vs 매뉴얼 인쇄 페이지(`-N-` 표기) 오프셋

페이지 번호는 모두 PDF의 ----- p.X ----- 마커의 X (시스템 페이지)로 답합니다.
매뉴얼 인쇄 페이지(-N-)는 별도 필드로 표기.

JSON 외의 텍스트는 출력 금지."""


USER_TEMPLATE = """\
다음은 [{manual_name}] 매뉴얼 전체 텍스트입니다 (PDF 페이지 마커 포함).

매뉴얼 전체를 정독한 뒤, 아래 JSON 스키마로 구조를 반환하세요.

```json
{{
  "manual_name": "{manual_name}",
  "page_offset_info": "PDF p.X = 매뉴얼 인쇄 -N- 의 관계 (예: -454-가 PDF p.453에 있으면 offset=-1)",
  "page_offset": -1,
  "chapters": [
    {{"chapter_no": "Ⅰ", "title": "...", "page_from": 1, "page_to": 100}}
  ],
  "qualities": {{
    "F-5": {{
      "name": "영주",
      "page_from": 425,
      "page_to": 470,
      "sub_categories": [
        {{"code": "F-5-1", "name": "국민의 배우자·자녀", "page_from": 445, "page_to": 452}},
        {{"code": "F-5-4", "name": "일반 영주자의 배우자/미성년 자녀", "page_from": 453, "page_to": 453}}
      ],
      "actions": {{
        "CHANGE":       {{"page_from": 428, "page_to": 470}},
        "REGISTRATION": null,
        "EXTEND":       null,
        "REENTRY":      null,
        "VISA_CONFIRM": null,
        "EXTRA_WORK":   null,
        "WORKPLACE":    null,
        "GRANT":        {{"page_from": 428, "page_to": 433}},
        "DOMESTIC_RESIDENCE_REPORT": null,
        "ACTIVITY_EXTRA": null
      }}
    }}
  }},
  "code_aliases": [
    {{"manual_code": "F-5-2", "manual_meaning": "결혼이민자", "note": "매뉴얼은 시행령 별표 1의3 호수 기반"}},
    {{"manual_code": "F-5-4", "manual_meaning": "영주자의 배우자/미성년 자녀"}}
  ],
  "notes": "매뉴얼 구조 특이사항 (선택)"
}}
```

자격코드는 매뉴얼에 등장하는 모든 sub-category까지 포함. action별 페이지가 명확히 갈라지지 않으면 null.

매뉴얼 텍스트:
{manual_text}"""


def call_anthropic(prompt: str, raw_save_path: Path, *,
                   model: str = "claude-sonnet-4-6",
                   max_tokens: int = 64000) -> dict:
    """매뉴얼 분석 LLM 호출 (streaming).

    - streaming API: 10분 이상 걸려도 처리 가능
    - max_tokens 64000: 매뉴얼 분석 결과 잘림 방지
    - 텍스트 누적하며 raw 즉시 저장
    - prompt caching 활용: 재호출 시 입력 비용 90% 할인
    """
    import anthropic
    client = anthropic.Anthropic()
    raw_save_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"  → API 호출 (model={model}, prompt {len(prompt):,} chars, streaming)")
    t0 = time.time()
    full_text = []
    last_print = t0

    with client.messages.stream(
        model=model,
        max_tokens=max_tokens,
        system=[
            {"type": "text", "text": SYSTEM_PROMPT,
             "cache_control": {"type": "ephemeral"}},
        ],
        messages=[{"role": "user", "content": [
            {"type": "text", "text": prompt,
             "cache_control": {"type": "ephemeral"}},
        ]}],
    ) as stream:
        for chunk in stream.text_stream:
            full_text.append(chunk)
            now = time.time()
            if now - last_print > 5:  # 5초마다 진행 표시
                acc = "".join(full_text)
                print(f"     ... {len(acc):,} chars 누적 ({now - t0:.0f}s 경과)")
                last_print = now
                # 부분 저장 (스트림 중단되어도 복구 가능)
                raw_save_path.write_text(acc, encoding="utf-8")
        final_message = stream.get_final_message()

    text = "".join(full_text)
    elapsed = time.time() - t0
    usage = final_message.usage
    print(f"  ← 응답 완료 ({elapsed:.1f}s, in={usage.input_tokens:,} out={usage.output_tokens:,})")
    cache_read = getattr(usage, "cache_read_input_tokens", 0) or 0
    cache_write = getattr(usage, "cache_creation_input_tokens", 0) or 0
    if cache_read or cache_write:
        print(f"     캐시 read={cache_read:,}, write={cache_write:,}")
    cost = (usage.input_tokens * 3e-6 + usage.output_tokens * 15e-6
            + cache_write * 3.75e-6 + cache_read * 0.3e-6)
    print(f"     예상비용: ~${cost:.3f}")

    # raw 최종 저장
    raw_save_path.write_text(text, encoding="utf-8")
    print(f"     raw 저장: {raw_save_path.name}")

    # JSON 추출
    body = text.strip()
    if body.startswith("```"):
        body = body.split("```", 2)[1]
        if body.startswith("json\n"):
            body = body[5:]
        body = body.rsplit("```", 1)[0]
    body = body.strip()

    try:
        return json.loads(body)
    except json.JSONDecodeError as e:
        print(f"  [WARN] JSON 파싱 실패 — 잘림 복구 시도: {e}")
        last_brace = body.rfind("}")
        if last_brace > 0:
            try:
                return json.loads(body[:last_brace + 1])
            except json.JSONDecodeError:
                pass
        raise RuntimeError(
            f"JSON 파싱 실패. raw 응답을 {raw_save_path}에서 직접 확인 후 수동 복구 필요"
        )


def write_excel(structure: dict, xlsx_path: Path):
    """구조 분석 결과를 Excel로 저장."""
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    fill_h = PatternFill(start_color="FFFFCC00", end_color="FFFFCC00", fill_type="solid")

    # 1. Chapters 시트
    ws = wb.active; ws.title = "Chapters"
    ws.append(["chapter_no", "title", "page_from", "page_to"])
    for c in ws[1]: c.font = bold; c.fill = fill_h
    for ch in structure.get("chapters", []):
        ws.append([ch.get("chapter_no", ""), ch.get("title", ""),
                   ch.get("page_from", 0), ch.get("page_to", 0)])

    # 2. Qualities 시트
    ws = wb.create_sheet("Qualities")
    ws.append(["code", "name", "page_from", "page_to"])
    for c in ws[1]: c.font = bold; c.fill = fill_h
    for code, q in structure.get("qualities", {}).items():
        ws.append([code, q.get("name", ""),
                   q.get("page_from", 0), q.get("page_to", 0)])

    # 3. Sub-categories 시트
    ws = wb.create_sheet("SubCategories")
    ws.append(["parent_code", "code", "name", "page_from", "page_to"])
    for c in ws[1]: c.font = bold; c.fill = fill_h
    for parent, q in structure.get("qualities", {}).items():
        for sc in q.get("sub_categories", []):
            ws.append([parent, sc.get("code", ""), sc.get("name", ""),
                       sc.get("page_from", 0), sc.get("page_to", 0)])

    # 4. Actions 시트
    ws = wb.create_sheet("Actions")
    ws.append(["code", "action_type", "page_from", "page_to"])
    for c in ws[1]: c.font = bold; c.fill = fill_h
    for code, q in structure.get("qualities", {}).items():
        for action, rng in (q.get("actions") or {}).items():
            if rng is None: continue
            ws.append([code, action, rng.get("page_from", 0), rng.get("page_to", 0)])

    # 5. Code Aliases 시트
    ws = wb.create_sheet("CodeAliases")
    ws.append(["manual_code", "manual_meaning", "note"])
    for c in ws[1]: c.font = bold; c.fill = fill_h
    for a in structure.get("code_aliases", []):
        ws.append([a.get("manual_code", ""), a.get("manual_meaning", ""),
                   a.get("note", "")])

    # 6. Meta 시트
    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    for c in ws[1]: c.font = bold; c.fill = fill_h
    ws.append(["manual_name", structure.get("manual_name", "")])
    ws.append(["page_offset", structure.get("page_offset", 0)])
    ws.append(["page_offset_info", structure.get("page_offset_info", "")])
    ws.append(["notes", structure.get("notes", "")])

    # 모든 시트 컬럼 너비 자동
    for sheet in wb.worksheets:
        for col in sheet.columns:
            mx = max((len(str(c.value)) for c in col if c.value), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(mx + 2, 80)

    wb.save(xlsx_path)
    print(f"  ✓ Excel 저장: {xlsx_path}")


def analyze(manual_label: str, force: bool = False):
    pdf_path  = MANUAL_PATHS[manual_label]
    json_path = OUT_DIR / f"{manual_label}_structure.json"
    xlsx_path = OUT_DIR / f"{manual_label}_structure.xlsx"

    if json_path.exists() and not force:
        print(f"[{manual_label}] 이미 분석됨: {json_path}")
        print(f"  → 재분석하려면 --force 옵션")
        return json.loads(json_path.read_text(encoding="utf-8"))

    print(f"\n=== {manual_label} 분석 시작 ===")
    print(f"  PDF: {pdf_path}")
    text = extract_full_text(pdf_path)
    print(f"  텍스트 추출: {len(text):,} chars")

    prompt = USER_TEMPLATE.format(manual_name=manual_label, manual_text=text)
    print(f"  prompt: {len(prompt):,} chars (~{len(prompt)//3:,} tokens)")

    raw_path = OUT_DIR / f"{manual_label}_raw_response.txt"
    structure = call_anthropic(prompt, raw_path)

    json_path.write_text(json.dumps(structure, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✓ JSON 저장: {json_path}")

    write_excel(structure, xlsx_path)

    return structure


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--manual", choices=list(MANUAL_PATHS), default=None,
                   help="특정 매뉴얼만 분석. 미지정 시 전체.")
    p.add_argument("--force", action="store_true", help="이미 분석된 결과도 재실행")
    args = p.parse_args()

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("[ERROR] ANTHROPIC_API_KEY 환경변수 필요"); sys.exit(1)

    targets = [args.manual] if args.manual else list(MANUAL_PATHS)
    for m in targets:
        analyze(m, force=args.force)
    print(f"\n[OK] 완료. 결과: {OUT_DIR}")
