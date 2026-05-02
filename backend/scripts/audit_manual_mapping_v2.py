"""
manual_ref 감사 v2 — action_type 우선 라우팅 계층 추가 (READ-ONLY)

v1 문제:
  E-9 시리즈 등이 conflict 폭증. 원인은 후보 결정 로직이 두 매뉴얼 중
  먼저 발견되는 쪽을 채택하면서 사증민원 structure JSON에만 sub_categories가
  있는 자격(E-9 등)을 잘못 사증민원으로 라우팅한 것.

v2 핵심:
  1. action_type → preferred_manual 결정을 가장 먼저.
  2. 두 매뉴얼에서 독립적으로 후보를 빌드한 뒤 preferred 쪽을 primary로.
  3. preferred에 후보가 없을 때만 secondary가 primary 자리에 들어가되
     confidence는 강제 하향 + routing_warning 기록.
  4. APPLICATION_CLAIM 또는 빈 detailed_code → unknown 라우팅 (수동 override만 인정).
  5. conflict는 좁게 정의: 두 매뉴얼이 모두 exact/high이거나, 현재 DB 매뉴얼이
     preferred와 다르고 preferred 후보가 exact/high일 때만.

출력:
  backend/data/manuals/manual_mapping_audit_v2.json
  backend/data/manuals/manual_mapping_audit_v2.xlsx (openpyxl 있으면)

쓰기 보장:
  위 두 파일 외 어떤 파일도 변경하지 않음. DB·인덱스·structure·watcher 모두 무수정.
"""
from __future__ import annotations
import json, re, sys, csv
from pathlib import Path
from collections import Counter

import fitz

ROOT     = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
MANUALS    = ROOT / "backend" / "data" / "manuals"
STRUCT     = MANUALS / "structure"
INDEX_V6   = MANUALS / "manual_index_v6.json"
CSV_CODES  = STRUCT / "법무부_체류자격 분류코드_20260331.csv"

V1_AUDIT   = MANUALS / "manual_mapping_audit.json"  # 비교용 (read-only)
OUT_JSON   = MANUALS / "manual_mapping_audit_v2.json"
OUT_XLSX   = MANUALS / "manual_mapping_audit_v2.xlsx"

PDF_PATHS = {
    "체류민원": MANUALS / "unlocked_체류민원.pdf",
    "사증민원": MANUALS / "unlocked_사증민원.pdf",
}

from backend.services.manual_indexer_v6 import (
    DB_TO_MANUAL_ALIAS,
    MANUAL_PAGE_OVERRIDE,
    ACTION_KEYWORDS,
)


# ─── 라우팅 규칙 ───────────────────────────────────────────────
# 사용자 스펙 그대로 키워드 집합으로 보관 (대소문자, 영문/한글 모두 인정).
_RESIDENCE_TOKENS = {
    "CHANGE", "EXTEND", "REGISTRATION", "STATUS_GRANT", "GRANT",
    "EXTRA_ACTIVITY", "EXTRA_WORK", "ACTIVITY_EXTRA",
    "WORKPLACE_CHANGE", "WORKPLACE",
    "REENTRY", "RESIDENCE_REPORT", "DOMESTIC_RESIDENCE_REPORT",
    "PERMISSION", "REPORT",
    "체류자격변경", "체류기간연장", "외국인등록", "체류자격부여",
    "체류자격외활동", "근무처변경", "재입국허가", "체류지변경",
    "신고", "허가",
}
_VISA_TOKENS = {
    "VISA", "VISA_ISSUANCE", "VISA_CONFIRMATION", "VISA_CONFIRM",
    "사증발급", "사증발급인정", "사증",
}
_UNKNOWN_TOKENS = {
    "APPLICATION_CLAIM", "", None,
}


def preferred_manual_for(action_type: str | None) -> tuple[str, str]:
    """(preferred_manual, routing_reason)."""
    a = (action_type or "").strip()
    if a in _UNKNOWN_TOKENS:
        return "unknown", "APPLICATION_CLAIM 또는 빈 action_type — 자동 라우팅 불가"
    if a in _VISA_TOKENS:
        return "사증민원", f"action_type={a} → 사증민원 (visa group)"
    if a in _RESIDENCE_TOKENS:
        return "체류민원", f"action_type={a} → 체류민원 (residence group)"
    return "unknown", f"action_type={a!r} 분류 규칙에 없음 — 라우팅 미적용"


# ─── 패턴 / 로딩 ───────────────────────────────────────────────
_BRACKET_FORMAL = re.compile(r"〔\s*([A-Z]-\d+(?:-(?:\d+|[A-Z][\w]*))?)\s*〕")


def load_pdf_pages(path: Path) -> tuple[dict[int, str], int]:
    doc = fitz.open(path)
    out = {p.number + 1: p.get_text() for p in doc}
    total = len(doc)
    doc.close()
    return out, total


def load_structure(name: str) -> dict | None:
    p = STRUCT / f"{name}_structure.json"
    if not p.exists():
        return None
    return json.loads(p.read_text(encoding="utf-8"))


def load_csv_codes(path: Path) -> set[str]:
    if not path.exists():
        return set()
    codes: set[str] = set()
    with open(path, encoding="cp949", errors="replace") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if row and row[0].strip():
                codes.add(row[0].strip())
    return codes


# ─── PDF 매칭 유틸 ──────────────────────────────────────────────
def find_bracket_section(pdf_pages, code, max_section_size=30):
    target = re.compile(rf"〔\s*{re.escape(code)}\s*〕")
    first = None
    for pno in sorted(pdf_pages):
        if target.search(pdf_pages[pno]):
            first = pno
            break
    if first is None:
        return None
    next_other = None
    for pno in sorted(p for p in pdf_pages if p > first):
        for m in _BRACKET_FORMAL.finditer(pdf_pages[pno]):
            if m.group(1) != code:
                next_other = pno
                break
        if next_other:
            break
    end = (next_other - 1) if next_other else (first + max_section_size)
    end = min(end, first + max_section_size, max(pdf_pages))
    return (first, end)


def find_paren_cluster(pdf_pages, code, min_count=3):
    pat = re.compile(rf"\(\s*{re.escape(code)}\s*\)")
    pages_score = {pno: len(pat.findall(t)) for pno, t in pdf_pages.items() if len(pat.findall(t)) >= min_count}
    if not pages_score:
        return None
    pages = sorted(pages_score)
    clusters, cur = [], [pages[0]]
    for p in pages[1:]:
        if p - cur[-1] <= 5:
            cur.append(p)
        else:
            clusters.append(cur); cur = [p]
    clusters.append(cur)
    best = max(clusters, key=lambda c: sum(pages_score[p] for p in c))
    return (best[0], best[-1])


def find_action_in_range(pdf_pages, pf, pt, action_type):
    kws = ACTION_KEYWORDS.get(action_type, [])
    if not kws:
        return None
    first = last = None; matched = None
    for p in range(pf, pt + 1):
        txt = pdf_pages.get(p, "")
        for k in kws:
            if k in txt:
                if first is None: first, matched = p, k
                last = p
                break
    return (first, last, matched or "") if first else None


# ─── 단일 매뉴얼 안에서만 후보 결정 ─────────────────────────────
def resolve_in_manual(row: dict, manual: str, pdf_pages: dict[int, str],
                      struct: dict | None, v6_index: dict | None) -> dict | None:
    """단일 매뉴얼 한정. 못 찾으면 None."""
    code = (row.get("detailed_code") or "").strip()
    at   = (row.get("action_type") or "").strip()
    if not code:
        return None

    # 1. MANUAL_PAGE_OVERRIDE — 이 매뉴얼로 등록된 것만
    ovr_key = f"{code}|{at}"
    if ovr_key in MANUAL_PAGE_OVERRIDE:
        for ovr in MANUAL_PAGE_OVERRIDE[ovr_key]:
            if ovr.get("manual") == manual:
                return {
                    "manual_ref": [{
                        "manual": manual,
                        "page_from": ovr["page_from"],
                        "page_to": ovr["page_to"],
                        "match_text": ovr.get("match_text", ""),
                        "match_type": "manual_override",
                    }],
                    "method": "manual_override",
                    "confidence": "exact",
                    "evidence": f"수동 override({manual}): {ovr.get('match_text','')}",
                }

    # 2. exact_bracket_heading
    section = find_bracket_section(pdf_pages, code)
    if section:
        pf, pt = section
        if at:
            hit = find_action_in_range(pdf_pages, pf, pt, at)
            if hit:
                af, al, kw = hit
                return {
                    "manual_ref": [{
                        "manual": manual, "page_from": af, "page_to": al,
                        "match_text": f"〔{code}〕 + '{kw}'",
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": "exact",
                    "evidence": f"{manual} p.{af}~{al}: 〔{code}〕 정식 sub-cat 안 '{kw}' 등장",
                }
        return {
            "manual_ref": [{
                "manual": manual, "page_from": pf, "page_to": pt,
                "match_text": f"〔{code}〕 정식 헤더",
                "match_type": "section_only",
            }],
            "method": "exact_bracket_heading",
            "confidence": "medium",
            "evidence": f"{manual} p.{pf}~{pt}: 〔{code}〕 정식 sub-cat (action 키워드 미발견)",
        }

    # 3. code_alias_exact
    alias = DB_TO_MANUAL_ALIAS.get(code)
    if alias and alias != code:
        section = find_bracket_section(pdf_pages, alias)
        if section:
            pf, pt = section
            if at:
                hit = find_action_in_range(pdf_pages, pf, pt, at)
                if hit:
                    af, al, kw = hit
                    return {
                        "manual_ref": [{
                            "manual": manual, "page_from": af, "page_to": al,
                            "match_text": f"〔{alias}〕(alias of {code}) + '{kw}'",
                            "match_type": "action_exact",
                        }],
                        "method": "code_alias_exact",
                        "confidence": "high",
                        "evidence": f"{manual} p.{af}~{al}: alias 〔{alias}〕 (DB:{code}) + '{kw}'",
                    }
            return {
                "manual_ref": [{
                    "manual": manual, "page_from": pf, "page_to": pt,
                    "match_text": f"〔{alias}〕(alias of {code})",
                    "match_type": "section_only",
                }],
                "method": "code_alias_exact",
                "confidence": "medium",
                "evidence": f"{manual} p.{pf}~{pt}: alias 〔{alias}〕 (DB:{code}) 섹션",
            }

    # 4. structure_heading
    if struct:
        quals = struct.get("qualities", {}) or {}
        # sub_categories 우선
        for q_code, q in quals.items():
            for sc in q.get("sub_categories", []) or []:
                if sc.get("code") == code:
                    pf = sc.get("page_from"); pt = sc.get("page_to") or pf
                    if not pf:
                        continue
                    if at:
                        hit = find_action_in_range(pdf_pages, pf, pt, at)
                        if hit:
                            af, al, kw = hit
                            return {
                                "manual_ref": [{
                                    "manual": manual, "page_from": af, "page_to": al,
                                    "match_text": f"structure {code} + '{kw}'",
                                    "match_type": "action_exact",
                                }],
                                "method": "structure_heading",
                                "confidence": "high",
                                "evidence": f"{manual} p.{af}~{al}: structure sub-cat {code} + '{kw}'",
                            }
                    return {
                        "manual_ref": [{
                            "manual": manual, "page_from": pf, "page_to": pt,
                            "match_text": f"structure {sc.get('name', code)}",
                            "match_type": "section_only",
                        }],
                        "method": "structure_heading",
                        "confidence": "medium",
                        "evidence": f"{manual} p.{pf}~{pt}: structure sub-cat {code} ({sc.get('name','')})",
                    }
        # quality(대분류) 일치
        if code in quals:
            q = quals[code]
            actions = q.get("actions") or {}
            act = actions.get(at) if at else None
            if act and act.get("page_from"):
                return {
                    "manual_ref": [{
                        "manual": manual,
                        "page_from": act["page_from"],
                        "page_to": act.get("page_to") or act["page_from"],
                        "match_text": f"structure {code}.actions.{at}",
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": "high",
                    "evidence": f"{manual} p.{act['page_from']}: structure quality {code}.actions.{at}",
                }
            pf = q.get("page_from"); pt = q.get("page_to") or pf
            if pf:
                return {
                    "manual_ref": [{
                        "manual": manual, "page_from": pf, "page_to": pt,
                        "match_text": f"structure {code}",
                        "match_type": "section_only",
                    }],
                    "method": "structure_heading",
                    "confidence": "medium",
                    "evidence": f"{manual} p.{pf}~{pt}: structure quality {code} 섹션",
                }

    # 5. v6 action_index — 이 매뉴얼 한정 entries만
    if v6_index:
        ai = v6_index.get("action_index", {}) or {}
        key = f"{code}|{at}"
        if key in ai:
            entries = [e for e in ai[key] if e.get("manual") == manual]
            if entries:
                best = max(entries, key=lambda e: e.get("score", 0))
                mk = best.get("match_kind", "")
                conf = "high" if mk in ("subcategory_with_action", "subcategory_change_default") else "medium"
                return {
                    "manual_ref": [{
                        "manual": manual,
                        "page_from": best["page_from"],
                        "page_to": best["page_to"],
                        "match_text": best.get("match_text", ""),
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": conf,
                    "evidence": f"v6 action_index {key} ({manual}): {best.get('match_text','')} kind={mk}",
                }

    # 6. (code) 본문 클러스터
    section = find_paren_cluster(pdf_pages, code)
    if section:
        pf, pt = section
        if at:
            hit = find_action_in_range(pdf_pages, pf, pt, at)
            if hit:
                af, al, kw = hit
                return {
                    "manual_ref": [{
                        "manual": manual, "page_from": af, "page_to": al,
                        "match_text": f"({code}) 본문 + '{kw}'",
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": "medium",
                    "evidence": f"{manual} p.{af}~{al}: ({code}) 본문 클러스터 + '{kw}'",
                }
        return {
            "manual_ref": [{
                "manual": manual, "page_from": pf, "page_to": pt,
                "match_text": f"({code}) 본문",
                "match_type": "section_only",
            }],
            "method": "section_only",
            "confidence": "medium",
            "evidence": f"{manual} p.{pf}~{pt}: ({code}) 자격 섹션",
        }

    # 7. prefix_fallback
    main = re.match(r"^([A-Z]-\d+)", code)
    if main and main.group(1) != code:
        prefix = main.group(1)
        section = find_paren_cluster(pdf_pages, prefix)
        if section:
            pf, pt = section
            return {
                "manual_ref": [{
                    "manual": manual, "page_from": pf, "page_to": pt,
                    "match_text": f"prefix {prefix}",
                    "match_type": "section_only",
                }],
                "method": "prefix_fallback",
                "confidence": "low",
                "evidence": f"{manual} p.{pf}~{pt}: prefix {prefix} 섹션",
            }

    return None


# ─── 라우팅 + primary/secondary 결정 ────────────────────────────
_CONF_RANK = {"none": 0, "low": 1, "medium": 2, "high": 3, "exact": 4}


def resolve_with_routing(row: dict, pdf_data: dict, struct_data: dict, v6_index: dict | None) -> dict:
    at   = (row.get("action_type") or "").strip()
    code = (row.get("detailed_code") or "").strip()
    preferred, reason = preferred_manual_for(at)

    # APPLICATION_CLAIM / 빈 코드 / 라우팅 미적용 — override 외엔 거절
    if preferred == "unknown" or not code:
        ovr_key = f"{code}|{at}"
        primary = None
        if ovr_key in MANUAL_PAGE_OVERRIDE and MANUAL_PAGE_OVERRIDE[ovr_key]:
            ovr = MANUAL_PAGE_OVERRIDE[ovr_key][0]
            primary = {
                "manual_ref": [{
                    "manual": ovr["manual"],
                    "page_from": ovr["page_from"], "page_to": ovr["page_to"],
                    "match_text": ovr.get("match_text", ""),
                    "match_type": "manual_override",
                }],
                "method": "manual_override",
                "confidence": "exact",
                "evidence": f"수동 override (라우팅 unknown 이지만 정답 등록됨)",
            }
        return {
            "routing": {
                "action_type": at,
                "preferred_manual": "unknown",
                "routing_reason": reason,
                "secondary_manual_allowed": False,
            },
            "primary_candidate": primary,
            "secondary_candidate": None,
            "routing_warning": "" if primary else "라우팅 미적용 — 자동 후보 결정 불가",
        }

    # 두 매뉴얼 독립 조회
    cand_c = resolve_in_manual(row, "체류민원", pdf_data.get("체류민원", {}),
                                struct_data.get("체류민원"), v6_index)
    cand_s = resolve_in_manual(row, "사증민원", pdf_data.get("사증민원", {}),
                                struct_data.get("사증민원"), v6_index)

    if preferred == "체류민원":
        primary, secondary = cand_c, cand_s
    else:  # 사증민원
        primary, secondary = cand_s, cand_c

    routing_warning = ""

    # primary가 비었고 secondary가 있으면 — 강제 하향 후 primary 자리에 옮김
    if primary is None and secondary is not None:
        old_conf = secondary["confidence"]
        # exact/high → low, medium/low/none → 그대로
        downgraded = "low" if _CONF_RANK[old_conf] >= _CONF_RANK["high"] else old_conf
        primary = {
            **secondary,
            "confidence": downgraded,
            "evidence": secondary["evidence"] + f" [routing 강제 하향 {old_conf}→{downgraded}: 비선호 매뉴얼 후보]",
        }
        secondary = None
        routing_warning = (
            f"action_type={at} 권장은 {preferred} 인데 {primary['manual_ref'][0]['manual']} 에서만 후보 검출 — "
            f"신뢰도 강제 하향({old_conf}→{downgraded})"
        )

    # primary가 비선호 매뉴얼인 경우 — primary 자체가 None이 아닌데 매뉴얼이 어긋남
    elif primary is not None and primary["manual_ref"]:
        prim_manual = primary["manual_ref"][0].get("manual")
        if prim_manual != preferred:
            old_conf = primary["confidence"]
            downgraded = "low" if _CONF_RANK[old_conf] >= _CONF_RANK["high"] else old_conf
            if downgraded != old_conf:
                primary = {
                    **primary,
                    "confidence": downgraded,
                    "evidence": primary["evidence"] + f" [routing 강제 하향 {old_conf}→{downgraded}]",
                }
                routing_warning = (
                    f"primary 후보가 비선호 매뉴얼({prim_manual}) — 신뢰도 하향({old_conf}→{downgraded})"
                )

    return {
        "routing": {
            "action_type": at,
            "preferred_manual": preferred,
            "routing_reason": reason,
            "secondary_manual_allowed": True,
        },
        "primary_candidate": primary,
        "secondary_candidate": secondary,
        "routing_warning": routing_warning,
    }


# ─── 비교 (좁은 conflict 정의) ──────────────────────────────────
def compare_v2(current: list, primary: dict | None, secondary: dict | None,
                preferred: str) -> dict:
    cur = current or []
    has_prim = primary is not None and primary.get("manual_ref")
    has_sec  = secondary is not None and secondary.get("manual_ref")

    if not cur and not has_prim:
        return {"status": "missing_both", "page_changed": False, "manual_changed": False}
    if not cur:
        return {"status": "missing_current", "page_changed": False, "manual_changed": False}
    if not has_prim:
        return {"status": "missing_candidate", "page_changed": False, "manual_changed": False}

    c0 = cur[0]
    p0 = primary["manual_ref"][0]
    page_changed   = (c0.get("page_from") != p0.get("page_from")) or (c0.get("page_to") != p0.get("page_to"))
    manual_changed = c0.get("manual") != p0.get("manual")

    # conflict (a): 두 매뉴얼이 모두 exact/high — 진짜 모호
    if has_sec and primary["confidence"] in ("exact", "high") and secondary["confidence"] in ("exact", "high"):
        return {
            "status": "conflict", "page_changed": page_changed, "manual_changed": manual_changed,
            "reason": "both_manuals_plausible",
        }

    # conflict (b): current 매뉴얼이 preferred와 다르고 primary가 preferred에서 exact/high
    if manual_changed and primary["confidence"] in ("exact", "high"):
        prim_manual = p0.get("manual")
        if prim_manual == preferred:
            return {
                "status": "conflict", "page_changed": page_changed, "manual_changed": True,
                "reason": "current_in_wrong_manual",
            }

    if not page_changed and not manual_changed:
        return {"status": "same", "page_changed": False, "manual_changed": False}
    return {"status": "changed", "page_changed": page_changed, "manual_changed": manual_changed}


def needs_review_v2(comp: dict, conf: str, routing_warning: str) -> tuple[bool, str]:
    reasons = []
    s = comp["status"]
    if s == "missing_current":   reasons.append("DB에 manual_ref 없음")
    if s == "missing_candidate": reasons.append("후보 페이지 없음")
    if s == "missing_both":      reasons.append("현재값/후보 모두 없음")
    if s == "conflict":          reasons.append(f"conflict ({comp.get('reason','')})")
    if s == "changed":           reasons.append("페이지 변경 감지")
    if conf == "low":            reasons.append("정확도 낮음 (prefix/하향)")
    if conf == "medium":         reasons.append("자격 섹션 단위 — action 미확정")
    if conf == "none":           reasons.append("매핑 불가")
    if routing_warning:          reasons.append("routing 경고")

    if conf == "exact" and s == "same" and not routing_warning:
        return False, ""
    if conf == "high" and s == "same" and not routing_warning:
        return False, ""
    if not reasons:
        return False, ""
    return True, " / ".join(reasons)


# ─── 메인 ──────────────────────────────────────────────────────
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    print("[audit-v2] DB 로드...")
    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rows = db.get("master_rows", []) or []
    print(f"  master_rows: {len(rows)}")

    print("[audit-v2] PDF 텍스트 추출...")
    pdf_data, pdf_meta = {}, {}
    for label, p in PDF_PATHS.items():
        if not p.exists():
            print(f"  [skip] {label}: 파일 없음"); continue
        pages, total = load_pdf_pages(p)
        pdf_data[label] = pages
        pdf_meta[label] = {"file": str(p.relative_to(ROOT)), "pages": total}
        print(f"  {label}: {total} pages")

    print("[audit-v2] structure JSON 로드...")
    struct_data = {"체류민원": load_structure("체류민원"), "사증민원": load_structure("사증민원")}

    print("[audit-v2] manual_index_v6 로드...")
    v6 = json.loads(INDEX_V6.read_text(encoding="utf-8")) if INDEX_V6.exists() else None

    print("[audit-v2] CSV 코드 로드...")
    csv_codes = load_csv_codes(CSV_CODES)
    print(f"  CSV 코드: {len(csv_codes)}")

    print("[audit-v2] 라우팅 + 후보 결정...")
    audit_rows = []
    for i, row in enumerate(rows):
        pkg = resolve_with_routing(row, pdf_data, struct_data, v6)
        primary   = pkg["primary_candidate"]
        secondary = pkg["secondary_candidate"]
        preferred = pkg["routing"]["preferred_manual"]

        comp = compare_v2(row.get("manual_ref") or [], primary, secondary, preferred)
        nr, reason = needs_review_v2(comp, (primary or {}).get("confidence", "none"), pkg["routing_warning"])

        code = (row.get("detailed_code") or "").strip()
        csv_known = (code in csv_codes) if code else None

        # 표면 호환: proposed_manual_ref / match 는 primary 기준
        proposed_ref = (primary or {}).get("manual_ref", [])
        match_summary = {
            "method":     (primary or {}).get("method", "none"),
            "confidence": (primary or {}).get("confidence", "none"),
            "evidence":   (primary or {}).get("evidence", "후보 없음"),
        }

        audit_rows.append({
            "row_id":              row.get("row_id"),
            "title":               row.get("business_name"),
            "action_type":         row.get("action_type"),
            "detailed_code":       code,
            "domain":              row.get("domain"),
            "csv_known_code":      csv_known,
            "current_manual_ref":  row.get("manual_ref") or [],
            "proposed_manual_ref": proposed_ref,
            "comparison":          comp,
            "match":               match_summary,
            "needs_human_review":  nr,
            "review_reason":       reason,
            # v2 신규
            "routing":             pkg["routing"],
            "primary_candidate":   primary,
            "secondary_candidate": secondary,
            "routing_warning":     pkg["routing_warning"],
        })

        if (i + 1) % 50 == 0:
            print(f"  [{i+1}/{len(rows)}]")

    # ── 통계 ──
    conf_dist     = Counter(a["match"]["confidence"]   for a in audit_rows)
    method_dist   = Counter(a["match"]["method"]       for a in audit_rows)
    status_dist   = Counter(a["comparison"]["status"]  for a in audit_rows)
    pref_dist     = Counter(a["routing"]["preferred_manual"] for a in audit_rows)
    routing_warns = sum(1 for a in audit_rows if a["routing_warning"])

    n_total = len(audit_rows)
    n_exact_high = conf_dist.get("exact", 0) + conf_dist.get("high", 0)
    n_medium = conf_dist.get("medium", 0)
    n_low_none = conf_dist.get("low", 0) + conf_dist.get("none", 0)
    n_changed = sum(1 for a in audit_rows if a["comparison"]["status"] == "changed")
    n_conflict = sum(1 for a in audit_rows if a["comparison"]["status"] == "conflict")
    n_missing_cur = sum(1 for a in audit_rows if a["comparison"]["status"] == "missing_current")
    n_missing_cand = sum(1 for a in audit_rows if a["comparison"]["status"] == "missing_candidate")
    n_review = sum(1 for a in audit_rows if a["needs_human_review"])

    # E-9 / F-1 conflict 비교 (v1 audit JSON 읽음)
    def count_conflict_for_prefix(rows_, prefix):
        return sum(1 for a in rows_ if a["comparison"]["status"] == "conflict"
                   and (a.get("detailed_code") or "").startswith(prefix))

    e9_after = count_conflict_for_prefix(audit_rows, "E-9")
    f1_after = count_conflict_for_prefix(audit_rows, "F-1")
    e9_before = f1_before = None
    if V1_AUDIT.exists():
        try:
            v1 = json.loads(V1_AUDIT.read_text(encoding="utf-8"))
            v1_rows = v1.get("rows", [])
            e9_before = count_conflict_for_prefix(v1_rows, "E-9")
            f1_before = count_conflict_for_prefix(v1_rows, "F-1")
        except Exception as e:
            print(f"  [v1 audit 읽기 실패] {e}")

    summary = {
        "total":               n_total,
        "preferred_manual":    dict(pref_dist),
        "exact_or_high":       n_exact_high,
        "medium":              n_medium,
        "low_or_none":         n_low_none,
        "current_changed":     n_changed,
        "conflict":            n_conflict,
        "missing_current":     n_missing_cur,
        "missing_candidate":   n_missing_cand,
        "needs_review":        n_review,
        "routing_warnings":    routing_warns,
        "confidence_dist":     dict(conf_dist),
        "method_dist":         dict(method_dist),
        "status_dist":         dict(status_dist),
        "e9_conflict_v1_v2":   {"v1": e9_before, "v2": e9_after},
        "f1_conflict_v1_v2":   {"v1": f1_before, "v2": f1_after},
    }

    out = {
        "audit_meta": {
            "version":   "v2",
            "db_path":   str(DB_PATH.relative_to(ROOT)),
            "manuals":   pdf_meta,
            "index_v6":  str(INDEX_V6.relative_to(ROOT)) if INDEX_V6.exists() else None,
            "structure": {k: bool(v) for k, v in struct_data.items()},
            "csv_codes": len(csv_codes),
            "v1_audit":  str(V1_AUDIT.relative_to(ROOT)) if V1_AUDIT.exists() else None,
        },
        "summary": summary,
        "rows":    audit_rows,
    }

    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX ──
    xlsx_written = False
    skip_reason = ""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "audit_v2"
        headers = [
            "row_id","title","action_type","detailed_code","domain",
            "preferred_manual","routing_reason","routing_warning",
            "status","page_changed","manual_changed",
            "method","confidence","needs_review","review_reason",
            "current_manual","current_pf","current_pt",
            "primary_manual","primary_pf","primary_pt","primary_conf",
            "secondary_manual","secondary_pf","secondary_pt","secondary_conf",
            "evidence",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="FFE599")
            cell.alignment = Alignment(horizontal="center")

        def first_ref(c):
            if not c or not c.get("manual_ref"): return ("", "", "")
            r = c["manual_ref"][0]
            return (r.get("manual",""), r.get("page_from",""), r.get("page_to",""))

        for a in audit_rows:
            cm, cpf, cpt = ("", "", "")
            if a["current_manual_ref"]:
                r = a["current_manual_ref"][0]
                cm, cpf, cpt = r.get("manual",""), r.get("page_from",""), r.get("page_to","")
            pm, ppf, ppt = first_ref(a["primary_candidate"])
            sm, spf, spt = first_ref(a["secondary_candidate"])
            p_conf = (a["primary_candidate"] or {}).get("confidence","")
            s_conf = (a["secondary_candidate"] or {}).get("confidence","")
            ws.append([
                a["row_id"], a["title"], a["action_type"], a["detailed_code"], a.get("domain",""),
                a["routing"]["preferred_manual"], a["routing"]["routing_reason"], a["routing_warning"],
                a["comparison"]["status"], a["comparison"]["page_changed"], a["comparison"]["manual_changed"],
                a["match"]["method"], a["match"]["confidence"], a["needs_human_review"], a["review_reason"],
                cm, cpf, cpt, pm, ppf, ppt, p_conf, sm, spf, spt, s_conf,
                a["match"]["evidence"],
            ])

        widths = {"A":10,"B":28,"C":14,"D":12,"E":10,"F":12,"G":36,"H":32,"I":18,
                  "J":8,"K":8,"L":22,"M":10,"N":8,"O":36,"P":10,"Q":8,"R":8,
                  "S":10,"T":8,"U":8,"V":8,"W":10,"X":8,"Y":8,"Z":8,"AA":50}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"

        ws2 = wb.create_sheet("summary")
        ws2.append(["metric","value"])
        for cell in ws2[1]:
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            if isinstance(v, dict):
                ws2.append([k, json.dumps(v, ensure_ascii=False)])
            else:
                ws2.append([k, v])
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 90

        wb.save(OUT_XLSX)
        xlsx_written = True
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        skip_reason = f"openpyxl import failed: {e}"
        print(f"[skip xlsx] {skip_reason}")
    except Exception as e:
        skip_reason = f"xlsx 생성 오류: {e}"
        print(f"[skip xlsx] {skip_reason}")

    # ── 콘솔 요약 ──
    print("\n" + "=" * 76)
    print("AUDIT SUMMARY v2")
    print("=" * 76)
    print(f"  total rows:                   {n_total}")
    print(f"  preferred_manual dist:        {dict(pref_dist)}")
    print(f"  exact + high:                 {n_exact_high}")
    print(f"  medium:                       {n_medium}")
    print(f"  low + none:                   {n_low_none}")
    print(f"  changed:                      {n_changed}")
    print(f"  conflict:                     {n_conflict}")
    print(f"  missing_current:              {n_missing_cur}")
    print(f"  missing_candidate:            {n_missing_cand}")
    print(f"  needs_review:                 {n_review}")
    print(f"  routing_warnings:             {routing_warns}")
    print(f"\n  confidence dist:  {dict(conf_dist)}")
    print(f"  status dist:      {dict(status_dist)}")
    print(f"  method dist:      {dict(method_dist)}")
    print(f"\n  E-9 conflict v1→v2: {e9_before} → {e9_after}")
    print(f"  F-1 conflict v1→v2: {f1_before} → {f1_after}")
    print(f"\n  Production data modified: NO (read-only audit)")

    print("\n=== Top 30 review needed ===")
    review_rows = [a for a in audit_rows if a["needs_human_review"]]
    priority = {"conflict":0,"missing_candidate":1,"missing_current":2,"missing_both":3,"changed":4,"same":9}
    review_rows.sort(key=lambda a: (
        priority.get(a["comparison"]["status"], 5),
        _CONF_RANK.get(a["match"]["confidence"], 0),
        a["row_id"] or "",
    ))
    for a in review_rows[:30]:
        title = (a["title"] or "")[:24]
        print(
            f"  {a['row_id']:8s} {a['detailed_code']:8s} {a['action_type']:13s} "
            f"pref={a['routing']['preferred_manual']:6s} conf={a['match']['confidence']:6s} "
            f"status={a['comparison']['status']:18s} {title:24s} :: {a['review_reason']}"
        )

    print("\n=== Files created (only) ===")
    print(f"  - {OUT_JSON.relative_to(ROOT)}")
    if xlsx_written:
        print(f"  - {OUT_XLSX.relative_to(ROOT)}")
    else:
        print(f"  - (xlsx skipped: {skip_reason})")

    print("\n=== Files NOT modified ===")
    print(f"  - {DB_PATH.relative_to(ROOT)}")
    print(f"  - {INDEX_V6.relative_to(ROOT)}")
    print(f"  - {V1_AUDIT.relative_to(ROOT)}  (v1 audit untouched)")
    print(f"  - manual PDFs / structure / CSV / indexer_v6 / watcher / frontend / **")


if __name__ == "__main__":
    main()
