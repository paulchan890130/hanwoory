"""
Patch v8 — 9 rows (WRONG judge + PDF verification + Step 1 confirmed).

DEFAULT: dry-run. `--apply` required for DB write.

Targets:
  M1-0086  F-4-R   CHANGE  → 체류 p.618-620 (지역특화 재외동포 자격변경 요건+서류)
  M1-0112  D-8-2   EXTEND  → 체류 p.122    (벤처투자 D-8-2 체류기간연장허가)
  M1-0113  D-8-3   EXTEND  → 체류 p.122    (개인기업투자 D-8-3 동일 섹션)
  M1-0364  F-5-1   CHANGE  → 체류 p.446    (국민 배우자·자녀 5년 합법체류 요건+서류)
  M1-0366  F-5-6   CHANGE  → 체류 p.447+p.449 (non-contig: 허가요건+제출서류)
  M1-0284  F-3-3R  CHANGE  → 체류 p.614  match_text 업데이트 (verifier ACCEPTED 명시)
  M1-0091  F-2-71  GRANT   → 체류 p.745    (K-STAR 동반가족 F-2-71 체류자격 부여)
  M1-0256  F-2-71  GRANT   → 체류 p.745    (동상)
  M1-0365  F-5-2   CHANGE  → 체류 p.453+p.454 (non-contig: 가.대상·나.요건 + 다.제출서류)
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

ROOT       = Path(__file__).parent.parent.parent
DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"
DRY_JSON   = ROOT / "backend" / "data" / "manuals" / "v8_patch_dryrun.json"
DRY_XLSX   = ROOT / "backend" / "data" / "manuals" / "v8_patch_dryrun.xlsx"

PATCHES: dict[str, dict] = {
    "M1-0086": {
        "detailed_code": "F-4-R",
        "action_type":   "CHANGE",
        "title":         "지역특화형 재외동포 체류자격 변경",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.618 '나. 요건' F-4-R 자격변경 신청요건; "
            "p.619 해외전입자 요건·학령기 자녀 요건; "
            "p.620 제출서류 (추천서·여권·범죄경력증명서·재학증명서 등)"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  618,
                "page_to":    620,
                "match_type": "manual_override",
                "match_text": "F-4-R 지역특화형 재외동포 CHANGE 자격변경허가 요건+제출서류 (체류민원 p.618-620) — 수동 검증",
            },
        ],
    },
    "M1-0112": {
        "detailed_code": "D-8-2",
        "action_type":   "EXTEND",
        "title":         "벤처투자",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.122 '2. 벤처 투자(D-8-2) 외국인에 체류기간 연장허가' — "
            "heading에 D-8-2 명시 + req=True + 제출서류 포함"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  122,
                "page_to":    122,
                "match_type": "manual_override",
                "match_text": "D-8-2 벤처투자 EXTEND 체류기간연장허가 (체류민원 p.122) — 수동 검증",
            },
        ],
    },
    "M1-0113": {
        "detailed_code": "D-8-3",
        "action_type":   "EXTEND",
        "title":         "개인기업투자",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.122 '3. 개인기업 투자(D-8-3)에 대한 체류기간 연장허가' — "
            "D-8-3 heading 직접 확인 + D-8-2 섹션과 동일 페이지 공존"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  122,
                "page_to":    122,
                "match_type": "manual_override",
                "match_text": "D-8-3 개인기업투자 EXTEND 체류기간연장허가 (체류민원 p.122) — 수동 검증",
            },
        ],
    },
    "M1-0364": {
        "detailed_code": "F-5-1",
        "action_type":   "CHANGE",
        "title":         "영주 변경 — 국민의 배우자·자녀 (5년 이상 합법체류)",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.446 F-5-1=True, req=True, doc=True — "
            "'에서 5년 이상 계속 체류 / 완전출국 없이 상기 대상 체류자격으로만 연속 체류' "
            "나. 요건 세부 + 제출서류 포함. LLM judge recommended_page=446."
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  446,
                "page_to":    446,
                "match_type": "manual_override",
                "match_text": "F-5-1 영주 CHANGE 신청요건+제출서류 (체류민원 p.446) — LLM judge + 수동 검증",
            },
        ],
    },
    "M1-0366": {
        "detailed_code": "F-5-6",
        "action_type":   "CHANGE",
        "title":         "결혼이민자영주",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.447 '1) 허가 요건 / 가) 한국인 배우자와 정상적인 혼인생활 유지 / "
            "나) 결혼이민(F-6) 자격으로 2년 이상 국내 계속 체류'; "
            "p.449 '2) 제출 서류 / 공통 서류' — 허가요건+제출서류 non-contiguous"
        ),
        "non_contiguous": True,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  447,
                "page_to":    447,
                "match_type": "manual_override",
                "match_text": "F-5-6 결혼이민 CHANGE 허가요건 (체류민원 p.447) — 수동 검증",
            },
            {
                "manual":     "체류민원",
                "page_from":  449,
                "page_to":    449,
                "match_type": "manual_override",
                "match_text": "F-5-6 결혼이민 CHANGE 제출서류 (체류민원 p.449) — 수동 검증",
            },
        ],
    },
    "M1-0284": {
        "detailed_code": "F-3-3R",
        "action_type":   "CHANGE",
        "title":         "지역특화숙련기능인력가족",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.614 '동반가족 > 1. 체류자격변경' — 대상·신청요건·제출서류 존재. "
            "LLM judge EXACT + verifier ACCEPTED. page_from/page_to 불변, match_text 업데이트."
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  614,
                "page_to":    614,
                "match_type": "manual_override",
                "match_text": "F-3-3R 지역특화숙련기능인력 CHANGE (체류민원 p.614) — LLM judge + verifier 검증",
            },
        ],
    },
    "M1-0091": {
        "detailed_code": "F-2-71",
        "action_type":   "GRANT",
        "title":         "K-STAR 거주자의 국내출생 미성년자녀",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.745 'K-STAR의 동반가족(F-2-71, F-5-S2)' 섹션 / "
            "'체류자격 부여' heading + F-2-71 명시 + doc=True (제출서류 포함)"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  745,
                "page_to":    745,
                "match_type": "manual_override",
                "match_text": "F-2-71 K-STAR 동반가족 GRANT 체류자격 부여 (체류민원 p.745) — 수동 검증",
            },
        ],
    },
    "M1-0256": {
        "detailed_code": "F-2-71",
        "action_type":   "GRANT",
        "title":         "점수제 우수인재 국내출생 자녀 거주",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.745 'K-STAR의 동반가족(F-2-71, F-5-S2)' 섹션 / "
            "'체류자격 부여' heading + F-2-71 명시 + doc=True (제출서류 포함)"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  745,
                "page_to":    745,
                "match_type": "manual_override",
                "match_text": "F-2-71 점수제 우수인재 국내출생 자녀 GRANT 체류자격 부여 (체류민원 p.745) — 수동 검증",
            },
        ],
    },
    "M1-0365": {
        "detailed_code": "F-5-2",
        "action_type":   "CHANGE",
        "title":         "영주 변경 — 영주권자의 미성년 자녀",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.453 '3. 일반 영주자의 배우자 또는 미성년 자녀〔F-5-4〕 / 가. 대상 / 나. 요건' — "
            "DB 코드 F-5-2 = 매뉴얼 〔F-5-4〕; "
            "p.454 '다. 제출서류 / ○가족관계 입증 서류, 출생증명서 등' — "
            "req+doc non-contiguous"
        ),
        "non_contiguous": True,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  453,
                "page_to":    453,
                "match_type": "manual_override",
                "match_text": "F-5-2 영주권자 미성년자녀 CHANGE 가.대상·나.요건 (체류민원 p.453) — 수동 검증",
            },
            {
                "manual":     "체류민원",
                "page_from":  454,
                "page_to":    454,
                "match_type": "manual_override",
                "match_text": "F-5-2 영주권자 미성년자녀 CHANGE 다.제출서류 (체류민원 p.454) — 수동 검증",
            },
        ],
    },
}


# ── Safety primitives (identical to v7) ──────────────────────────
def _abort(msg: str) -> None:
    print(f"\n[ABORT] {msg}", file=sys.stderr)
    sys.exit(1)


def make_backup(db_bytes: bytes) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"immigration_guidelines_db_v2.v8_patch_backup_{ts}.json"
    bk.write_bytes(db_bytes)
    if not bk.exists() or bk.stat().st_size != len(db_bytes):
        _abort(f"백업 작성 검증 실패: {bk}")
    try:
        json.loads(bk.read_text(encoding="utf-8"))
    except Exception as e:
        _abort(f"백업 JSON readback 실패: {e}")
    return bk


def diff_rows(before: list, after: list) -> list:
    bi = {r.get("row_id"): r for r in before}
    ai = {r.get("row_id"): r for r in after}
    out = []
    for rid in set(bi.keys()) | set(ai.keys()):
        b = bi.get(rid); a = ai.get(rid)
        if b is None or a is None:
            out.append({"row_id": rid, "added_or_removed": True}); continue
        diff_fields = [k for k in set(b.keys()) | set(a.keys()) if b.get(k) != a.get(k)]
        if diff_fields:
            out.append({"row_id": rid, "fields": sorted(diff_fields)})
    return out


def load_db() -> tuple:
    raw = DB_PATH.read_bytes()
    return json.loads(raw.decode("utf-8")), raw


def build_patched_db(db: dict) -> tuple:
    new_db = json.loads(json.dumps(db, ensure_ascii=False))
    rows = new_db.get("master_rows") or []
    found = {}
    for r in rows:
        if r.get("row_id") in PATCHES:
            found[r["row_id"]] = r
    missing = [rid for rid in PATCHES if rid not in found]
    if missing:
        _abort(f"target row_ids 누락: {missing}")
    patch_report = []
    for rid, spec in PATCHES.items():
        row = found[rid]
        if row.get("detailed_code") != spec["detailed_code"]:
            _abort(f"{rid} detailed_code 불일치: {row.get('detailed_code')} != {spec['detailed_code']}")
        if row.get("action_type") != spec["action_type"]:
            _abort(f"{rid} action_type 불일치: {row.get('action_type')} != {spec['action_type']}")
        old_ref = row.get("manual_ref") or []
        row["manual_ref"] = spec["new_manual_ref"]
        patch_report.append({
            "row_id":           rid,
            "detailed_code":    spec["detailed_code"],
            "action_type":      spec["action_type"],
            "title":            spec["title"],
            "preferred_manual": spec["preferred_manual"],
            "old_manual_ref":   old_ref,
            "new_manual_ref":   spec["new_manual_ref"],
            "page_evidence":    spec["page_evidence"],
            "non_contiguous":   spec["non_contiguous"],
            "page_count":       len(spec["new_manual_ref"]),
        })
    return new_db, patch_report


def verify_patch_safety(before_db: dict, after_db: dict) -> list:
    errors = []
    bm = before_db.get("master_rows") or []; am = after_db.get("master_rows") or []
    if len(bm) != 369: errors.append(f"백업 master_rows 개수가 369 아님: {len(bm)}")
    if len(am) != len(bm): errors.append(f"패치 후 개수 불일치: {len(am)} != {len(bm)}")
    bi = {r.get("row_id"): r for r in bm}; ai = {r.get("row_id"): r for r in am}
    if set(bi.keys()) != set(ai.keys()): errors.append("row_id 집합 변경됨")
    for rid in bi:
        b = bi[rid]; a = ai.get(rid, {})
        check_keys = set(b.keys()) | set(a.keys())
        if rid in PATCHES:
            for k in check_keys:
                if k == "manual_ref": continue
                if b.get(k) != a.get(k):
                    errors.append(f"{rid} non-manual_ref 필드 변경됨: {k}")
        else:
            for k in check_keys:
                if b.get(k) != a.get(k):
                    errors.append(f"비-target row {rid} 필드 변경됨: {k}")
    for k in set(before_db.keys()) | set(after_db.keys()):
        if k == "master_rows": continue
        if before_db.get(k) != after_db.get(k):
            errors.append(f"top-level 키 변경됨: {k}")
    return errors


def write_dryrun_reports(patch_report: list, summary: dict) -> None:
    DRY_JSON.write_text(json.dumps({
        "patch_meta": {"version": "v8", "mode": "DRY_RUN",
                       "scope": "9 rows: F-4-R/D-8-2/D-8-3/F-5-1/F-5-6/F-3-3R/F-2-71×2/F-5-2"},
        "summary": summary, "rows": patch_report,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {DRY_JSON.relative_to(ROOT)} ({DRY_JSON.stat().st_size:,} bytes)")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("patches")
        headers = ["row_id","code","action","title","preferred","pages","non_contig",
                   "old_ref","new_ref","evidence"]
        widths = [10,12,14,36,14,6,12,80,100,80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for x in patch_report:
            ws.append([x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                       x["preferred_manual"], x["page_count"], x["non_contiguous"],
                       json.dumps(x["old_manual_ref"], ensure_ascii=False)[:600],
                       json.dumps(x["new_manual_ref"], ensure_ascii=False)[:800],
                       x["page_evidence"][:500]])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCB")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(DRY_XLSX)
        print(f"[OUT] {DRY_XLSX.relative_to(ROOT)} ({DRY_XLSX.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[skip xlsx] {e}")


def main() -> int:
    try: sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    is_apply = args.apply
    if not DB_PATH.exists(): _abort(f"DB 없음: {DB_PATH}")
    print(f"[patch] mode = {'APPLY' if is_apply else 'DRY-RUN'}")
    db, raw = load_db()
    bytes_before = len(raw)
    print(f"[patch] DB byte size before: {bytes_before:,}  master_rows: {len(db.get('master_rows') or [])}")
    if len(db.get("master_rows") or []) != 369: _abort("master_rows 개수 비정상")
    print(f"[patch] target rows ({len(PATCHES)}): {sorted(PATCHES.keys())}")
    new_db, patch_report = build_patched_db(db)
    errors = verify_patch_safety(db, new_db)
    if errors:
        for e in errors: print(f"  [error] {e}")
        _abort("safety check 실패")
    diffs = diff_rows(db.get("master_rows") or [], new_db.get("master_rows") or [])
    diff_ids = {d["row_id"] for d in diffs}
    if diff_ids != set(PATCHES.keys()):
        _abort(f"diff 불일치: diff={sorted(diff_ids)} expected={sorted(PATCHES.keys())}")
    for d in diffs:
        if d.get("fields") != ["manual_ref"]:
            _abort(f"{d['row_id']} manual_ref 외 필드 변경: {d.get('fields')}")
    summary = {
        "mode": "APPLY" if is_apply else "DRY_RUN",
        "production_db_modified": False,
        "patch_candidate_count": len(patch_report),
        "diff_row_ids": sorted(diff_ids),
        "diff_count": len(diff_ids),
        "non_target_row_diffs": 0,
        "non_manual_ref_field_diffs": 0,
        "non_contiguous_rows": [x["row_id"] for x in patch_report if x["non_contiguous"]],
        "db_byte_size_before": bytes_before,
    }
    write_dryrun_reports(patch_report, summary)
    if not is_apply:
        print("\n" + "="*78)
        print(f"PATCH v8 — DRY-RUN  ({len(PATCHES)} rows)")
        print("="*78)
        print(f"  target row_ids: {sorted(diff_ids)}")
        print(f"  non-target row diffs: 0  /  non-manual_ref field diffs: 0")
        print(f"  non-contiguous rows: {summary['non_contiguous_rows']}")
        print(f"  DB byte size: {bytes_before:,} → {bytes_before:,} (unchanged)")
        print()
        print("  Old → New (compact):")
        for x in patch_report:
            print(f"    {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}  pages={x['page_count']}  non_contig={x['non_contiguous']}")
            print(f"      OLD: " + json.dumps(x["old_manual_ref"], ensure_ascii=False)[:160])
            print(f"      NEW: " + json.dumps(x["new_manual_ref"], ensure_ascii=False)[:200])
        print()
        print("  Apply command (NOT EXECUTED):")
        print("    python backend/scripts/patch_v8.py --apply")
        return 0
    # APPLY
    print("\n[apply] 백업 생성...")
    bk = make_backup(raw)
    print(f"  backup: {bk.relative_to(ROOT)}")
    new_json = json.dumps(new_db, ensure_ascii=False, indent=2)
    DB_PATH.write_text(new_json, encoding="utf-8")
    readback = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rb_rows = readback.get("master_rows") or []
    if len(rb_rows) != 369:
        _abort(f"readback master_rows 개수 비정상: {len(rb_rows)}\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    rb_idx = {r.get("row_id"): r for r in rb_rows}
    for rid, spec in PATCHES.items():
        if rb_idx.get(rid, {}).get("manual_ref") != spec["new_manual_ref"]:
            _abort(f"{rid} readback 불일치\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    errs2 = verify_patch_safety(json.loads(bk.read_text(encoding="utf-8")), readback)
    if errs2:
        for e in errs2: print(f"  [error] {e}")
        _abort(f"사후 verify 실패\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    bytes_after = DB_PATH.stat().st_size
    print(f"\n{'='*78}\nPATCH v8 — APPLIED\n{'='*78}")
    print(f"  backup: {bk.relative_to(ROOT)}")
    print(f"  DB updated: {len(patch_report)} rows")
    print(f"  byte size before: {bytes_before:,}  after: {bytes_after:,}")
    print(f"  rollback: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
