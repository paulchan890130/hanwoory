"""체류자격 canonical(V=v_status) 계약 — **실제 PostgreSQL** 통합 테스트.

``TEST_DATABASE_URL`` 설정 시에만 실행(미설정 → skip). 검증:
- 엑셀 일괄등록 F4 → customers.v_status='F-4' 저장(visa_status 로 분리되지 않음), 고객카드 GET V=F-4.
- 레거시로 visa_status 에만 값이 있는 행 → 읽기(V)·재추출 모두 그 값으로 통일(DB 원문 불변).
- 신규(V=F-4) → 추출 → 재파싱 시 F-4 유지(round-trip).

실제 고객 원문은 쓰지 않고 합성 값만 사용한다.
"""
from __future__ import annotations

import io
import os

import pytest
from cryptography.fernet import Fernet
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import make_url
from sqlalchemy.orm import Session, sessionmaker

os.environ.setdefault("CUSTOMER_PII_ENCRYPTION_KEY", Fernet.generate_key().decode())
os.environ.setdefault("PII_HASH_SECRET", "visa-it-secret")

TEST_DB_URL = os.environ.get("TEST_DATABASE_URL", "").strip()
pytestmark = pytest.mark.skipif(
    not TEST_DB_URL, reason="TEST_DATABASE_URL 미설정 — 실제 PostgreSQL 통합 테스트 skip")

_TID = "t-visa"


def _ensure_database(url_str: str) -> None:
    url = make_url(url_str)
    admin_url = url.set(database="postgres")
    eng = create_engine(admin_url, isolation_level="AUTOCOMMIT", future=True)
    try:
        with eng.connect() as c:
            exists = c.execute(
                text("SELECT 1 FROM pg_database WHERE datname = :n"), {"n": url.database}
            ).first()
            if not exists:
                c.execute(text(f'CREATE DATABASE "{url.database}"'))
    finally:
        eng.dispose()


@pytest.fixture(scope="module")
def engine():
    _ensure_database(TEST_DB_URL)
    eng = create_engine(TEST_DB_URL, future=True)
    from backend.db.base import Base
    from backend.db.models.customer import Customer
    from backend.db.models.tenant import Tenant
    tables = [Tenant.__table__, Customer.__table__]
    Base.metadata.drop_all(eng, tables=list(reversed(tables)))
    Base.metadata.create_all(eng, tables=tables)
    yield eng
    Base.metadata.drop_all(eng, tables=list(reversed(tables)))
    eng.dispose()


@pytest.fixture
def db(engine, monkeypatch):
    SessionLocal = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False, class_=Session)
    import backend.db.session as dbs
    monkeypatch.setattr(dbs, "get_sessionmaker", lambda: SessionLocal)
    monkeypatch.setattr(dbs, "is_configured", lambda: True)
    with engine.begin() as c:
        c.execute(text("TRUNCATE customers, tenants RESTART IDENTITY CASCADE"))
    from backend.db.models.tenant import Tenant
    with SessionLocal() as s:
        s.add(Tenant(tenant_id=_TID, office_name="T", is_active=True))
        s.commit()
    return SessionLocal


def test_bulk_import_visa_saves_to_v_status(db):
    """엑셀 일괄등록 F4 → v_status=F-4 저장(visa_status 아님), 고객카드 GET V=F-4."""
    from backend.services import customer_bulk_service as bulk
    from backend.services.customer_pg_service import find_customer

    blob, _ = bulk.build_export_workbook_bytes([{"한글": "비자합성", "V": "F4", "등록증": "001010"}])
    res = bulk.commit(blob, _TID, include_duplicates=True)
    assert res["registered"] == 1, res
    cid = res["new_customer_ids"][0]
    assert find_customer(_TID, cid)["V"] == "F-4"          # 고객카드/quick-doc 정본
    with db() as s:
        row = s.execute(text("SELECT v_status, visa_status FROM customers WHERE customer_id=:c")
                        .bindparams(c=cid)).first()
    assert (row[0] or "") == "F-4"                          # v_status 저장
    assert (row[1] or "") == ""                             # visa_status 로 분리되지 않음


def test_legacy_visa_status_read_fallback_and_export(db):
    """레거시로 visa_status 에만 값이 있는 행 → 읽기(V)·추출 모두 그 값으로 통일(원문 불변)."""
    from backend.services import customer_bulk_service as bulk
    from backend.services.customer_pg_service import find_customer

    with db() as s:
        s.execute(text(
            "INSERT INTO customers (tenant_id, customer_id, korean_name, visa_status) "
            "VALUES (:t, :c, :n, :v)"
        ).bindparams(t=_TID, c="9500", n="레거시비자", v="F-5"))
        s.commit()
    got = find_customer(_TID, "9500")
    assert got["V"] == "F-5"                                # V fallback(고객카드/검색/quick-doc 통일)
    blob, _n = bulk.build_export_workbook_bytes([got])
    ws = load_workbook(io.BytesIO(blob))[bulk.SHEET_NAME]
    v_idx = bulk.STD_KEYS.index("V") + 1
    assert ws.cell(row=bulk.DATA_START_ROW, column=v_idx).value == "F-5"   # 재추출도 F-5
    with db() as s:                                          # DB 원문 비파괴
        row = s.execute(text("SELECT v_status, visa_status FROM customers WHERE customer_id='9500'")).first()
    assert (row[0] or "") == "" and (row[1] or "") == "F-5"


def test_reexport_preserves_visa(db):
    """신규(V=F-4) 등록 → 추출 → 다시 파싱 시 F-4 유지."""
    from backend.services import customer_bulk_service as bulk
    from backend.services.customer_pg_service import create_customer, list_customers

    create_customer(_TID, {"한글": "재추출", "등록증": "001010", "V": "F-4"})
    records = list_customers(_TID, reveal=True)
    blob, _n = bulk.build_export_workbook_bytes(records)
    rows = bulk._read_rows(blob)
    data, msgs, _t, _w = bulk._row_to_customer(rows[0])
    assert not msgs and data["V"] == "F-4"
