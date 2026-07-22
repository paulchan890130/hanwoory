"""고객 고정폭 식별값(외국인등록번호 앞자리 reg_front) 선행 0 손실 방어 단위테스트.

DB 불필요 — 정규화 함수, ``_row_to_dict`` 읽기 방어선(SimpleNamespace), 쓰기 payload
정규화, 엑셀 일괄 import 행 변환, 엑셀 추출 라운드트립, 연결(숙소/보증인) 읽기·쓰기
방어를 검증한다. 실제 고객 원문은 사용하지 않고 합성 값(001010/900101 등)만 쓴다.

실행: pytest backend/tests/test_customer_identifier_normalize.py
"""
import io
import os

import pytest
from cryptography.fernet import Fernet

# _row_to_dict 가 import 하는 pii_crypto 용 키/시크릿(모듈 import 전에).
os.environ.setdefault("CUSTOMER_PII_ENCRYPTION_KEY", Fernet.generate_key().decode())
os.environ.setdefault("PII_HASH_SECRET", "unit-test-secret")

from backend.services.customer_identifier_normalize import (  # noqa: E402
    RegFrontValidationError,
    canonical_reg_front,
    canonical_reg_front_for_legacy_read,
    century_prefix_from_reg_back,
    derive_birth_date,
    normalize_reg_front,
    normalize_reg_front_from_excel,
    validate_reg_front_for_write,
)


# ── 1) 정규화 함수 ─────────────────────────────────────────────────────────────
@pytest.mark.parametrize("raw,expected", [
    ("1010", "001010"),      # 선행 0 손실(문자열) → 복구
    ("001010", "001010"),    # 이미 정상
    (1010, "001010"),        # int 강제변환 손실 → 복구
    (1010.0, "001010"),      # float(.0) → 복구
    ("900101", "900101"),    # 정상 6자리
    ("", ""),                # 빈 값
    ("  ", ""),              # 공백만
])
def test_canonical_recovers_leading_zeros(raw, expected):
    assert canonical_reg_front(raw) == expected


@pytest.mark.parametrize("raw", [
    "1234567",   # 7자리(too_long) — 절단·추측 금지
    "abc",       # 영숫자 오염
    "99999",     # 5자리지만 zfill→099999(mm=99) 무효
    "1010.5",    # 소수부 0 아님
    "-1010",     # 음수
])
def test_canonical_non_destructive_on_invalid(raw):
    # 유효 복구 불가 → 원문(strip)을 그대로 반환(파괴적 변경 없음).
    assert canonical_reg_front(raw) == str(raw).strip()


def test_normalize_reason_flags():
    assert normalize_reg_front("1010").reason == "recovered"
    assert normalize_reg_front("900101").reason == "ok"
    assert normalize_reg_front("1234567").valid is False
    # 엄격 모드: 정확히 6자리만 허용.
    assert normalize_reg_front("1010", allow_numeric_recovery=False).valid is False
    assert normalize_reg_front("001010", allow_numeric_recovery=False).valid is True


@pytest.mark.parametrize("d6,valid", [
    ("000229", True),    # 세기미상 → Feb29 허용
    ("000230", False),   # Feb30 무효
    ("994566", False),   # mm=45 무효
    ("001301", False),   # mm=13 무효
    ("001000", False),   # dd=00 무효
])
def test_yymmdd_validation(d6, valid):
    assert normalize_reg_front(d6).valid is valid


# ── 2) 세기 판정(등록번호 뒷자리 첫 숫자) ─────────────────────────────────────
@pytest.mark.parametrize("back,prefix", [
    ("1020304", "19"), ("2020304", "19"), ("5020304", "19"), ("6020304", "19"),
    ("3020304", "20"), ("4020304", "20"), ("7020304", "20"), ("8020304", "20"),
])
def test_century_prefix(back, prefix):
    assert century_prefix_from_reg_back(back) == prefix


def test_derive_birth_date_uses_back_first_digit_not_yy():
    # 앞자리 00(=2000/1900 모호) 이지만 뒷자리로 세기 확정.
    assert derive_birth_date("1010", "3020304") == "2000-10-10"
    assert derive_birth_date("1010", "1020304") == "1900-10-10"
    assert derive_birth_date("bad", "1020304") == ""


@pytest.mark.parametrize("front,back,expected", [
    ("001010", "7020304", "2000-10-10"),  # 정상 2000
    ("001010", "1020304", "1900-10-10"),  # 정상 1900
    ("000229", "3020304", "2000-02-29"),  # 2000 윤년 → 존재
    ("000229", "1020304", ""),            # 1900 평년 → 2/29 없음 → 빈 값
    ("990229", "1020304", ""),            # 1999 평년 → 2/29 없음
    ("040229", "3020304", "2004-02-29"),  # 2004 윤년 → 존재
])
def test_derive_birth_date_real_gregorian(front, back, expected):
    assert derive_birth_date(front, back) == expected


# ── 1b) 경계별 정책: 레거시 읽기 / 웹 쓰기 / Excel ──────────────────────────────
@pytest.mark.parametrize("raw,expected", [
    ("1010", "001010"), ("101", "000101"), ("10101", "010101"),
    ("001010", "001010"), ("900101", "900101"),
])
def test_legacy_read_recovers(raw, expected):
    assert canonical_reg_front_for_legacy_read(raw) == expected


@pytest.mark.parametrize("raw", ["12345678", "abc", "1010.5", "-1010"])
def test_legacy_read_non_destructive_on_invalid(raw):
    assert canonical_reg_front_for_legacy_read(raw) == str(raw).strip()


@pytest.mark.parametrize("raw,expected", [("001010", "001010"), ("", ""), ("000229", "000229")])
def test_write_validation_accepts_valid_six(raw, expected):
    assert validate_reg_front_for_write(raw) == expected


@pytest.mark.parametrize("raw", [
    "1010",       # 4자리 — 조용히 001010 으로 바꾸지 않음
    "12345678",   # 8자리
    "abcdef",     # 문자
    "1010.5",     # 소수
    "-1010",      # 음수
    "991332",     # 월·일 무효
])
def test_write_validation_rejects_invalid(raw):
    with pytest.raises(RegFrontValidationError) as ei:
        validate_reg_front_for_write(raw)
    assert ei.value.code == "INVALID_REG_FRONT"


def test_excel_normalizer_recovers_and_flags_invalid():
    assert normalize_reg_front_from_excel(1010).canonical_value == "001010"
    assert normalize_reg_front_from_excel(1010.0).canonical_value == "001010"
    assert normalize_reg_front_from_excel("001010").valid is True
    assert normalize_reg_front_from_excel("12345678").valid is False
    assert normalize_reg_front_from_excel("991332").valid is False


# ── 3) _row_to_dict 읽기 방어선(레거시 '1010' → '001010') ─────────────────────
def test_row_to_dict_read_defense_recovers_reg_front():
    from types import SimpleNamespace

    from backend.services.customer_pg_service import _row_to_dict

    row = SimpleNamespace(
        customer_id="0001", tenant_id="t1", korean_name="김테스트",
        reg_front="1010",  # 레거시 선행 0 손실 원문
        reg_back="", reg_back_encrypted=None, reg_back_last4=None,
    )
    out = _row_to_dict(row)
    assert out["등록증"] == "001010"  # 화면·문서에는 복구된 6자리


def test_row_to_dict_keeps_valid_reg_front():
    from types import SimpleNamespace

    from backend.services.customer_pg_service import _row_to_dict

    row = SimpleNamespace(
        customer_id="0002", tenant_id="t1", korean_name="이테스트",
        reg_front="900101", reg_back="", reg_back_encrypted=None, reg_back_last4=None,
    )
    assert _row_to_dict(row)["등록증"] == "900101"


# ── 4) 쓰기 payload 정책: create=엄격 / upsert=grandfather ────────────────────
def test_create_payload_strict_rejects_recoverable():
    # 신규(create) payload 는 '1010' 을 조용히 복구하지 않고 예외.
    from backend.services.customer_pg_service import _validate_reg_front_in_payload

    p = {"reg_front": "001010", "korean_name": "x"}
    _validate_reg_front_in_payload(p)
    assert p["reg_front"] == "001010"

    p_bad = {"reg_front": "1010"}
    with pytest.raises(RegFrontValidationError):
        _validate_reg_front_in_payload(p_bad)

    p2 = {"korean_name": "x"}  # reg_front 없음 → 무변경
    _validate_reg_front_in_payload(p2)
    assert "reg_front" not in p2


def test_upsert_payload_grandfathers_legacy():
    # 수정/복원(upsert) payload 는 화면 canonical/레거시 값을 비파괴 복구로 통과.
    from backend.services.customer_pg_service import (
        _legacy_canonicalize_reg_front_in_payload,
    )

    p = {"reg_front": "1010"}
    _legacy_canonicalize_reg_front_in_payload(p)
    assert p["reg_front"] == "001010"

    p_bad = {"reg_front": "12345678"}  # 복구 불가 → 원문 보존(예외 없음)
    _legacy_canonicalize_reg_front_in_payload(p_bad)
    assert p_bad["reg_front"] == "12345678"


# ── 5) 엑셀 일괄 import 행 변환 ───────────────────────────────────────────────
def test_bulk_import_numeric_reg_front_recovered():
    from backend.services.customer_bulk_service import _row_to_customer

    raw = {"한글": "박테스트", "등록증": "1010", "_eng_name": "", "_phone": ""}
    data, errs, transforms, warnings = _row_to_customer(raw)
    assert data["등록증"] == "001010"
    assert any("선행 0" in t for t in transforms)


def test_bulk_import_string_reg_front_preserved():
    from backend.services.customer_bulk_service import _row_to_customer

    raw = {"한글": "박테스트", "등록증": "001010", "_eng_name": "", "_phone": ""}
    data, errs, transforms, warnings = _row_to_customer(raw)
    assert data["등록증"] == "001010"


@pytest.mark.parametrize("bad", ["12345678", "991332", "abcdef"])
def test_bulk_import_invalid_reg_front_is_error_not_warning(bad):
    from backend.services.customer_bulk_service import _row_to_customer

    raw = {"한글": "박테스트", "등록증": bad, "_eng_name": "", "_phone": ""}
    data, msgs, transforms, warnings = _row_to_customer(raw)
    # 경고가 아니라 오류(msgs) → validate/commit 이 행을 차단한다.
    assert any("6자리" in m for m in msgs)
    assert data["등록증"] == ""  # 잘못된 원문을 저장하지 않음


def test_bulk_validate_marks_error_row_and_commit_skips(monkeypatch):
    from backend.services import customer_bulk_service as bulk

    good = {"한글": "정상", "등록증": "001010", "_eng_name": "", "_phone": "", "_row_no": 4}
    bad = {"한글": "불량", "등록증": "12345678", "_eng_name": "", "_phone": "", "_row_no": 5}
    monkeypatch.setattr(bulk, "_read_rows", lambda *_a, **_k: [good, bad])
    monkeypatch.setattr(bulk, "_existing_index", lambda *_a, **_k: ({}, {}))
    res = bulk.validate(b"", "t1")
    statuses = {r["row_no"]: r["status"] for r in res["rows"]}
    assert statuses[5] == "error"          # 잘못된 앞자리 행은 오류
    assert res["counts"]["error"] >= 1


# ── 6) 엑셀 추출 라운드트립(Text(@) + 값 보존) ────────────────────────────────
def test_export_reg_front_text_format_and_value():
    from openpyxl import load_workbook

    from backend.services.customer_excel_service import (
        _EXPORT_COLUMNS,
        build_export_bytes_from_records,
    )

    reg_col_idx = next(i for i, (_h, k, _w) in enumerate(_EXPORT_COLUMNS, start=1) if k == "등록증")
    records = [{"고객ID": "0001", "한글": "최테스트", "등록증": "001010"}]
    blob, count = build_export_bytes_from_records(records)
    assert count == 1
    wb = load_workbook(io.BytesIO(blob))
    ws = wb["고객목록"]
    cell = ws.cell(row=2, column=reg_col_idx)
    assert cell.value == "001010"           # 선행 0 보존
    assert cell.number_format == "@"        # 텍스트 서식(엑셀 재해석 방지)


# ── 7) 연결(숙소/보증인) 읽기 방어 + 쓰기 정규화 ─────────────────────────────
def test_relationship_row_read_defense():
    from types import SimpleNamespace

    from backend.services.relationship_pg_service import _ACCOM_FIELDS, _row_to_dict

    row = SimpleNamespace(created_at=None, updated_at=None,
                          **{f: "" for f in _ACCOM_FIELDS})
    row.provider_reg_front = "1010"
    out = _row_to_dict(row, _ACCOM_FIELDS)
    assert out["provider_reg_front"] == "001010"


def test_relationship_write_normalization_in_place():
    from backend.services.relationship_pg_service import _canonicalize_reg_front_fields

    payload = {"provider_reg_front": "1010", "guarantor_reg_front": "1225"}
    _canonicalize_reg_front_fields(payload)
    assert payload["provider_reg_front"] == "001010"
    assert payload["guarantor_reg_front"] == "001225"


# ── 8) OCR 등록증 필드 계약(실제 Tesseract 미의존 — _digits_ocr monkeypatch) ────
def test_ocr_reg_front_contract(monkeypatch):
    from PIL import Image

    from backend.services import roi_ocr_service as roi

    img = Image.new("RGB", (240, 140), "white")
    roi_box = {"x": 0.30, "y": 0.15, "w": 0.12, "h": 0.04}

    def run(inject):
        monkeypatch.setattr(roi, "_digits_ocr", lambda *a, **k: inject)
        return roi.extract_arc_field(img, "등록증", roi_box)

    v, dbg = run("000101")   # OCR 이 읽은 선행 0 은 절대 유실 안 됨
    assert v == "000101" and dbg["failure_reason"] == ""
    v, dbg = run("001010")
    assert v == "001010"
    v, dbg = run("991332")   # 무효 날짜 → 자동저장 가능한 값으로 반환하지 않음 + 사유
    assert v == "" and dbg["failure_reason"]
    v, dbg = run("")         # 숫자 없음
    assert v == "" and dbg["failure_reason"]


# ── 9) QuickDoc/HWPX/PDF payload(공통 build_field_values) — 선행 0 반영 ─────────
def test_quick_doc_build_field_values_reg_front():
    from backend.routers.quick_doc import build_field_values

    # read-defense 를 거친 값(canonical) 을 그대로 받았다고 가정.
    row = {"성": "KIM", "명": "TEST", "등록증": "001010", "번호": "7020304"}
    fv = build_field_values(row)
    assert fv["yyyy"] == "2000"          # 뒷자리 7 → 2000
    assert fv["mm"] == "10" and fv["dd"] == "10"
    assert fv["fnumber"] == "001010"     # 6칸 앞자리 선행 0 유지 (PDF·HWPX 공통)


def test_quick_doc_guarantor_reg_front():
    from backend.routers.quick_doc import build_field_values

    row = {"성": "A", "명": "B", "등록증": "900101", "번호": "1020304"}
    g = {"등록증": "001010", "번호": "7020304"}
    fv = build_field_values(row, guarantor=g)
    assert fv["byyyy"] == "2000"
    assert fv["bfnumber"] == "001010"    # 보증인 앞자리 선행 0 유지
