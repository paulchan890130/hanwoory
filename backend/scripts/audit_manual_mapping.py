"""
manual_ref 감사 스크립트 (READ-ONLY, DB 변경 없음)

목적:
  immigration_guidelines_db_v2.json 의 각 row가 보유한 manual_ref(현재값)와
  PDF/structure/CSV 기반으로 새로 도출한 후보(proposed)를 비교하여
  audit JSON + xlsx + 콘솔 요약을 생성.

출력:
  backend/data/manuals/manual_mapping_audit.json
  backend/data/manuals/manual_mapping_audit.xlsx (openpyxl 있으면)

쓰기 보장:
  위 두 출력 파일 외에는 어떤 파일도 변경하지 않음.
  DB는 절대 수정하지 않음.
"""
from __future__ import annotations
import json, re, sys, csv
from pathlib import Path
from collections import Counter

import fitz

ROOT     = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
MANUALS    = ROOT / "backend" / "data" / "manuals"
STRUCT     = MANUALS / "structure"
INDEX_V6   = MANUALS / "manual_index_v6.json"
CSV_CODES  = STRUCT / "법무부_체류자격 분류코드_20260331.csv"

OUT_JSON   = MANUALS / "manual_mapping_audit.json"
OUT_XLSX   = MANUALS / "manual_mapping_audit.xlsx"

PDF_PATHS = {
    "체류민원": MANUALS / "unlocked_체류민원.pdf",
    "사증민원": MANUALS / "unlocked_사증민원.pdf",
}

# indexer_v6의 상수 import (READ-ONLY)
from backend.services.manual_indexer_v6 import (
    DB_TO_MANUAL_ALIAS,
    MANUAL_PAGE_OVERRIDE,
    ACTION_KEYWORDS,
)


# ─── 패턴 ──────────────────────────────────────────────────────
_BRACKET_FORMAL = re.compile(r"〔\s*([A-Z]-\d+(?:-(?:\d+|[A-Z][\w]*))?)\s*〕")


# ─── 데이터 로딩 ────────────────────────────────────────────────
def load_pdf_pages(path: Path) -> tuple[dict[int, str], int]:
    """{page_no(1-based): text}"""
    doc = fitz.open(path)
    out = {p.number + 1: p.get_text() for p in doc}
    total = len(doc)
    doc.close()
    return out, total


def load_structure(name: str) -> dict | None:
    p = STRUCT / f"{name}_structure.json"
    if not p.exists():
        return None
    return json.loads(p.read_text(encoding="utf-8"))


def load_csv_codes(path: Path) -> set[str]:
    """법무부 체류자격 분류코드 CSV(cp949) → 코드 집합."""
    if not path.exists():
        return set()
    codes: set[str] = set()
    with open(path, encoding="cp949", errors="replace") as f:
        reader = csv.reader(f)
        next(reader, None)  # header
        for row in reader:
            if row and row[0].strip():
                codes.add(row[0].strip())
    return codes


# ─── PDF 매칭 유틸 ──────────────────────────────────────────────
def find_bracket_section(pdf_pages: dict[int, str], code: str,
                         max_section_size: int = 30) -> tuple[int, int] | None:
    """〔code〕 정식 헤더가 처음 나오는 페이지 ~ 다음 다른 〔??〕 직전."""
    target = re.compile(rf"〔\s*{re.escape(code)}\s*〕")
    first = None
    for pno, txt in sorted(pdf_pages.items()):
        if target.search(txt):
            first = pno
            break
    if first is None:
        return None
    next_other = None
    for pno in sorted(p for p in pdf_pages if p > first):
        for m in _BRACKET_FORMAL.finditer(pdf_pages[pno]):
            if m.group(1) != code:
                next_other = pno
                break
        if next_other:
            break
    end = (next_other - 1) if next_other else (first + max_section_size)
    end = min(end, first + max_section_size, max(pdf_pages))
    return (first, end)


def find_paren_cluster(pdf_pages: dict[int, str], code: str,
                       min_count: int = 3) -> tuple[int, int] | None:
    """(code) 본문 클러스터 — 헤더에 자주 등장하는 자격 섹션."""
    pages_score: dict[int, int] = {}
    pat = re.compile(rf"\(\s*{re.escape(code)}\s*\)")
    for pno, txt in pdf_pages.items():
        cnt = len(pat.findall(txt))
        if cnt >= min_count:
            pages_score[pno] = cnt
    if not pages_score:
        return None
    pages = sorted(pages_score)
    clusters: list[list[int]] = []
    cur = [pages[0]]
    for p in pages[1:]:
        if p - cur[-1] <= 5:
            cur.append(p)
        else:
            clusters.append(cur)
            cur = [p]
    clusters.append(cur)
    best = max(clusters, key=lambda c: sum(pages_score[p] for p in c))
    return (best[0], best[-1])


def find_action_in_range(pdf_pages: dict[int, str], pf: int, pt: int,
                         action_type: str) -> tuple[int, int, str] | None:
    """[pf..pt] 범위에서 action_type 키워드가 처음/마지막 나오는 페이지."""
    kws = ACTION_KEYWORDS.get(action_type, [])
    if not kws:
        return None
    first = last = None
    matched_kw = None
    for p in range(pf, pt + 1):
        txt = pdf_pages.get(p, "")
        for k in kws:
            if k in txt:
                if first is None:
                    first = p
                    matched_kw = k
                last = p
                break
    if first is None:
        return None
    return (first, last, matched_kw or "")


# ─── 우선 매뉴얼 결정 ──────────────────────────────────────────
def primary_manual_for(row: dict) -> str:
    domain = (row.get("domain") or "").strip()
    at = (row.get("action_type") or "").strip()
    if "사증" in domain or at == "VISA_CONFIRM":
        return "사증민원"
    return "체류민원"


# ─── 후보 결정 ─────────────────────────────────────────────────
def resolve_candidate(row: dict, pdf_data: dict, struct_data: dict,
                      v6_index: dict | None) -> dict:
    """
    단일 row → {manual_ref, method, confidence, evidence}.
    Confidence:
      exact   = 〔code〕 또는 manual_override + action 키워드 적중
      high    = code alias 적용 후 정식/액션 적중, structure 정식 sub-cat + action,
                또는 v6 action_index 정확 매칭 (단, prefix 폴백 아님)
      medium  = 자격 섹션은 잡혔으나 action 키워드 없음
      low     = prefix 폴백 (F-4-19 → F-4 자격 섹션)
      none    = 어떤 방법으로도 못 찾음
    """
    code = (row.get("detailed_code") or "").strip()
    at = (row.get("action_type") or "").strip()
    primary = primary_manual_for(row)
    secondary = "체류민원" if primary == "사증민원" else "사증민원"

    # ── 1. MANUAL_PAGE_OVERRIDE — 사람이 검증한 정답 ──
    ovr_key = f"{code}|{at}"
    if ovr_key in MANUAL_PAGE_OVERRIDE:
        ovr = MANUAL_PAGE_OVERRIDE[ovr_key][0]
        return {
            "manual_ref": [{
                "manual": ovr["manual"],
                "page_from": ovr["page_from"],
                "page_to": ovr["page_to"],
                "match_text": ovr.get("match_text", ""),
                "match_type": "manual_override",
            }],
            "method": "manual_override",
            "confidence": "exact",
            "evidence": f"수동 override: {ovr.get('match_text', '')}",
        }

    if not code:
        return {
            "manual_ref": [],
            "method": "none",
            "confidence": "none",
            "evidence": "detailed_code 비어있음",
        }

    manuals_to_try = [m for m in (primary, secondary) if m in pdf_data]

    # ── 2. exact_bracket_heading — 〔code〕 PDF 직접 탐색 ──
    for manual in manuals_to_try:
        section = find_bracket_section(pdf_data[manual], code)
        if section is None:
            continue
        pf, pt = section
        if at:
            hit = find_action_in_range(pdf_data[manual], pf, pt, at)
            if hit:
                af, al, kw = hit
                return {
                    "manual_ref": [{
                        "manual": manual, "page_from": af, "page_to": al,
                        "match_text": f"〔{code}〕 + '{kw}'",
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": "exact",
                    "evidence": f"{manual} p.{af}~{al}: 〔{code}〕 정식 sub-cat 안 '{kw}' 등장",
                }
        return {
            "manual_ref": [{
                "manual": manual, "page_from": pf, "page_to": pt,
                "match_text": f"〔{code}〕 정식 헤더",
                "match_type": "section_only",
            }],
            "method": "exact_bracket_heading",
            "confidence": "medium",
            "evidence": f"{manual} p.{pf}~{pt}: 〔{code}〕 정식 sub-cat 섹션 (action 키워드 미발견)",
        }

    # ── 3. code_alias_exact — DB_TO_MANUAL_ALIAS ──
    alias = DB_TO_MANUAL_ALIAS.get(code)
    if alias and alias != code:
        for manual in manuals_to_try:
            section = find_bracket_section(pdf_data[manual], alias)
            if section is None:
                continue
            pf, pt = section
            if at:
                hit = find_action_in_range(pdf_data[manual], pf, pt, at)
                if hit:
                    af, al, kw = hit
                    return {
                        "manual_ref": [{
                            "manual": manual, "page_from": af, "page_to": al,
                            "match_text": f"〔{alias}〕(alias of {code}) + '{kw}'",
                            "match_type": "action_exact",
                        }],
                        "method": "code_alias_exact",
                        "confidence": "high",
                        "evidence": f"{manual} p.{af}~{al}: alias 〔{alias}〕 (DB:{code}) + '{kw}'",
                    }
            return {
                "manual_ref": [{
                    "manual": manual, "page_from": pf, "page_to": pt,
                    "match_text": f"〔{alias}〕(alias of {code})",
                    "match_type": "section_only",
                }],
                "method": "code_alias_exact",
                "confidence": "medium",
                "evidence": f"{manual} p.{pf}~{pt}: alias 〔{alias}〕 (DB:{code}) 섹션",
            }

    # ── 4. structure_heading — structure JSON sub_categories / qualities ──
    for manual in manuals_to_try:
        s = struct_data.get(manual)
        if not s:
            continue
        quals = s.get("qualities", {}) or {}
        # sub_categories 안에 있는지
        for q_code, q in quals.items():
            for sc in q.get("sub_categories", []) or []:
                if sc.get("code") == code:
                    pf = sc.get("page_from")
                    pt = sc.get("page_to") or pf
                    if not pf:
                        continue
                    if at:
                        hit = find_action_in_range(pdf_data[manual], pf, pt, at)
                        if hit:
                            af, al, kw = hit
                            return {
                                "manual_ref": [{
                                    "manual": manual, "page_from": af, "page_to": al,
                                    "match_text": f"structure {code} '{kw}'",
                                    "match_type": "action_exact",
                                }],
                                "method": "structure_heading",
                                "confidence": "high",
                                "evidence": f"{manual} p.{af}~{al}: structure sub-cat {code} + '{kw}'",
                            }
                    return {
                        "manual_ref": [{
                            "manual": manual, "page_from": pf, "page_to": pt,
                            "match_text": f"structure {sc.get('name', code)}",
                            "match_type": "section_only",
                        }],
                        "method": "structure_heading",
                        "confidence": "medium",
                        "evidence": f"{manual} p.{pf}~{pt}: structure sub-cat {code} ({sc.get('name', '')})",
                    }
        # quality(대분류) 자체와 일치하는지
        if code in quals:
            q = quals[code]
            actions = q.get("actions") or {}
            act = actions.get(at) if at else None
            if act and act.get("page_from"):
                return {
                    "manual_ref": [{
                        "manual": manual,
                        "page_from": act["page_from"],
                        "page_to": act.get("page_to") or act["page_from"],
                        "match_text": f"structure {code} action {at}",
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": "high",
                    "evidence": f"{manual} p.{act['page_from']}: structure quality {code}.actions.{at}",
                }
            pf = q.get("page_from")
            pt = q.get("page_to") or pf
            if pf:
                return {
                    "manual_ref": [{
                        "manual": manual, "page_from": pf, "page_to": pt,
                        "match_text": f"structure {code}",
                        "match_type": "section_only",
                    }],
                    "method": "structure_heading",
                    "confidence": "medium",
                    "evidence": f"{manual} p.{pf}~{pt}: structure quality {code} 섹션 전체",
                }

    # ── 5. v6 action_index — 기존 인덱스 정확 매칭 ──
    if v6_index:
        ai = v6_index.get("action_index", {}) or {}
        key = f"{code}|{at}"
        if key in ai and ai[key]:
            entries_pri = [e for e in ai[key] if e.get("manual") == primary]
            entries = entries_pri or ai[key]
            best = max(entries, key=lambda e: e.get("score", 0))
            mk = best.get("match_kind", "")
            conf = "high" if mk in ("subcategory_with_action", "subcategory_change_default") else "medium"
            return {
                "manual_ref": [{
                    "manual": best["manual"],
                    "page_from": best["page_from"],
                    "page_to": best["page_to"],
                    "match_text": best.get("match_text", ""),
                    "match_type": "action_exact",
                }],
                "method": "exact_action_heading",
                "confidence": conf,
                "evidence": f"v6 action_index {key}: {best.get('match_text', '')} (kind={mk})",
            }

    # ── 6. (code) 본문 클러스터 — 일반 자격 섹션 ──
    for manual in manuals_to_try:
        section = find_paren_cluster(pdf_data[manual], code)
        if section is None:
            continue
        pf, pt = section
        if at:
            hit = find_action_in_range(pdf_data[manual], pf, pt, at)
            if hit:
                af, al, kw = hit
                return {
                    "manual_ref": [{
                        "manual": manual, "page_from": af, "page_to": al,
                        "match_text": f"({code}) 본문 + '{kw}'",
                        "match_type": "action_exact",
                    }],
                    "method": "exact_action_heading",
                    "confidence": "medium",
                    "evidence": f"{manual} p.{af}~{al}: ({code}) 본문 클러스터 + '{kw}'",
                }
        return {
            "manual_ref": [{
                "manual": manual, "page_from": pf, "page_to": pt,
                "match_text": f"({code}) 본문",
                "match_type": "section_only",
            }],
            "method": "section_only",
            "confidence": "medium",
            "evidence": f"{manual} p.{pf}~{pt}: ({code}) 자격 섹션 (action 미확정)",
        }

    # ── 7. prefix_fallback — F-4-19 → F-4 자격 섹션 ──
    main = re.match(r"^([A-Z]-\d+)", code)
    if main and main.group(1) != code:
        prefix = main.group(1)
        for manual in manuals_to_try:
            section = find_paren_cluster(pdf_data[manual], prefix)
            if section is None:
                continue
            pf, pt = section
            return {
                "manual_ref": [{
                    "manual": manual, "page_from": pf, "page_to": pt,
                    "match_text": f"prefix {prefix}",
                    "match_type": "section_only",
                }],
                "method": "prefix_fallback",
                "confidence": "low",
                "evidence": f"{manual} p.{pf}~{pt}: prefix {prefix} 섹션 (정확 페이지 미확정)",
            }

    return {
        "manual_ref": [],
        "method": "none",
        "confidence": "none",
        "evidence": "어떤 방법으로도 후보 페이지를 찾지 못함",
    }


# ─── 비교 ──────────────────────────────────────────────────────
def compare_refs(current: list, proposed_pkg: dict) -> dict:
    cur = current or []
    prop = (proposed_pkg or {}).get("manual_ref", []) or []

    if not cur and not prop:
        return {"status": "missing_both", "page_changed": False, "manual_changed": False}
    if not cur:
        return {"status": "missing_current", "page_changed": False, "manual_changed": False}
    if not prop:
        return {"status": "missing_candidate", "page_changed": False, "manual_changed": False}

    c0, p0 = cur[0], prop[0]
    page_changed = (c0.get("page_from") != p0.get("page_from")) or (c0.get("page_to") != p0.get("page_to"))
    manual_changed = c0.get("manual") != p0.get("manual")

    if not page_changed and not manual_changed:
        return {"status": "same", "page_changed": False, "manual_changed": False}
    if manual_changed:
        return {"status": "conflict", "page_changed": page_changed, "manual_changed": True}
    return {"status": "changed", "page_changed": True, "manual_changed": False}


def needs_human_review(comp: dict, conf: str) -> tuple[bool, str]:
    reasons: list[str] = []
    status = comp["status"]

    if status == "missing_current":
        reasons.append("DB에 manual_ref 없음")
    if status == "missing_candidate":
        reasons.append("후보 페이지 없음")
    if status == "missing_both":
        reasons.append("현재값/후보 모두 없음")
    if status == "conflict":
        reasons.append("매뉴얼이 다름 (체류 ↔ 사증)")
    if status == "changed":
        reasons.append("페이지 변경 감지")

    if conf == "low":
        reasons.append("prefix 폴백 — 정확도 낮음")
    if conf == "medium":
        reasons.append("자격 섹션 단위 — action 페이지 미확정")
    if conf == "none":
        reasons.append("매핑 불가")

    if not reasons and conf in ("exact", "high"):
        return False, ""

    if conf == "exact" and status == "same":
        return False, ""
    return True, " / ".join(reasons) if reasons else "신뢰도 검토 필요"


# ─── 메인 ──────────────────────────────────────────────────────
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    print("[audit] DB 로드...")
    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rows = db.get("master_rows", []) or []
    print(f"  master_rows: {len(rows)}")

    print("[audit] PDF 텍스트 추출...")
    pdf_data: dict[str, dict[int, str]] = {}
    pdf_meta: dict[str, dict] = {}
    for label, p in PDF_PATHS.items():
        if not p.exists():
            print(f"  [skip] {label}: 파일 없음 ({p.name})")
            continue
        pages, total = load_pdf_pages(p)
        pdf_data[label] = pages
        pdf_meta[label] = {"file": str(p.relative_to(ROOT)), "pages": total}
        print(f"  {label}: {total} pages")

    print("[audit] structure JSON 로드...")
    struct_data = {
        "체류민원": load_structure("체류민원"),
        "사증민원": load_structure("사증민원"),
    }
    for k, v in struct_data.items():
        print(f"  {k}: {'OK' if v else 'NONE'}")

    print("[audit] manual_index_v6 로드...")
    v6 = json.loads(INDEX_V6.read_text(encoding="utf-8")) if INDEX_V6.exists() else None
    print(f"  v6: {'OK (action_index=' + str(len(v6.get('action_index', {}))) + ')' if v6 else 'NONE'}")

    print("[audit] CSV 코드 로드...")
    csv_codes = load_csv_codes(CSV_CODES)
    print(f"  CSV 코드: {len(csv_codes)}")

    print("[audit] 각 row 후보 결정 + 비교...")
    audit_rows: list[dict] = []
    for i, row in enumerate(rows):
        proposed = resolve_candidate(row, pdf_data, struct_data, v6)
        comp = compare_refs(row.get("manual_ref") or [], proposed)
        nr, reason = needs_human_review(comp, proposed["confidence"])

        code = (row.get("detailed_code") or "").strip()
        csv_known = (code in csv_codes) if code else None

        audit_rows.append({
            "row_id":               row.get("row_id"),
            "title":                row.get("business_name"),
            "action_type":          row.get("action_type"),
            "detailed_code":        code,
            "domain":               row.get("domain"),
            "csv_known_code":       csv_known,
            "current_manual_ref":   row.get("manual_ref") or [],
            "proposed_manual_ref":  proposed["manual_ref"],
            "comparison":           comp,
            "match": {
                "method":     proposed["method"],
                "confidence": proposed["confidence"],
                "evidence":   proposed["evidence"],
            },
            "needs_human_review":   nr,
            "review_reason":        reason,
        })

        if (i + 1) % 50 == 0:
            print(f"  [{i + 1}/{len(rows)}]")

    # ── 통계 ──
    conf_dist  = Counter(a["match"]["confidence"]   for a in audit_rows)
    method_dist = Counter(a["match"]["method"]      for a in audit_rows)
    status_dist = Counter(a["comparison"]["status"] for a in audit_rows)

    n_total = len(audit_rows)
    n_exact_high = conf_dist.get("exact", 0) + conf_dist.get("high", 0)
    n_medium = conf_dist.get("medium", 0)
    n_low_none = conf_dist.get("low", 0) + conf_dist.get("none", 0)
    n_changed = sum(1 for a in audit_rows
                    if a["comparison"]["page_changed"] or a["comparison"]["manual_changed"])
    n_no_current = sum(1 for a in audit_rows if not a["current_manual_ref"])
    n_no_proposed = sum(1 for a in audit_rows if not a["proposed_manual_ref"])
    n_review = sum(1 for a in audit_rows if a["needs_human_review"])

    summary = {
        "total":               n_total,
        "exact_or_high":       n_exact_high,
        "medium":              n_medium,
        "low_or_none":         n_low_none,
        "current_changed":     n_changed,
        "no_current":          n_no_current,
        "no_proposed":         n_no_proposed,
        "needs_review":        n_review,
        "confidence_dist":     dict(conf_dist),
        "method_dist":         dict(method_dist),
        "status_dist":         dict(status_dist),
    }

    out = {
        "audit_meta": {
            "db_path":   str(DB_PATH.relative_to(ROOT)),
            "manuals":   pdf_meta,
            "index_v6":  str(INDEX_V6.relative_to(ROOT)) if INDEX_V6.exists() else None,
            "structure": {k: bool(v) for k, v in struct_data.items()},
            "csv_codes": len(csv_codes),
        },
        "summary": summary,
        "rows":    audit_rows,
    }

    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX (openpyxl 가능 시) ──
    xlsx_written = False
    xlsx_skip_reason = ""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "audit"
        headers = [
            "row_id", "title", "action_type", "detailed_code", "domain",
            "status", "page_changed", "manual_changed",
            "method", "confidence", "needs_review", "review_reason",
            "current_manual", "current_pf", "current_pt",
            "proposed_manual", "proposed_pf", "proposed_pt",
            "evidence",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFE599")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def first_ref(refs):
            if not refs:
                return ("", "", "")
            r = refs[0]
            return (r.get("manual", ""), r.get("page_from", ""), r.get("page_to", ""))

        for a in audit_rows:
            cm, cpf, cpt = first_ref(a["current_manual_ref"])
            pm, ppf, ppt = first_ref(a["proposed_manual_ref"])
            ws.append([
                a["row_id"], a["title"], a["action_type"], a["detailed_code"], a.get("domain", ""),
                a["comparison"]["status"], a["comparison"]["page_changed"], a["comparison"]["manual_changed"],
                a["match"]["method"], a["match"]["confidence"], a["needs_human_review"], a["review_reason"],
                cm, cpf, cpt, pm, ppf, ppt,
                a["match"]["evidence"],
            ])

        widths = {
            "A": 10, "B": 30, "C": 14, "D": 12, "E": 10,
            "F": 18, "G": 8, "H": 8, "I": 22, "J": 10, "K": 8, "L": 36,
            "M": 10, "N": 8, "O": 8, "P": 10, "Q": 8, "R": 8, "S": 50,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"

        # summary 시트
        ws2 = wb.create_sheet("summary")
        ws2.append(["metric", "value"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            if isinstance(v, dict):
                ws2.append([k, json.dumps(v, ensure_ascii=False)])
            else:
                ws2.append([k, v])
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 80

        wb.save(OUT_XLSX)
        xlsx_written = True
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        xlsx_skip_reason = f"openpyxl import failed: {e}"
        print(f"[skip xlsx] {xlsx_skip_reason}")
    except Exception as e:
        xlsx_skip_reason = f"xlsx 생성 오류: {e}"
        print(f"[skip xlsx] {xlsx_skip_reason}")

    # ── 콘솔 요약 ──
    print("\n" + "=" * 72)
    print("AUDIT SUMMARY")
    print("=" * 72)
    print(f"  1. 전체 row 수:                    {n_total}")
    print(f"  2. exact + high (신뢰):            {n_exact_high}")
    print(f"  3. medium (자격 섹션 단위):        {n_medium}")
    print(f"  4. low + none (검토 필수):         {n_low_none}")
    print(f"  5. 현재 manual_ref 변경 감지:      {n_changed}")
    print(f"  6. 현재 manual_ref 없음:           {n_no_current}")
    print(f"  7. 후보 페이지 없음:               {n_no_proposed}")
    print(f"  8. human_review 필요:              {n_review}")
    print(f"\n  confidence dist: {dict(conf_dist)}")
    print(f"  method dist:     {dict(method_dist)}")
    print(f"  status dist:     {dict(status_dist)}")

    print("\n=== Top 30 review needed ===")
    review_rows = [a for a in audit_rows if a["needs_human_review"]]
    # 우선순위: conflict > missing_candidate > missing_current > changed > low > medium
    priority = {
        "conflict": 0, "missing_candidate": 1, "missing_current": 2,
        "missing_both": 3, "changed": 4, "same": 9,
    }
    review_rows.sort(key=lambda a: (
        priority.get(a["comparison"]["status"], 5),
        {"none": 0, "low": 1, "medium": 2, "high": 3, "exact": 4}.get(a["match"]["confidence"], 5),
        a["row_id"] or "",
    ))
    for a in review_rows[:30]:
        title = (a["title"] or "")[:24]
        print(
            f"  {a['row_id']:8s} {a['detailed_code']:8s} {a['action_type']:13s} "
            f"conf={a['match']['confidence']:6s} status={a['comparison']['status']:18s} "
            f"{title:24s} :: {a['review_reason']}"
        )

    print("\n=== Files created ===")
    print(f"  - {OUT_JSON.relative_to(ROOT)}")
    if xlsx_written:
        print(f"  - {OUT_XLSX.relative_to(ROOT)}")
    else:
        print(f"  - (xlsx skipped: {xlsx_skip_reason or 'unknown'})")

    print("\n=== Files NOT modified (read-only) ===")
    print(f"  - {DB_PATH.relative_to(ROOT)}")
    print(f"  - {INDEX_V6.relative_to(ROOT)}")
    for label, p in PDF_PATHS.items():
        print(f"  - {p.relative_to(ROOT)}")
    print("  - backend/data/manuals/structure/*")
    print("  - backend/services/manual_indexer_v6.py")
    print("  - backend/services/manual_watcher.py")
    print("  - frontend/**")


if __name__ == "__main__":
    main()
