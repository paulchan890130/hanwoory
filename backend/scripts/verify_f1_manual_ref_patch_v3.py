"""
F-1 manual_ref surgical patch — 사후 검증 (READ-ONLY)

검증 항목:
  1. immigration_guidelines_db_v2.json 가 valid JSON 으로 로드됨
  2. master_rows 개수 == 369
  3. 정확히 4 row 만 pre-patch backup 대비 변경됨
  4. 변경된 row 는 manual_ref 외 다른 필드 무수정
  5. 4 target row 의 manual_ref 가 기대값과 일치
  6. 나머지 365 row 는 pre-patch backup 과 bit-identical
  7. 이전 12 blocklist row 는 무수정 유지
  8. backup 파일 존재
  9. rollback 명령 출력

쓰기 보장: 어떤 파일도 수정/이동/삭제하지 않음. 출력 리포트(JSON/XLSX) 2건만 새로 작성.
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

ROOT      = Path(__file__).parent.parent.parent
DB_PATH   = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUPS   = ROOT / "backend" / "data" / "backups"
BLOCKLIST = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_blocklist_v3.json"

OUT_JSON = ROOT / "backend" / "data" / "manuals" / "f1_manual_ref_patch_verify_v3.json"
OUT_XLSX = ROOT / "backend" / "data" / "manuals" / "f1_manual_ref_patch_verify_v3.xlsx"

TARGETS = ["M1-0209", "M1-0211", "M1-0227", "M1-0230"]

EXPECTED_REFS: dict[str, list[dict]] = {
    "M1-0209": [
        {
            "manual": "사증민원", "page_from": 298, "page_to": 298,
            "match_type": "manual_override",
            "match_text": "F-1-21 주한 외국공관원 가사보조인 단수사증 (사증민원 p.298 #2) — 수동 검증",
        },
        {
            "manual": "사증민원", "page_from": 306, "page_to": 307,
            "match_type": "manual_override",
            "match_text": "F-1-21 사증발급인정서 첨부서류 (사증민원 p.306-307 #3) — 수동 검증",
        },
    ],
    "M1-0211": [
        {
            "manual": "체류민원", "page_from": 349, "page_to": 349,
            "match_type": "manual_override",
            "match_text": "F-1-15 우수인재·투자자·유학생 부모 체류기간 연장 (체류민원 p.349 #8) — 수동 검증",
        },
    ],
    "M1-0227": [
        {
            "manual": "체류민원", "page_from": 348, "page_to": 348,
            "match_type": "manual_override",
            "match_text": "F-1-22 고액투자가 가사보조인 체류기간 연장 (체류민원 p.348) — 수동 검증",
        },
    ],
    "M1-0230": [
        {
            "manual": "체류민원", "page_from": 348, "page_to": 348,
            "match_type": "manual_override",
            "match_text": "F-1-24 해외우수인재 가사보조인 체류기간 연장 (체류민원 p.348) — 수동 검증",
        },
        {
            "manual": "체류민원", "page_from": 668, "page_to": 669,
            "match_type": "manual_override",
            "match_text": "F-1-24 최우수인재 부모·가사보조인 신청서류 (체류민원 p.668-669) — 수동 검증",
        },
    ],
}


def find_f1_backup() -> Path | None:
    bs = sorted(BACKUPS.glob("immigration_guidelines_db_v2.f1_manual_ref_patch_backup_*.json"))
    return bs[-1] if bs else None


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    checks: list[dict] = []

    def check(name: str, ok: bool, detail: str = "") -> bool:
        checks.append({"name": name, "ok": bool(ok), "detail": detail})
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] {name}" + (f" — {detail}" if detail else ""))
        return ok

    print("[verify] DB 로드...")
    db = None
    db_load_err = ""
    try:
        db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    except Exception as e:
        db_load_err = f"{type(e).__name__}: {e}"

    check("1. DB JSON 로드 가능", db is not None,
          f"path={DB_PATH.relative_to(ROOT)}" + (f", error={db_load_err}" if db_load_err else ""))

    if db is None:
        # 실패 시 abort 하기 전에 리포트 작성
        out = {"verify_meta": {"status": "FAIL", "step": "json_load", "error": db_load_err},
               "checks": checks}
        OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        sys.exit(1)

    rows = db.get("master_rows") or []
    check("2. master_rows == 369", len(rows) == 369, f"actual={len(rows)}")

    print("\n[verify] backup 탐색...")
    bk_path = find_f1_backup()
    backup_exists = bk_path is not None and bk_path.exists()
    check("8. F-1 patch backup 존재", backup_exists,
          f"path={bk_path.relative_to(ROOT)}" if bk_path else "no f1_manual_ref_patch_backup_*.json found")

    backup_db = None
    if backup_exists:
        try:
            backup_db = json.loads(bk_path.read_text(encoding="utf-8"))
        except Exception as e:
            check("8a. backup JSON parse", False, f"{type(e).__name__}: {e}")
            backup_exists = False

    if backup_db is None:
        out = {"verify_meta": {"status": "FAIL", "step": "backup_load"},
               "checks": checks}
        OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        sys.exit(1)

    bk_rows = backup_db.get("master_rows") or []
    bk_idx = {r.get("row_id"): r for r in bk_rows}
    cur_idx = {r.get("row_id"): r for r in rows}

    # 3. 정확히 4 row 변경 (pre-patch backup 대비)
    print("\n[verify] backup 대비 diff 계산...")
    changed_ids: set[str] = set()
    only_manual_ref_changed: list[str] = []
    non_manual_ref_diffs: list[dict] = []
    for rid, cur in cur_idx.items():
        bk = bk_idx.get(rid)
        if bk is None:
            changed_ids.add(rid)
            non_manual_ref_diffs.append({"row_id": rid, "issue": "row not in backup"})
            continue
        diff_fields = []
        for k in set(cur.keys()) | set(bk.keys()):
            if cur.get(k) != bk.get(k):
                diff_fields.append(k)
        if diff_fields:
            changed_ids.add(rid)
            if diff_fields == ["manual_ref"]:
                only_manual_ref_changed.append(rid)
            else:
                non_manual_ref_diffs.append({
                    "row_id": rid,
                    "fields": sorted(diff_fields),
                })

    check("3. 정확히 4 row 변경", len(changed_ids) == 4,
          f"changed={len(changed_ids)} rows={sorted(changed_ids)}")
    check("3a. 변경된 row 가 정확히 target 4건", set(changed_ids) == set(TARGETS),
          f"diff_set={sorted(changed_ids)} expected={TARGETS}")
    check("4. 변경된 row 는 manual_ref 외 필드 무수정",
          len(non_manual_ref_diffs) == 0 and set(only_manual_ref_changed) == set(TARGETS),
          f"manual_ref_only={sorted(only_manual_ref_changed)} other_diffs={non_manual_ref_diffs}")

    # 5. target row manual_ref 기대값
    print("\n[verify] target row manual_ref 확인...")
    target_pass = 0
    target_results = []
    for rid in TARGETS:
        cur_ref = cur_idx.get(rid, {}).get("manual_ref")
        exp_ref = EXPECTED_REFS[rid]
        ok = cur_ref == exp_ref
        if ok:
            target_pass += 1
        target_results.append({
            "row_id": rid,
            "ok": ok,
            "current": cur_ref,
            "expected": exp_ref,
        })
        check(f"5.{rid} manual_ref 기대값 일치", ok,
              "" if ok else f"current={json.dumps(cur_ref, ensure_ascii=False)[:200]}")

    # 6. 나머지 365 row 는 backup 과 bit-identical
    print("\n[verify] 비-target row identical 확인...")
    non_target_changed = sorted(rid for rid in changed_ids if rid not in TARGETS)
    check("6. 비-target 365 row 는 backup 과 동일",
          len(non_target_changed) == 0,
          f"unexpected_changes={non_target_changed}")

    # 7. blocklist 12 row 무수정 확인
    print("\n[verify] blocklist 12 row 확인...")
    bl_doc = json.loads(BLOCKLIST.read_text(encoding="utf-8"))
    blocked_ids = list(bl_doc.get("blocked_row_ids") or [])
    blocked_diffs = []
    for rid in blocked_ids:
        cur = cur_idx.get(rid, {})
        bk = bk_idx.get(rid, {})
        if cur != bk:
            diff_fields = sorted(set(cur.keys()) | set(bk.keys())
                                 if cur.get(k:=None) is None else [])
            # 정밀 diff
            df = [k for k in set(cur.keys()) | set(bk.keys())
                  if cur.get(k) != bk.get(k)]
            blocked_diffs.append({"row_id": rid, "fields": df})
    check(f"7. blocklist {len(blocked_ids)} row 무수정",
          len(blocked_diffs) == 0,
          f"diffs={blocked_diffs}" if blocked_diffs else f"all {len(blocked_ids)} rows identical to backup")

    # 종합
    all_pass = all(c["ok"] for c in checks)
    rollback_cmd = (
        f"cp {bk_path.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json"
        if bk_path else "(no backup found)"
    )

    summary = {
        "status":                       "PASS" if all_pass else "FAIL",
        "backup_path":                  str(bk_path.relative_to(ROOT)) if bk_path else None,
        "rollback_command":             rollback_cmd,
        "master_rows":                  len(rows),
        "updated_row_count":            len(changed_ids),
        "target_rows_verified_count":   target_pass,
        "expected_target_count":        len(TARGETS),
        "non_target_row_diff_count":    len(non_target_changed),
        "non_manual_ref_field_diff_count": len(non_manual_ref_diffs),
        "blocklist_row_count":          len(blocked_ids),
        "blocklist_diff_count":         len(blocked_diffs),
        "checks_total":                 len(checks),
        "checks_passed":                sum(1 for c in checks if c["ok"]),
        "checks_failed":                sum(1 for c in checks if not c["ok"]),
    }

    out = {
        "verify_meta": {
            "version":     "v3",
            "scope":       "F-1 manual_ref surgical patch — 사후 검증",
            "db_path":     str(DB_PATH.relative_to(ROOT)),
            "backup_path": str(bk_path.relative_to(ROOT)) if bk_path else None,
        },
        "summary":       summary,
        "checks":        checks,
        "target_results": target_results,
        "non_target_changed_rows":      non_target_changed,
        "non_manual_ref_diffs":         non_manual_ref_diffs,
        "blocklist_diffs":              blocked_diffs,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
        ws_s.column_dimensions["A"].width = 36
        ws_s.column_dimensions["B"].width = 80

        ws_c = wb.create_sheet("checks")
        ws_c.append(["#", "name", "ok", "detail"])
        for c in ws_c[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for i, c in enumerate(checks, start=1):
            ws_c.append([i, c["name"], c["ok"], c["detail"][:500]])
            row_fill = PatternFill("solid", fgColor="C6F6D5" if c["ok"] else "FFCCCB")
            for cell in ws_c[ws_c.max_row]:
                cell.fill = row_fill
        ws_c.column_dimensions["A"].width = 4
        ws_c.column_dimensions["B"].width = 50
        ws_c.column_dimensions["C"].width = 8
        ws_c.column_dimensions["D"].width = 100
        ws_c.freeze_panes = "A2"

        ws_t = wb.create_sheet("target_rows")
        ws_t.append(["row_id", "ok", "current_manual_ref", "expected_manual_ref"])
        for c in ws_t[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for x in target_results:
            ws_t.append([
                x["row_id"], x["ok"],
                json.dumps(x["current"], ensure_ascii=False)[:600],
                json.dumps(x["expected"], ensure_ascii=False)[:600],
            ])
            for cell in ws_t[ws_t.max_row]:
                cell.fill = PatternFill("solid", fgColor="C6F6D5" if x["ok"] else "FFCCCB")
        ws_t.column_dimensions["A"].width = 10
        ws_t.column_dimensions["B"].width = 8
        ws_t.column_dimensions["C"].width = 80
        ws_t.column_dimensions["D"].width = 80
        ws_t.freeze_panes = "A2"

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # 콘솔 요약
    print("\n" + "=" * 78)
    print("F-1 MANUAL_REF PATCH — POST-APPLY VERIFY")
    print("=" * 78)
    print(f"  status:                          {summary['status']}")
    print(f"  backup path:                     {summary['backup_path']}")
    print(f"  master_rows:                     {summary['master_rows']}")
    print(f"  updated row count:               {summary['updated_row_count']}")
    print(f"  target rows verified count:      "
          f"{summary['target_rows_verified_count']} / {summary['expected_target_count']}")
    print(f"  non-target row diffs:            {summary['non_target_row_diff_count']}")
    print(f"  non-manual_ref field diffs:      {summary['non_manual_ref_field_diff_count']}")
    print(f"  blocklist rows checked:          {summary['blocklist_row_count']}")
    print(f"  blocklist diffs:                 {summary['blocklist_diff_count']}")
    print(f"  checks: {summary['checks_passed']}/{summary['checks_total']} passed, "
          f"{summary['checks_failed']} failed")
    print(f"\n  rollback command:")
    print(f"    {rollback_cmd}")
    print(f"\n  final status: {'✅ PASS — F-1 패치 안전하게 적용됨' if all_pass else '❌ FAIL — 위 항목 확인'}")

    sys.exit(0 if all_pass else 1)


if __name__ == "__main__":
    main()
