"""Reconcile PG ``users.password_hash`` with the authoritative legacy hash.

Some accounts (notably ``jpup``) were imported into local PostgreSQL with an
empty / stale / placeholder ``password_hash`` (e.g. the throwaway beta password
created by ``_ensure_admin_user``, or a skipped row). This script restores the
**exact legacy hash** from the migration snapshot's ``Accounts`` tab so the user
logs in with the *same password as before* — no reset, no new password.

Source of truth
---------------
The legacy ``password_hash`` is read from the local Excel snapshot
(``migration_input/*.xlsx`` → ``Accounts`` tab) — the same source the importer
already uses. Google Sheets is never contacted. The hash format is the project's
standard PBKDF2-HMAC-SHA256 + base64(salt+dk), which ``verify_password`` accepts
unchanged.

Safety contract
---------------
* **Dry-run by default.** ``--apply`` is required for any DB write.
* Reconciles a PG row only when its hash is **empty** or **mismatched** vs the
  authoritative legacy hash. If they already match → skipped (idempotent).
* **Refuses** to touch a login whose legacy source hash is missing (never
  invents or resets a password).
* Row-level ``UPDATE`` only — never rewrites all accounts, never clears tables.
* ``--apply`` requires a **local loopback** ``DATABASE_URL`` (local_guard).
* Backs up affected rows (old hash) to a local, git-ignored backups file.
* Prints only ``login_id`` + masked hash prefix/suffix. **Never prints a raw
  password** (there is no raw password anywhere in this flow).

CLI
---
    python backend/scripts/migrate_account_password_hashes.py --login-id jpup --dry-run
    python backend/scripts/migrate_account_password_hashes.py --login-id jpup --apply
    python backend/scripts/migrate_account_password_hashes.py --dry-run            # all logins in snapshot
    python backend/scripts/migrate_account_password_hashes.py --source "migration_input/신 고객 데이터.xlsx" --apply
"""
from __future__ import annotations

import argparse
import datetime
import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

MIGRATION_INPUT = ROOT / "migration_input"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"


def _mask(s: str) -> str:
    s = s or ""
    if len(s) <= 10:
        return f"<len={len(s)}>"
    return f"{s[:4]}...{s[-4:]} <len={len(s)}>"


def _discover_source() -> Path | None:
    """Find the first snapshot xlsx that contains an ``Accounts`` tab."""
    from openpyxl import load_workbook

    if not MIGRATION_INPUT.is_dir():
        return None
    for p in sorted(MIGRATION_INPUT.glob("*.xlsx")):
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
            if "Accounts" in wb.sheetnames:
                return p
        except Exception:
            continue
    return None


def _read_legacy_hashes(source: Path) -> dict[str, str]:
    """Return {login_id: password_hash} from the snapshot ``Accounts`` tab."""
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb["Accounts"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        li = header.index("login_id")
        pi = header.index("password_hash")
    except ValueError:
        return {}
    out: dict[str, str] = {}
    for r in rows[1:]:
        login = str(r[li]).strip() if li < len(r) and r[li] is not None else ""
        ph = str(r[pi]).strip() if pi < len(r) and r[pi] is not None else ""
        if login:
            out[login] = ph
    return out


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
        sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except Exception:
        pass

    ap = argparse.ArgumentParser(description="Reconcile PG users.password_hash from legacy snapshot.")
    ap.add_argument("--login-id", default=None, help="Target a single login_id (e.g. jpup).")
    ap.add_argument("--source", default=None, help="Path to snapshot xlsx with an Accounts tab.")
    ap.add_argument("--apply", action="store_true", help="Actually UPDATE PG. Without it: dry-run.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Explicit dry-run (default behavior; no-op unless --apply is also absent).")
    ap.add_argument("--only-empty", action="store_true",
                    help="Only fill empty PG hashes; do NOT touch mismatched ones.")
    args = ap.parse_args()

    source = Path(args.source) if args.source else _discover_source()
    if not source or not source.exists():
        print("[abort] no snapshot Accounts source found (migration_input/*.xlsx with an 'Accounts' tab).")
        return 2
    print(f"[source] {source.name}")

    legacy = _read_legacy_hashes(source)
    if not legacy:
        print("[abort] could not read login_id/password_hash from the Accounts tab.")
        return 2

    targets = [args.login_id] if args.login_id else sorted(legacy.keys())

    # Refuse early if a targeted single login has no source hash.
    if args.login_id and not legacy.get(args.login_id):
        print(f"[blocker] login_id={args.login_id!r}: legacy source password_hash is MISSING. "
              f"Refusing to invent or reset. (No change made.)")
        return 3

    # Build plan against PG.
    from sqlalchemy import select
    from backend.db.models.user import AccountUser
    from backend.db.session import get_sessionmaker, is_configured

    if not is_configured():
        print("[abort] DATABASE_URL not configured — cannot reach local PostgreSQL.")
        return 2

    SessionLocal = get_sessionmaker()
    plan: list[dict] = []   # {login_id, reason, old, new}
    with SessionLocal() as session:
        for login in targets:
            legacy_hash = legacy.get(login, "")
            if not legacy_hash:
                print(f"  - {login}: SKIP (no legacy source hash — not inventing)")
                continue
            row = session.scalar(select(AccountUser).where(AccountUser.login_id == login))
            if row is None:
                print(f"  - {login}: SKIP (no PG users row — run the importer first)")
                continue
            current = (row.password_hash or "").strip()
            if current == legacy_hash:
                print(f"  - {login}: OK (PG hash already matches legacy) {_mask(current)}")
                continue
            if current and args.only_empty:
                print(f"  - {login}: SKIP (mismatch but --only-empty set) pg={_mask(current)}")
                continue
            reason = "empty" if not current else "mismatch (clearly wrong vs authoritative legacy)"
            plan.append({"login_id": login, "reason": reason, "old": current, "new": legacy_hash})
            print(f"  - {login}: WILL UPDATE [{reason}]  pg={_mask(current)} -> legacy={_mask(legacy_hash)}")

    print()
    print(f"[plan] {len(plan)} account(s) to reconcile")
    if not plan:
        print("[done] nothing to do.")
        return 0

    if not args.apply:
        print("[dry-run] no DB changes made. Re-run with --apply to write.")
        return 0

    # ── apply ──
    from backend.db.local_guard import assert_local_database_url
    assert_local_database_url(os.environ.get("DATABASE_URL"))
    print("[guard] DATABASE_URL host check: PASS (local loopback only)")

    # backup affected rows (old hashes) to a local, git-ignored file
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"account_pw_hash_backup_{ts}.json"
    bk.write_text(json.dumps(
        [{"login_id": p["login_id"], "old_password_hash": p["old"]} for p in plan],
        ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[backup] affected old hashes saved → {bk.relative_to(ROOT)}")

    updated = 0
    with SessionLocal() as session:
        for p in plan:
            row = session.scalar(select(AccountUser).where(AccountUser.login_id == p["login_id"]))
            if row is None:
                continue
            row.password_hash = p["new"]
            updated += 1
        session.commit()
    print(f"[apply] updated {updated} account(s).")
    print("[done] affected logins can now sign in with their original password.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
