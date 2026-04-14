"""
정리.xlsx → immigration_guidelines_db_v2.json 재생성 스크립트

데이터를 수정할 때마다 이 파일을 실행하면
서버용 JSON 파일이 자동으로 갱신됩니다.

실행 방법:
    python 엑셀_json_변환.py
"""

import openpyxl
import json
import os
from datetime import datetime

# ── 경로 설정 ──────────────────────────────────────────────────
현재폴더 = os.path.dirname(os.path.abspath(__file__))
엑셀경로 = os.path.join(현재폴더, '정리.xlsx')
json경로 = os.path.join(현재폴더, 'immigration_guidelines_db_v2.json')


def 시트_딕셔너리_변환(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        row = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            row[h] = v if v is not None else ''
        rows.append(row)
    return rows


def 변환실행():
    print('=' * 50)
    print('  정리.xlsx → JSON 변환 시작')
    print('=' * 50)

    if not os.path.exists(엑셀경로):
        print(f'[오류] 파일을 찾을 수 없습니다: {엑셀경로}')
        print('정리.xlsx 파일을 이 폴더에 넣어주세요.')
        return False

    print(f'[1/4] 엑셀 파일 읽는 중...')
    wb = openpyxl.load_workbook(엑셀경로)

    print(f'[2/4] 데이터 추출 중...')
    master_rows = 시트_딕셔너리_변환(wb['MASTER_ROWS'])
    rules       = 시트_딕셔너리_변환(wb['RULES'])
    exceptions  = 시트_딕셔너리_변환(wb['EXCEPTIONS'])
    doc_dict    = 시트_딕셔너리_변환(wb['DOC_DICTIONARY'])
    search_keys = 시트_딕셔너리_변환(wb['SEARCH_KEYS'])
    legacy_map  = 시트_딕셔너리_변환(wb['LEGACY_UI_MAP'])

    # 검색 키를 각 업무 행에 병합
    sk_map = {}
    for sk in search_keys:
        rid = sk.get('row_id', '')
        if rid:
            sk_map.setdefault(rid, []).append({
                'key_type': sk.get('key_type', ''),
                'key_value': sk.get('key_value', ''),
            })
    for row in master_rows:
        row['search_keys'] = sk_map.get(row['row_id'], [])

    from collections import Counter
    업무유형통계 = dict(Counter(r['action_type'] for r in master_rows).most_common())

    print(f'[3/4] JSON 파일 생성 중...')
    output = {
        "설명": "출입국 업무관리 시스템 실무지침 데이터베이스",
        "버전": "2.0",
        "갱신일": datetime.now().strftime('%Y-%m-%d %H:%M'),
        "통계": {
            "업무항목": len(master_rows),
            "공통규칙": len(rules),
            "예외조건": len(exceptions),
            "서류명사전": len(doc_dict),
            "검색인덱스": len(search_keys),
            "메뉴매핑": len(legacy_map),
            "업무유형분포": 업무유형통계,
        },
        "master_rows": master_rows,
        "rules":       rules,
        "exceptions":  exceptions,
        "doc_dictionary": doc_dict,
        "legacy_ui_map":  legacy_map,
    }

    with open(json경로, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    크기 = os.path.getsize(json경로) / 1024
    print(f'[4/4] 저장 완료')
    print()
    print(f'  저장 위치: {json경로}')
    print(f'  파일 크기: {크기:.1f} KB')
    print(f'  업무 항목: {len(master_rows)}건')
    print(f'  갱신 일시: {output["갱신일"]}')
    print()
    print('변환 완료. 서버를 재시작하면 새 데이터가 적용됩니다.')
    return True


if __name__ == '__main__':
    변환실행()
    input('\n엔터를 누르면 창이 닫힙니다...')
