"""
v3 적용 후 manual_ref 이상치 감사 (READ-ONLY, REPORT-ONLY)

5가지 패턴 검출:
  1. SAME_PAGE_CLUSTER       — 다른 detailed_code/action_type가 같은 manual_ref 공유 (size>=3)
  2. DUAL_MANUAL_SUSPICIOUS  — manual_ref 가 체류+사증 동시 포함, action_type이 dual-purpose 아님
  3. WRONG_PRIMARY_MANUAL    — action_type 의 preferred 매뉴얼이 빠짐
  4. BROAD_FAMILY_PAGE       — 같은 자격 family(F-1/F-2/F-3/F-4/F-5/D-8) 내 다른 subcode가 같은 페이지
  5. APPLIED_ROW_ONLY        — 변경 여부 표기 (current vs backup)

특별 규칙:
  사용자 명시 4건 (F-1-15/21/22/24) 은 HIGH priority 강제 surfacing.

쓰기 보장:
  DB·blocklist·triage·verify 어떤 파일도 변경하지 않음.
"""
from __future__ import annotations
import json, sys
from pathlib import Path
from collections import defaultdict

ROOT      = Path(__file__).parent.parent.parent
DB_PATH   = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUPS   = ROOT / "backend" / "data" / "backups"
TRIAGE    = ROOT / "backend" / "data" / "manuals" / "manual_mapping_triage_v3.json"
BLOCKLIST = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_blocklist_v3.json"

OUT_JSON  = ROOT / "backend" / "data" / "manuals" / "manual_ref_post_apply_anomalies_v3.json"
OUT_XLSX  = ROOT / "backend" / "data" / "manuals" / "manual_ref_post_apply_anomalies_v3.xlsx"

# 사용자 명시 — 정답 페이지 + HIGH priority
USER_SPECIFIED_HIGH = {
    ("F-1-15", "EXTEND"): {
        "title": "우수인재·투자자·유학생 부모 방문동거",
        "expected": [
            {"manual": "체류민원", "page_from": 543, "page_to": 543},
            {"manual": "사증민원", "page_from": 404, "page_to": 404},
        ],
    },
    ("F-1-21", "VISA_CONFIRM"): {
        "title": "주한 외국공관원 가사보조인",
        "expected": [
            {"manual": "체류민원", "page_from": 543, "page_to": 543},
            {"manual": "사증민원", "page_from": 404, "page_to": 404},
        ],
    },
    ("F-1-22", "EXTEND"): {
        "title": "고액투자가 가사보조인 방문동거",
        "expected": [
            {"manual": "체류민원", "page_from": 543, "page_to": 543},
            {"manual": "사증민원", "page_from": 404, "page_to": 404},
        ],
    },
    ("F-1-24", "EXTEND"): {
        "title": "해외우수인재 가사보조인 방문동거",
        "expected": [
            {"manual": "체류민원", "page_from": 543, "page_to": 543},
            {"manual": "사증민원", "page_from": 404, "page_to": 404},
        ],
    },
}

VISA_PREF = {"VISA_CONFIRM"}
RES_PREF = {"CHANGE", "EXTEND", "REGISTRATION", "REENTRY",
            "EXTRA_WORK", "GRANT", "WORKPLACE",
            "ACTIVITY_EXTRA", "DOMESTIC_RESIDENCE_REPORT"}
BROAD_FAMILIES = {"F-1", "F-2", "F-3", "F-4", "F-5", "D-8"}


def manual_ref_sig(refs: list) -> tuple:
    if not refs:
        return ()
    return tuple(sorted(
        (r.get("manual", ""), r.get("page_from", 0), r.get("page_to", 0))
        for r in refs
    ))


def family_of(code: str) -> str:
    parts = (code or "").split("-")
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return code or ""


def manuals_in(refs: list) -> set:
    return {r.get("manual", "") for r in (refs or [])}


def find_latest_backup() -> Path | None:
    if not BACKUPS.exists():
        return None
    bs = sorted(BACKUPS.glob("immigration_guidelines_db_v2.manual_ref_backup_*.json"))
    return bs[-1] if bs else None


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not DB_PATH.exists():
        print(f"[ABORT] DB 없음: {DB_PATH}"); sys.exit(1)
    backup = find_latest_backup()
    if not backup:
        print(f"[ABORT] backup 없음"); sys.exit(1)

    print("[anomaly] DB 로드...")
    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rows = db.get("master_rows") or []
    cur_index = {r.get("row_id"): r for r in rows}

    print(f"[anomaly] backup 로드: {backup.relative_to(ROOT)}")
    backup_db = json.loads(backup.read_text(encoding="utf-8"))
    bk_index = {r.get("row_id"): r for r in (backup_db.get("master_rows") or [])}

    print("[anomaly] triage / blocklist 로드...")
    triage = json.loads(TRIAGE.read_text(encoding="utf-8"))
    triage_index = {t.get("row_id"): t for t in (triage.get("rows") or [])}
    blocklist = json.loads(BLOCKLIST.read_text(encoding="utf-8"))
    blocked_ids = set(blocklist.get("blocked_row_ids") or [])

    # ── 변경된 row 식별 (current vs backup) ──
    changed_in_apply: set[str] = set()
    for rid, cur in cur_index.items():
        bk = bk_index.get(rid)
        if bk is None:
            continue
        if cur.get("manual_ref") != bk.get("manual_ref"):
            changed_in_apply.add(rid)
    print(f"  total rows: {len(rows)}")
    print(f"  changed_in_apply (DB vs backup): {len(changed_in_apply)}")

    # ── 1. SAME_PAGE_CLUSTER ──
    sig_groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        sig = manual_ref_sig(r.get("manual_ref"))
        if sig:
            sig_groups[sig].append(r)

    same_page_clusters = []
    for sig, group in sig_groups.items():
        if len(group) < 3:
            continue
        codes = {r.get("detailed_code") for r in group}
        ats = {r.get("action_type") for r in group}
        if len(codes) >= 2 or len(ats) >= 2:
            same_page_clusters.append((sig, group))

    # ── 2. DUAL_MANUAL_SUSPICIOUS ──
    dual_suspicious: set[str] = set()
    for rid, r in cur_index.items():
        ms = manuals_in(r.get("manual_ref"))
        if "체류민원" in ms and "사증민원" in ms:
            dual_suspicious.add(rid)

    # ── 3. WRONG_PRIMARY_MANUAL ──
    wrong_primary: set[str] = set()
    for rid, r in cur_index.items():
        ms = manuals_in(r.get("manual_ref"))
        if not ms:
            continue
        at = r.get("action_type")
        if at in VISA_PREF and "사증민원" not in ms:
            wrong_primary.add(rid)
        elif at in RES_PREF and "체류민원" not in ms:
            wrong_primary.add(rid)

    # ── 4. BROAD_FAMILY_PAGE ──
    family_sig_groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        fam = family_of(r.get("detailed_code"))
        if fam not in BROAD_FAMILIES:
            continue
        sig = manual_ref_sig(r.get("manual_ref"))
        if sig:
            family_sig_groups[(fam, sig)].append(r)

    broad_family_clusters = []
    for (fam, sig), group in family_sig_groups.items():
        if len(group) >= 2:
            codes = {r.get("detailed_code") for r in group}
            if len(codes) >= 2:
                broad_family_clusters.append((fam, sig, group))

    # ── 빌드: row별 anomaly 누적 ──
    row_anomalies: dict[str, dict] = defaultdict(lambda: {
        "anomaly_types": set(),
        "cluster_keys":  [],
        "cluster_size":  0,
        "reasons":       [],
    })

    for sig, group in same_page_clusters:
        sig_str = json.dumps([list(s) for s in sig], ensure_ascii=False)
        cluster_key = f"signature:{sig_str}"
        members_codes = sorted({r.get("detailed_code") for r in group})
        members_ats = sorted({r.get("action_type") for r in group})
        for r in group:
            rid = r["row_id"]
            row_anomalies[rid]["anomaly_types"].add("SAME_PAGE_CLUSTER")
            row_anomalies[rid]["cluster_keys"].append(cluster_key)
            row_anomalies[rid]["cluster_size"] = max(row_anomalies[rid]["cluster_size"], len(group))
            row_anomalies[rid]["reasons"].append(
                f"동일 manual_ref 공유 클러스터 size={len(group)}, codes={members_codes}, actions={members_ats}"
            )

    for rid in dual_suspicious:
        cur = cur_index[rid]
        row_anomalies[rid]["anomaly_types"].add("DUAL_MANUAL_SUSPICIOUS")
        row_anomalies[rid]["reasons"].append(
            f"manual_ref 가 체류·사증 양쪽 포함 (action_type={cur.get('action_type')})"
        )

    for rid in wrong_primary:
        cur = cur_index[rid]
        ms = manuals_in(cur.get("manual_ref"))
        row_anomalies[rid]["anomaly_types"].add("WRONG_PRIMARY_MANUAL")
        row_anomalies[rid]["reasons"].append(
            f"action_type={cur.get('action_type')} 의 preferred 매뉴얼이 manual_ref 에 없음 (현재: {sorted(ms)})"
        )

    for fam, sig, group in broad_family_clusters:
        sig_str = json.dumps([list(s) for s in sig], ensure_ascii=False)
        cluster_key = f"family:{fam}|sig:{sig_str}"
        member_codes = sorted({r.get("detailed_code") for r in group})
        for r in group:
            rid = r["row_id"]
            row_anomalies[rid]["anomaly_types"].add("BROAD_FAMILY_PAGE")
            row_anomalies[rid]["cluster_keys"].append(cluster_key)
            row_anomalies[rid]["reasons"].append(
                f"family={fam} 내 {len(group)}개 다른 subcode가 같은 페이지 (codes={member_codes})"
            )

    # 사용자 명시 HIGH 4건 — 단, 이미 manual_override 로 패치된 경우 skip (stale flag 방지)
    user_specified_matches = {}  # rid -> spec
    for (code, at), spec in USER_SPECIFIED_HIGH.items():
        for r in rows:
            if r.get("detailed_code") == code and r.get("action_type") == at:
                rid = r["row_id"]
                cur_ref = r.get("manual_ref") or []
                # If already patched via manual_override → SKIP (no longer stale)
                already_patched = any(e.get("match_type") == "manual_override" for e in cur_ref)
                if already_patched:
                    continue
                user_specified_matches[rid] = spec
                row_anomalies[rid]["anomaly_types"].add("USER_SPECIFIED_HIGH")
                row_anomalies[rid]["reasons"].append(
                    f"사용자 검증: 정답 페이지는 체류 p.543 + 사증 p.404 / 현재 DB: "
                    + (json.dumps(cur_ref, ensure_ascii=False) if cur_ref else "비어있음")
                )
                break

    # ── recommended_action 결정 ──
    def decide(rid: str) -> tuple[str, str]:
        ans = row_anomalies[rid]["anomaly_types"]
        size = row_anomalies[rid]["cluster_size"]

        # 사용자 명시 → CREATE_MANUAL_OVERRIDE + HIGH
        if "USER_SPECIFIED_HIGH" in ans:
            return "CREATE_MANUAL_OVERRIDE", "HIGH"

        # 큰 same-page 클러스터(size>=4) → MANUAL_REVIEW HIGH
        if "SAME_PAGE_CLUSTER" in ans and size >= 4:
            return "MANUAL_REVIEW", "HIGH"

        # 작은 same-page 클러스터(size==3) → MANUAL_REVIEW MEDIUM
        if "SAME_PAGE_CLUSTER" in ans:
            return "MANUAL_REVIEW", "MEDIUM"

        if "WRONG_PRIMARY_MANUAL" in ans:
            return "MANUAL_REVIEW", "MEDIUM"

        if "DUAL_MANUAL_SUSPICIOUS" in ans:
            return "MANUAL_REVIEW", "MEDIUM"

        if "BROAD_FAMILY_PAGE" in ans:
            return "MANUAL_REVIEW", "MEDIUM"

        return "KEEP", "LOW"

    # ── 결과 ──
    anomaly_rows = []
    for rid, info in row_anomalies.items():
        cur = cur_index.get(rid)
        if not cur:
            continue
        bk = bk_index.get(rid, {})
        action, priority = decide(rid)
        was_changed = rid in changed_in_apply
        anomaly_rows.append({
            "row_id":              rid,
            "detailed_code":       cur.get("detailed_code"),
            "action_type":         cur.get("action_type"),
            "title":               cur.get("business_name"),
            "current_manual_ref":  cur.get("manual_ref") or [],
            "backup_manual_ref":   bk.get("manual_ref") or [],
            "was_changed_in_apply": was_changed,
            "anomaly_types":       sorted(info["anomaly_types"]),
            "cluster_key":         info["cluster_keys"][0] if info["cluster_keys"] else "",
            "cluster_size":        info["cluster_size"],
            "reason":              " // ".join(info["reasons"]),
            "recommended_action":  action,
            "priority":            priority,
            "is_blocked":          rid in blocked_ids,
            "expected_from_user":  user_specified_matches.get(rid, {}).get("expected"),
        })

    # 정렬: HIGH 먼저, 그 다음 was_changed_in_apply DESC, 그 다음 cluster_size DESC, row_id
    pri_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    anomaly_rows.sort(key=lambda x: (
        pri_order.get(x["priority"], 9),
        not x["was_changed_in_apply"],
        -x["cluster_size"],
        x["row_id"] or "",
    ))

    # ── 통계 ──
    total_anomalies = len(anomaly_rows)
    applied_anomalies = sum(1 for x in anomaly_rows if x["was_changed_in_apply"])
    same_cluster_count = sum(1 for x in anomaly_rows if "SAME_PAGE_CLUSTER" in x["anomaly_types"])
    dual_count = sum(1 for x in anomaly_rows if "DUAL_MANUAL_SUSPICIOUS" in x["anomaly_types"])
    wrong_count = sum(1 for x in anomaly_rows if "WRONG_PRIMARY_MANUAL" in x["anomaly_types"])
    broad_count = sum(1 for x in anomaly_rows if "BROAD_FAMILY_PAGE" in x["anomaly_types"])
    high_count = sum(1 for x in anomaly_rows if x["priority"] == "HIGH")
    restore_rows = [x for x in anomaly_rows if x["recommended_action"] == "RESTORE_FROM_BACKUP"]
    review_rows  = [x for x in anomaly_rows if x["recommended_action"] == "MANUAL_REVIEW"]
    override_rows = [x for x in anomaly_rows if x["recommended_action"] == "CREATE_MANUAL_OVERRIDE"]

    summary = {
        "total_rows_scanned":                len(rows),
        "total_anomalies":                   total_anomalies,
        "anomalies_in_applied_139":          applied_anomalies,
        "same_page_cluster_count":           same_cluster_count,
        "dual_manual_suspicious_count":      dual_count,
        "wrong_primary_manual_count":        wrong_count,
        "broad_family_page_count":           broad_count,
        "high_priority_count":               high_count,
        "recommended_KEEP":                  sum(1 for x in anomaly_rows if x["recommended_action"] == "KEEP"),
        "recommended_MANUAL_REVIEW":         len(review_rows),
        "recommended_RESTORE_FROM_BACKUP":   len(restore_rows),
        "recommended_CREATE_MANUAL_OVERRIDE": len(override_rows),
    }

    out = {
        "anomaly_meta": {
            "version":     "v3",
            "db_path":     str(DB_PATH.relative_to(ROOT)),
            "backup_path": str(backup.relative_to(ROOT)),
            "rules": {
                "SAME_PAGE_CLUSTER":      "size>=3 + 다양한 detailed_code or action_type",
                "DUAL_MANUAL_SUSPICIOUS": "manual_ref 가 체류+사증 동시 포함",
                "WRONG_PRIMARY_MANUAL":   "action_type 의 preferred 매뉴얼이 manual_ref 에 없음",
                "BROAD_FAMILY_PAGE":      "F-1/F-2/F-3/F-4/F-5/D-8 family 내 다른 subcode가 같은 페이지",
                "USER_SPECIFIED_HIGH":    "사용자 명시 4건 (F-1-15/21/22/24) — 강제 HIGH",
            },
        },
        "summary": summary,
        "rows":    anomaly_rows,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # summary
        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, v])
        ws_s.column_dimensions["A"].width = 38
        ws_s.column_dimensions["B"].width = 16

        # anomalies
        headers = [
            "row_id","detailed_code","action_type","title",
            "priority","recommended_action","anomaly_types",
            "was_changed_in_apply","is_blocked","cluster_size",
            "current_manual_ref","backup_manual_ref",
            "expected_from_user","reason","cluster_key",
        ]
        widths = [10,12,14,28,8,26,30,12,8,8,40,40,40,80,40]

        def write_sheet(name: str, items: list[dict]):
            ws = wb.create_sheet(name)
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for x in items:
                ws.append([
                    x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                    x["priority"], x["recommended_action"],
                    ", ".join(x["anomaly_types"]),
                    x["was_changed_in_apply"], x["is_blocked"], x["cluster_size"],
                    json.dumps(x["current_manual_ref"], ensure_ascii=False)[:200],
                    json.dumps(x["backup_manual_ref"], ensure_ascii=False)[:200],
                    json.dumps(x.get("expected_from_user") or [], ensure_ascii=False)[:200],
                    x["reason"][:1000],
                    x["cluster_key"][:200],
                ])
                color = {"HIGH":"FFCCCB","MEDIUM":"FFE599","LOW":"C6F6D5"}.get(x["priority"])
                if color:
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill("solid", fgColor=color)
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

        write_sheet("anomalies_all", anomaly_rows)
        write_sheet("HIGH",       [x for x in anomaly_rows if x["priority"] == "HIGH"])
        write_sheet("MEDIUM",     [x for x in anomaly_rows if x["priority"] == "MEDIUM"])
        write_sheet("applied_139", [x for x in anomaly_rows if x["was_changed_in_apply"]])
        write_sheet("user_specified", [x for x in anomaly_rows if "USER_SPECIFIED_HIGH" in x["anomaly_types"]])

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # ── 콘솔 ──
    print("\n" + "=" * 78)
    print("POST-APPLY ANOMALY REPORT")
    print("=" * 78)
    print(f"  total rows scanned:              {summary['total_rows_scanned']}")
    print(f"  total anomalies:                 {summary['total_anomalies']}")
    print(f"  anomalies in applied 139:        {summary['anomalies_in_applied_139']}")
    print(f"  same-page cluster count:         {summary['same_page_cluster_count']}")
    print(f"  dual-manual suspicious:          {summary['dual_manual_suspicious_count']}")
    print(f"  wrong-primary-manual:            {summary['wrong_primary_manual_count']}")
    print(f"  broad-family-page:               {summary['broad_family_page_count']}")
    print(f"  HIGH priority:                   {summary['high_priority_count']}")
    print(f"\n  recommend KEEP:                  {summary['recommended_KEEP']}")
    print(f"  recommend MANUAL_REVIEW:         {summary['recommended_MANUAL_REVIEW']}")
    print(f"  recommend RESTORE_FROM_BACKUP:   {summary['recommended_RESTORE_FROM_BACKUP']}")
    print(f"  recommend CREATE_MANUAL_OVERRIDE:{summary['recommended_CREATE_MANUAL_OVERRIDE']}")

    print("\n=== HIGH priority rows ===")
    for x in [x for x in anomaly_rows if x["priority"] == "HIGH"]:
        print(
            f"  {x['row_id']:8s} {x['detailed_code']:8s} {x['action_type']:13s} "
            f"changed={x['was_changed_in_apply']!s:5s} "
            f"action={x['recommended_action']:24s} "
            f"types={','.join(x['anomaly_types'])}"
        )
        print(f"           current: {json.dumps(x['current_manual_ref'], ensure_ascii=False)}")
        if x.get("expected_from_user"):
            print(f"           expected: {json.dumps(x['expected_from_user'], ensure_ascii=False)}")

    if restore_rows:
        print("\n=== RESTORE_FROM_BACKUP recommended ===")
        for x in restore_rows:
            print(f"  {x['row_id']:8s} {x['detailed_code']:8s} {x['action_type']:13s}  {x['title']}")

    if override_rows:
        print("\n=== CREATE_MANUAL_OVERRIDE recommended ===")
        for x in override_rows:
            print(f"  {x['row_id']:8s} {x['detailed_code']:8s} {x['action_type']:13s}  {x['title']}")

    print("\n=== MANUAL_REVIEW (top 30) ===")
    for x in review_rows[:30]:
        print(
            f"  {x['row_id']:8s} {x['detailed_code']:8s} {x['action_type']:13s} "
            f"pri={x['priority']:6s} cluster_size={x['cluster_size']:>2}  "
            f"types={','.join(x['anomaly_types'])}"
        )

    print(f"\n=== Production DB modified or not ===")
    print(f"  No — DB / blocklist / triage / verify 모두 무수정 (이번 단계는 report-only)")


if __name__ == "__main__":
    main()
