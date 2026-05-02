"""
Stage 1 — Quality Gate v4 (READ-ONLY, NO LLM, NO API CALL)

Classifies all 369 master_rows by manual_ref quality risk.
Identifies which rows need LLM judge (next stage) vs manual review.

Inputs (read-only):
  backend/data/immigration_guidelines_db_v2.json
  backend/data/manuals/manual_mapping_triage_v3.json
  backend/data/manuals/manual_mapping_apply_blocklist_v3.json
  backend/data/manuals/manual_ref_post_apply_anomalies_v3.json
  backend/data/manuals/f1_manual_ref_validation_v3.json
  backend/data/manuals/manual_mapping_extra_work_review_v3.json
  backend/data/manuals/unlocked_체류민원.pdf
  backend/data/manuals/unlocked_사증민원.pdf

Outputs:
  backend/data/manuals/manual_ref_quality_gate_v4.json
  backend/data/manuals/manual_ref_quality_gate_v4.xlsx

ABSOLUTE PROHIBITIONS:
  - Never modifies the DB or any other input file.
  - Never calls any LLM API.
  - Never connects to apply scripts.
"""
from __future__ import annotations
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import fitz  # PyMuPDF

ROOT     = Path(__file__).parent.parent.parent
DATA     = ROOT / "backend" / "data"
MANUALS  = DATA / "manuals"
DB_PATH  = DATA / "immigration_guidelines_db_v2.json"

PDF_RES  = MANUALS / "unlocked_체류민원.pdf"
PDF_VIS  = MANUALS / "unlocked_사증민원.pdf"

TRIAGE   = MANUALS / "manual_mapping_triage_v3.json"
BLOCK    = MANUALS / "manual_mapping_apply_blocklist_v3.json"
ANOM     = MANUALS / "manual_ref_post_apply_anomalies_v3.json"
F1VAL    = MANUALS / "f1_manual_ref_validation_v3.json"
EXTRA    = MANUALS / "manual_mapping_extra_work_review_v3.json"

OUT_JSON = MANUALS / "manual_ref_quality_gate_v4.json"
OUT_XLSX = MANUALS / "manual_ref_quality_gate_v4.xlsx"

# ───────────────────────────────────────────────────────────────
# Confirmed-correct overrides — manually verified, heading heuristic miss
# These rows are correct mappings; rule2 fails because the chapter heading
# uses a non-bracket title-box format or the DB code differs from the
# manual's bracket code (e.g. F-5-2 DB → 〔F-5-4〕 manual).
# Adding a row here forces heading_match → FOUND_AS_HEADING and removes
# all heading-related risk BEFORE scoring. Never add without direct PDF
# verification.
# ───────────────────────────────────────────────────────────────
CONFIRMED_CORRECT_OVERRIDES: dict[str, str] = {
    "M1-0114": "LLM judge + 수동 검증 완료 — D-8-1 법인투자 VISA_CONFIRM 사증 p.108-109",
    "M1-0134": "LLM judge + verifier 검증 완료 — D-10-2 EXTEND 체류 p.155",
    "M1-0199": "수동 검증 완료 — E-10 VISA_CONFIRM 사증 p.294-296",
    "M1-0284": "LLM judge + verifier 검증 완료 — F-3-3R CHANGE 체류 p.614",
    "M1-0365": "수동 검증 완료 — F-5-2 CHANGE 체류 p.453-454",
    "M1-0366": "수동 검증 완료 — F-5-6 CHANGE 체류 p.447+p.449",
    "M1-0320": "수동 검증 — E-7-S CHANGE 체류민원 p.291 통합 체류관리 섹션",
    "M1-0321": "수동 검증 — E-7-S EXTEND 체류민원 p.291 통합 체류관리 섹션",
    "M1-0324": "수동 검증 — E-7-S GRANT 체류민원 p.291 통합 체류관리 섹션",
    "M1-0159": "LLM verifier ACCEPTED — E-4 REENTRY 체류민원 p.197 (재입국허가 섹션, 테이블셀 헤더라 top30 미검출)",
    "M1-0175": "수동 검증 — E-7-4 CHANGE 체류민원 p.296-297 숙련기능인력 K-point E74 체류자격변경 섹션",
    "M1-0369": "수동 검증 — F-5-14 방문취업 H-2 영주 CHANGE 체류민원 p.539-540",
    "M1-0002": "수동 검증 — D-1 EXTRA_WORK 체류민원 p.32 테이블셀 헤더 미검출 (신청서류 4번항목 있음)",
    "M1-0312": "수동 검증 — E-7-2 REGISTRATION 체류민원 p.215 공통첨부서류+초청자요건 포함 (외국인등록 heading 없으나 동일 섹션)",
    "M1-0317": "수동 검증 — E-7-3 REGISTRATION 체류민원 p.215 공통첨부서류+초청자요건 포함",
    "M1-0329": "수동 검증 — E-7-T REGISTRATION 체류민원 p.215 공통첨부서류+초청자요건 포함",
    "M1-0125": "수동 검증 — D-9-1 VISA_CONFIRM 사증민원 p.114-121 한·인도협정 D-9 사증발급인정서 관련 (body_keyword_only 맞으나 D-9 관련 맞는 페이지)",
    # F-2 EXTEND: F-2 섹션 내 체류기간연장 언급은 body-only이나 해당 F-2 챕터가 정확
    "M1-0235": "수동 검증 — F-2-3 EXTEND 체류민원 p.384 F-2 영주자 배우자자녀 연장 (별도 heading 없이 CHANGE섹션에 통합)",
    "M1-0239": "수동 검증 — F-2-4 EXTEND 체류민원 p.384 F-2 난민인정자 연장 (CHANGE섹션 통합)",
    "M1-0247": "수동 검증 — F-2-5 EXTEND 체류민원 p.384 F-2 고액투자자 연장 (CHANGE섹션 통합)",
    "M1-0251": "수동 검증 — F-2-6 EXTEND 체류민원 p.384 F-2 숙련생산기능 연장 (CHANGE섹션 통합)",
    "M1-0092": "수동 검증 — F-2-71 EXTEND 체류민원 p.384 K-STAR 거주 배우자 연장 (CHANGE섹션 통합)",
    "M1-0255": "수동 검증 — F-2-71 EXTEND 체류민원 p.384 점수제 우수인재 배우자자녀 연장 (CHANGE섹션 통합)",
    "M1-0095": "수동 검증 — F-2-T EXTEND 체류민원 p.384 최우수인재 거주 연장 (CHANGE섹션 통합)",
    "M1-0243": "수동 검증 — F-2-7 EXTEND 체류민원 p.363 점수제 우수인재 연장 (CHANGE섹션 통합)",
    # F-2 REGISTRATION: p.362-373은 F-2 등록 관련 섹션의 정확한 위치
    "M1-0236": "수동 검증 — F-2-3 REGISTRATION 체류민원 p.362 F-2 영주자 배우자자녀 등록 (heading 없으나 해당 F-2 등록섹션)",
    "M1-0240": "수동 검증 — F-2-4 REGISTRATION 체류민원 p.362 F-2 난민인정자 등록",
    "M1-0248": "수동 검증 — F-2-5 REGISTRATION 체류민원 p.362 F-2 고액투자자 등록",
    "M1-0252": "수동 검증 — F-2-6 REGISTRATION 체류민원 p.373 F-2 숙련생산기능 등록",
    "M1-0244": "수동 검증 — F-2-7 REGISTRATION 체류민원 p.373 점수제 우수인재 등록",
    "M1-0258": "수동 검증 — F-2-99 EXTEND 체류민원 p.389 기타장기체류자 변경허가 섹션 내 연장 포함 body-only 정상",
    "M1-0282": "수동 검증 — F-2-R EXTEND 체류민원 p.586 지역특화형 섹션 내 연장 포함 body-only 정상",
    "M1-0262": "수동 검증 — F-4 EXTEND 체류민원 p.530 재외동포 체류기간연장허가 섹션 (PDF 직접 확인, footer p.10)",
    # E-7 VISA_CONFIRM — 사증민원 p.185 E-7 사증발급인정서 섹션 내, heading 없으나 관련 내용
    "M1-0171": "수동 검증 — E-7-1 전문인력 VISA_CONFIRM 사증민원 p.185 E-7 사증발급인정서 (heading heuristic miss)",
    "M1-0314": "수동 검증 — E-7-2 준전문인력 VISA_CONFIRM 사증민원 p.185 E-7 사증발급인정서",
    "M1-0319": "수동 검증 — E-7-3 일반기능인력 VISA_CONFIRM 사증민원 p.185 E-7 사증발급인정서",
    # F-1-11 — 방문동거 방문취업자 가족 CHANGE/EXTEND p.543-544 broad_family_page false positive
    "M1-0214": "수동 검증 — F-1-11 CHANGE 체류민원 p.543-544 방문취업자가족 방문동거 체류관리 (broad_family false positive)",
    "M1-0216": "수동 검증 — F-1-11 EXTEND 체류민원 p.543-544 방문취업자가족 방문동거 체류관리 (broad_family false positive)",
    # F-4 cluster — p.528(CHANGE) p.530(DOMESTIC) 공통 제출서류 있음, heading heuristic miss
    "M1-0007": "수동 검증 — F-4-19 CHANGE 체류민원 p.528 재외동포 자격부여 세부절차 변경 제출서류",
    "M1-0261": "수동 검증 — F-4 CHANGE 체류민원 p.528 재외동포 자격부여 세부절차 변경 제출서류",
    "M1-0263": "수동 검증 — F-4 DOMESTIC_RESIDENCE_REPORT 체류민원 p.530 거소신고 절차 (heading 정상)",
    "M1-0093": "수동 검증 — F-5-S1 K-STAR 영주 CHANGE 체류민원 p.740 (Ⅲ. K-STAR 영주 섹션 시작, 체류자격변경 heading heuristic miss)",
    "M1-0242": "수동 검증 — F-2-7 CHANGE 체류민원 p.291 E-7-S→F-2-7 점수제 특례 섹션 (heading body-only 정상)",
}


# ───────────────────────────────────────────────────────────────
# Rule 1 — PDF Text Normalization
# ───────────────────────────────────────────────────────────────
_NORMALIZE_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"체류자격\s*부\s*여"),                "체류자격부여"),
    (re.compile(r"근무처의?\s*변경[·\s]*추가"),         "근무처변경추가"),
    (re.compile(r"체류자격\s*외\s*활동"),               "체류자격외활동"),
    (re.compile(r"사증\s*발급\s*인정서"),               "사증발급인정서"),
    (re.compile(r"사증\s*발급\s*확인서"),               "사증발급확인서"),
    (re.compile(r"외국인\s*등록"),                       "외국인등록"),
    (re.compile(r"재입국\s*허가"),                       "재입국허가"),
    (re.compile(r"체류기간\s*연장"),                     "체류기간연장"),
    (re.compile(r"체류자격\s*변경"),                     "체류자격변경"),
]


def normalize_pdf_text(text: str) -> str:
    """모든 매칭 작업 전 적용. HWP→PDF 분할 아티팩트를 합쳐 단일 토큰으로."""
    if not text:
        return ""
    out = text
    for pat, repl in _NORMALIZE_PATTERNS:
        out = pat.sub(repl, out)
    # token-level whitespace collapse (단어 간 공백은 보존, 줄 내 다중 공백만 압축)
    out = re.sub(r"[ \t　]+", " ", out)
    return out


def _normalize_self_test() -> None:
    cases = [
        ("체류자격\n부\n여",   "체류자격부여"),
        ("체류자격 부 여",     "체류자격부여"),
        ("근무처의 변경\n추가", "근무처변경추가"),
        ("체류자격외 활동",     "체류자격외활동"),
        ("사증 발급 인정서",   "사증발급인정서"),
        ("외국인 등록",         "외국인등록"),
        ("재입국 허가",         "재입국허가"),
        ("체류기간 연장",       "체류기간연장"),
    ]
    for src, expected in cases:
        out = normalize_pdf_text(src)
        if expected not in out:
            raise AssertionError(f"normalize self-test FAIL: {src!r} → {out!r} (expected substring {expected!r})")


# ───────────────────────────────────────────────────────────────
# Action-type heading keyword map (post-normalization)
# ───────────────────────────────────────────────────────────────
ACTION_KEYWORDS: dict[str, list[str]] = {
    "EXTRA_WORK":                ["체류자격외활동", "시간제취업"],
    "EXTEND":                    ["체류기간연장", "연장허가"],
    "CHANGE":                    ["체류자격변경", "변경허가"],
    "REGISTRATION":              ["외국인등록"],
    "REENTRY":                   ["재입국허가"],
    "GRANT":                     ["체류자격부여", "자격부여"],
    "VISA_CONFIRM":              ["사증발급인정서", "사증발급확인서"],
    "WORKPLACE":                 ["근무처변경추가", "근무처변경", "근무처추가"],
    # 다음 두 action 은 매뉴얼에 단독 섹션이 거의 없음 → no-strict
    "ACTIVITY_EXTRA":            ["활동범위 확대", "단순노무 특례"],
    "DOMESTIC_RESIDENCE_REPORT": ["거소신고", "국내거소신고"],
    "APPLICATION_CLAIM":         ["사실증명", "직접신청"],
}

# 우선 매뉴얼 라우팅
VISA_PREF = {"VISA_CONFIRM"}
RES_PREF = {"CHANGE", "EXTEND", "REGISTRATION", "REENTRY",
            "EXTRA_WORK", "GRANT", "WORKPLACE",
            "ACTIVITY_EXTRA", "DOMESTIC_RESIDENCE_REPORT",
            "APPLICATION_CLAIM"}

BROAD_FAMILY_HEADINGS = [
    "외국국적동포가족", "동포가족", "동포 가족",
    "방문동거(F-1) 자격 기본", "방문동거 자격 기본",
    "동반(F-3) 자격 기본",
    "외국국적동포의 가족",
]


# ───────────────────────────────────────────────────────────────
# PDF page text cache + heading extraction
# ───────────────────────────────────────────────────────────────
class PdfPool:
    def __init__(self, pdf_paths: dict[str, Path]):
        self.docs: dict[str, fitz.Document] = {
            name: fitz.open(str(p)) for name, p in pdf_paths.items()
        }
        # cache key: (manual, page_no_1based) → {full_text, top30_text, short_lines}
        self._cache: dict[tuple[str, int], dict] = {}

    def page(self, manual: str, page_no: int) -> dict:
        key = (manual, page_no)
        if key in self._cache:
            return self._cache[key]
        doc = self.docs.get(manual)
        if doc is None or page_no < 1 or page_no > doc.page_count:
            self._cache[key] = {"full": "", "top30": "", "short_lines": []}
            return self._cache[key]
        page = doc.load_page(page_no - 1)
        h = page.rect.height
        # Use blocks to get bbox info
        blocks = page.get_text("blocks") or []
        full_raw = page.get_text() or ""
        top30_parts: list[str] = []
        short_lines: list[str] = []
        for b in blocks:
            # b = (x0, y0, x1, y1, text, block_no, type)
            try:
                x0, y0, x1, y1, t = b[0], b[1], b[2], b[3], b[4]
            except Exception:
                continue
            if not t:
                continue
            # 단독 짧은 줄 (heading 후보)
            for line in t.splitlines():
                line_strip = line.strip()
                if 0 < len(line_strip) <= 20:
                    short_lines.append(line_strip)
            if y0 / h <= 0.30:
                top30_parts.append(t)
        full_norm  = normalize_pdf_text(full_raw)
        top30_norm = normalize_pdf_text("\n".join(top30_parts))
        short_norm = [normalize_pdf_text(s) for s in short_lines]
        self._cache[key] = {
            "full": full_norm,
            "top30": top30_norm,
            "short_lines": short_norm,
        }
        return self._cache[key]


# ───────────────────────────────────────────────────────────────
# Validators
# ───────────────────────────────────────────────────────────────
# Generic application-procedure indicators per action_type group.
# If a page contains detailed_code/title AND any of these as heading-style markers,
# treat it as FOUND_AS_HEADING even when the literal action_keyword isn't a heading.
GENERIC_HEADING_INDICATORS_RES = [
    "신청요건", "신청 요건",
    "허가요건", "허가 요건",
    "허용대상", "허용 대상",
    "제출서류", "필수서류", "첨부서류",
]
GENERIC_HEADING_INDICATORS_VISA = [
    "발급대상", "발급 대상",
    "초청자 요건", "초청자요건",
    "첨부서류",
    "사증발급인정서",
]


def _has_indicator(text: str, indicators: list[str]) -> bool:
    return any(ind in text for ind in indicators)


def heading_match_for_action(
    page_data: dict,
    action_type: str,
    detailed_code: str = "",
    title: str = "",
) -> str:
    """FOUND_AS_HEADING | FOUND_IN_BODY_ONLY | NOT_FOUND

    Strict path: literal action_keyword in heading area or short line → FOUND_AS_HEADING.
    Relaxed path: detailed_code OR title-token on page AND a generic procedure indicator
    (신청요건/제출서류/허가요건 for residence; 발급대상/첨부서류 for visa) appears anywhere
    → FOUND_AS_HEADING. Catches cases where the chapter heading at the top is the
    sub-title (e.g. "5년 이내 지역특화 숙련기능인력...") instead of the literal action keyword.
    """
    keywords = ACTION_KEYWORDS.get(action_type, [])
    full = page_data["full"]
    top30 = page_data["top30"]
    short_lines = page_data["short_lines"]

    # Strict heading match — original behavior
    for kw in keywords or []:
        kw_n = normalize_pdf_text(kw)
        if kw_n in top30:
            return "FOUND_AS_HEADING"
        for s in short_lines:
            if kw_n in s:
                return "FOUND_AS_HEADING"

    # Relaxed heading match — generic procedure indicators + detailed_code/title presence
    if action_type in VISA_PREF:
        indicators = GENERIC_HEADING_INDICATORS_VISA
    else:
        indicators = GENERIC_HEADING_INDICATORS_RES

    has_indicator_in_full = _has_indicator(full, indicators)
    if has_indicator_in_full:
        # Check if the row's identity is on the page — detailed_code or 4-char+ title token
        identity_on_page = False
        if detailed_code and detailed_code in full:
            identity_on_page = True
        elif title:
            tokens = re.findall(r"[가-힣]{4,}", title)
            distinguishing = [t for t in tokens if t not in _GENERIC_TITLE_TOKENS]
            if distinguishing and any(t in full for t in distinguishing):
                identity_on_page = True
        if identity_on_page:
            return "FOUND_AS_HEADING"

    # Body-keyword fallback — original behavior
    for kw in keywords or []:
        kw_n = normalize_pdf_text(kw)
        if kw_n in full:
            return "FOUND_IN_BODY_ONLY"
    return "NOT_FOUND"


# 제목에 자주 등장하지만 broad family 페이지에서도 흔히 나오는 일반어 (구별 불가)
_GENERIC_TITLE_TOKENS = {
    "방문동거", "동반자격", "체류자격", "동포가족", "외국국적동포",
    "체류기간", "자격변경", "외국인등록", "재입국허가", "신청서류",
    "체류관리", "사증발급", "기본대상", "세부절차", "활동범위",
    "체류자격외활동",
}


def broad_family_check(page_data: dict, detailed_code: str, title: str) -> bool:
    """heading 영역에 broad family marker + 구별 가능한 자세한 코드/제목 토큰 부재 → True."""
    top30 = page_data["top30"]
    full = page_data["full"]
    has_broad_marker = any(m in top30 for m in BROAD_FAMILY_HEADINGS)
    if not has_broad_marker:
        return False
    # 1) detailed_code 가 페이지 본문에 등장하면 broad 아님
    #    (sub-code 가 명시적으로 헤딩에 박혀있는 정상 매핑)
    if detailed_code and detailed_code in full:
        return False
    # 2) 제목의 *구별 가능* 토큰 (일반 가족동거 어휘 제외) 이 페이지에 있으면 broad 아님
    if title:
        title_tokens = re.findall(r"[가-힣]{4,}", title)
        distinguishing = [t for t in title_tokens if t not in _GENERIC_TITLE_TOKENS]
        for tok in distinguishing:
            if tok in full:
                return False
    return True


def manuals_in(refs: list) -> set[str]:
    return {r.get("manual", "") for r in (refs or []) if r.get("manual")}


def get_preferred_manual(action_type: str) -> str:
    if action_type in VISA_PREF:
        return "사증민원"
    if action_type in RES_PREF:
        return "체류민원"
    return "체류민원"  # default


def preferred_validator(refs: list, action_type: str) -> tuple[bool, int, str]:
    """
    Return: (preferred_manual_ok, risk, risk_type)
    """
    pref = get_preferred_manual(action_type)
    ms = manuals_in(refs)
    if not ms:
        return False, 40, "preferred_manual_missing"
    if pref not in ms:
        return False, 40, "preferred_manual_missing"

    pref_entries = [e for e in (refs or []) if e.get("manual") == pref]
    non_pref_entries = [e for e in (refs or []) if e.get("manual") and e.get("manual") != pref]
    if not non_pref_entries:
        return True, 0, ""

    # preferred 가 manual_override 이면 사람-검증된 것이므로 non-preferred 동시 존재 OK
    pref_types = {e.get("match_type") for e in pref_entries}
    if "manual_override" in pref_types:
        return True, 0, ""

    # 둘 다 가지고 있고 preferred 이 권위적이지 않을 때
    non_pref_types = {e.get("match_type") for e in non_pref_entries}
    pref_scores = [e.get("score") for e in pref_entries if e.get("score") is not None]
    non_pref_scores = [e.get("score") for e in non_pref_entries if e.get("score") is not None]
    same_kind = pref_types == non_pref_types
    same_score = (
        bool(pref_scores) and bool(non_pref_scores)
        and max(pref_scores) == max(non_pref_scores)
    )
    # 두 매뉴얼 모두 section_only & 점수 동등 → 동등 상태
    if same_kind and same_score and pref_types == {"section_only"}:
        return True, 20, "dual_manual_equal"
    # 동일 match_type + 점수 동등 → 사실상 동등
    if same_kind and same_score:
        return True, 20, "dual_manual_equal"
    # match_type 구분이 없는 동시 존재 → 30
    return True, 30, "non_preferred_no_distinction"


def manual_ref_signature(refs: list) -> tuple:
    if not refs:
        return ()
    return tuple(sorted(
        (r.get("manual", ""), r.get("page_from", 0), r.get("page_to", 0))
        for r in refs
    ))


# ───────────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────────
def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    print("[gate] normalize self-test...")
    _normalize_self_test()
    print("  PASS")

    if not DB_PATH.exists():
        print(f"[ABORT] DB 없음: {DB_PATH}", file=sys.stderr)
        return 1

    print("[gate] DB / 보조 파일 로드...")
    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rows = db.get("master_rows") or []
    print(f"  rows: {len(rows)}")

    blocklist = set(json.loads(BLOCK.read_text(encoding="utf-8")).get("blocked_row_ids") or [])
    anom_doc = json.loads(ANOM.read_text(encoding="utf-8"))
    stale_ids: set[str] = set()
    for r in (anom_doc.get("rows") or []):
        types = r.get("anomaly_types") or []
        if "USER_SPECIFIED_HIGH" in types:
            stale_ids.add(r.get("row_id"))

    # ── PDF pool ──
    print("[gate] PDF pool 초기화...")
    pool = PdfPool({"체류민원": PDF_RES, "사증민원": PDF_VIS})

    # ── primary entry per row ──
    def primary_entry(row: dict) -> dict | None:
        refs = row.get("manual_ref") or []
        if not refs:
            return None
        pref = get_preferred_manual(row.get("action_type", ""))
        # 1) preferred + manual_override
        for e in refs:
            if e.get("manual") == pref and e.get("match_type") == "manual_override":
                return e
        # 2) any preferred
        for e in refs:
            if e.get("manual") == pref:
                return e
        # 3) first
        return refs[0]

    # ── cluster sizes ──
    sig_groups: dict[tuple, list[str]] = defaultdict(list)
    for r in rows:
        sig = manual_ref_signature(r.get("manual_ref"))
        if sig:
            sig_groups[sig].append(r.get("row_id"))

    out_rows: list[dict] = []
    for r in rows:
        rid = r.get("row_id")
        code = r.get("detailed_code", "")
        action = r.get("action_type", "")
        title = r.get("business_name", "")
        refs = r.get("manual_ref") or []
        pref = get_preferred_manual(action)

        risk = 0
        types: list[str] = []

        # Rule 7 (BLOCKED check first; risk_score ignored if blocked)
        is_blocked = rid in blocklist

        # Rule 2 — heading match for action_type
        prim = primary_entry(r)
        heading_match = "NOT_FOUND"
        broad_family = False
        body_keyword_only = False
        # Confirmed-correct override — applied BEFORE risk scoring
        confirmed_correct = rid in CONFIRMED_CORRECT_OVERRIDES

        if prim:
            page_data = pool.page(prim.get("manual", ""), int(prim.get("page_from") or 0))
            heading_match = heading_match_for_action(page_data, action, code, title)
            # Override forces FOUND_AS_HEADING; heading risk is never added
            if confirmed_correct:
                heading_match = "FOUND_AS_HEADING"
            elif heading_match == "FOUND_IN_BODY_ONLY":
                risk += 30
                types.append("heading_body_only(rule2)")
            elif heading_match == "NOT_FOUND":
                risk += 50
                types.append("heading_not_found(rule2)")

            # Rule 5 — broad family (skipped when confirmed_correct — mapping manually verified)
            broad_family = broad_family_check(page_data, code, title)
            if broad_family and not confirmed_correct:
                risk += 50
                types.append("broad_family_page(rule5)")

            # Rule 6 — body keyword only (overlap with rule2 body case is intentional per spec)
            # Skipped when confirmed_correct (heading already overridden to FOUND_AS_HEADING)
            if heading_match == "FOUND_IN_BODY_ONLY" and not confirmed_correct:
                body_keyword_only = True
                risk += 40
                types.append("body_keyword_only(rule6)")

        # Rule 3 — preferred manual
        pref_ok, pref_risk, pref_type = preferred_validator(refs, action)
        if pref_risk:
            risk += pref_risk
            types.append(f"{pref_type}(rule3)")

        # Rule 4 — same-page cluster
        sig = manual_ref_signature(refs)
        cluster_size = len(sig_groups.get(sig, [])) if sig else 0
        # Exception: if heading_match is FOUND_AS_HEADING for current row's action_type,
        # downgrade by 1 level (allowed per rule).
        downgrade_cluster = (heading_match == "FOUND_AS_HEADING")
        if cluster_size >= 5:
            level_risk = 40
            label = "same_page_cluster_large(rule4)"
            if downgrade_cluster:
                level_risk = 20
                label = "same_page_cluster_large(downgraded)(rule4)"
            risk += level_risk
            types.append(label)
        elif 3 <= cluster_size <= 4:
            level_risk = 20
            label = "same_page_cluster_small(rule4)"
            if downgrade_cluster:
                level_risk = 0
                label = "same_page_cluster_small(downgraded)(rule4)"
            risk += level_risk
            if level_risk > 0:
                types.append(label)
            else:
                types.append(label + ":waived")

        # Rule 7 — stale audit flag
        stale = rid in stale_ids
        if stale:
            risk += 15
            types.append("stale_audit_flag(rule7)")

        # quality_status
        if is_blocked:
            quality_status = "BLOCKED"
            recommended = "BLOCKED_NO_ACTION"
        elif risk == 0:
            quality_status = "PASS"
            recommended = "AUTO_SAFE"
        elif risk <= 29:
            quality_status = "WARN"
            recommended = "LLM_JUDGE_REQUIRED"
        elif risk <= 59:
            quality_status = "FAIL"
            recommended = "MANUAL_REVIEW"
        else:
            quality_status = "NEEDS_OVERRIDE"
            recommended = "MANUAL_REVIEW"

        out_rows.append({
            "row_id":                  rid,
            "detailed_code":           code,
            "action_type":             action,
            "title":                   title,
            "current_manual_ref":      refs,
            "quality_status":          quality_status,
            "risk_score":              risk if not is_blocked else 0,
            "risk_types":              types,
            "heading_match":           heading_match,
            "preferred_manual_ok":     pref_ok,
            "preferred_manual":        pref,
            "same_page_cluster_size":  cluster_size,
            "broad_family_page":       broad_family,
            "body_keyword_only":       body_keyword_only,
            "stale_flag":              stale,
            "recommended_next_action": recommended,
            "is_blocked":              is_blocked,
            "confirmed_correct":       confirmed_correct,
            "primary_entry":           prim,
        })

    # ── 통계 ──
    total = len(out_rows)
    counts = defaultdict(int)
    next_counts = defaultdict(int)
    flag_counts = defaultdict(int)
    for x in out_rows:
        counts[x["quality_status"]] += 1
        next_counts[x["recommended_next_action"]] += 1
        if x["broad_family_page"]:
            flag_counts["broad_family_page"] += 1
        if x["body_keyword_only"]:
            flag_counts["body_keyword_only"] += 1
        if x["stale_flag"]:
            flag_counts["stale_flag"] += 1
        if x["same_page_cluster_size"] >= 5:
            flag_counts["cluster_ge_5"] += 1

    summary = {
        "total":                  total,
        "PASS":                   counts.get("PASS", 0),
        "WARN":                   counts.get("WARN", 0),
        "FAIL":                   counts.get("FAIL", 0),
        "NEEDS_OVERRIDE":         counts.get("NEEDS_OVERRIDE", 0),
        "BLOCKED":                counts.get("BLOCKED", 0),
        "LLM_JUDGE_REQUIRED":     next_counts.get("LLM_JUDGE_REQUIRED", 0),
        "MANUAL_REVIEW":          next_counts.get("MANUAL_REVIEW", 0),
        "AUTO_SAFE":              next_counts.get("AUTO_SAFE", 0),
        "BLOCKED_NO_ACTION":      next_counts.get("BLOCKED_NO_ACTION", 0),
        "broad_family_page":      flag_counts.get("broad_family_page", 0),
        "body_keyword_only":      flag_counts.get("body_keyword_only", 0),
        "stale_flag":             flag_counts.get("stale_flag", 0),
        "cluster_ge_5":           flag_counts.get("cluster_ge_5", 0),
    }

    # ── Output JSON ──
    out_doc = {
        "meta": {
            "version": "v4",
            "stage":   "1_quality_gate",
            "db_path": str(DB_PATH.relative_to(ROOT)),
            "rules_applied": [
                "rule1_normalize_pdf_text",
                "rule2_action_heading_validator",
                "rule3_preferred_manual",
                "rule4_same_page_cluster",
                "rule5_broad_family_page",
                "rule6_body_keyword_only",
                "rule7_stale_audit_flag",
                "confirmed_correct_overrides",
            ],
            "confirmed_correct_count": len(CONFIRMED_CORRECT_OVERRIDES),
            "no_db_modification": True,
            "no_llm_call":        True,
        },
        "summary": summary,
        "rows":    out_rows,
    }
    OUT_JSON.write_text(json.dumps(out_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, v])
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 14

        ws = wb.create_sheet("quality_gate")
        headers = [
            "row_id", "detailed_code", "action_type", "title",
            "quality_status", "risk_score", "risk_types",
            "heading_match", "preferred_manual_ok",
            "same_page_cluster_size", "broad_family_page",
            "body_keyword_only", "stale_flag", "recommended_next_action",
        ]
        widths = [10, 12, 14, 32, 16, 10, 60, 22, 14, 10, 10, 10, 10, 22]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        color_map = {
            "NEEDS_OVERRIDE": "FFB3B3",
            "FAIL":           "FFCC99",
            "WARN":           "FFE599",
            "BLOCKED":        "CCCCCC",
            "PASS":           "FFFFFF",
        }
        for x in out_rows:
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                x["quality_status"], x["risk_score"], ", ".join(x["risk_types"]),
                x["heading_match"], x["preferred_manual_ok"],
                x["same_page_cluster_size"], x["broad_family_page"],
                x["body_keyword_only"], x["stale_flag"], x["recommended_next_action"],
            ])
            color = color_map.get(x["quality_status"], "FFFFFF")
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # ── stdout summary ──
    print("\n" + "=" * 60)
    print("MANUAL_REF QUALITY GATE v4 — STAGE 1 SUMMARY")
    print("=" * 60)
    print(f"  Total rows        : {summary['total']}")
    print(f"  PASS              : {summary['PASS']}")
    print(f"  WARN              : {summary['WARN']}")
    print(f"  FAIL              : {summary['FAIL']}")
    print(f"  NEEDS_OVERRIDE    : {summary['NEEDS_OVERRIDE']}")
    print(f"  BLOCKED           : {summary['BLOCKED']}")
    print("  ─────────────────────")
    print(f"  LLM_JUDGE_REQUIRED: {summary['LLM_JUDGE_REQUIRED']}   (= WARN count)")
    print(f"  MANUAL_REVIEW     : {summary['MANUAL_REVIEW']}   (= FAIL + NEEDS_OVERRIDE)")
    print("  ─────────────────────")
    print(f"  broad_family_page : {summary['broad_family_page']}")
    print(f"  body_keyword_only : {summary['body_keyword_only']}")
    print(f"  stale_flag        : {summary['stale_flag']}")
    print(f"  cluster ≥ 5       : {summary['cluster_ge_5']}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
