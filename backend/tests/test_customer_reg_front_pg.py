"""외국인등록번호 앞자리(reg_front) 선행 0 보존 — **실제 PostgreSQL** 통합 테스트.

``TEST_DATABASE_URL`` 설정 시에만 실행(미설정 → skip). 검증:
- 수동 API 로 ``"001010"`` / ``"1010"`` / int-형 저장 후 **DB raw 컬럼**이 ``"001010"``.
- 선행 0 이 손실된 레거시 행(raw ``"1010"``)을 직접 INSERT 한 뒤 ``find_customer`` 가
  읽기 방어선으로 ``"001010"`` 을 돌려주되 **DB 원문은 그대로**(비파괴)임을 확인.

실제 고객 원문은 사용하지 않고 합성 값만 쓴다.
"""
from __future__ import annotations

import os

import pytest
from cryptography.fernet import Fernet
from sqlalchemy import create_engine, select, text
from sqlalchemy.engine import make_url
from sqlalchemy.orm import Session, sessionmaker

os.environ.setdefault("CUSTOMER_PII_ENCRYPTION_KEY", Fernet.generate_key().decode())
os.environ.setdefault("PII_HASH_SECRET", "reg-front-it-secret")

TEST_DB_URL = os.environ.get("TEST_DATABASE_URL", "").strip()
pytestmark = pytest.mark.skipif(
    not TEST_DB_URL, reason="TEST_DATABASE_URL 미설정 — 실제 PostgreSQL 통합 테스트 skip")


def _ensure_database(url_str: str) -> None:
    url = make_url(url_str)
    admin_url = url.set(database="postgres")
    eng = create_engine(admin_url, isolation_level="AUTOCOMMIT", future=True)
    try:
        with eng.connect() as c:
            exists = c.execute(
                text("SELECT 1 FROM pg_database WHERE datname = :n"),
                {"n": url.database}).first()
            if not exists:
                c.execute(text(f'CREATE DATABASE "{url.database}"'))
    finally:
        eng.dispose()


@pytest.fixture(scope="module")
def engine():
    _ensure_database(TEST_DB_URL)
    eng = create_engine(TEST_DB_URL, future=True)
    from backend.db.base import Base
    from backend.db.models.customer import Customer
    from backend.db.models.tenant import Tenant
    tables = [Tenant.__table__, Customer.__table__]
    Base.metadata.drop_all(eng, tables=list(reversed(tables)))
    Base.metadata.create_all(eng, tables=tables)
    yield eng
    Base.metadata.drop_all(eng, tables=list(reversed(tables)))
    eng.dispose()


@pytest.fixture
def db(engine, monkeypatch):
    SessionLocal = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False, class_=Session)
    import backend.db.session as dbs
    monkeypatch.setattr(dbs, "get_sessionmaker", lambda: SessionLocal)
    monkeypatch.setattr(dbs, "is_configured", lambda: True)
    with engine.begin() as c:
        c.execute(text("TRUNCATE customers, tenants RESTART IDENTITY CASCADE"))
    # 테넌트 seed(FK).
    from backend.db.models.tenant import Tenant
    with SessionLocal() as s:
        s.add(Tenant(tenant_id="t-regfront", office_name="T", is_active=True))
        s.commit()
    return SessionLocal


def _raw_reg_front(db, customer_id):
    with db() as s:
        return s.scalar(text("SELECT reg_front FROM customers WHERE customer_id = :c")
                        .bindparams(c=customer_id))


def test_new_write_stores_valid_six_digits(db):
    from backend.services.customer_pg_service import create_customer

    out = create_customer("t-regfront", {"한글": "김합성", "등록증": "001010"})
    cid = out["고객ID"]
    assert out["등록증"] == "001010"
    assert _raw_reg_front(db, cid) == "001010"  # DB raw 6자리 저장


@pytest.mark.parametrize("bad", ["1010", "12345678", "991332", "abcdef"])
def test_new_write_rejects_invalid_reg_front(db, bad):
    # 웹/API 신규 등록 = 엄격: 잘못된 앞자리는 조용히 001010 으로 바꾸지 않고 거부.
    from backend.services.customer_pg_service import create_customer
    from backend.services.customer_identifier_normalize import RegFrontValidationError

    with pytest.raises(RegFrontValidationError):
        create_customer("t-regfront", {"한글": "불량", "등록증": bad})
    # 저장 자체가 일어나지 않음.
    with db() as s:
        cnt = s.scalar(text("SELECT count(*) FROM customers WHERE tenant_id='t-regfront'"))
    assert cnt == 0


def test_legacy_row_read_defense_non_destructive(db):
    from backend.services.customer_pg_service import find_customer

    # 방어선을 우회해 손상된 레거시 원문('1010')을 직접 INSERT.
    with db() as s:
        s.execute(text(
            "INSERT INTO customers (tenant_id, customer_id, korean_name, reg_front) "
            "VALUES (:t, :c, :n, :r)"
        ).bindparams(t="t-regfront", c="9001", n="이레거시", r="1010"))
        s.commit()

    got = find_customer("t-regfront", "9001")
    assert got["등록증"] == "001010"          # 읽기 방어선으로 복구
    assert _raw_reg_front(db, "9001") == "1010"  # DB 원문은 불변(비파괴)


# ── Excel 실제 왕복: import 양식 workbook → _read_rows → _row_to_customer → create → PG raw ──
def _import_workbook_bytes(reg_value, *, number_format=None):
    """import 양식(고객 시트, DATA_START_ROW=4, STD_KEYS 위치)으로 1행 workbook 생성.
    한글=col1(STD_KEYS[0]), 등록증=col6(STD_KEYS[5])."""
    import io as _io

    from openpyxl import Workbook

    from backend.services.customer_bulk_service import (
        DATA_START_ROW, HEADER_ROW, HEADERS, SHEET_NAME, STD_KEYS,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # 2행 헤더(파서의 헤더 정확 일치 검증 통과용).
    for _c, _h in enumerate(HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=_c, value=_h)
    name_col = STD_KEYS.index("한글") + 1
    reg_col = STD_KEYS.index("등록증") + 1
    ws.cell(row=DATA_START_ROW, column=name_col, value="엑셀합성")
    c = ws.cell(row=DATA_START_ROW, column=reg_col, value=reg_value)
    if number_format:
        c.number_format = number_format
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.parametrize("reg_value,number_format", [
    ("001010", None),        # Text 문자열
    (1010, None),            # 숫자 셀(선행 0 손실) → 복구
    (1010.0, None),          # float 셀 → 복구
    ("001010", "@"),         # Text(@) 서식
    (1010, "000000"),        # custom number format
])
def test_excel_roundtrip_recovers_to_canonical(db, reg_value, number_format):
    from backend.services import customer_bulk_service as bulk

    blob = _import_workbook_bytes(reg_value, number_format=number_format)
    res = bulk.commit(blob, "t-regfront", include_duplicates=True)
    assert res["registered"] == 1, res
    cid = res["new_customer_ids"][0]
    assert _raw_reg_front(db, cid) == "001010"          # PG raw 6자리
    from backend.services.customer_pg_service import find_customer
    assert find_customer("t-regfront", cid)["등록증"] == "001010"  # API 값


@pytest.mark.parametrize("bad", [12345678, "991332", "abcdef"])
def test_excel_roundtrip_blocks_invalid_row(db, bad):
    from backend.services import customer_bulk_service as bulk

    blob = _import_workbook_bytes(bad)
    res = bulk.commit(blob, "t-regfront", include_duplicates=True)
    assert res["registered"] == 0
    assert res["skipped_error"] >= 1
    with db() as s:
        cnt = s.scalar(text("SELECT count(*) FROM customers WHERE tenant_id='t-regfront'"))
    assert cnt == 0


# ── 검색: 레거시 '1010' 행을 canonical '001010' / '1010' 둘 다로 찾음 ──────────
def test_search_finds_legacy_reg_front(db):
    from backend.routers.search import _search_customers

    with db() as s:
        s.execute(text(
            "INSERT INTO customers (tenant_id, customer_id, korean_name, reg_front) "
            "VALUES ('t-regfront','7001','검색합성',:r)").bindparams(r="1010"))
        s.commit()
    ids_canon = [r.id for r in _search_customers("001010", "t-regfront")]
    ids_short = [r.id for r in _search_customers("1010", "t-regfront")]
    assert "7001" in ids_canon      # canonical 로 검색됨
    assert "7001" in ids_short      # substring(레거시 입력)으로도 검색됨


# ── 만기알림 생년월일: 레거시 '1010' + 뒷자리 세기코드 → 2000-10-10 ───────────
def test_expiry_alert_birth_from_legacy_reg_front(db):
    import datetime

    from backend.routers.customers import _CACHE_EXPIRY, get_expiry_alerts
    from backend.services.cache_service import cache_invalidate

    exp = (datetime.date.today() + datetime.timedelta(days=60)).isoformat()
    with db() as s:
        s.execute(text(
            "INSERT INTO customers (tenant_id, customer_id, korean_name, reg_front, "
            "reg_back, card_expiry_date) VALUES ('t-regfront','7100','만기합성',:r,:b,:e)"
        ).bindparams(r="1010", b="7020304", e=exp))
        s.commit()
    cache_invalidate("t-regfront", _CACHE_EXPIRY)
    res = get_expiry_alerts(user={"tenant_id": "t-regfront", "sub": "x"})
    births = [row["생년월일"] for row in res["card_alerts"] if row["고객ID"] == "7100"]
    assert births and births[0] == "2000-10-10"  # 뒷자리 7 → 2000, 앞자리 복구
