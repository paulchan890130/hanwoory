"""
EXTRA_WORK 엄격 검증 보고서 (READ-ONLY)

목적:
  AUTO_SAFE 카테고리에 들어간 EXTRA_WORK row의 proposed page가 정말로
  '체류자격외 활동허가' 업무 섹션을 가리키는지 페이지 텍스트 분석으로 평가.

분류:
  KEEP_AUTO_SAFE  — 페이지가 EXTRA_WORK 섹션 헤더를 포함
  BLOCK           — 페이지가 자격해당자/활동범위 등 자격 설명 페이지에 그침
  NEEDS_REVIEW    — 신호가 혼합되어 자동 판단 불가

규칙 (사용자 명시 기준):
  본문에 '체류자격외 활동허가'가 인용된 것만으로는 통과시키지 않음.
  실제 EXTRA_WORK 섹션이거나, 시간제취업/학생 자격외 활동 표 등으로
  판별 가능한 페이지여야 함.

출력:
  backend/data/manuals/manual_mapping_extra_work_review_v3.json
  backend/data/manuals/manual_mapping_extra_work_review_v3.xlsx
"""
from __future__ import annotations
import json, re, sys
from pathlib import Path
from collections import Counter

import fitz

ROOT     = Path(__file__).parent.parent.parent
MANUALS  = ROOT / "backend" / "data" / "manuals"
TRIAGE   = MANUALS / "manual_mapping_triage_v3.json"

OUT_JSON = MANUALS / "manual_mapping_extra_work_review_v3.json"
OUT_XLSX = MANUALS / "manual_mapping_extra_work_review_v3.xlsx"

PDF_PATHS = {
    "체류민원": MANUALS / "unlocked_체류민원.pdf",
    "사증민원": MANUALS / "unlocked_사증민원.pdf",
}

# 페이지 헤더 영역 (자격 설명 카드) 시그널 — BLOCK 후보
QUALIFICATION_MARKERS = [
    "자격해당자", "자격 해당자",
    "1회에부여할수", "1회에 부여할 수",
    "체류기간상한", "체류기간 상한",
    "별도허가없이", "별도 허가 없이",
    "유학활동이가능한", "유학활동이 가능한",
    "별도의허가없이", "별도의 허가 없이",
]

# EXTRA_WORK 진짜 섹션 시그널 — KEEP 후보
EXTRA_WORK_HEADERS = [
    "체류자격외 활동",
    "체류자격외활동",
    "자격외 활동",
    "자격외활동",
    "시간제취업",
]

# EXTRA_WORK 헤더가 실제로 'section heading' 형태인지 확인하는 보조 정규식
# 출입국관리법 법 제20조(체류자격외 활동) 같은 본문 인용을 거를 수 있도록
# 단독 줄 또는 ▶/❍/○/◆ 마커 + EXTRA_WORK 키워드 패턴을 추가로 확인
SECTION_HEADER_PATTERN = re.compile(
    r"(?:^|[\n\r])\s*[▶❍○◆■□1-9\.]?\s*(?:체류자격외\s*활동|시간제취업)",
    re.MULTILINE,
)

# 법조문 인용 패턴 — 이 패턴만 있으면 본문 인용으로 간주 (KEEP 시그널 아님)
LAW_CITATION_PATTERN = re.compile(
    r"(?:법\s*제20조|시행령\s*제25조|시행규칙\s*제29조)\s*\(\s*체류자격외\s*활동",
)


def first_page_text(doc: fitz.Document, page_no: int, max_chars: int = 1500) -> str:
    if page_no < 1 or page_no > len(doc):
        return ""
    return doc[page_no - 1].get_text()[:max_chars]


def classify_extra_work_page(page_text: str) -> tuple[str, str]:
    """
    페이지 텍스트 → (KEEP_AUTO_SAFE | BLOCK | NEEDS_REVIEW, reason)

    판단 영역:
      head    = 처음 600자 (페이지 상단 제목/카드 영역)
      window  = 처음 1500자 (본문 일부 포함)
    """
    if not page_text:
        return "NEEDS_REVIEW", "page_text 없음 (PDF 추출 실패 가능)"

    head   = page_text[:600]
    window = page_text[:1500]

    qualif_in_head = any(kw in head for kw in QUALIFICATION_MARKERS)

    # EXTRA_WORK 섹션 헤더가 실제 헤더 형태로 등장하는지
    has_section_header = bool(SECTION_HEADER_PATTERN.search(window))

    # 단순 키워드 등장 (법조문 인용 포함)
    has_keyword_anywhere = any(kw in window for kw in EXTRA_WORK_HEADERS)

    # 법조문 인용만 있는지 (헤더가 아닌 본문 인용)
    only_law_citation = (
        LAW_CITATION_PATTERN.search(window)
        and not has_section_header
    )

    # ── 판정 ──
    # Case A: 자격 설명 페이지 (qualif markers in head) AND 진짜 섹션 헤더 없음
    if qualif_in_head and not has_section_header:
        return "BLOCK", (
            "자격 설명 페이지로 보임 — head 영역에 '자격해당자/활동범위/체류기간상한/"
            "별도허가없이' 등 자격 설명 키워드만 존재하고 EXTRA_WORK 섹션 헤더가 없음"
        )

    # Case B: EXTRA_WORK 섹션 헤더 존재 → KEEP
    if has_section_header:
        if qualif_in_head:
            return "KEEP_AUTO_SAFE", (
                "qualification 카드 + EXTRA_WORK 섹션이 같은 페이지에 공존 "
                "(짧은 자격이거나 페이지 경계로 자격 설명이 carry-over된 케이스)"
            )
        return "KEEP_AUTO_SAFE", "EXTRA_WORK 섹션 헤더 등장"

    # Case C: 키워드는 있지만 섹션 헤더 형태가 아니거나 법조문 인용
    if has_keyword_anywhere:
        if only_law_citation:
            return "BLOCK", "EXTRA_WORK 키워드가 법조문 인용 형태로만 등장 — 실제 섹션 아님"
        return "NEEDS_REVIEW", "EXTRA_WORK 키워드 등장하나 섹션 헤더 형태 불명확"

    # Case D: 키워드 없음
    return "BLOCK", "페이지에 EXTRA_WORK / 시간제취업 / 자격외활동 키워드 자체가 없음"


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not TRIAGE.exists():
        print(f"[ABORT] triage 파일 없음: {TRIAGE}")
        sys.exit(1)

    print("[review] triage v3 로드...")
    triage = json.loads(TRIAGE.read_text(encoding="utf-8"))
    rows = triage.get("rows") or []

    # AUTO_SAFE + EXTRA_WORK 만
    targets = [r for r in rows
               if r.get("triage_category") == "AUTO_SAFE"
               and r.get("action_type") == "EXTRA_WORK"]
    print(f"  AUTO_SAFE + EXTRA_WORK: {len(targets)}")

    print("[review] PDF 로드...")
    docs: dict[str, fitz.Document] = {}
    for label, p in PDF_PATHS.items():
        if p.exists():
            docs[label] = fitz.open(p)
            print(f"  {label}: {len(docs[label])} pages")

    review: list[dict] = []
    for r in targets:
        proposed = r.get("proposed_manual_ref") or []
        if not proposed:
            review.append({
                "row_id": r.get("row_id"),
                "title": r.get("title"),
                "detailed_code": r.get("detailed_code"),
                "manual": "",
                "page_from": None,
                "page_to": None,
                "page_heading_top": "",
                "new_page_text_preview": "",
                "is_actual_extra_work_section": False,
                "suggested_action": "BLOCK",
                "reason": "proposed_manual_ref 비어있음",
            })
            continue

        p0 = proposed[0]
        manual = p0.get("manual")
        pf = p0.get("page_from")
        pt = p0.get("page_to")
        doc = docs.get(manual)
        page_text = first_page_text(doc, pf, max_chars=1500) if (doc and isinstance(pf, int)) else ""
        # 페이지 상단(첫 250자) — heading area 미리보기
        heading_top = page_text[:250]

        action, reason = classify_extra_work_page(page_text)

        review.append({
            "row_id": r.get("row_id"),
            "title": r.get("title"),
            "detailed_code": r.get("detailed_code"),
            "manual": manual,
            "page_from": pf,
            "page_to": pt,
            "page_heading_top": heading_top,
            "new_page_text_preview": page_text[:1500],
            "is_actual_extra_work_section": action == "KEEP_AUTO_SAFE",
            "suggested_action": action,
            "reason": reason,
        })

    for d in docs.values():
        d.close()

    dist = Counter(x["suggested_action"] for x in review)
    summary = {
        "total_extra_work_auto_safe": len(targets),
        "suggested_KEEP_AUTO_SAFE":   dist.get("KEEP_AUTO_SAFE", 0),
        "suggested_BLOCK":            dist.get("BLOCK", 0),
        "suggested_NEEDS_REVIEW":     dist.get("NEEDS_REVIEW", 0),
    }

    out = {
        "review_meta": {
            "version": "v3",
            "rule":    "qualification page only → BLOCK; section header → KEEP; mixed → NEEDS_REVIEW",
            "triage":  str(TRIAGE.relative_to(ROOT)),
        },
        "summary": summary,
        "rows":    review,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "extra_work_review"
        headers = ["row_id","detailed_code","title","manual","page_from","page_to",
                   "is_actual_extra_work_section","suggested_action","reason",
                   "page_heading_top(~250)","new_page_text_preview(~700)"]
        widths  = [10, 12, 28, 10, 8, 8, 12, 18, 60, 50, 80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 정렬: BLOCK > NEEDS_REVIEW > KEEP_AUTO_SAFE
        order = {"BLOCK": 0, "NEEDS_REVIEW": 1, "KEEP_AUTO_SAFE": 2}
        review_sorted = sorted(review, key=lambda x: (order.get(x["suggested_action"], 9), x["row_id"]))
        for x in review_sorted:
            ws.append([
                x["row_id"], x["detailed_code"], x["title"], x["manual"],
                x["page_from"], x["page_to"],
                x["is_actual_extra_work_section"], x["suggested_action"], x["reason"],
                x["page_heading_top"][:250],
                (x["new_page_text_preview"] or "")[:700],
            ])
            # 색상
            color = {"BLOCK":"FFCCCB", "NEEDS_REVIEW":"FFE599", "KEEP_AUTO_SAFE":"C6F6D5"}.get(x["suggested_action"])
            if color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=color)

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric","value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, v])
        ws_s.column_dimensions["A"].width = 32
        ws_s.column_dimensions["B"].width = 12

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # 콘솔 요약
    print("\n" + "=" * 72)
    print("EXTRA_WORK STRICT REVIEW")
    print("=" * 72)
    for k, v in summary.items():
        print(f"  {k:32s} {v}")

    print("\n=== Suggested BLOCK (자동 차단 후보) ===")
    for x in [x for x in review if x["suggested_action"] == "BLOCK"]:
        print(f"  {x['row_id']:8s} {x['detailed_code']:8s} {x['manual']:6s} p.{x['page_from']:>3}~{x['page_to']:<3}  {x['title'][:24]}  :: {x['reason']}")

    print("\n=== Suggested NEEDS_REVIEW (사람 검토 필요) ===")
    for x in [x for x in review if x["suggested_action"] == "NEEDS_REVIEW"]:
        print(f"  {x['row_id']:8s} {x['detailed_code']:8s} {x['manual']:6s} p.{x['page_from']:>3}~{x['page_to']:<3}  {x['title'][:24]}  :: {x['reason']}")


if __name__ == "__main__":
    main()
