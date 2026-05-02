"""
Stage 5 — Apply Candidate Generator v1 (READ-ONLY, NO API CALL, NO DB WRITE)

Joins judge + verifier + quality gate + blocklist results to produce a
human-review list. NEVER writes to the DB. The apply step (if any) is a
separate, future task.

Inputs:
  --judge FILE [--judge FILE ...]      (Stage 3 results, repeatable)
  --verifier FILE [--verifier FILE ...] (Stage 4 results, repeatable)

  Default: pick latest *.json in
    backend/data/manuals/llm_judge_results/ and llm_verifier_results/

Outputs:
  backend/data/manuals/apply_candidates_v1.json
  backend/data/manuals/apply_candidates_v1.xlsx

A row is an apply candidate ONLY if ALL of:
  judge.decision        == EXACT
  judge.confidence      ∈ {high, medium}
  verifier.decision     == ACCEPTED
  preferred_manual_ok   == true   (from quality gate)
  same_page_cluster_size < 3      (no cluster risk)
  row_id NOT in blocklist
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path
from typing import Any

ROOT     = Path(__file__).parent.parent.parent
MANUALS  = ROOT / "backend" / "data" / "manuals"
QG_PATH  = MANUALS / "manual_ref_quality_gate_v4.json"
BLOCK    = MANUALS / "manual_mapping_apply_blocklist_v3.json"
JUDGE_DIR = MANUALS / "llm_judge_results"
VERI_DIR  = MANUALS / "llm_verifier_results"
OUT_JSON = MANUALS / "apply_candidates_v1.json"
OUT_XLSX = MANUALS / "apply_candidates_v1.xlsx"


def latest_files(dir_path: Path, pattern: str = "*.json") -> list[Path]:
    if not dir_path.exists():
        return []
    return sorted(dir_path.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)


def merge_judge_results(paths: list[Path]) -> tuple[dict[str, dict], dict]:
    """row_id → judge_result. Later files override earlier ones for same row_id."""
    merged: dict[str, dict] = {}
    meta: dict[str, list] = {"sources": [], "models": []}
    for p in paths:
        doc = json.loads(p.read_text(encoding="utf-8"))
        m = doc.get("meta") or {}
        meta["sources"].append(str(p.relative_to(ROOT)))
        meta["models"].append(m.get("model", ""))
        for r in (doc.get("results") or []):
            rid = r.get("row_id")
            if rid:
                merged[rid] = r
    return merged, meta


def merge_verifier_results(paths: list[Path]) -> dict[str, dict]:
    merged: dict[str, dict] = {}
    for p in paths:
        doc = json.loads(p.read_text(encoding="utf-8"))
        for r in (doc.get("results") or []):
            rid = r.get("row_id")
            if rid:
                merged[rid] = r
    return merged


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--judge", action="append", default=None,
                    help="Stage 3 result JSON (repeatable). Default: all in llm_judge_results/")
    ap.add_argument("--verifier", action="append", default=None,
                    help="Stage 4 result JSON (repeatable). Default: all in llm_verifier_results/")
    args = ap.parse_args()

    judge_paths = (
        [Path(p) if Path(p).is_absolute() else ROOT / p for p in args.judge]
        if args.judge else latest_files(JUDGE_DIR)
    )
    veri_paths = (
        [Path(p) if Path(p).is_absolute() else ROOT / p for p in args.verifier]
        if args.verifier else latest_files(VERI_DIR)
    )

    if not judge_paths:
        print(f"[ABORT] judge result 없음", file=sys.stderr); return 1
    if not veri_paths:
        print(f"[WARN] verifier result 없음 — 모든 EXACT 후보가 NEEDS_HUMAN 처리됨")

    print(f"[apply-cand] judge files    : {len(judge_paths)}")
    for p in judge_paths:
        print(f"               {p.relative_to(ROOT)}")
    print(f"[apply-cand] verifier files : {len(veri_paths)}")
    for p in veri_paths:
        print(f"               {p.relative_to(ROOT)}")

    judges, judge_meta = merge_judge_results(judge_paths)
    verifiers = merge_verifier_results(veri_paths)
    print(f"  unique judge row_ids   : {len(judges)}")
    print(f"  unique verifier row_ids: {len(verifiers)}")

    qg = json.loads(QG_PATH.read_text(encoding="utf-8"))
    qg_by_row = {r["row_id"]: r for r in (qg.get("rows") or [])}
    blocked = set(json.loads(BLOCK.read_text(encoding="utf-8")).get("blocked_row_ids") or [])

    # Build candidates
    candidates: list[dict] = []
    excluded_blocklist = 0
    excluded_cluster = 0
    excluded_pref = 0
    excluded_judge_not_exact = 0
    excluded_judge_low_conf = 0
    excluded_no_verifier = 0
    excluded_verifier_rejected = 0

    page_correction_count = 0
    keep_same_count = 0
    needs_human_count = 0

    for rid, jr in judges.items():
        # Filter chain (count exclusion reason for the first failure)
        if jr.get("decision") != "EXACT":
            excluded_judge_not_exact += 1
            continue
        if jr.get("confidence") not in ("high", "medium"):
            excluded_judge_low_conf += 1
            continue

        if rid in blocked:
            excluded_blocklist += 1
            continue

        qg_row = qg_by_row.get(rid, {})
        if not qg_row.get("preferred_manual_ok", False):
            excluded_pref += 1
            continue
        cluster_size = qg_row.get("same_page_cluster_size") or 0
        if cluster_size >= 3:
            excluded_cluster += 1
            continue

        vr = verifiers.get(rid)
        if not vr:
            excluded_no_verifier += 1
            continue
        if vr.get("verifier_decision") != "ACCEPTED":
            excluded_verifier_rejected += 1
            continue

        # Apply candidate — determine action
        primary = qg_row.get("primary_entry") or {}
        current_manual = primary.get("manual", "")
        current_page = primary.get("page_from") or 0
        recommended_page = jr.get("recommended_page")

        if recommended_page in (None, "", current_page):
            apply_action = "KEEP_SAME"
            keep_same_count += 1
        else:
            apply_action = "PAGE_CORRECTION"
            page_correction_count += 1

        candidates.append({
            "row_id":              rid,
            "detailed_code":       qg_row.get("detailed_code"),
            "action_type":         qg_row.get("action_type"),
            "title":               qg_row.get("title"),
            "current_manual":      current_manual,
            "current_page":        current_page,
            "judge_decision":      jr.get("decision"),
            "judge_confidence":    jr.get("confidence"),
            "verifier_decision":   vr.get("verifier_decision"),
            "apply_action":        apply_action,
            "positive_evidence":   jr.get("positive_evidence", ""),
            "negative_evidence":   jr.get("negative_evidence", ""),
            "reject_reason":       vr.get("reject_reason"),
            "recommended_page":    recommended_page,
            "page_role":           jr.get("page_role"),
        })

    # NEEDS_HUMAN: judge=EXACT but verifier missing or rejected (still surfaced for review)
    needs_human_rows: list[dict] = []
    for rid, jr in judges.items():
        if jr.get("decision") != "EXACT":
            continue
        if rid in blocked:
            continue
        qg_row = qg_by_row.get(rid, {})
        if not qg_row.get("preferred_manual_ok", False):
            continue
        cluster_size = qg_row.get("same_page_cluster_size") or 0
        if cluster_size >= 3:
            continue
        vr = verifiers.get(rid)
        if vr and vr.get("verifier_decision") == "ACCEPTED":
            continue  # already in candidates
        # Surfaces all EXACT-judge rows blocked by missing/rejected verifier
        primary = qg_row.get("primary_entry") or {}
        needs_human_rows.append({
            "row_id":              rid,
            "detailed_code":       qg_row.get("detailed_code"),
            "action_type":         qg_row.get("action_type"),
            "title":               qg_row.get("title"),
            "current_manual":      primary.get("manual", ""),
            "current_page":        primary.get("page_from") or 0,
            "judge_decision":      jr.get("decision"),
            "judge_confidence":    jr.get("confidence"),
            "verifier_decision":   (vr or {}).get("verifier_decision"),
            "apply_action":        "NEEDS_HUMAN",
            "positive_evidence":   jr.get("positive_evidence", ""),
            "negative_evidence":   jr.get("negative_evidence", ""),
            "reject_reason":       (vr or {}).get("reject_reason"),
            "recommended_page":    jr.get("recommended_page"),
            "page_role":           jr.get("page_role"),
        })
    needs_human_count = len(needs_human_rows)

    all_rows = candidates + needs_human_rows

    summary = {
        "judge_files":                       [str(p.relative_to(ROOT)) for p in judge_paths],
        "verifier_files":                    [str(p.relative_to(ROOT)) for p in veri_paths],
        "judge_models":                      judge_meta["models"],
        "total_judge_rows":                  len(judges),
        "total_verifier_rows":               len(verifiers),
        "total_apply_candidates":            len(candidates),
        "KEEP_SAME":                         keep_same_count,
        "PAGE_CORRECTION":                   page_correction_count,
        "NEEDS_HUMAN":                       needs_human_count,
        "excluded_judge_not_exact":          excluded_judge_not_exact,
        "excluded_judge_low_confidence":     excluded_judge_low_conf,
        "excluded_blocklist":                excluded_blocklist,
        "excluded_cluster_risk":             excluded_cluster,
        "excluded_preferred_manual_fail":    excluded_pref,
        "excluded_no_verifier_result":       excluded_no_verifier,
        "excluded_verifier_rejected":        excluded_verifier_rejected,
        "no_db_modification":                True,
    }

    out_doc = {
        "meta": {
            "stage":   "5_apply_candidates",
            "version": "v1",
            "rules": {
                "judge_decision":      "EXACT",
                "judge_confidence":    "high or medium",
                "verifier_decision":   "ACCEPTED",
                "preferred_manual_ok": True,
                "cluster_size":        "< 3",
                "blocklist":           "excluded",
            },
            "no_db_modification": True,
        },
        "summary":     summary,
        "candidates":  candidates,
        "needs_human": needs_human_rows,
    }
    OUT_JSON.write_text(json.dumps(out_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
        ws_s.column_dimensions["A"].width = 36
        ws_s.column_dimensions["B"].width = 80

        ws = wb.create_sheet("apply_candidates")
        headers = [
            "row_id", "detailed_code", "action_type", "title",
            "current_manual", "current_page",
            "judge_decision", "judge_confidence", "verifier_decision",
            "apply_action",
            "positive_evidence", "negative_evidence", "reject_reason",
            "recommended_page", "page_role",
        ]
        widths = [10, 12, 14, 32, 12, 10, 12, 12, 14, 18, 60, 60, 50, 14, 22]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        color_map = {
            "KEEP_SAME":       "C6F6D5",
            "PAGE_CORRECTION": "FFE599",
            "NEEDS_HUMAN":     "FFCC99",
        }
        for x in all_rows:
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                x["current_manual"], x["current_page"],
                x["judge_decision"], x["judge_confidence"], x["verifier_decision"],
                x["apply_action"],
                (x.get("positive_evidence") or "")[:600],
                (x.get("negative_evidence") or "")[:600],
                (x.get("reject_reason") or "") if x.get("reject_reason") else "",
                x.get("recommended_page"),
                x.get("page_role"),
            ])
            color = color_map.get(x["apply_action"])
            if color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=color)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    print("\n" + "=" * 60)
    print("APPLY CANDIDATES — STAGE 5 SUMMARY")
    print("=" * 60)
    print(f"  Total EXACT + ACCEPTED candidates : {summary['total_apply_candidates']}")
    print(f"  KEEP_SAME                         : {summary['KEEP_SAME']}")
    print(f"  PAGE_CORRECTION                   : {summary['PAGE_CORRECTION']}")
    print(f"  NEEDS_HUMAN                       : {summary['NEEDS_HUMAN']}")
    print(f"  Excluded (blocklist)              : {summary['excluded_blocklist']}")
    print(f"  Excluded (cluster risk ≥3)        : {summary['excluded_cluster_risk']}")
    print(f"  Excluded (preferred_manual fail)  : {summary['excluded_preferred_manual_fail']}")
    print(f"  Excluded (judge ≠ EXACT)          : {summary['excluded_judge_not_exact']}")
    print(f"  Excluded (judge low confidence)   : {summary['excluded_judge_low_confidence']}")
    print(f"  Excluded (no verifier result)     : {summary['excluded_no_verifier_result']}")
    print(f"  Excluded (verifier REJECTED)      : {summary['excluded_verifier_rejected']}")
    print()
    print("  ※ This file is READ-ONLY input for human review.")
    print("  ※ No DB write. No apply. Apply is a separate future task.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
