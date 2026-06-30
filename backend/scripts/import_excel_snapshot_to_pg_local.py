"""Local Excel snapshot → local PostgreSQL importer.

Reads ``migration_input/*.xlsx`` (user-supplied snapshots of the live Google
xlsx workbooks) and upserts them into the local Docker Postgres.

Safety contract
---------------
* Refuses to run unless ``DATABASE_URL`` host is loopback.
* Default mode is **dry-run** (no DB writes).
* ``--execute`` is required for actual writes.
* **Never imports external API libraries.** Excel is the only source.
* Every domain is wrapped in try/except so one failure does not kill the rest.

CLI
---
    python backend/scripts/import_excel_snapshot_to_pg_local.py
    python backend/scripts/import_excel_snapshot_to_pg_local.py --execute
    python backend/scripts/import_excel_snapshot_to_pg_local.py --execute --reset-local-pg
    python backend/scripts/import_excel_snapshot_to_pg_local.py --only hanwoory --execute
    python backend/scripts/import_excel_snapshot_to_pg_local.py --only jpup --execute
    python backend/scripts/import_excel_snapshot_to_pg_local.py --only templates
"""
from __future__ import annotations

import argparse
import hashlib
import os
import sys
import traceback
from pathlib import Path
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parent.parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv(ROOT / ".env")
except ImportError:
    pass


_REPORT: list[str] = []
_STATS: dict[str, dict[str, int]] = {}  # domain → {inserted, updated, skipped}
_SKIPPED_REASONS: list[str] = []

# Default tenants whose Accounts rows should be imported. Anything else
# (e.g. ``asd`` from the live system, or synthetic ``test_*`` users) is
# skipped unless the user passes --include-experimental.
DEFAULT_ALLOWED_TENANTS = ("hanwoory", "jpup")

# Synthetic users that earlier seed scripts created. The default import
# does NOT seed these and the default reset wipes them so the final
# verification view stays clean.
SYNTHETIC_LOGINS = ("test_admin", "test_user", "inactive_user")
SYNTHETIC_TENANTS = ("test_admin", "test_user", "inactive_user")


def _utf8() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
        sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except Exception:
        pass


def _say(msg: str) -> None:
    print(msg, flush=True)
    _REPORT.append(msg)


def _bump(domain: str, key: str, n: int = 1) -> None:
    bucket = _STATS.setdefault(domain, {"inserted": 0, "updated": 0, "skipped": 0})
    bucket[key] = bucket.get(key, 0) + n


def _skip(domain: str, reason: str) -> None:
    _bump(domain, "skipped", 1)
    if len(_SKIPPED_REASONS) < 50:
        _SKIPPED_REASONS.append(f"  [{domain}] {reason}")


# ── Excel readers ─────────────────────────────────────────────────────────


def _open_xlsx(path: Path):
    from openpyxl import load_workbook
    return load_workbook(path, read_only=True, data_only=True)


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        # openpyxl returns ints when the cell is integer-shaped
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _read_tab(wb, sheet_name: str) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) for ``sheet_name``. Empty if missing."""
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    raw_header = rows[0]
    header: list[str] = []
    used: dict[str, int] = {}
    for idx, h in enumerate(raw_header):
        name = _norm_cell(h) or f"col_{idx + 1}"
        if name in used:
            used[name] += 1
            name = f"{name}__{used[name]}"
        else:
            used[name] = 1
        header.append(name)
    data: list[list[str]] = []
    for raw in rows[1:]:
        vals = [_norm_cell(c) for c in raw]
        if any(v for v in vals):
            # pad / truncate to header width
            if len(vals) < len(header):
                vals = vals + [""] * (len(header) - len(vals))
            elif len(vals) > len(header):
                vals = vals[: len(header)]
            data.append(vals)
    return header, data


def _rows_to_dicts(header: list[str], rows: list[list[str]]) -> list[dict]:
    return [dict(zip(header, r)) for r in rows]


# ── Workbook discovery ────────────────────────────────────────────────────


WORKBOOK_LAYOUT = [
    # (role, relative_path, tenant_id_or_None_for_template)
    ("hanwoory_customers", "신 고객 데이터.xlsx", "hanwoory"),
    ("hanwoory_work", "신 업무정리.xlsx", "hanwoory"),
    ("template_customers", "기준 고객 데이터.xlsx", None),
    ("template_work", "기준 업무정리.xlsx", None),
]


def _discover_tenant_files(input_dir: Path) -> list[tuple[str, str, str]]:
    """Find tenants/{고객 데이터|업무정리} - {tenant}.xlsx pairs."""
    tdir = input_dir / "tenants"
    if not tdir.exists():
        return []
    out: list[tuple[str, str, str]] = []
    for p in sorted(tdir.glob("*.xlsx")):
        name = p.name
        if "고객 데이터 - " in name:
            tenant = name.replace("고객 데이터 - ", "").replace(".xlsx", "").strip()
            out.append((f"{tenant}_customers", f"tenants/{name}", tenant))
        elif "업무정리 - " in name:
            tenant = name.replace("업무정리 - ", "").replace(".xlsx", "").strip()
            out.append((f"{tenant}_work", f"tenants/{name}", tenant))
    return out


# ── Customer column mapping ───────────────────────────────────────────────


SHEET_TO_PG_CUSTOMER = {
    "고객ID": "customer_id",
    "한글": "korean_name",
    "성": "surname_en",
    "명": "given_en",
    "여권": "passport_no",
    "국적": "nationality",
    "성별": "gender",
    "등록증": "reg_front",
    "번호": "reg_back",
    "발급일": "card_issue_date",
    "만기일": "card_expiry_date",
    "발급": "passport_issue_date",
    "만기": "passport_expiry_date",
    "주소": "address",
    "연": "phone1",
    "락": "phone2",
    "처": "phone3",
    "V": "v_status",
    "체류자격": "visa_status",
    "비자종류": "visa_type",
    "메모": "memo",
    "폴더": "folder_id",
    "위임내역": "delegation_history",
}


def _customer_row_to_pg(d: dict) -> dict:
    """Map Korean-keyed sheet row → English PG columns. 비고+기타 → memo."""
    out: dict[str, Any] = {}
    for sk, pg in SHEET_TO_PG_CUSTOMER.items():
        if sk in d:
            v = d.get(sk, "")
            out[pg] = v if v != "" else None
    # 날짜 컬럼은 'YYYY-MM-DD' 로 정규화한다. openpyxl 이 날짜 셀을 datetime 으로
    # 돌려주면 str() 가 'YYYY-MM-DD 00:00:00' 을 만들어 DB 에 그대로 들어가던 버그
    # (고객 상세 화면 datetime 노출의 근본 원인)를 import 단계에서 차단한다.
    from backend.services.date_normalize import normalize_date_only
    for _date_col in ("card_issue_date", "card_expiry_date", "passport_issue_date", "passport_expiry_date"):
        if _date_col in out and out[_date_col] is not None:
            out[_date_col] = normalize_date_only(out[_date_col])
    # 비고 / 기타 are not in mapping above — concat into memo if absent
    extra_bits: list[str] = []
    for k in ("비고", "기타"):
        v = (d.get(k) or "").strip()
        if v:
            extra_bits.append(v)
    if extra_bits:
        joined = "\n".join(extra_bits)
        if not out.get("memo"):
            out["memo"] = joined
        else:
            out["memo"] = f"{out['memo']}\n{joined}".strip()
    return out


# ── Deterministic ID fallback ─────────────────────────────────────────────


def _det_id(tenant: str, role: str, sheet: str, row_no: int, *parts: str) -> str:
    raw = "|".join([tenant, role, sheet, str(row_no)] + [str(p or "") for p in parts])
    return "det-" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


# ── Local guard + reset ───────────────────────────────────────────────────


_RESET_TABLES = [
    "customer_signatures",
    "temp_signature_slots",
    "agent_signatures",
    "accommodation_providers",
    "guarantor_connections",
    "completed_tasks",
    "active_tasks",
    "planned_tasks",
    "daily_entries",
    "daily_balances",
    "memos",
    "events",
    "customers",
    "board_comments",
    "board_posts",
    "marketing_posts",
    "cert_prices",
    "cert_regions",
    "cert_groups",
    "cert_directions",
    "cert_vendors",
    "work_reference_rows",
    "work_reference_sheets",
]


def _reset_local_pg(session, allowed_tenants: set[str] | None = None,
                    include_experimental: bool = False) -> None:
    """Reset business tables and prune unintended tenants/users.

    - Business-table rows (customers, tasks, events, ...) are wiped entirely.
    - Synthetic accounts (test_admin, test_user, inactive_user and their
      tenants) are deleted unconditionally — leftovers from earlier seed
      scripts are NOT preserved.
    - When ``allowed_tenants`` is given and ``include_experimental`` is
      False, any tenant/user not in that allow-list is also pruned.
    """
    from sqlalchemy import text as sql_text
    for tbl in _RESET_TABLES:
        try:
            session.execute(sql_text(f"DELETE FROM {tbl}"))
        except Exception as e:
            _say(f"  [reset][warn] {tbl}: {type(e).__name__}: {e}")
    session.commit()
    _say(f"[reset] deleted rows from {len(_RESET_TABLES)} business tables")

    # Always purge synthetic accounts from earlier seed rounds.
    synth_logins = list(SYNTHETIC_LOGINS)
    synth_tenants = list(SYNTHETIC_TENANTS)
    try:
        session.execute(
            sql_text("DELETE FROM users WHERE login_id = ANY(:logins)"),
            {"logins": synth_logins},
        )
        session.execute(
            sql_text("DELETE FROM tenants WHERE tenant_id = ANY(:tids)"),
            {"tids": synth_tenants},
        )
        session.commit()
        _say(f"[reset] purged synthetic users/tenants: {synth_logins}")
    except Exception as e:
        _say(f"  [reset.synthetic][warn] {type(e).__name__}: {e}")
        session.rollback()

    # Purge auto-created fallback admin logins (``{tenant}_admin``). These
    # are created only by --create-local-admins; if that flag is not
    # passed in this run, leftover autos from previous runs must go so the
    # final account list stays clean.
    try:
        r = session.execute(
            sql_text("DELETE FROM users WHERE login_id LIKE '%\\_admin' ESCAPE '\\' "
                     "AND login_id <> 'test_admin'"),
        )
        if r.rowcount:
            _say(f"[reset] purged {r.rowcount} fallback *_admin users (use --create-local-admins to recreate)")
        session.commit()
    except Exception as e:
        _say(f"  [reset.fallback_admin][warn] {type(e).__name__}: {e}")
        session.rollback()

    # If we have an allow-list, prune anything outside it.
    if allowed_tenants and not include_experimental:
        try:
            allowed_list = sorted(allowed_tenants)
            r = session.execute(
                sql_text("DELETE FROM users WHERE tenant_id <> ALL(:tids)"),
                {"tids": allowed_list},
            )
            _say(f"[reset] pruned {r.rowcount} users outside allowed tenants {allowed_list}")
            r2 = session.execute(
                sql_text("DELETE FROM tenants WHERE tenant_id <> ALL(:tids)"),
                {"tids": allowed_list},
            )
            _say(f"[reset] pruned {r2.rowcount} tenants outside allowed tenants {allowed_list}")
            session.commit()
        except Exception as e:
            _say(f"  [reset.allowlist][warn] {type(e).__name__}: {e}")
            session.rollback()


# ── Tenant / user provisioning ────────────────────────────────────────────


def _ensure_tenant(session, tenant_id: str, office_name: str | None = None) -> None:
    from backend.db.models.tenant import Tenant
    from sqlalchemy import select
    row = session.scalar(select(Tenant).where(Tenant.tenant_id == tenant_id))
    if row is None:
        session.add(Tenant(
            tenant_id=tenant_id,
            office_name=office_name or tenant_id,
        ))
        session.flush()


def _seed_synthetic_accounts(session) -> None:
    """Opt-in synthetic seed (test_admin / test_user / inactive_user).

    Off by default. Only runs when --seed-synthetic is passed.
    """
    from backend.db.models.tenant import Tenant
    from backend.db.models.user import AccountUser
    from backend.services.accounts_service import hash_password
    from sqlalchemy import select

    seeds = [
        ("test_admin", "test_admin", "Beta Admin Office", True, True),
        ("test_user", "test_user", "Beta User Office", False, True),
        ("inactive_user", "inactive_user", "Inactive Office", False, False),
    ]
    for login, tid, office, is_admin, is_active in seeds:
        if not session.scalar(select(Tenant).where(Tenant.tenant_id == tid)):
            session.add(Tenant(tenant_id=tid, office_name=office, is_active=is_active))
        session.flush()
        if not session.scalar(select(AccountUser).where(AccountUser.login_id == login)):
            session.add(AccountUser(
                login_id=login, tenant_id=tid,
                password_hash=hash_password("beta_test_password_123"),
                is_admin=is_admin, is_active=is_active,
            ))
    _say("  [seed.synthetic] inserted test_admin / test_user / inactive_user")


def _ensure_admin_user(session, tenant_id: str, fallback_login: str) -> None:
    """If no admin user exists for this tenant, create a local one with the default beta password."""
    from backend.db.models.user import AccountUser
    from backend.services.accounts_service import hash_password
    from sqlalchemy import select
    has_admin = session.scalar(
        select(AccountUser).where(
            AccountUser.tenant_id == tenant_id,
            AccountUser.is_admin.is_(True),
            AccountUser.is_active.is_(True),
        )
    )
    if has_admin:
        return
    # Pick a login: prefer fallback_login if free
    chosen = fallback_login
    existing = session.scalar(select(AccountUser).where(AccountUser.login_id == chosen))
    if existing:
        if existing.tenant_id == tenant_id:
            existing.is_admin = True
            existing.is_active = True
            session.flush()
            return
        # collision with different tenant; pick a suffixed name
        chosen = f"{fallback_login}_local"
    session.add(AccountUser(
        login_id=chosen,
        tenant_id=tenant_id,
        password_hash=hash_password("beta_test_password_123"),
        is_admin=True,
        is_active=True,
    ))
    session.flush()
    _say(f"  [accounts] created local admin login_id={chosen} tenant_id={tenant_id} password=beta_test_password_123")


# ── Domain importers ──────────────────────────────────────────────────────


def imp_accounts(session, wb, allowed_tenants: set[str] | None = None) -> None:
    """Import Accounts tab → tenants + users, filtered by ``allowed_tenants``.

    Default behavior (allowed_tenants={'hanwoory','jpup'}) skips experimental
    rows like ``asd`` so the final verification view contains only the two
    intended tenants. Pass an empty set or None to disable the filter.
    """
    from backend.db.models.tenant import Tenant
    from backend.db.models.user import AccountUser
    from sqlalchemy import select

    header, rows = _read_tab(wb, "Accounts")
    if not rows:
        _say("  [accounts] no Accounts sheet — skipping")
        return
    dicts = _rows_to_dicts(header, rows)
    for d in dicts:
        login = (d.get("login_id") or "").strip()
        tid = (d.get("tenant_id") or login).strip()
        if not login or not tid:
            _skip("accounts", f"empty login_id/tenant_id row={d}")
            continue
        if allowed_tenants and tid not in allowed_tenants:
            _skip("accounts", f"tenant={tid} login={login} not in allowed_tenants — skipped")
            continue
        # tenant upsert
        tenant_row = session.scalar(select(Tenant).where(Tenant.tenant_id == tid))
        tenant_fields = {
            "office_name": d.get("office_name") or tid,
            "office_adr": d.get("office_adr") or None,
            "biz_reg_no": d.get("biz_reg_no") or None,
            "agent_rrn_hash": d.get("agent_rrn") or None,
            "folder_id": d.get("folder_id") or None,
            "customer_sheet_key": d.get("customer_sheet_key") or None,
            "work_sheet_key": d.get("work_sheet_key") or None,
            "is_active": (d.get("is_active") or "TRUE").upper() != "FALSE",
        }
        if tenant_row is None:
            session.add(Tenant(tenant_id=tid, **tenant_fields))
            _bump("accounts.tenants", "inserted")
        else:
            for k, v in tenant_fields.items():
                if v is not None or k == "is_active":
                    setattr(tenant_row, k, v)
            _bump("accounts.tenants", "updated")
        session.flush()
        # user upsert
        u = session.scalar(select(AccountUser).where(AccountUser.login_id == login))
        user_fields = {
            "tenant_id": tid,
            "password_hash": (d.get("password_hash") or "").strip(),
            "contact_name": d.get("contact_name") or None,
            "contact_tel": d.get("contact_tel") or None,
            "is_admin": (d.get("is_admin") or "FALSE").upper() == "TRUE",
            "is_active": (d.get("is_active") or "TRUE").upper() != "FALSE",
        }
        if not user_fields["password_hash"]:
            _skip("accounts", f"missing password_hash for {login}")
            continue
        if u is None:
            session.add(AccountUser(login_id=login, **user_fields))
            _bump("accounts.users", "inserted")
        else:
            for k, v in user_fields.items():
                setattr(u, k, v)
            _bump("accounts.users", "updated")
    session.flush()


def imp_customers(session, wb, tenant_id: str) -> None:
    """Import 고객 데이터 sheet → customers table."""
    from backend.db.models.customer import Customer
    from sqlalchemy import select

    header, rows = _read_tab(wb, "고객 데이터")
    seen: set[str] = set()
    for d in _rows_to_dicts(header, rows):
        cid = (d.get("고객ID") or "").strip()
        if not cid:
            _skip("customers", f"tenant={tenant_id} empty 고객ID")
            continue
        if cid in seen:
            _skip("customers", f"tenant={tenant_id} duplicate 고객ID={cid}")
            continue
        seen.add(cid)
        pg = _customer_row_to_pg(d)
        pg.pop("customer_id", None)  # passed explicitly below
        existing = session.scalar(select(Customer).where(
            Customer.tenant_id == tenant_id, Customer.customer_id == cid,
        ))
        if existing is None:
            session.add(Customer(tenant_id=tenant_id, customer_id=cid, **pg))
            _bump(f"customers[{tenant_id}]", "inserted")
        else:
            for k, v in pg.items():
                setattr(existing, k, v)
            existing.deleted_at = None
            _bump(f"customers[{tenant_id}]", "updated")
    session.flush()


def imp_events(session, wb, tenant_id: str) -> None:
    from backend.db.models.event import Event
    from sqlalchemy import delete
    header, rows = _read_tab(wb, "일정")
    # events have no natural unique key per row → wipe-and-replace per tenant
    session.execute(delete(Event).where(Event.tenant_id == tenant_id))
    for idx, d in enumerate(_rows_to_dicts(header, rows)):
        ds = (d.get("date_str") or "").strip()
        et = (d.get("event_text") or "").strip()
        if not ds or not et:
            _skip("events", f"tenant={tenant_id} row={idx} empty")
            continue
        session.add(Event(tenant_id=tenant_id, date_str=ds, event_text=et, sort_order=idx))
        _bump(f"events[{tenant_id}]", "inserted")
    session.flush()


def imp_planned(session, wb, tenant_id: str, role: str) -> None:
    from backend.db.models.task import PlannedTask
    from sqlalchemy import select
    header, rows = _read_tab(wb, "예정업무")
    seen: set[str] = set()
    for idx, d in enumerate(_rows_to_dicts(header, rows)):
        tid = (d.get("id") or "").strip() or _det_id(tenant_id, role, "예정업무", idx, d.get("date"), d.get("content"))
        if tid in seen:
            tid = _det_id(tenant_id, role, "예정업무", idx, d.get("date"), d.get("content"), "dup")
        seen.add(tid)
        existing = session.scalar(select(PlannedTask).where(
            PlannedTask.tenant_id == tenant_id, PlannedTask.task_id == tid,
        ))
        fields = {
            "date": d.get("date") or None,
            "period": d.get("period") or None,
            "content": d.get("content") or None,
            "note": d.get("note") or None,
        }
        if existing is None:
            session.add(PlannedTask(tenant_id=tenant_id, task_id=tid, **fields))
            _bump(f"planned[{tenant_id}]", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump(f"planned[{tenant_id}]", "updated")
    session.flush()


def imp_active(session, wb, tenant_id: str, role: str) -> None:
    from backend.db.models.task import ActiveTask
    from sqlalchemy import select
    header, rows = _read_tab(wb, "진행업무")
    seen: set[str] = set()
    for idx, d in enumerate(_rows_to_dicts(header, rows)):
        tid = (d.get("id") or "").strip() or _det_id(
            tenant_id, role, "진행업무", idx,
            d.get("date"), d.get("name"), d.get("work"), d.get("category"),
        )
        if tid in seen:
            tid = _det_id(tenant_id, role, "진행업무", idx,
                          d.get("date"), d.get("name"), d.get("work"), "dup")
        seen.add(tid)
        existing = session.scalar(select(ActiveTask).where(
            ActiveTask.tenant_id == tenant_id, ActiveTask.task_id == tid,
        ))
        proc_raw = (d.get("processed") or "").strip().upper()
        fields = {
            "category": d.get("category") or None,
            "date": d.get("date") or None,
            "name": d.get("name") or None,
            "work": d.get("work") or None,
            "details": d.get("details") or None,
            "transfer": d.get("transfer") or None,
            "cash": d.get("cash") or None,
            "card": d.get("card") or None,
            "stamp": d.get("stamp") or None,
            "receivable": d.get("receivable") or None,
            "planned_expense": d.get("planned_expense") or None,
            "processed": proc_raw in ("TRUE", "1", "Y", "YES"),
            "processed_timestamp": d.get("processed_timestamp") or None,
            "reception": d.get("reception") or None,
            "processing": d.get("processing") or None,
            "storage": d.get("storage") or None,
            "customer_id": d.get("customer_id") or None,
            "source_daily_id": d.get("source_daily_id") or None,
        }
        if existing is None:
            session.add(ActiveTask(tenant_id=tenant_id, task_id=tid, **fields))
            _bump(f"active[{tenant_id}]", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump(f"active[{tenant_id}]", "updated")
    session.flush()


def imp_completed(session, wb, tenant_id: str, role: str) -> None:
    """Import 완료업무. Headers may contain duplicates (complete_date twice).
    Use deterministic ID fallback for missing/duplicate id."""
    from backend.db.models.task import CompletedTask
    from sqlalchemy import select
    header, rows = _read_tab(wb, "완료업무")
    seen: set[str] = set()
    for idx, d in enumerate(_rows_to_dicts(header, rows)):
        tid_raw = (d.get("id") or "").strip()
        tid = tid_raw or _det_id(
            tenant_id, role, "완료업무", idx,
            d.get("date"), d.get("name"), d.get("work"), d.get("category"),
        )
        if tid in seen:
            tid = _det_id(tenant_id, role, "완료업무", idx,
                          d.get("date"), d.get("name"), d.get("work"), d.get("category"), "dup")
            _skip(f"completed[{tenant_id}]", f"duplicate task_id={tid_raw} row={idx} → regenerated")
        seen.add(tid)
        existing = session.scalar(select(CompletedTask).where(
            CompletedTask.tenant_id == tenant_id, CompletedTask.task_id == tid,
        ))
        fields = {
            "category": d.get("category") or None,
            "date": d.get("date") or None,
            "name": d.get("name") or None,
            "work": d.get("work") or None,
            "details": d.get("details") or None,
            "complete_date": d.get("complete_date") or None,
            "reception": d.get("reception") or None,
            "processing": d.get("processing") or None,
            "storage": d.get("storage") or None,
            "customer_id": d.get("customer_id") or None,
        }
        if existing is None:
            session.add(CompletedTask(tenant_id=tenant_id, task_id=tid, **fields))
            _bump(f"completed[{tenant_id}]", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump(f"completed[{tenant_id}]", "updated")
    session.flush()


def imp_daily(session, wb, tenant_id: str, role: str) -> None:
    from backend.db.models.daily import DailyEntry, DailyBalance
    from sqlalchemy import select

    header, rows = _read_tab(wb, "일일결산")
    seen: set[str] = set()
    for idx, d in enumerate(_rows_to_dicts(header, rows)):
        eid = (d.get("id") or "").strip() or _det_id(
            tenant_id, role, "일일결산", idx,
            d.get("date"), d.get("name"), d.get("task"),
        )
        if eid in seen:
            eid = _det_id(tenant_id, role, "일일결산", idx,
                          d.get("date"), d.get("name"), d.get("task"), "dup")
        seen.add(eid)

        def to_int(v: Any) -> int:
            s = str(v or "").strip().replace(",", "")
            if not s or s in ("-", "."):
                return 0
            try:
                return int(float(s))
            except Exception:
                return 0

        existing = session.scalar(select(DailyEntry).where(
            DailyEntry.tenant_id == tenant_id, DailyEntry.entry_id == eid,
        ))
        fields = {
            "date": d.get("date") or None,
            "time": d.get("time") or None,
            "category": d.get("category") or None,
            "name": d.get("name") or None,
            "task": d.get("task") or None,
            "income_cash": to_int(d.get("income_cash")),
            "income_etc": to_int(d.get("income_etc")),
            "exp_cash": to_int(d.get("exp_cash")),
            "exp_etc": to_int(d.get("exp_etc")),
            "cash_out": to_int(d.get("cash_out")),
            "memo": d.get("memo") or None,
            "customer_id": d.get("customer_id") or None,
        }
        if existing is None:
            session.add(DailyEntry(tenant_id=tenant_id, entry_id=eid, **fields))
            _bump(f"daily[{tenant_id}]", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump(f"daily[{tenant_id}]", "updated")

    # 잔액 → daily_balances
    header2, rows2 = _read_tab(wb, "잔액")
    cash_val = 0
    profit_val = 0
    for d in _rows_to_dicts(header2, rows2):
        k = (d.get("key") or "").strip().lower()
        v = d.get("value") or "0"
        try:
            n = int(float(str(v).replace(",", "").strip() or "0"))
        except Exception:
            n = 0
        if k in ("cash", "현금"):
            cash_val = n
        elif k in ("profit", "수익", "이익"):
            profit_val = n
    bal = session.scalar(select(DailyBalance).where(DailyBalance.tenant_id == tenant_id))
    if bal is None:
        session.add(DailyBalance(tenant_id=tenant_id, cash=cash_val, profit=profit_val))
        _bump(f"balance[{tenant_id}]", "inserted")
    else:
        bal.cash = cash_val
        bal.profit = profit_val
        _bump(f"balance[{tenant_id}]", "updated")
    session.flush()


def imp_memos(session, wb, tenant_id: str) -> None:
    from backend.db.models.memo import Memo
    from sqlalchemy import select
    pairs = [("장기메모", "long"), ("중기메모", "mid"), ("단기메모", "short")]
    for sheet_name, kind in pairs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            content = ""
        else:
            # Concatenate all non-empty cells across all rows
            parts: list[str] = []
            for r in rows:
                for c in r:
                    s = _norm_cell(c)
                    if s:
                        parts.append(s)
            content = "\n".join(parts).strip()
        existing = session.scalar(select(Memo).where(
            Memo.tenant_id == tenant_id, Memo.kind == kind,
        ))
        if existing is None:
            session.add(Memo(tenant_id=tenant_id, kind=kind, content=content))
            _bump(f"memos[{tenant_id}]", "inserted")
        else:
            existing.content = content
            _bump(f"memos[{tenant_id}]", "updated")
    session.flush()


def imp_customer_signatures(session, wb, tenant_id: str) -> None:
    from backend.db.models.signature import CustomerSignature
    from sqlalchemy import select
    header, rows = _read_tab(wb, "고객서명")
    seen: set[str] = set()
    for d in _rows_to_dicts(header, rows):
        cid = (d.get("고객ID") or "").strip()
        data = (d.get("서명데이터") or "").strip()
        if not cid or not data:
            _skip("customer_signatures", f"tenant={tenant_id} empty 고객ID/서명데이터")
            continue
        if cid in seen:
            _skip("customer_signatures", f"tenant={tenant_id} dup 고객ID={cid}")
            continue
        seen.add(cid)
        existing = session.scalar(select(CustomerSignature).where(
            CustomerSignature.tenant_id == tenant_id,
            CustomerSignature.customer_id == cid,
        ))
        if existing is None:
            session.add(CustomerSignature(
                tenant_id=tenant_id, customer_id=cid, signature_data=data,
            ))
            _bump(f"customer_sig[{tenant_id}]", "inserted")
        else:
            existing.signature_data = data
            _bump(f"customer_sig[{tenant_id}]", "updated")
    session.flush()


def imp_relationships(session, wb, tenant_id: str) -> None:
    from backend.db.models.relationship import AccommodationProvider, GuarantorConnection
    from sqlalchemy import select

    def _flatten(model, sheet_name: str, fieldset: list[str]) -> None:
        header, rows = _read_tab(wb, sheet_name)
        if not rows:
            return
        seen: set[str] = set()
        for d in _rows_to_dicts(header, rows):
            target = (d.get("target_customer_id") or "").strip()
            if not target:
                _skip(f"{sheet_name}[{tenant_id}]", "empty target_customer_id")
                continue
            if target in seen:
                _skip(f"{sheet_name}[{tenant_id}]", f"dup target_customer_id={target}")
                continue
            seen.add(target)
            existing = session.scalar(select(model).where(
                model.tenant_id == tenant_id, model.target_customer_id == target,
            ))
            fields = {k: (d.get(k) or None) for k in fieldset}
            if existing is None:
                session.add(model(tenant_id=tenant_id, target_customer_id=target, **fields))
                _bump(f"{sheet_name}[{tenant_id}]", "inserted")
            else:
                for k, v in fields.items():
                    setattr(existing, k, v)
                _bump(f"{sheet_name}[{tenant_id}]", "updated")
        session.flush()

    _flatten(AccommodationProvider, "숙소제공자연결", [
        "provider_type", "provider_customer_id", "provider_name",
        "provider_last_name", "provider_first_name", "provider_nation",
        "provider_reg_front", "provider_reg_back", "provider_birth",
        "provider_phone", "provider_address", "provider_relation",
        "provide_start_date", "provide_end_date", "housing_type",
    ])
    _flatten(GuarantorConnection, "신원보증인연결", [
        "guarantor_type", "guarantor_customer_id", "guarantor_name",
        "guarantor_last_name", "guarantor_first_name", "guarantor_nation",
        "guarantor_reg_front", "guarantor_reg_back", "guarantor_birth",
        "guarantor_phone", "guarantor_address", "guarantor_relation",
        "guarantor_workplace", "guarantor_extra",
    ])


def imp_board(session, wb) -> None:
    from backend.db.models.board import BoardPost, BoardComment
    from sqlalchemy import select
    h, rows = _read_tab(wb, "게시판")
    for d in _rows_to_dicts(h, rows):
        pid = (d.get("id") or "").strip()
        if not pid:
            _skip("board_posts", "empty id")
            continue
        existing = session.scalar(select(BoardPost).where(BoardPost.id == pid))
        fields = {
            "tenant_id": d.get("tenant_id") or None,
            "author_login": d.get("author_login") or None,
            "office_name": d.get("office_name") or None,
            "is_notice": d.get("is_notice") or None,
            "category": d.get("category") or None,
            "title": d.get("title") or None,
            "content": d.get("content") or None,
            "created_at": d.get("created_at") or None,
            "updated_at": d.get("updated_at") or None,
            "popup_yn": d.get("popup_yn") or None,
            "link_url": d.get("link_url") or None,
            "comment_count": int(float(d.get("comment_count") or "0")) if (d.get("comment_count") or "").strip() else 0,
        }
        if existing is None:
            session.add(BoardPost(id=pid, **fields))
            _bump("board_posts", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump("board_posts", "updated")

    h, rows = _read_tab(wb, "게시판댓글")
    for d in _rows_to_dicts(h, rows):
        cid = (d.get("id") or "").strip()
        post_id = (d.get("post_id") or "").strip()
        if not cid or not post_id:
            _skip("board_comments", "empty id/post_id")
            continue
        existing = session.scalar(select(BoardComment).where(BoardComment.id == cid))
        fields = {
            "post_id": post_id,
            "tenant_id": d.get("tenant_id") or None,
            "author_login": d.get("author_login") or None,
            "office_name": d.get("office_name") or None,
            "content": d.get("content") or None,
            "created_at": d.get("created_at") or None,
            "updated_at": d.get("updated_at") or None,
        }
        if existing is None:
            session.add(BoardComment(id=cid, **fields))
            _bump("board_comments", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump("board_comments", "updated")
    session.flush()


def imp_marketing(session, wb) -> None:
    from backend.db.models.marketing import MarketingPost
    from sqlalchemy import select
    h, rows = _read_tab(wb, "홈페이지게시물")
    seen: set[str] = set()
    for d in _rows_to_dicts(h, rows):
        pid = (d.get("id") or "").strip()
        if not pid:
            _skip("marketing", "empty id")
            continue
        if pid in seen:
            _skip("marketing", f"dup id={pid}")
            continue
        seen.add(pid)
        existing = session.scalar(select(MarketingPost).where(MarketingPost.id == pid))
        fields = {k: (d.get(k) or None) for k in (
            "title", "slug", "category", "summary", "content",
            "thumbnail_url", "is_published", "is_featured", "created_by",
            "created_at", "updated_at", "image_file_id", "image_url",
            "image_alt", "meta_description", "tags",
        )}
        if existing is None:
            session.add(MarketingPost(id=pid, **fields))
            _bump("marketing", "inserted")
        else:
            for k, v in fields.items():
                setattr(existing, k, v)
            _bump("marketing", "updated")
    session.flush()


def imp_agent_signature(session, wb) -> None:
    from backend.db.models.signature import AgentSignature
    from sqlalchemy import select
    h, rows = _read_tab(wb, "행정사서명")
    for d in _rows_to_dicts(h, rows):
        tid = (d.get("tenant_id") or "").strip()
        data = (d.get("서명데이터") or "").strip()
        if not tid or not data:
            _skip("agent_sig", "empty tenant_id/서명데이터")
            continue
        # ensure tenant exists before FK insert
        from backend.db.models.tenant import Tenant
        if not session.scalar(select(Tenant).where(Tenant.tenant_id == tid)):
            _skip("agent_sig", f"tenant_id={tid} not provisioned — skipped")
            continue
        existing = session.scalar(select(AgentSignature).where(AgentSignature.tenant_id == tid))
        if existing is None:
            session.add(AgentSignature(tenant_id=tid, signature_data=data))
            _bump("agent_sig", "inserted")
        else:
            existing.signature_data = data
            _bump("agent_sig", "updated")
    session.flush()


def imp_temp_signatures(session, wb) -> None:
    from backend.db.models.signature import TempSignatureSlot
    from backend.db.models.tenant import Tenant
    from sqlalchemy import select
    h, rows = _read_tab(wb, "서명임시저장")
    for d in _rows_to_dicts(h, rows):
        tid = (d.get("tenant_id") or "").strip()
        slot_raw = (d.get("slot") or "").strip()
        data = (d.get("서명데이터") or "").strip()
        if not tid or not slot_raw or not data:
            _skip("temp_sig", "empty fields")
            continue
        try:
            slot = int(float(slot_raw))
        except Exception:
            _skip("temp_sig", f"non-int slot={slot_raw}")
            continue
        if slot not in (1, 2, 3):
            _skip("temp_sig", f"slot out of range: {slot}")
            continue
        if not session.scalar(select(Tenant).where(Tenant.tenant_id == tid)):
            _skip("temp_sig", f"tenant_id={tid} not provisioned")
            continue
        existing = session.scalar(select(TempSignatureSlot).where(
            TempSignatureSlot.tenant_id == tid, TempSignatureSlot.slot == slot,
        ))
        if existing is None:
            session.add(TempSignatureSlot(
                tenant_id=tid, slot=slot, signature_data=data,
                note=d.get("비고") or None,
            ))
            _bump("temp_sig", "inserted")
        else:
            existing.signature_data = data
            existing.note = d.get("비고") or None
            _bump("temp_sig", "updated")
    session.flush()


def imp_certification(session, wb, tenant_id: str) -> None:
    from backend.db.models.certification import (
        CertVendor, CertDirection, CertGroup, CertRegion, CertPrice,
    )
    from sqlalchemy import select

    mapping = [
        ("각종공인증_업체", CertVendor, ["name", "contact", "memo"]),
        ("각종공인증_대분류", CertDirection, ["name", "sort_order"]),
        ("각종공인증_중분류", CertGroup, [
            "group_name", "aliases", "default_direction",
            "applicable_directions", "sort_order",
        ]),
        ("각종공인증_소분류지역", CertRegion, [
            "name", "applicable_directions", "applicable_group_ids", "sort_order",
        ]),
        ("각종공인증_가격조건", CertPrice, [
            "vendor_id", "group_id", "direction", "region", "condition",
            "price", "possible", "documents", "lead_time", "strength",
            "risk", "source", "last_checked",
        ]),
    ]
    for sheet, model, cols in mapping:
        h, rows = _read_tab(wb, sheet)
        seen: set[str] = set()
        for idx, d in enumerate(_rows_to_dicts(h, rows)):
            rid = (d.get("id") or "").strip() or _det_id(tenant_id, "cert", sheet, idx, d.get("name") or d.get("group_name"))
            if rid in seen:
                _skip(f"{sheet}[{tenant_id}]", f"dup id row={idx}")
                continue
            seen.add(rid)
            existing = session.scalar(select(model).where(model.id == rid))
            fields: dict[str, Any] = {c: (d.get(c) or None) for c in cols}
            fields["active"] = d.get("active") or "TRUE"
            fields["created_at"] = d.get("created_at") or None
            fields["updated_at"] = d.get("updated_at") or None
            if existing is None:
                session.add(model(id=rid, tenant_id=tenant_id, **fields))
                _bump(f"{sheet}[{tenant_id}]", "inserted")
            else:
                existing.tenant_id = tenant_id
                for k, v in fields.items():
                    setattr(existing, k, v)
                _bump(f"{sheet}[{tenant_id}]", "updated")
        session.flush()


_CERT_TABS = {
    "각종공인증_업체", "각종공인증_대분류", "각종공인증_중분류",
    "각종공인증_소분류지역", "각종공인증_가격조건",
}


def imp_work_reference(session, wb, tenant_id: str) -> None:
    """All non-cert tabs → work_reference_sheets + work_reference_rows."""
    from backend.db.models.work_data import WorkReferenceSheet, WorkReferenceRow
    from sqlalchemy import select, delete

    for sheet_name in wb.sheetnames:
        if sheet_name in _CERT_TABS:
            continue
        header, rows = _read_tab(wb, sheet_name)
        if not header:
            continue
        existing = session.scalar(select(WorkReferenceSheet).where(
            WorkReferenceSheet.tenant_id == tenant_id,
            WorkReferenceSheet.sheet_name == sheet_name,
        ))
        if existing is None:
            session.add(WorkReferenceSheet(
                tenant_id=tenant_id, sheet_name=sheet_name, headers=header,
            ))
            _bump(f"work_ref_sheet[{tenant_id}]", "inserted")
        else:
            existing.headers = header
            _bump(f"work_ref_sheet[{tenant_id}]", "updated")
        # wipe + reinsert rows
        session.execute(delete(WorkReferenceRow).where(
            WorkReferenceRow.tenant_id == tenant_id,
            WorkReferenceRow.sheet_name == sheet_name,
        ))
        for idx, r in enumerate(rows):
            data = {k: v for k, v in zip(header, r)}
            session.add(WorkReferenceRow(
                tenant_id=tenant_id, sheet_name=sheet_name,
                row_index=idx, data=data,
            ))
            _bump(f"work_ref_rows[{tenant_id}]", "inserted")
    session.flush()


# ── Template inspection ───────────────────────────────────────────────────


def inspect_template(wb, label: str) -> dict:
    info = {"label": label, "sheets": []}
    for s in wb.sheetnames:
        header, rows = _read_tab(wb, s)
        info["sheets"].append({"name": s, "headers": header, "row_count": len(rows)})
    return info


# ── Remote (Render) execution gate ──────────────────────────────────────────
# The default safety contract is unchanged: local loopback DBs are allowed,
# everything else is blocked. Importing into a real Render PostgreSQL requires
# THREE explicit signals together (--execute + --allow-remote-render-pg +
# exact --confirm). Destructive reset is NEVER allowed against a remote DB.
# The policy + confirmation string live in backend.db.local_guard (single
# source of truth), shared with migrate_account_password_hashes.py.
from backend.db.local_guard import REMOTE_CONFIRM_STRING  # noqa: E402

_REMOTE_COMMAND_HINT = (
    "python backend/scripts/import_excel_snapshot_to_pg_local.py \\\n"
    "  --only all --execute --allow-remote-render-pg \\\n"
    f'  --confirm "{REMOTE_CONFIRM_STRING}"\n'
    "(Do NOT pass --reset-local-pg against Render.)"
)


# ── Main ──────────────────────────────────────────────────────────────────


def main() -> int:
    _utf8()
    parser = argparse.ArgumentParser(description="Local Excel snapshot → local PG importer")
    parser.add_argument("--execute", action="store_true", help="Actually write to PG. Default is dry-run.")
    parser.add_argument("--reset-local-pg", action="store_true",
                        help="DELETE business-table rows before import, purge synthetic accounts, "
                             "and prune tenants outside --allowed-tenants.")
    parser.add_argument("--only", choices=["hanwoory", "jpup", "templates", "all"], default="all")
    parser.add_argument("--tenant", default=None, help="Specific tenant_id (default = all discovered)")
    parser.add_argument("--input-dir", default=str(ROOT / "migration_input"))
    parser.add_argument(
        "--allowed-tenants",
        default=",".join(DEFAULT_ALLOWED_TENANTS),
        help="Comma-separated tenant allow-list for Accounts import + reset prune. "
             "Default: hanwoory,jpup. Pass empty string to disable filtering.",
    )
    parser.add_argument(
        "--include-experimental", action="store_true",
        help="Disable the tenant allow-list filter and import every Accounts row "
             "(including 'asd' and similar). Off by default.",
    )
    parser.add_argument(
        "--seed-synthetic", action="store_true",
        help="Create test_admin / test_user / inactive_user synthetic accounts. "
             "Off by default — the clean local view does not include these.",
    )
    parser.add_argument(
        "--create-local-admins", action="store_true",
        help="When a tenant has no is_admin=true user, create a local-only "
             "{tenant_id}_admin login with password beta_test_password_123. "
             "Off by default — real admins from Accounts (e.g. wkdwhfl) are kept as-is.",
    )
    parser.add_argument(
        "--allow-remote-render-pg", action="store_true",
        help="Explicitly permit importing into a NON-LOCAL (e.g. Render) PostgreSQL. "
             "Must be combined with --execute and the exact --confirm string. "
             "Off by default — the local-only guard blocks remote DBs otherwise.",
    )
    parser.add_argument(
        "--confirm", default="",
        help=f'Required confirmation string for remote execution. Must equal exactly: '
             f'"{REMOTE_CONFIRM_STRING}".',
    )
    args = parser.parse_args()

    allowed_tenants: set[str] = set()
    if args.allowed_tenants.strip() and not args.include_experimental:
        allowed_tenants = {t.strip() for t in args.allowed_tenants.split(",") if t.strip()}

    input_dir = Path(args.input_dir).resolve()
    _say(f"[INPUT]   {input_dir}  exists={input_dir.exists()}")
    if not input_dir.exists():
        _say("[FATAL]   input dir not found")
        return 4

    # Discover all workbooks
    layout = list(WORKBOOK_LAYOUT)
    layout.extend(_discover_tenant_files(input_dir))

    # Filter by --only / --tenant
    def keep(role: str, tenant: str | None) -> bool:
        if args.only == "templates":
            return role.startswith("template_")
        if args.tenant:
            return tenant == args.tenant
        if args.only == "hanwoory":
            return tenant == "hanwoory" or role.startswith("template_")
        if args.only == "jpup":
            return tenant == "jpup" or role.startswith("template_")
        return True

    files_present: list[tuple[str, Path, str | None]] = []
    files_missing: list[str] = []
    for role, rel, tenant in layout:
        p = input_dir / rel
        if not p.exists():
            files_missing.append(f"{role}: {rel}")
            continue
        if not keep(role, tenant):
            continue
        files_present.append((role, p, tenant))

    _say("[ROLES]")
    for role, p, tenant in files_present:
        wb_tmp = _open_xlsx(p)
        try:
            _say(f"  {role:24s} tenant={tenant or '(template)':10s} {p.name}  sheets={len(wb_tmp.sheetnames)}")
        finally:
            wb_tmp.close()
    if files_missing:
        _say("[MISSING]")
        for m in files_missing:
            _say(f"  - {m}")

    # Dry-run short-circuit
    if not args.execute:
        _say("\n[DRY-RUN] No DB connection opened. Pass --execute to actually import.")
        return 0

    # Local guard (LOCAL allowed by default; REMOTE requires explicit confirmation).
    from backend.db.local_guard import resolve_execution_mode
    mode = resolve_execution_mode(
        os.environ.get("DATABASE_URL"),
        allow_remote=args.allow_remote_render_pg,
        confirm=args.confirm,
        reset_requested=args.reset_local_pg,
        command_hint=_REMOTE_COMMAND_HINT,
        say=_say,
    )
    if mode is None:
        return 3  # gate printed a clear abort message
    if mode == "REMOTE_RENDER_CONFIRMED":
        _say("[guard]   (destructive reset disabled for remote; row-level upserts only)")

    from backend.db.session import get_sessionmaker
    SessionLocal = get_sessionmaker()

    with SessionLocal() as session:
        if args.reset_local_pg:
            _reset_local_pg(
                session,
                allowed_tenants=allowed_tenants,
                include_experimental=args.include_experimental,
            )

        # Optional synthetic seed (off by default)
        if args.seed_synthetic:
            try:
                _seed_synthetic_accounts(session)
                session.commit()
            except Exception as e:
                _say(f"  [seed.synthetic][error] {type(e).__name__}: {e}")
                session.rollback()

        template_info: list[dict] = []

        for role, p, tenant in files_present:
            _say(f"\n[FILE] role={role} tenant={tenant or '(template)'} file={p.name}")
            wb = _open_xlsx(p)
            try:
                if role.startswith("template_"):
                    template_info.append(inspect_template(wb, role))
                    continue

                # Ensure tenant exists before any FK insert
                _ensure_tenant(session, tenant or "hanwoory")
                session.commit()

                if role == "hanwoory_customers":
                    # admin-shared tabs first
                    try: imp_accounts(session, wb, allowed_tenants=allowed_tenants)
                    except Exception as e:
                        _say(f"  [accounts][error] {type(e).__name__}: {e}")
                        traceback.print_exc()
                        session.rollback()
                    session.commit()

                    for fn, label in [
                        (imp_customers, "customers"),
                        (imp_events, "events"),
                        (imp_memos, "memos"),
                        (imp_customer_signatures, "customer_sig"),
                        (imp_relationships, "relationships"),
                    ]:
                        try:
                            fn(session, wb, tenant)
                            session.commit()
                        except Exception as e:
                            _say(f"  [{label}][error] {type(e).__name__}: {e}")
                            traceback.print_exc()
                            session.rollback()
                    for fn, label in [
                        (imp_planned, "planned"),
                        (imp_active, "active"),
                        (imp_completed, "completed"),
                        (imp_daily, "daily"),
                    ]:
                        try:
                            fn(session, wb, tenant, role)
                            session.commit()
                        except Exception as e:
                            _say(f"  [{label}][error] {type(e).__name__}: {e}")
                            traceback.print_exc()
                            session.rollback()

                    for fn, label in [
                        (imp_board, "board"),
                        (imp_marketing, "marketing"),
                        (imp_agent_signature, "agent_sig"),
                        (imp_temp_signatures, "temp_sig"),
                    ]:
                        try:
                            fn(session, wb)
                            session.commit()
                        except Exception as e:
                            _say(f"  [{label}][error] {type(e).__name__}: {e}")
                            traceback.print_exc()
                            session.rollback()

                elif role == "hanwoory_work" or role.endswith("_work"):
                    for fn, label in [
                        (imp_certification, "certification"),
                        (imp_work_reference, "work_reference"),
                    ]:
                        try:
                            fn(session, wb, tenant)
                            session.commit()
                        except Exception as e:
                            _say(f"  [{label}][error] {type(e).__name__}: {e}")
                            traceback.print_exc()
                            session.rollback()

                elif role.endswith("_customers"):
                    # tenant customer workbook (e.g. jpup_customers) — no admin tabs
                    for fn, label in [
                        (imp_customers, "customers"),
                        (imp_events, "events"),
                        (imp_memos, "memos"),
                        (imp_customer_signatures, "customer_sig"),
                        (imp_relationships, "relationships"),
                    ]:
                        try:
                            fn(session, wb, tenant)
                            session.commit()
                        except Exception as e:
                            _say(f"  [{label}][error] {type(e).__name__}: {e}")
                            traceback.print_exc()
                            session.rollback()
                    for fn, label in [
                        (imp_planned, "planned"),
                        (imp_active, "active"),
                        (imp_completed, "completed"),
                        (imp_daily, "daily"),
                    ]:
                        try:
                            fn(session, wb, tenant, role)
                            session.commit()
                        except Exception as e:
                            _say(f"  [{label}][error] {type(e).__name__}: {e}")
                            traceback.print_exc()
                            session.rollback()
            finally:
                wb.close()

        # Local admin fallback creation is OFF by default — preserve real
        # admins like ``wkdwhfl`` and avoid spawning ``test_admin`` /
        # ``jpup_admin``. Pass --create-local-admins to opt back in.
        if args.create_local_admins:
            try:
                for tid in sorted(allowed_tenants or DEFAULT_ALLOWED_TENANTS):
                    _ensure_admin_user(session, tid, f"{tid}_admin")
                session.commit()
            except Exception as e:
                _say(f"  [accounts.ensure][error] {type(e).__name__}: {e}")
                session.rollback()

        if template_info:
            _say("\n[TEMPLATE INSPECTION]")
            for info in template_info:
                _say(f"  {info['label']}: {len(info['sheets'])} sheets")
                for s in info["sheets"]:
                    _say(f"    - {s['name']:30s} rows={s['row_count']:5d} cols={len(s['headers'])}")

    # Final summary
    _say("\n[SUMMARY]")
    tot_i = tot_u = tot_s = 0
    for domain, stats in sorted(_STATS.items()):
        i, u, s = stats.get("inserted", 0), stats.get("updated", 0), stats.get("skipped", 0)
        tot_i += i; tot_u += u; tot_s += s
        _say(f"  {domain:40s} inserted={i:5d}  updated={u:5d}  skipped={s:5d}")
    _say(f"  {'-'*40}")
    _say(f"  {'TOTAL':40s} inserted={tot_i:5d}  updated={tot_u:5d}  skipped={tot_s:5d}")

    if _SKIPPED_REASONS:
        _say("\n[SKIPPED ROWS — first 50]")
        for r in _SKIPPED_REASONS:
            _say(r)

    return 0


if __name__ == "__main__":
    sys.exit(main())
