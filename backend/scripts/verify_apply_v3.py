"""
v3 apply 직후 검증 스크립트 (READ-ONLY)

검증 항목:
  1. 백업 존재 + JSON 파싱 가능
  2. DB JSON 파싱 가능
  3. master_rows 개수 보존 (369)
  4. 139건 manual_ref 갱신 확인
  5. blocklist 12건 무변경 확인
  6. 보호 카테고리(ROUTING_SAFE_REVIEW/MANUAL_REVIEW/LOW_CONFIDENCE/NO_CANDIDATE/APPLICATION_CLAIM_REVIEW) 무변경
  7. 갱신된 manual_ref 가 proposed_manual_ref 와 일치
  8. manual_ref 외 다른 필드 변경 없음 (백업 vs 현재 diff)

출력:
  backend/data/manuals/manual_mapping_apply_verify_v3.json
  backend/data/manuals/manual_mapping_apply_verify_v3.xlsx
"""
from __future__ import annotations
import json, sys
from pathlib import Path
from collections import Counter

ROOT     = Path(__file__).parent.parent.parent
DB_PATH  = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
TRIAGE   = ROOT / "backend" / "data" / "manuals" / "manual_mapping_triage_v3.json"
BLOCKLST = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_blocklist_v3.json"
BACKUPS  = ROOT / "backend" / "data" / "backups"

OUT_JSON = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_verify_v3.json"
OUT_XLSX = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_verify_v3.xlsx"

PROTECTED_CATS = {"ROUTING_SAFE_REVIEW","MANUAL_REVIEW","LOW_CONFIDENCE","NO_CANDIDATE","APPLICATION_CLAIM_REVIEW"}


def find_latest_backup() -> Path | None:
    if not BACKUPS.exists():
        return None
    bs = sorted(BACKUPS.glob("immigration_guidelines_db_v2.manual_ref_backup_*.json"))
    return bs[-1] if bs else None


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    findings = {
        "backup_path":              None,
        "backup_size":              None,
        "backup_valid_json":        False,
        "db_valid_json":            False,
        "db_master_rows":           None,
        "expected_master_rows":     369,
        "updated_count_actual":     0,
        "expected_updated":         139,
        "blocked_unchanged":        0,
        "blocked_total":            0,
        "blocked_changed_unexpect": [],
        "protected_unchanged":      0,
        "protected_total":          0,
        "protected_changed_unexpect": [],
        "match_proposed_count":     0,
        "match_proposed_total":     0,
        "mismatch_proposed_rows":   [],
        "non_manual_ref_field_diffs": [],
        "actual_mismatch_after_apply": 0,
        "errors":                   [],
    }

    # ── 1. 백업 ──
    print("[verify] backup 확인...")
    backup = find_latest_backup()
    if not backup or not backup.exists():
        findings["errors"].append("backup 파일 없음")
        print("  [ERR] backup 없음")
    else:
        findings["backup_path"] = str(backup.relative_to(ROOT))
        findings["backup_size"] = backup.stat().st_size
        print(f"  found: {findings['backup_path']} ({findings['backup_size']:,} bytes)")
        try:
            backup_db = json.loads(backup.read_text(encoding="utf-8"))
            findings["backup_valid_json"] = True
        except Exception as e:
            findings["errors"].append(f"backup 파싱 실패: {e}")
            backup_db = None

    # ── 2. DB ──
    print("[verify] DB JSON 검증...")
    try:
        cur_db = json.loads(DB_PATH.read_text(encoding="utf-8"))
        findings["db_valid_json"] = True
    except Exception as e:
        findings["errors"].append(f"DB 파싱 실패: {e}")
        print(f"  [ERR] {e}")
        OUT_JSON.write_text(json.dumps(findings, ensure_ascii=False, indent=2), encoding="utf-8")
        sys.exit(2)

    cur_rows = cur_db.get("master_rows") or []
    findings["db_master_rows"] = len(cur_rows)
    print(f"  master_rows: {len(cur_rows)} (expected 369)")
    if len(cur_rows) != 369:
        findings["errors"].append(f"master_rows 개수 불일치: {len(cur_rows)} != 369")

    # 인덱스
    cur_index = {r.get("row_id"): r for r in cur_rows}
    bk_index = {r.get("row_id"): r for r in (backup_db.get("master_rows") if backup_db else [])}

    # ── 3. triage / blocklist 로드 ──
    print("[verify] triage v3 + blocklist 로드...")
    triage = json.loads(TRIAGE.read_text(encoding="utf-8"))
    triage_rows = triage.get("rows") or []
    blocklist = json.loads(BLOCKLST.read_text(encoding="utf-8"))
    blocked_ids = set(blocklist.get("blocked_row_ids") or [])

    # ── 4. 갱신된 row 확인 (= AUTO_SAFE + changed + not blocked) ──
    print("[verify] manual_ref 갱신 검증...")
    expected_to_update = []
    for t in triage_rows:
        if t.get("triage_category") != "AUTO_SAFE":     continue
        if not t.get("apply_candidate"):                continue
        if t.get("comparison_status") != "changed":     continue
        if t.get("row_id") in blocked_ids:              continue
        expected_to_update.append(t)
    findings["match_proposed_total"] = len(expected_to_update)

    updated_actual = 0
    matched = 0
    mismatches = []
    for t in expected_to_update:
        rid = t["row_id"]
        cur = cur_index.get(rid)
        bk  = bk_index.get(rid) if backup_db else None
        if not cur:
            mismatches.append({"row_id": rid, "issue": "DB에서 row 사라짐"})
            continue
        cur_ref = cur.get("manual_ref") or []
        bk_ref  = (bk or {}).get("manual_ref") or []
        proposed = t.get("proposed_manual_ref") or []

        # 갱신 확인: backup 과 비교해 다르면 갱신된 것
        if cur_ref != bk_ref:
            updated_actual += 1
        else:
            mismatches.append({
                "row_id": rid, "issue": "갱신 안 된 row (backup과 동일)",
                "current": cur_ref, "proposed": proposed,
            })
            continue

        # proposed와 일치?
        # 비교: list of dicts. 단순 ==로 비교
        if cur_ref == proposed:
            matched += 1
        else:
            mismatches.append({
                "row_id": rid, "issue": "current_manual_ref != proposed_manual_ref",
                "current": cur_ref, "proposed": proposed,
            })
    findings["updated_count_actual"] = updated_actual
    findings["match_proposed_count"] = matched
    findings["mismatch_proposed_rows"] = mismatches

    print(f"  updated rows (vs backup): {updated_actual} (expected 139)")
    print(f"  matched proposed:          {matched}")

    # ── 5. blocked rows 무변경 ──
    print("[verify] blocked 12건 무변경 확인...")
    findings["blocked_total"] = len(blocked_ids)
    blocked_unchanged = 0
    blocked_unexpect = []
    for rid in blocked_ids:
        cur = cur_index.get(rid)
        bk  = bk_index.get(rid) if backup_db else None
        if not cur or not bk:
            blocked_unexpect.append({"row_id": rid, "issue": "DB or backup에 row 없음"})
            continue
        # 모든 필드 동일?
        if cur == bk:
            blocked_unchanged += 1
        else:
            # 어느 필드가 다른지
            diff_fields = sorted(set(cur.keys()) | set(bk.keys()))
            actually_diff = [k for k in diff_fields if cur.get(k) != bk.get(k)]
            blocked_unexpect.append({
                "row_id": rid, "diff_fields": actually_diff,
            })
    findings["blocked_unchanged"] = blocked_unchanged
    findings["blocked_changed_unexpect"] = blocked_unexpect
    print(f"  blocked unchanged: {blocked_unchanged}/{len(blocked_ids)}")

    # ── 6. 보호 카테고리 무변경 ──
    print("[verify] 보호 카테고리(5종) 무변경 확인...")
    protected_ids = {t["row_id"] for t in triage_rows
                     if t.get("triage_category") in PROTECTED_CATS}
    findings["protected_total"] = len(protected_ids)
    protected_unchanged = 0
    protected_unexpect = []
    for rid in protected_ids:
        cur = cur_index.get(rid)
        bk  = bk_index.get(rid) if backup_db else None
        if not cur or not bk:
            protected_unexpect.append({"row_id": rid, "issue": "DB or backup에 row 없음"})
            continue
        if cur == bk:
            protected_unchanged += 1
        else:
            diff_fields = [k for k in (set(cur.keys()) | set(bk.keys())) if cur.get(k) != bk.get(k)]
            protected_unexpect.append({"row_id": rid, "diff_fields": sorted(diff_fields)})
    findings["protected_unchanged"] = protected_unchanged
    findings["protected_changed_unexpect"] = protected_unexpect
    print(f"  protected unchanged: {protected_unchanged}/{len(protected_ids)}")

    # ── 7. manual_ref 외 다른 필드 변경 없음 ──
    print("[verify] manual_ref 외 다른 필드 변경 검사 (전체 369 row)...")
    non_mref_diffs = []
    for rid, cur in cur_index.items():
        bk = bk_index.get(rid)
        if not bk:
            non_mref_diffs.append({"row_id": rid, "issue": "backup에 row 없음 (신규)"})
            continue
        cur_keys = set(cur.keys()); bk_keys = set(bk.keys())
        all_keys = cur_keys | bk_keys
        for k in all_keys:
            if k == "manual_ref":
                continue
            if cur.get(k) != bk.get(k):
                non_mref_diffs.append({"row_id": rid, "field": k,
                                       "before": bk.get(k), "after": cur.get(k)})
    findings["non_manual_ref_field_diffs"] = non_mref_diffs
    print(f"  non-manual_ref field diffs: {len(non_mref_diffs)} (expect 0)")

    # ── 8. 실제 mismatch (DB의 manual_ref vs proposed_manual_ref) ──
    print("[verify] 현재 DB ↔ proposed_manual_ref mismatch 카운트...")
    actual_mismatch = 0
    for t in expected_to_update:
        rid = t["row_id"]
        cur = cur_index.get(rid)
        if not cur: continue
        if cur.get("manual_ref") != (t.get("proposed_manual_ref") or []):
            actual_mismatch += 1
    findings["actual_mismatch_after_apply"] = actual_mismatch
    print(f"  actual mismatch: {actual_mismatch} (expect 0)")

    # ── 9. 결과 저장 ──
    OUT_JSON.write_text(json.dumps(findings, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "verify"
        ws.append(["check","value","status"])
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")

        def status(ok): return "OK" if ok else "FAIL"

        rows = [
            ("backup_path",               findings["backup_path"], findings["backup_path"] is not None),
            ("backup_size",               findings["backup_size"], (findings["backup_size"] or 0) > 1000),
            ("backup_valid_json",         findings["backup_valid_json"], findings["backup_valid_json"]),
            ("db_valid_json",             findings["db_valid_json"], findings["db_valid_json"]),
            ("master_rows",               f"{findings['db_master_rows']} (exp 369)", findings["db_master_rows"] == 369),
            ("updated_count",             f"{findings['updated_count_actual']} (exp 139)", findings["updated_count_actual"] == 139),
            ("match_proposed",            f"{findings['match_proposed_count']}/{findings['match_proposed_total']}", findings["match_proposed_count"] == findings["match_proposed_total"]),
            ("blocked_unchanged",         f"{findings['blocked_unchanged']}/{findings['blocked_total']}", findings["blocked_unchanged"] == findings["blocked_total"]),
            ("protected_unchanged",       f"{findings['protected_unchanged']}/{findings['protected_total']}", findings["protected_unchanged"] == findings["protected_total"]),
            ("non_manual_ref_field_diffs", len(findings["non_manual_ref_field_diffs"]), len(findings["non_manual_ref_field_diffs"]) == 0),
            ("actual_mismatch_after_apply", findings["actual_mismatch_after_apply"], findings["actual_mismatch_after_apply"] == 0),
            ("errors",                    len(findings["errors"]), len(findings["errors"]) == 0),
        ]
        for k, v, ok in rows:
            ws.append([k, str(v), status(ok)])
            color = "C6F6D5" if ok else "FFCCCB"
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 8

        if findings["mismatch_proposed_rows"]:
            ws2 = wb.create_sheet("mismatches")
            ws2.append(["row_id","issue","current","proposed"])
            for c in ws2[1]:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFCCCB")
            for m in findings["mismatch_proposed_rows"][:200]:
                ws2.append([m.get("row_id"), m.get("issue",""),
                           json.dumps(m.get("current",""), ensure_ascii=False)[:200],
                           json.dumps(m.get("proposed",""), ensure_ascii=False)[:200]])

        if findings["non_manual_ref_field_diffs"]:
            ws3 = wb.create_sheet("non_mref_diffs")
            ws3.append(["row_id","field","before","after"])
            for c in ws3[1]:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFCCCB")
            for d in findings["non_manual_ref_field_diffs"][:200]:
                ws3.append([d.get("row_id"), d.get("field",""),
                           str(d.get("before",""))[:200], str(d.get("after",""))[:200]])

        if findings["blocked_changed_unexpect"]:
            ws4 = wb.create_sheet("blocked_unexpected")
            ws4.append(["row_id","diff_fields_or_issue"])
            for b in findings["blocked_changed_unexpect"]:
                ws4.append([b.get("row_id"), str(b)])

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # ── 콘솔 ──
    print("\n" + "=" * 76)
    print("APPLY VERIFICATION REPORT")
    print("=" * 76)
    def line(label, value, ok):
        print(f"  {'✓' if ok else '✗'} {label:36s} {value}")
    line("backup_path",                findings["backup_path"], findings["backup_path"] is not None)
    line("backup_size",                f"{findings['backup_size']:,} bytes" if findings["backup_size"] else "—", (findings["backup_size"] or 0) > 1000)
    line("backup_valid_json",          findings["backup_valid_json"], findings["backup_valid_json"])
    line("db_valid_json",              findings["db_valid_json"], findings["db_valid_json"])
    line("master_rows",                f"{findings['db_master_rows']} (exp 369)", findings["db_master_rows"] == 369)
    line("updated_count",              f"{findings['updated_count_actual']} (exp 139)", findings["updated_count_actual"] == 139)
    line("match_proposed",             f"{findings['match_proposed_count']}/{findings['match_proposed_total']}", findings["match_proposed_count"] == findings["match_proposed_total"])
    line("blocked_unchanged",          f"{findings['blocked_unchanged']}/{findings['blocked_total']}", findings["blocked_unchanged"] == findings["blocked_total"])
    line("protected_unchanged",        f"{findings['protected_unchanged']}/{findings['protected_total']}", findings["protected_unchanged"] == findings["protected_total"])
    line("non_manual_ref_field_diffs", len(findings["non_manual_ref_field_diffs"]), len(findings["non_manual_ref_field_diffs"]) == 0)
    line("actual_mismatch_after_apply", findings["actual_mismatch_after_apply"], findings["actual_mismatch_after_apply"] == 0)
    line("errors",                     len(findings["errors"]), len(findings["errors"]) == 0)

    # 종합
    all_ok = (
        findings["backup_valid_json"] and
        findings["db_valid_json"] and
        findings["db_master_rows"] == 369 and
        findings["updated_count_actual"] == 139 and
        findings["match_proposed_count"] == findings["match_proposed_total"] and
        findings["blocked_unchanged"] == findings["blocked_total"] and
        findings["protected_unchanged"] == findings["protected_total"] and
        len(findings["non_manual_ref_field_diffs"]) == 0 and
        findings["actual_mismatch_after_apply"] == 0 and
        len(findings["errors"]) == 0
    )
    print("\n" + ("=" * 30 + " ALL CHECKS PASSED " + "=" * 27 if all_ok else "=" * 30 + " ISSUES FOUND " + "=" * 32))


if __name__ == "__main__":
    main()
