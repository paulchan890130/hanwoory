"""엑셀 일괄 고객등록 — 양식 생성 / 파싱·검증 / 커밋 (PG-only).

설계 원칙:
  - 고객ID 는 양식에 없음 — 시스템이 자동 발번(create_customer).
  - 민감정보(외국인등록번호 뒷자리)는 반드시 create_customer 경유 →
    기존 Fernet 암호화/HMAC/마스킹 로직을 동일하게 탄다(평문 별도 저장 금지).
  - 검증/오류 메시지·로그에 reg_back/passport/주소 평문을 절대 넣지 않는다.
  - 검증과 커밋은 모두 "파일 업로드"로 동작(stateless) — 평문 민감정보를 응답으로
    프론트에 돌려보내지 않는다(미리보기는 마스킹).
  - 트랜잭션 정책: 검증 통과 행만 등록(부분 성공). 중복 의심은 기본 제외(선택 시 등록).
  - 모델에 컬럼 없는 '중국명/한자명', '생년월일'은 양식에서 제외.

양식 행 구조:
  1행 = 안내문(병합·잠금), 2행 = 헤더(잠금), 3행 = 예시(잠금·회색·업로드 시 무시),
  4행~ = 실제 입력(잠금 해제). freeze panes = A4.

정규화(저장/미리보기 표시는 항상 통일형):
  - 날짜  : yyyy-mm-dd / yyyymmdd / yyyy.mm.dd / yyyy/mm/dd / Excel 날짜셀 → yyyy-mm-dd
  - 체류자격: F4/f4/F-4 → F-4, H2→H-2, F5→F-5, C38/C-3-8→C-3-8, G1→G-1 (알 수 없으면 원문+경고)
  - 국적  : 중국/CHN/CN/China → 중국 등 (알 수 없으면 원문+경고). 둘 다 필수 아님(행 차단 안 함).
"""
from __future__ import annotations

import datetime as _dt
import io
import re
from typing import Optional


class BulkTemplateMismatch(Exception):
    """업로드 파일이 고객 일괄등록 양식(고객 시트 + 2행 15개 헤더)과 일치하지 않음.
    위치 기반으로 조용히 등록하지 않도록 파싱 전에 차단 → 라우터에서 400 으로 변환한다."""

# 엑셀 열(표시 헤더, 한글 표준키, 필수). 고객ID/중국명/생년월일 제외.
# 영문명 → 성+명 분해, 전화번호 → 연/락/처 분해(아래 _row_to_customer).
EXCEL_COLUMNS = [
    ("성명(필수)", "한글", True),
    ("영문명(여권)", "_eng_name", False),
    ("성별(남/여)", "성별", False),
    ("국적(예: 중국 또는 CHN)", "국적", False),
    ("전화번호", "_phone", False),
    ("외국인등록번호 앞자리", "등록증", False),
    ("외국인등록번호 뒷자리", "번호", False),
    ("등록증 발급일(예: 2026-06-24 또는 20260624)", "발급일", False),
    ("체류 만료일(예: 2026-06-24 또는 20260624)", "만기일", False),
    ("여권번호", "여권", False),
    ("여권 발급일(예: 2026-06-24 또는 20260624)", "발급", False),
    ("여권 만기일(예: 2026-06-24 또는 20260624)", "만기", False),
    # 체류자격 애플리케이션 정본 키는 "V"(= customers.v_status, 고객카드/검색/quick-doc 과 동일).
    # 표시 헤더는 "체류자격…" 그대로. (구 표준키 "체류자격"→visa_status 로 분리되던 버그 수정)
    ("체류자격(예: F-4 또는 F4)", "V", False),
    ("주소", "주소", False),
    ("메모", "비고", False),
]

HEADERS = [h for h, _, _ in EXCEL_COLUMNS]
STD_KEYS = [k for _, k, _ in EXCEL_COLUMNS]
DATE_STD_KEYS = {"발급일", "만기일", "발급", "만기"}
# 표준키 → 표시 라벨(미리보기 transform 문구용)
LABEL_BY_KEY = {k: h for h, k, _ in EXCEL_COLUMNS}

SHEET_NAME = "고객"
HEADER_ROW = 2     # 1행 안내, 2행 헤더
EXAMPLE_ROW = 3    # 3행 예시(업로드 시 항상 무시)
DATA_START_ROW = 4  # 4행~ 실제 입력

GUIDE_TEXT = (
    "※ 1~3행은 수정하지 마세요. 실제 고객 정보는 4행부터 입력하세요.\n"
    "※ 3행은 작성 예시이며 업로드 시 등록되지 않습니다.\n"
    "※ 고객ID는 입력하지 않습니다. 시스템에서 자동 생성됩니다.\n"
    "※ 성명은 필수입니다.\n"
    "※ 날짜는 2026-06-24 또는 20260624처럼 입력할 수 있으며, 저장 시 yyyy-mm-dd로 자동 변환됩니다.\n"
    "※ 체류자격은 F4 또는 F-4 모두 입력 가능하며, 저장 시 F-4 형식으로 통일됩니다.\n"
    "※ 국적은 중국 또는 CHN 모두 입력 가능하며, 저장 시 한글 국적명으로 통일됩니다.\n"
    "※ 외국인등록번호 뒷자리는 암호화되어 저장됩니다."
)

# 예시 행(3행) — 표준키 순서. 업로드 시 무시됨.
EXAMPLE_BY_KEY = {
    "한글": "장삼",
    "_eng_name": "ZHANG SAN",
    "성별": "남",
    "국적": "중국",
    "_phone": "01012345678",
    "등록증": "900101",
    "번호": "1234567",
    "발급일": "20260624",
    "만기일": "20271231",
    "여권": "E12345678",
    "발급": "20240101",
    "만기": "20340101",
    "V": "F4",
    "주소": "경기도 시흥시 정왕동",
    "비고": "예시 행입니다. 업로드 시 등록되지 않습니다.",
}

_YMD = re.compile(r"^(\d{4})-(\d{2})-(\d{2})([ T].*)?$")
_DOTTED = re.compile(r"^(\d{4})[./](\d{2})[./](\d{2})([ T].*)?$")
_YYYYMMDD = re.compile(r"^\d{8}$")
_DIGITS = re.compile(r"\d")


def _norm_date(v: str) -> tuple[str, bool]:
    """(정규화값, ok). 빈값은 ('', True). 판독불가는 (원본, False)."""
    s = str(v or "").strip()
    if not s:
        return "", True
    m = _YMD.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}", True
    m = _DOTTED.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}", True
    if _YYYYMMDD.match(s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}", True
    return s, False


# ── 체류자격 정규화 ─────────────────────────────────────────────────────────────
# 다단계(C-3-8 등) 특수형 명시 맵. 그 외 단일문자+한자리 숫자는 L-D 로 통일.
_VISA_SPECIAL = {"C38": "C-3-8"}


def normalize_visa(raw: str) -> tuple[str, bool, bool]:
    """(정규화값, changed, known). 알 수 없으면 (원문, False, False) — 행 차단 안 함."""
    orig = str(raw or "").strip()
    if not orig:
        return "", False, True
    s = orig.upper().replace(" ", "")
    compact = s.replace("-", "")
    if compact in _VISA_SPECIAL:
        canon = _VISA_SPECIAL[compact]
    elif re.match(r"^[A-Z]\d$", compact):          # 단일문자 + 한자리: F4→F-4, H2→H-2 …
        canon = f"{compact[0]}-{compact[1]}"
    elif re.match(r"^[A-Z](-\d)+$", s):            # 이미 표준형(F-4, C-3-8 등)
        canon = s
    else:
        return orig, False, False                  # 알 수 없음 → 원문 유지 + 경고
    return canon, (canon != orig), True


# ── 국적 정규화 ────────────────────────────────────────────────────────────────
_NAT_CANON = {
    "중국": ["중국", "중화인민공화국", "chn", "cn", "china"],
    "대한민국": ["한국", "대한민국", "kor", "kr", "korea"],
    "미국": ["미국", "usa", "us", "united states", "unitedstates", "america"],
    "베트남": ["베트남", "vnm", "vn", "vietnam"],
    "러시아": ["러시아", "rus", "ru", "russia"],
    "우즈베키스탄": ["우즈베키스탄", "uzb", "uz", "uzbekistan"],
    "카자흐스탄": ["카자흐스탄", "kaz", "kz", "kazakhstan"],
}
_NAT_LOOKUP = {alias.lower(): canon for canon, aliases in _NAT_CANON.items() for alias in aliases}


def normalize_nationality(raw: str) -> tuple[str, bool, bool]:
    """(정규화값, changed, known). 매핑 없으면 (원문, False, False) — 행 차단 안 함."""
    orig = str(raw or "").strip()
    if not orig:
        return "", False, True
    canon = _NAT_LOOKUP.get(orig.lower().replace(" ", ""))
    if canon is None:
        canon = _NAT_LOOKUP.get(orig.lower())
    if canon is None:
        return orig, False, False
    return canon, (canon != orig), True


def _split_name(value: str) -> tuple[str, str]:
    """영문명 → (성, 명). 공백 기준: 첫 토큰=성, 나머지=명. 단일 토큰은 명에."""
    parts = str(value or "").split()
    if len(parts) >= 2:
        return parts[0], " ".join(parts[1:])
    if len(parts) == 1:
        return "", parts[0]
    return "", ""


def _split_phone(value: str) -> tuple[str, str, str]:
    """전화번호 → (연, 락, 처). '-' 구분 우선, 없으면 숫자만 추출해 3-?-4 분해."""
    s = str(value or "").strip()
    if not s:
        return "", "", ""
    if "-" in s:
        parts = [p.strip() for p in s.split("-")]
        parts = (parts + ["", "", ""])[:3]
        return parts[0], parts[1], parts[2]
    d = "".join(_DIGITS.findall(s))
    if len(d) >= 10:
        return d[:3], d[3:-4], d[-4:]
    return s, "", ""


def _cell_to_str(v) -> str:
    """openpyxl 셀값 → 문자열. 날짜셀은 yyyy-mm-dd, 정수형 float 는 '.0' 제거."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


# ── 양식 생성 ──────────────────────────────────────────────────────────────────

def _record_to_form_cells(record: dict) -> list[str]:
    """고객 레코드(list_customers 한글키 dict) → 등록 양식 15개 셀 값(STD_KEYS 순서).
    영문명=성+공백+명(하나만 있으면 그 값), 전화=연-락-처(하이픈), 체류자격=canonical resolver(V 우선)."""
    from backend.services.customer_pg_service import resolve_customer_visa
    sur = str(record.get("성", "") or "").strip()
    given = str(record.get("명", "") or "").strip()
    eng = " ".join(p for p in (sur, given) if p)
    phone = "-".join(p for p in (
        str(record.get("연", "") or "").strip(),
        str(record.get("락", "") or "").strip(),
        str(record.get("처", "") or "").strip(),
    ) if p)
    by_key = {
        "한글": record.get("한글", ""),
        "_eng_name": eng,
        "성별": record.get("성별", ""),
        "국적": record.get("국적", ""),
        "_phone": phone,
        "등록증": record.get("등록증", ""),
        "번호": record.get("번호", ""),
        "발급일": record.get("발급일", ""),
        "만기일": record.get("만기일", ""),
        "여권": record.get("여권", ""),
        "발급": record.get("발급", ""),
        "만기": record.get("만기", ""),
        "V": resolve_customer_visa(record),
        "주소": record.get("주소", ""),
        "비고": record.get("비고", ""),
    }
    return [str(by_key.get(k, "") or "") for k in STD_KEYS]


def build_customer_bulk_workbook_bytes(records=None, *, mode: str = "template") -> bytes:
    """고객 일괄등록 양식 / 일괄추출을 **동일 구조**로 만드는 단일 정본 빌더.

    - mode="template": 데이터 없음(1~3행 + 4행~ 빈 입력영역).
    - mode="export":  4행부터 records 기록(등록 양식과 동일 시트/헤더/예시/서식) + 읽기 전용
      "추출 부가정보" 시트(고객ID·성명·위임내역·폴더).

    공통: 시트명 '고객', 1행 안내·2행 헤더·3행 예시, freeze A4, Text(@) 서식, 사용법 시트.
    헤더 배열(HEADERS/EXCEL_COLUMNS)은 이 함수 한 곳에서만 사용(중복 선언 없음). export 결과는
    그대로 bulk validate/commit 에 넣을 수 있다(round-trip)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Protection
    from openpyxl.utils import get_column_letter

    data_rows = list(records or []) if mode == "export" else []

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    n = len(HEADERS)

    # 1행: 안내문(병합)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=GUIDE_TEXT)
    c.font = Font(bold=True, color="7A5C10", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = PatternFill("solid", fgColor="FFF8E6")
    ws.row_dimensions[1].height = 150

    # 2행: 헤더
    header_fill = PatternFill("solid", fgColor="F0E6C8")
    for i, (h, key, required) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = 16 if len(h) <= 8 else 26
    ws.row_dimensions[HEADER_ROW].height = 42

    # 3행: 예시(회색) — 항상 3행에 존재(업로드 시 무시).
    example_fill = PatternFill("solid", fgColor="EDEDED")
    for i, key in enumerate(STD_KEYS, start=1):
        cell = ws.cell(row=EXAMPLE_ROW, column=i, value=EXAMPLE_BY_KEY.get(key, ""))
        cell.font = Font(italic=True, color="888888", size=10)
        cell.fill = example_fill

    # 4행~: 데이터(export 만) — 등록 양식과 동일 열 순서로 기록.
    for offset, record in enumerate(data_rows):
        for col, val in enumerate(_record_to_form_cells(record), start=1):
            ws.cell(row=DATA_START_ROW + offset, column=col, value=val)

    # 입력/기록 영역 행 수: template 1000행, export 는 데이터 수 확보(최소 1000).
    capacity = max(len(data_rows), 1000)
    last_row = DATA_START_ROW + capacity - 1

    # 잠금: 1~3행 locked, 4행~ unlocked
    for r in (1, HEADER_ROW, EXAMPLE_ROW):
        for col in range(1, n + 1):
            ws.cell(row=r, column=col).protection = Protection(locked=True)
    for r in range(DATA_START_ROW, last_row + 1):
        for col in range(1, n + 1):
            ws.cell(row=r, column=col).protection = Protection(locked=False)

    # Excel "일반" 서식은 앞자리 0을 지운다(전화 01099240490→1099240490, 등록증 000101→101).
    # 예시행 + 입력영역 전 열을 Text(@) 로 고정해 입력·추출 값이 그대로 문자열로 보존되게 한다.
    for r in range(EXAMPLE_ROW, last_row + 1):
        for col in range(1, n + 1):
            ws.cell(row=r, column=col).number_format = "@"

    ws.freeze_panes = "A4"
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    # 사용법 시트
    info = wb.create_sheet("사용법")
    lines = [
        "[엑셀 일괄 고객등록 사용법]",
        "",
        "1. '고객' 시트 4행부터 한 행에 한 명씩 입력합니다 (1~3행은 수정 금지).",
        "2. 3행은 작성 예시이며 업로드 시 등록되지 않습니다.",
        "3. 고객ID는 입력하지 않습니다 — 시스템이 자동으로 부여합니다.",
        "4. 성명은 필수입니다. 나머지는 비워둘 수 있습니다.",
        "5. 날짜는 2026-06-24 또는 20260624 등으로 입력하면 yyyy-mm-dd로 자동 변환됩니다.",
        "6. 체류자격은 F4/F-4 모두 가능하며 F-4 형식으로 통일됩니다.",
        "7. 국적은 중국/CHN 모두 가능하며 한글 국적명으로 통일됩니다.",
        "8. 영문명은 '성 이름' 순서로 입력하면 성/이름으로 분리 저장됩니다.",
        "9. 외국인등록번호 뒷자리는 암호화되어 저장되며 화면에는 마스킹됩니다.",
        "10. 업로드 후 미리보기에서 신규/중복의심/오류와 정규화 결과를 확인한 뒤 등록합니다.",
        "11. 일괄추출 파일의 '고객' 시트는 이 양식과 동일하며, 그대로 다시 업로드(재등록·검증)할 수 있습니다.",
    ]
    for idx, line in enumerate(lines, start=1):
        cell = info.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=12)
    info.column_dimensions["A"].width = 84

    # 추출 부가정보 시트(export 만) — bulk importer 는 '고객' 시트만 읽으므로 등록에 영향 없음.
    # 하이코리아/소시넷 ID/PW 는 어느 시트에도 포함하지 않는다.
    if mode == "export":
        aux = wb.create_sheet("추출 부가정보")
        aux.append(["고객ID", "성명", "위임내역", "폴더"])
        aux.cell(row=1, column=1).font = Font(bold=True, size=10)
        for record in data_rows:
            aux.append([
                str(record.get("고객ID", "") or ""),
                str(record.get("한글", "") or ""),
                str(record.get("위임내역", "") or ""),
                str(record.get("폴더", "") or ""),
            ])
        for col, w in ((1, 12), (2, 14), (3, 40), (4, 28)):
            aux.column_dimensions[get_column_letter(col)].width = w
        for r in range(2, len(data_rows) + 2):
            aux.cell(row=r, column=1).number_format = "@"   # 고객ID 선행 0 보존

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_bytes() -> bytes:
    """입력 양식 xlsx(bytes) — build_customer_bulk_workbook_bytes(mode='template') wrapper."""
    return build_customer_bulk_workbook_bytes(None, mode="template")


def build_export_workbook_bytes(records) -> tuple[bytes, int]:
    """일괄추출 xlsx(bytes) + 건수 — 등록 양식과 동일한 '고객' 시트(단일 정본)."""
    safe = list(records or [])
    return build_customer_bulk_workbook_bytes(safe, mode="export"), len(safe)


# ── 파싱 / 검증 ────────────────────────────────────────────────────────────────

def _read_rows(file_bytes: bytes) -> list[dict]:
    """업로드 xlsx → 행 dict 목록(표준키). 4행~ 만 읽음(3행 예시 무시). 빈 행 제외.

    파싱 전에 **고객 시트 존재 + 2행 헤더가 기대 15개 헤더와 정확히 일치**하는지 검증한다.
    불일치(구형 19열 추출·순서 변경·다른 헤더 등)면 위치 기반으로 조용히 읽지 않고
    ``BulkTemplateMismatch`` 를 올려 열 밀림 등록을 차단한다."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise BulkTemplateMismatch("no-customer-sheet")
    ws = wb[SHEET_NAME]

    # 2행 헤더 정확 일치 검증. 초과 열은 앞 15개만 비교(19열 추출도 앞 15개가 다르므로 차단됨).
    header_cells = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True), ())
    header_vals = [(_cell_to_str(c) if c is not None else "") for c in (header_cells or ())][: len(HEADERS)]
    if header_vals != HEADERS:
        wb.close()
        raise BulkTemplateMismatch("header-mismatch")

    rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        cells = list(row) if row else []
        cells = (cells + [None] * len(STD_KEYS))[: len(STD_KEYS)]
        values = [_cell_to_str(v) for v in cells]
        if not any(values):
            continue
        raw = {STD_KEYS[i]: values[i] for i in range(len(STD_KEYS))}
        raw["_row_no"] = r_idx
        rows.append(raw)
    wb.close()
    return rows


def _row_to_customer(raw: dict):
    """엑셀 행(표준키) → (고객 표준키 dict, 오류[], 변환내역[], 경고[]). 민감정보 비노출."""
    msgs: list[str] = []
    transforms: list[str] = []
    warnings: list[str] = []
    data: dict = {}

    name = raw.get("한글", "").strip()
    if not name:
        msgs.append("성명(필수)이 비어 있습니다.")
    data["한글"] = name

    sur, given = _split_name(raw.get("_eng_name", ""))
    data["성"], data["명"] = sur, given
    data["성별"] = raw.get("성별", "")

    # 국적 정규화
    nat_in = raw.get("국적", "")
    nat_out, nat_changed, nat_known = normalize_nationality(nat_in)
    data["국적"] = nat_out
    if nat_changed:
        transforms.append(f"국적 {nat_in.strip()} → {nat_out}")
    elif nat_in.strip() and not nat_known:
        warnings.append(f"국적 '{nat_in.strip()}'을(를) 인식하지 못해 원문으로 저장합니다.")

    yeon, rak, cheo = _split_phone(raw.get("_phone", ""))
    data["연"], data["락"], data["처"] = yeon, rak, cheo
    from backend.services.customer_identifier_normalize import normalize_reg_front_from_excel
    _reg_in = raw.get("등록증", "")
    _reg_r = normalize_reg_front_from_excel(_reg_in)
    if _reg_r.valid:
        data["등록증"] = _reg_r.canonical_value
        if _reg_r.changed:
            transforms.append("외국인등록번호 앞자리 선행 0 복구(6자리)")
    else:
        # 잘못된 앞자리는 경고로 저장하지 않고 **오류로 행 등록 차단**(원문 값 미노출).
        data["등록증"] = ""
        _reg_orig = _reg_in.strip() if isinstance(_reg_in, str) else str(_reg_in).strip()
        if _reg_orig:
            msgs.append("외국인등록번호 앞자리는 생년월일 6자리(YYMMDD)여야 합니다. 예: 001010")
    data["번호"] = raw.get("번호", "")  # create_customer 가 암호화
    data["여권"] = raw.get("여권", "")

    # 체류자격 정규화 — 정본 키는 "V"(= customers.v_status). visa_status(체류자격)로 쓰지 않는다.
    visa_in = raw.get("V", "")
    visa_out, visa_changed, visa_known = normalize_visa(visa_in)
    data["V"] = visa_out
    if visa_changed:
        transforms.append(f"체류자격 {visa_in.strip()} → {visa_out}")
    elif visa_in.strip() and not visa_known:
        warnings.append(f"체류자격 '{visa_in.strip()}'을(를) 인식하지 못해 원문으로 저장합니다.")

    data["주소"] = raw.get("주소", "")
    data["비고"] = raw.get("비고", "")

    # 날짜 4개
    for std_key in ("발급일", "만기일", "발급", "만기"):
        orig = raw.get(std_key, "")
        norm, ok = _norm_date(orig)
        data[std_key] = norm
        if not ok:
            msgs.append(f"{LABEL_BY_KEY[std_key].split('(')[0].strip()} 날짜 형식을 인식할 수 없습니다.")
        elif orig.strip() and norm != orig.strip():
            transforms.append(f"{LABEL_BY_KEY[std_key].split('(')[0].strip()} {orig.strip()} → {norm}")

    # 외국인등록번호 뒷자리 형식(있을 때만) — 값은 메시지에 넣지 않음.
    rb = "".join(_DIGITS.findall(data["번호"]))
    if data["번호"] and len(rb) not in (0, 7):
        msgs.append("외국인등록번호 뒷자리는 숫자 7자리여야 합니다.")

    return data, msgs, transforms, warnings


def _existing_index(tenant_id: str):
    """(passport_no→cust_id, reg_back_hash→cust_id) 기존 고객 인덱스(비삭제)."""
    from backend.db.models.customer import Customer
    from backend.db.session import get_sessionmaker
    from sqlalchemy import select

    by_passport: dict = {}
    by_hash: dict = {}
    SessionLocal = get_sessionmaker()
    with SessionLocal() as session:
        rows = session.execute(
            select(Customer.customer_id, Customer.passport_no, Customer.reg_back_hash).where(
                Customer.tenant_id == tenant_id, Customer.deleted_at.is_(None)
            )
        ).all()
    for cid, passport, h in rows:
        if passport:
            by_passport.setdefault(str(passport).strip(), str(cid))
        if h:
            by_hash.setdefault(str(h), str(cid))
    return by_passport, by_hash


def _find_duplicate(data: dict, tenant_id: str, by_passport: dict, by_hash: dict) -> Optional[str]:
    """중복 의심 기존 고객ID. ① reg_back_hash → ② passport. 없으면 None."""
    from backend.services import pii_crypto as _pii

    rb = "".join(_DIGITS.findall(data.get("번호", "")))
    if rb:
        try:
            h = _pii.hash_pii(tenant_id, rb)
        except Exception:
            h = ""
        if h and h in by_hash:
            return by_hash[h]
    passport = str(data.get("여권", "")).strip()
    if passport and passport in by_passport:
        return by_passport[passport]
    return None


def _mask_tail(v: str) -> str:
    """여권번호 등 표시 마스킹: 앞 3자리만 노출."""
    s = str(v or "").strip()
    if len(s) <= 3:
        return s
    return s[:3] + "*" * (len(s) - 3)


def validate(file_bytes: bytes, tenant_id: str) -> dict:
    """파싱+검증(등록하지 않음). 미리보기용(민감정보 마스킹, 정규화 전/후 표시)."""
    rows = _read_rows(file_bytes)
    by_passport, by_hash = _existing_index(tenant_id)

    preview = []
    n_new = n_dup = n_err = 0
    for raw in rows:
        data, msgs, transforms, warnings = _row_to_customer(raw)
        if msgs:
            status = "error"
            n_err += 1
            dup_id = None
        else:
            dup_id = _find_duplicate(data, tenant_id, by_passport, by_hash)
            if dup_id:
                status = "duplicate"
                n_dup += 1
            else:
                status = "new"
                n_new += 1
        preview.append({
            "row_no": raw.get("_row_no"),
            "status": status,
            "name": data.get("한글", ""),
            "nationality": data.get("국적", ""),
            "visa": data.get("V", ""),
            "passport_masked": _mask_tail(data.get("여권", "")),
            "messages": msgs,
            "transforms": transforms,
            "warnings": warnings,
            "dup_customer_id": dup_id,
        })

    return {
        "total": len(rows),
        "counts": {"new": n_new, "duplicate": n_dup, "error": n_err},
        "rows": preview,
    }


def commit(file_bytes: bytes, tenant_id: str, include_duplicates: bool) -> dict:
    """검증 통과 행 등록(부분 성공). 신규는 항상, 중복 의심은 include_duplicates 시.

    오류 행은 등록하지 않는다. 각 행은 create_customer 경유(자동 발번 + 암호화).
    예외 메시지/로그에 민감정보를 넣지 않는다.
    """
    from backend.services.customer_pg_service import (
        create_customer, CustomerIdConflict, TenantNotProvisioned,
    )
    from backend.services.customer_identifier_normalize import RegFrontValidationError
    from backend.services.pii_crypto import PiiKeyMissing

    rows = _read_rows(file_bytes)
    by_passport, by_hash = _existing_index(tenant_id)

    registered = 0
    skipped_dup = 0
    skipped_err = 0
    failed = 0
    fail_rows: list[int] = []
    new_ids: list[str] = []

    for raw in rows:
        data, msgs, _transforms, _warnings = _row_to_customer(raw)
        if msgs:
            skipped_err += 1
            continue
        dup_id = _find_duplicate(data, tenant_id, by_passport, by_hash)
        if dup_id and not include_duplicates:
            skipped_dup += 1
            continue
        try:
            result = create_customer(tenant_id, data)
            registered += 1
            cid = result.get("고객ID", "")
            if cid:
                new_ids.append(cid)
                passport = str(data.get("여권", "")).strip()
                if passport:
                    by_passport.setdefault(passport, cid)
        except (CustomerIdConflict, TenantNotProvisioned, PiiKeyMissing,
                RegFrontValidationError):
            # RegFrontValidationError 는 _row_to_customer 에서 이미 차단되지만(방어적 이중).
            failed += 1
            fail_rows.append(raw.get("_row_no"))
        except Exception:
            failed += 1
            fail_rows.append(raw.get("_row_no"))

    return {
        "registered": registered,
        "skipped_duplicate": skipped_dup,
        "skipped_error": skipped_err,
        "failed": failed,
        "failed_rows": fail_rows,
        "new_customer_ids": new_ids,
    }
