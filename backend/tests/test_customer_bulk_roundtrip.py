"""고객 일괄등록·일괄추출 단일 Excel 계약 + 체류자격 canonical(V) round-trip.

DB 불필요(openpyxl workbook + 순수 함수). PG 지속(V=v_status 저장/legacy fallback/문서)은
test_customer_reg_front_pg.py 참조.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook

from backend.services import customer_bulk_service as bulk
from backend.services.customer_pg_service import resolve_customer_visa


FIX = {
    "한글": "홍길동", "성": "HONG", "명": "GILDONG", "성별": "남", "국적": "중국",
    "연": "010", "락": "1234", "처": "5678", "등록증": "001010", "번호": "1234567",
    "발급일": "2025-01-02", "만기일": "2027-03-04", "여권": "E12345678",
    "발급": "2024-01-02", "만기": "2034-01-01", "V": "F-4", "주소": "서울시", "비고": "메모",
}


def _headers_of(blob: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(blob))
    ws = wb[bulk.SHEET_NAME]
    return [ws.cell(row=bulk.HEADER_ROW, column=i).value for i in range(1, len(bulk.HEADERS) + 1)]


# ── PART E 14: 등록 양식 / 추출 헤더 동일성 ──────────────────────────────────────
def test_template_and_export_headers_identical():
    tmpl = bulk.build_template_bytes()
    export, _n = bulk.build_export_workbook_bytes([FIX])
    th = _headers_of(tmpl)
    eh = _headers_of(export)
    assert th == bulk.HEADERS == eh                     # 셀 값 기준 정확 일치
    assert len(bulk.HEADERS) == 15                       # A:O
    wb = load_workbook(io.BytesIO(export))
    assert wb.sheetnames[0] == bulk.SHEET_NAME == "고객"   # 첫 시트 = 등록 양식 '고객'


# ── PART E 15: 파일 구조(안내/헤더/예시/데이터/freeze/Text/사용법) ───────────────
@pytest.mark.parametrize("mode", ["template", "export"])
def test_file_structure(mode):
    blob = bulk.build_template_bytes() if mode == "template" else bulk.build_export_workbook_bytes([FIX])[0]
    wb = load_workbook(io.BytesIO(blob))
    ws = wb[bulk.SHEET_NAME]
    assert ws.cell(row=1, column=1).value and "4행" in ws.cell(row=1, column=1).value       # 1행 안내
    assert ws.cell(row=bulk.HEADER_ROW, column=1).value == bulk.HEADERS[0]                    # 2행 헤더
    assert ws.cell(row=bulk.EXAMPLE_ROW, column=1).value == bulk.EXAMPLE_BY_KEY["한글"]        # 3행 예시
    assert ws.freeze_panes == "A4"
    # 4행 첫 열 Text(@) 서식(선행 0 보존)
    assert ws.cell(row=bulk.DATA_START_ROW, column=1).number_format == "@"
    assert "사용법" in wb.sheetnames
    if mode == "export":
        assert "추출 부가정보" in wb.sheetnames
    else:
        assert "추출 부가정보" not in wb.sheetnames


# ── PART E 16: export → validate round-trip ─────────────────────────────────────
def test_export_roundtrip_parses_back():
    export, n = bulk.build_export_workbook_bytes([FIX])
    assert n == 1
    rows = bulk._read_rows(export)          # 헤더 검증 통과 + 4행부터, 3행 예시 무시
    assert len(rows) == 1
    data, msgs, _t, _w = bulk._row_to_customer(rows[0])
    assert not msgs
    assert data["성"] == "HONG" and data["명"] == "GILDONG"
    assert (data["연"], data["락"], data["처"]) == ("010", "1234", "5678")
    assert data["V"] == "F-4"
    assert data["등록증"] == "001010"       # 선행 0 유지
    assert data["발급일"] == "2025-01-02" and data["만기일"] == "2027-03-04"
    assert data["발급"] == "2024-01-02" and data["만기"] == "2034-01-01"
    assert data["주소"] == "서울시" and data["비고"] == "메모"


def test_export_ignores_example_row():
    # 데이터 없는 export 는 예시(3행)만 있고 실제 파싱 행 0.
    export, n = bulk.build_export_workbook_bytes([])
    assert n == 0
    assert bulk._read_rows(export) == []


# ── PART E 17: 체류자격 canonical(V) ─────────────────────────────────────────────
def test_visa_resolver_priority():
    assert resolve_customer_visa({"V": "F-4", "체류자격": ""}) == "F-4"
    assert resolve_customer_visa({"V": "", "체류자격": "F-5"}) == "F-5"   # legacy fallback
    assert resolve_customer_visa({"V": "F-4", "체류자격": "F-4"}) == "F-4"
    assert resolve_customer_visa({"V": "F-4", "체류자격": "F-5"}) == "F-4"  # 불일치 → V 정본
    assert resolve_customer_visa({}) == ""


def test_export_visa_uses_resolver():
    # record V 만 → export F-4; record 체류자격(legacy)만 → export F-5.
    v_idx = bulk.STD_KEYS.index("V")
    e1 = bulk.build_export_workbook_bytes([{**FIX, "V": "F-4", "체류자격": ""}])[0]
    e2 = bulk.build_export_workbook_bytes([{**FIX, "V": "", "체류자격": "F-5"}])[0]
    ws1 = load_workbook(io.BytesIO(e1))[bulk.SHEET_NAME]
    ws2 = load_workbook(io.BytesIO(e2))[bulk.SHEET_NAME]
    assert ws1.cell(row=bulk.DATA_START_ROW, column=v_idx + 1).value == "F-4"
    assert ws2.cell(row=bulk.DATA_START_ROW, column=v_idx + 1).value == "F-5"


@pytest.mark.parametrize("raw_v,expect", [("F4", "F-4"), ("H2", "H-2"), ("F-4", "F-4"), ("C38", "C-3-8")])
def test_bulk_import_normalizes_visa_into_V(raw_v, expect):
    raw = {"한글": "x", "V": raw_v, "_eng_name": "", "_phone": "", "등록증": ""}
    data, _m, _t, _w = bulk._row_to_customer(raw)
    assert data["V"] == expect
    assert "체류자격" not in data          # 정본은 V 하나(visa_status 로 쓰지 않음)


# ── PART E 18 / PART F 20: 잘못된 양식 차단 ──────────────────────────────────────
_OLD_19_HEADERS = [
    "고객ID", "한글명", "영문 성", "영문 이름", "성별", "국적", "전화번호",
    "등록증 앞자리", "등록번호 뒷자리", "등록증 발급일", "등록증 만기일", "여권번호",
    "여권 발급일", "여권 만기일", "체류자격", "주소", "비고", "위임내역", "폴더",
]


def _wb_bytes(sheet_name: str, header_row: list[str], data_row: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for c, h in enumerate(header_row, start=1):
        ws.cell(row=bulk.HEADER_ROW, column=c, value=h)
    if data_row:
        for c, v in enumerate(data_row, start=1):
            ws.cell(row=bulk.DATA_START_ROW, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_old_export_sheet_name_blocked():
    # 실제 구형 추출: 시트명 '고객목록' → '고객' 시트 없음 → 차단.
    blob = _wb_bytes("고객목록", _OLD_19_HEADERS)
    with pytest.raises(bulk.BulkTemplateMismatch):
        bulk._read_rows(blob)


def test_old_19col_headers_blocked_even_if_sheet_renamed():
    # 시트명을 '고객'으로 바꿔도 앞 15개 헤더가 다르므로 차단(열 밀림 등록 방지).
    blob = _wb_bytes("고객", _OLD_19_HEADERS,
                     data_row=["0001", "홍길동", "HONG", "GILDONG", "남", "중국", "010-1-2"])
    with pytest.raises(bulk.BulkTemplateMismatch):
        bulk._read_rows(blob)


def test_reordered_headers_blocked():
    swapped = list(bulk.HEADERS)
    swapped[2], swapped[3] = swapped[3], swapped[2]     # 순서 변경
    with pytest.raises(bulk.BulkTemplateMismatch):
        bulk._read_rows(_wb_bytes("고객", swapped))


def test_15col_wrong_headers_blocked():
    wrong = [f"열{i}" for i in range(len(bulk.HEADERS))]
    with pytest.raises(bulk.BulkTemplateMismatch):
        bulk._read_rows(_wb_bytes("고객", wrong))


def test_valid_template_headers_pass():
    # 정상 양식은 통과(빈 데이터 → 0행).
    assert bulk._read_rows(_wb_bytes("고객", bulk.HEADERS)) == []


# ── PART F 21: 대용량(1,510행) ───────────────────────────────────────────────────
def test_export_1510_rows():
    records = [{**FIX, "한글": f"고객{i}", "여권": f"E{i:07d}"} for i in range(1510)]
    export, n = bulk.build_export_workbook_bytes(records)
    assert n == 1510
    wb = load_workbook(io.BytesIO(export))
    ws = wb[bulk.SHEET_NAME]
    v_idx = bulk.STD_KEYS.index("V") + 1
    # 헤더/예시 제외, 4행부터 1510개 데이터 + 체류자격 누락 0(V 값 있는 fixture 기준)
    filled = 0
    for r in range(bulk.DATA_START_ROW, bulk.DATA_START_ROW + 1510):
        if ws.cell(row=r, column=1).value:
            filled += 1
            assert ws.cell(row=r, column=v_idx).value == "F-4"   # 체류자격 비지 않음
    assert filled == 1510
    # round-trip 파싱도 1510건
    assert len(bulk._read_rows(export)) == 1510
