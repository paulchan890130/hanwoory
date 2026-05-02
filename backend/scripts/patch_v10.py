"""
Patch v10 — 6 rows (3 pre-confirmed HIGH + 2 Step 1 HIGH + 1 MEDIUM).

DEFAULT: dry-run. `--apply` required for DB write.

Targets:
  Pre-confirmed:
    M1-0134  D-10-2  EXTEND        → 체류 p.155 (judge+verifier ACCEPTED)
    M1-0115  D-8-2   VISA_CONFIRM  → add 사증 p.110 (벤처투자 사증발급인정서)
    M1-0114  D-8-1   VISA_CONFIRM  → add 사증 p.108-109 (법인투자 사증발급인정서 범위)

  Step 1 HIGH:
    M1-0146  E-2     VISA_CONFIRM  → add 사증 p.146 (회화지도 원어민교사 첨부서류)
    M1-0205  F-1-5   VISA_CONFIRM  → add 사증 p.303-304 (결혼이민자 부모 공관장재량 사증)

  Step 1 MEDIUM:
    M1-0132  D-10-3  VISA_CONFIRM  → add 사증 p.130 (첨단기술인턴 사증발급인정서 범위)

For M1-0134: replace full manual_ref (current 사증 p.129 is wrong manual + wrong page).
For M1-0115/0114/0146/0205/0132: add preferred-manual entry while keeping existing.
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

ROOT       = Path(__file__).parent.parent.parent
DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"
DRY_JSON   = ROOT / "backend" / "data" / "manuals" / "v10_patch_dryrun.json"
DRY_XLSX   = ROOT / "backend" / "data" / "manuals" / "v10_patch_dryrun.xlsx"

PATCHES: dict[str, dict] = {
    # ── Pre-confirmed HIGH ────────────────────────────────────────
    "M1-0134": {
        "detailed_code": "D-10-2",
        "action_type":   "EXTEND",
        "title":         "기술창업준비",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.155: '체류기간연장허가 가. 첨부서류 / 2) 기술창업준비(D-10-2)' heading + "
            "기술창업활동계획서·체류지입증서류 등 제출서류. LLM judge EXACT + verifier ACCEPTED."
        ),
        "replace_all": True,
        "add_entries": [
            {
                "manual":     "체류민원",
                "page_from":  155,
                "page_to":    155,
                "match_type": "manual_override",
                "match_text": "D-10-2 기술창업준비 EXTEND 체류기간연장허가 (체류민원 p.155) — LLM judge + verifier 검증",
            },
        ],
    },
    "M1-0115": {
        "detailed_code": "D-8-2",
        "action_type":   "VISA_CONFIRM",
        "title":         "벤처투자",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.110: '사증발급인정서 / 발급대상 / 2. 벤처기업 육성에 관한 특별조치법에 따라 "
            "벤처기업을 설립한 자 및 예비벤처기업 확인을 받은 자' — D-8-2 item 명시 + req + doc. "
            "LLM judge recommended_page='p.110'."
        ),
        "replace_all": False,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  110,
                "page_to":    110,
                "match_type": "manual_override",
                "match_text": "D-8-2 벤처투자 VISA_CONFIRM 사증발급인정서 (사증민원 p.110) — LLM judge + 수동 검증",
            },
        ],
    },
    "M1-0114": {
        "detailed_code": "D-8-1",
        "action_type":   "VISA_CONFIRM",
        "title":         "법인투자",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.108-109: D-8-1 코드 확인 (p.109) + ina=True + req + doc. "
            "p.108 '공관장재량으로 발급할수있는사증 / 대한민국 법인을 설립하고 법인등기 및 사업자등록을 완료하였을 것' + "
            "p.109 '7. 한·우즈벡 협정 대상자' (D-8-1 법인투자 세부). LLM judge recommended_page=109."
        ),
        "replace_all": False,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  108,
                "page_to":    109,
                "match_type": "manual_override",
                "match_text": "D-8-1 법인투자 VISA_CONFIRM 사증발급인정서 발급대상 (사증민원 p.108-109) — 수동 검증",
            },
        ],
    },
    # ── Step 1 HIGH ───────────────────────────────────────────────
    "M1-0146": {
        "detailed_code": "E-2",
        "action_type":   "VISA_CONFIRM",
        "title":         "회화지도",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.146: '회화지도(E-2)' 코드 + ina=True + req=True — "
            "'원어민 영어 보조교사 또는 원어민 중국어보조교사 합격증 / "
            "Talk 장학생 초청장(국립국제교육원장 발급)' E-2 전용 첨부서류 명시."
        ),
        "replace_all": False,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  146,
                "page_to":    146,
                "match_type": "manual_override",
                "match_text": "E-2 회화지도 VISA_CONFIRM 사증발급인정서 (사증민원 p.146) — 수동 검증",
            },
        ],
    },
    "M1-0205": {
        "detailed_code": "F-1-5",
        "action_type":   "VISA_CONFIRM",
        "title":         "결혼이민자의 부모 등 가족 방문동거",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.303-304: F-1-5 명시 + 첨부서류/공통서류 — "
            "결혼이민자의 부모 공관장재량 사증 섹션. "
            "p.303 'F-1-5 비자 발급 제한 / 4) 초청 인원', p.304 '첨부서류 / 공통서류'."
        ),
        "replace_all": False,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  303,
                "page_to":    304,
                "match_type": "manual_override",
                "match_text": "F-1-5 결혼이민자 부모 VISA_CONFIRM 공관장재량 사증 (사증민원 p.303-304) — 수동 검증",
            },
        ],
    },
    # ── Step 1 MEDIUM ─────────────────────────────────────────────
    "M1-0132": {
        "detailed_code": "D-10-3",
        "action_type":   "VISA_CONFIRM",
        "title":         "첨단기술인턴",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.130: D-10-3/첨단기술인턴 코드 + ina=True + req=True — "
            "사증민원 D-10 사증발급인정서 섹션 내 첨단기술인턴 부분. "
            "p.131 계속 '첨단기술인턴 초청이 가능한 기업(기관)임을 입증할 수 있는 서류'."
        ),
        "replace_all": False,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  130,
                "page_to":    131,
                "match_type": "manual_override",
                "match_text": "D-10-3 첨단기술인턴 VISA_CONFIRM 사증발급인정서 (사증민원 p.130-131) — 수동 검증",
            },
        ],
    },
}


# ── Safety primitives ────────────────────────────────────────────
def _abort(msg: str) -> None:
    print(f"\n[ABORT] {msg}", file=sys.stderr)
    sys.exit(1)


def make_backup(db_bytes: bytes) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"immigration_guidelines_db_v2.v10_patch_backup_{ts}.json"
    bk.write_bytes(db_bytes)
    if not bk.exists() or bk.stat().st_size != len(db_bytes):
        _abort(f"백업 작성 검증 실패: {bk}")
    try:
        json.loads(bk.read_text(encoding="utf-8"))
    except Exception as e:
        _abort(f"백업 JSON readback 실패: {e}")
    return bk


def diff_rows(before: list, after: list) -> list:
    bi = {r.get("row_id"): r for r in before}
    ai = {r.get("row_id"): r for r in after}
    out = []
    for rid in set(bi) | set(ai):
        b = bi.get(rid); a = ai.get(rid)
        if b is None or a is None:
            out.append({"row_id": rid, "added_or_removed": True}); continue
        diff = [k for k in set(b) | set(a) if b.get(k) != a.get(k)]
        if diff:
            out.append({"row_id": rid, "fields": sorted(diff)})
    return out


def load_db() -> tuple:
    raw = DB_PATH.read_bytes()
    return json.loads(raw.decode("utf-8")), raw


def build_patched_db(db: dict) -> tuple:
    new_db = json.loads(json.dumps(db, ensure_ascii=False))
    rows = new_db.get("master_rows") or []
    found = {r.get("row_id"): r for r in rows if r.get("row_id") in PATCHES}
    missing = [rid for rid in PATCHES if rid not in found]
    if missing:
        _abort(f"target row_ids 누락: {missing}")
    patch_report = []
    for rid, spec in PATCHES.items():
        row = found[rid]
        if row.get("detailed_code") != spec["detailed_code"]:
            _abort(f"{rid} detailed_code 불일치")
        if row.get("action_type") != spec["action_type"]:
            _abort(f"{rid} action_type 불일치")
        old_ref = row.get("manual_ref") or []
        if spec.get("replace_all"):
            new_ref = list(spec["add_entries"])
        else:
            existing_sigs = {(e.get("manual"), e.get("page_from"), e.get("page_to")) for e in old_ref}
            new_ref = list(old_ref)
            for ne in spec.get("add_entries") or []:
                sig = (ne["manual"], ne["page_from"], ne["page_to"])
                if sig not in existing_sigs:
                    new_ref.append(ne)
        row["manual_ref"] = new_ref
        patch_report.append({
            "row_id":           rid,
            "detailed_code":    spec["detailed_code"],
            "action_type":      spec["action_type"],
            "title":            spec["title"],
            "preferred_manual": spec["preferred_manual"],
            "old_manual_ref":   old_ref,
            "new_manual_ref":   new_ref,
            "added_entries":    spec.get("add_entries") or [],
            "replace_all":      spec.get("replace_all", False),
            "page_evidence":    spec["page_evidence"],
        })
    return new_db, patch_report


def verify_patch_safety(before_db: dict, after_db: dict) -> list:
    errors = []
    bm = before_db.get("master_rows") or []; am = after_db.get("master_rows") or []
    if len(bm) != 369: errors.append(f"master_rows 개수 이상: {len(bm)}")
    if len(am) != len(bm): errors.append(f"패치 후 개수 불일치: {len(am)}")
    bi = {r.get("row_id"): r for r in bm}; ai = {r.get("row_id"): r for r in am}
    if set(bi) != set(ai): errors.append("row_id 집합 변경됨")
    for rid in bi:
        b = bi[rid]; a = ai.get(rid, {})
        if rid in PATCHES:
            for k in set(b) | set(a):
                if k == "manual_ref": continue
                if b.get(k) != a.get(k):
                    errors.append(f"{rid} non-manual_ref 필드 변경됨: {k}")
        else:
            for k in set(b) | set(a):
                if b.get(k) != a.get(k):
                    errors.append(f"비-target row {rid} 필드 변경됨: {k}")
    for k in set(before_db) | set(after_db):
        if k == "master_rows": continue
        if before_db.get(k) != after_db.get(k):
            errors.append(f"top-level 키 변경됨: {k}")
    return errors


def write_dryrun_reports(patch_report: list, summary: dict) -> None:
    DRY_JSON.write_text(json.dumps({
        "patch_meta": {"version": "v10", "mode": "DRY_RUN",
                       "scope": "6 rows: D-10-2/D-8-2/D-8-1/E-2/F-1-5/D-10-3"},
        "summary": summary, "rows": patch_report,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {DRY_JSON.relative_to(ROOT)} ({DRY_JSON.stat().st_size:,} bytes)")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("patches")
        headers = ["row_id","code","action","title","preferred","replace_all",
                   "old_refs","added_entries","evidence"]
        widths  = [10,12,14,36,14,10,80,100,80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for x in patch_report:
            ws.append([x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                       x["preferred_manual"], x["replace_all"],
                       json.dumps(x["old_manual_ref"], ensure_ascii=False)[:600],
                       json.dumps(x["added_entries"], ensure_ascii=False)[:800],
                       x["page_evidence"][:400]])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCB")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(DRY_XLSX)
        print(f"[OUT] {DRY_XLSX.relative_to(ROOT)} ({DRY_XLSX.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[skip xlsx] {e}")


def main() -> int:
    try: sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    is_apply = args.apply
    if not DB_PATH.exists(): _abort(f"DB 없음: {DB_PATH}")
    db, raw = load_db()
    bytes_before = len(raw)
    bm = db.get("master_rows") or []
    if len(bm) != 369: _abort("master_rows 개수 비정상")
    print(f"[patch] mode={'APPLY' if is_apply else 'DRY-RUN'}  DB={bytes_before:,}  master_rows={len(bm)}")
    print(f"[patch] targets ({len(PATCHES)}): {sorted(PATCHES.keys())}")
    new_db, patch_report = build_patched_db(db)
    errors = verify_patch_safety(db, new_db)
    if errors:
        for e in errors: print(f"  [error] {e}")
        _abort("safety check 실패")
    diffs = diff_rows(bm, new_db.get("master_rows") or [])
    diff_ids = {d["row_id"] for d in diffs}
    if diff_ids != set(PATCHES.keys()):
        _abort(f"diff 불일치: diff={sorted(diff_ids)} expected={sorted(PATCHES.keys())}")
    for d in diffs:
        if d.get("fields") != ["manual_ref"]:
            _abort(f"{d['row_id']} manual_ref 외 필드 변경: {d.get('fields')}")
    summary = {
        "mode": "APPLY" if is_apply else "DRY_RUN",
        "production_db_modified": False,
        "patch_candidate_count": len(patch_report),
        "diff_row_ids": sorted(diff_ids),
        "diff_count": len(diff_ids),
        "non_target_row_diffs": 0,
        "non_manual_ref_field_diffs": 0,
        "db_byte_size_before": bytes_before,
    }
    write_dryrun_reports(patch_report, summary)
    if not is_apply:
        print(f"\n{'='*70}\nPATCH v10 — DRY-RUN ({len(PATCHES)} rows)\n{'='*70}")
        print(f"  targets: {sorted(diff_ids)}")
        print(f"  non-target diffs: 0  /  non-manual_ref field diffs: 0")
        print(f"  DB: {bytes_before:,} → {bytes_before:,} (unchanged)")
        print()
        for x in patch_report:
            mode = "REPLACE" if x["replace_all"] else "ADD"
            print(f"  {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}  [{mode}]")
            print(f"    old: {len(x['old_manual_ref'])} entries  →  new: {len(x['new_manual_ref'])} entries")
            for e in x["added_entries"]:
                print(f"    + {e['manual']} p.{e['page_from']}-{e['page_to']}  [{e['match_text'][:80]}]")
        print()
        print("  Apply command (NOT EXECUTED):\n    python backend/scripts/patch_v10.py --apply")
        return 0
    # APPLY
    print("\n[apply] 백업 생성...")
    bk = make_backup(raw)
    print(f"  backup: {bk.relative_to(ROOT)}")
    DB_PATH.write_text(json.dumps(new_db, ensure_ascii=False, indent=2), encoding="utf-8")
    readback = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rb_rows = readback.get("master_rows") or []
    if len(rb_rows) != 369:
        _abort(f"readback 이상: {len(rb_rows)}\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    errs2 = verify_patch_safety(json.loads(bk.read_text(encoding="utf-8")), readback)
    if errs2:
        for e in errs2: print(f"  [error] {e}")
        _abort(f"사후 verify 실패\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    bytes_after = DB_PATH.stat().st_size
    print(f"\n{'='*70}\nPATCH v10 — APPLIED\n{'='*70}")
    print(f"  backup: {bk.relative_to(ROOT)}")
    print(f"  DB: {bytes_before:,} → {bytes_after:,}  rows updated: {len(patch_report)}")
    print(f"  rollback: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
