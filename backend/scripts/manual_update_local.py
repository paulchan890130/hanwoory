"""manual_update_local.py — 로컬 매뉴얼 업데이트 파이프라인 v1 (rhwp 기반).

설계 원칙 (Manual Update v1)
---------------------------
* **기본 동작은 가볍다.** rhwp 는 1차로 텍스트/해시 추출 + diff + 영향 manual_ref
  후보 생성에만 사용한다. **기본 모드에서는 PDF 를 생성하지 않는다.**
* 변경된 페이지가 있을 때만, 변경 페이지 ± 이웃 페이지에 한해 *검토용* PDF 를 만든다.
* 전체(full) PDF 는 ``--full-pdf`` 를 명시할 때만 만든다.
* 365 개 manual_ref 를 전부 재검사하지 않는다 — 변경 페이지에 연결된 항목만 후보로 낸다.

워크플로우
---------
    1. backend/data/manuals/incoming/ 에 새 HWP/HWPX 를 둔다
    2. python backend/scripts/manual_update_local.py --version YYYYMMDD
    3. backend/data/manuals/staging/YYYYMMDD/ 안에 다음이 생성됨:
         input/             — 사용자 파일의 사본 (incoming → staging)
         rhwp_text/         — page-level JSONL + per-page text
         diff/              — baseline 대비 changed_pages.{json,md}
         candidates/        — manual_ref_update_candidates.{json,xlsx,md}
         review_pdf_pages/  — 변경 페이지 ± 이웃 검토용 PDF (변경 있을 때만)
         reports/           — 요약 보고서
         logs/              — node 실행 로그
         manifest.json      — 전체 메타 (pdf_mode 포함)
       그리고 ``--full-pdf`` 일 때만:
         rhwp_pdf/          — per-page PDF + merged PDF

운영 보호
---------
* immigration_guidelines_db_v2.json 절대 수정 안 함 (읽기 전용 — 후보 생성용).
* 운영 unlocked_*.pdf 절대 덮어쓰지 않음.
* 운영 PDF 뷰어( /api/guidelines/manual-pdf/{manual} ) 동작 변경 없음.
* row-level 만 (후보 생성도 영향 받은 manual_ref 만).

CLI
---
    --version YYYYMMDD           (필수)
    --input-dir DIR              (기본: backend/data/manuals/incoming)
    --baseline-version 260414    (기본)
    --dry-run                    (기본 True)
    --changed-pages-pdf          (변경 페이지 검토 PDF 강제 생성)
    --no-changed-pages-pdf       (기본 review 모드의 변경 페이지 PDF 생성 억제)
    --full-pdf                   (전체 staging PDF 생성 — 무거움)
    --neighbor N                 (변경 페이지 ± N, 기본 1)
    --label-only residence|visa|revision_history
    --max-pages N                (테스트용 — 추출 페이지 수 제한)
    --force                      (같은 버전 재실행 시 재추출 강제)
"""
from __future__ import annotations
import argparse, datetime, hashlib, json, os, pathlib, re, shutil, subprocess, sys

sys.stdout.reconfigure(encoding='utf-8')  # type: ignore[union-attr]

ROOT = pathlib.Path(__file__).resolve().parents[2]
NODE_TOOLS = ROOT / 'tools' / 'rhwp_manual_pipeline'
# 매뉴얼 데이터 디렉토리 — MANUALS_DATA_DIR(기본=backend/data/manuals)로 분리(Persistent Disk 대응).
# config.MANUALS_DATA_DIR 와 동일한 env 계약을 standalone 스크립트에서 직접 읽는다.
MANUALS = pathlib.Path(os.environ.get(
    'MANUALS_DATA_DIR', str(ROOT / 'backend' / 'data' / 'manuals')))
BASELINE_DIR_PARENT = MANUALS / 'baseline'
DB_PATH = ROOT / 'backend' / 'data' / 'immigration_guidelines_db_v2.json'  # 운영 DB — 이동 금지
# version-independent 사람-검토 결정 저장소 (row_id 기준). staging 재생성과 무관하게 유지.
# 절대 immigration_guidelines_db_v2.json 에 자동 반영되지 않는다 (참고/보존 전용).
DECISIONS_PATH = MANUALS / 'manual_review_decisions.json'


def _resolve_baseline_path(rel: str) -> pathlib.Path:
    """baseline manifest 의 repo-relative 경로를 현재 MANUALS 기준으로도 해석한다.

    로컬 기본(MANUALS=backend/data/manuals)에서는 ROOT/rel 그대로 존재한다.
    Persistent Disk(/data/manuals)로 분리된 경우 'manuals/' 이후 구간을 MANUALS 에 붙여
    재해석한다(baseline 이 디스크로 시드된 상황 대응)."""
    p = ROOT / rel
    if p.exists():
        return p
    norm = rel.replace('\\', '/')
    marker = 'manuals/'
    idx = norm.find(marker)
    if idx >= 0:
        return MANUALS / norm[idx + len(marker):]
    return p

LABEL_PATTERNS = [
    ('residence',        re.compile(r'체류민원|residence')),
    ('visa',             re.compile(r'사증민원|visa')),
    ('revision_history', re.compile(r'수정\s*이력|revision[_-]?history')),
]
LABEL_KR = {'residence': '체류민원', 'visa': '사증민원', 'revision_history': '수정이력'}


def classify_label(fname: str) -> str | None:
    for label, rx in LABEL_PATTERNS:
        if rx.search(fname):
            return label
    return None


def sha256(p: pathlib.Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def run_node(args: list[str], log_path: pathlib.Path) -> dict:
    """Run a Node script under tools/rhwp_manual_pipeline.

    Captures stdout JSON lines. Last 'done' line is returned as the meta dict.
    """
    log_path.parent.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        ['node', '--max-old-space-size=4096'] + args,
        capture_output=True, text=True, encoding='utf-8',
    )
    log_path.write_text(proc.stdout + ('\n[stderr]\n' + proc.stderr if proc.stderr else ''),
                        encoding='utf-8')
    if proc.returncode != 0:
        raise RuntimeError(f'node failed (rc={proc.returncode}): see {log_path}')
    last_done: dict = {}
    for line in proc.stdout.splitlines():
        line = line.strip()
        if not line.startswith('{'): continue
        try:
            row = json.loads(line)
            if row.get('event') == 'done':
                last_done = row
        except Exception:
            pass
    return last_done


def merge_per_page_pdfs(per_page_dir: pathlib.Path, out_path: pathlib.Path) -> dict:
    import fitz
    files = sorted(per_page_dir.glob('p*.pdf'),
                   key=lambda p: int(re.search(r'p(\d+)\.pdf$', p.name).group(1)))
    if not files:
        return {'status': 'no per-page PDFs', 'merged_path': None}
    expected = max(int(re.search(r'p(\d+)\.pdf$', f.name).group(1)) for f in files)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out = fitz.open()
    missing = []
    for i in range(1, expected + 1):
        p = per_page_dir / f'p{i:04d}.pdf'
        if not p.exists(): missing.append(i); continue
        with fitz.open(p) as src: out.insert_pdf(src)
    out.save(out_path, garbage=4, deflate=True); out.close()
    return {'status': 'ok' if not missing else 'partial',
            'merged_path': str(out_path),
            'merged_size': out_path.stat().st_size,
            'missing_pages': missing,
            'page_count': expected}


def load_jsonl(p: pathlib.Path) -> list[dict]:
    return [json.loads(l) for l in p.open(encoding='utf-8') if l.strip()]


def normalize(s: str) -> str:
    s = re.sub(r'[\s　]+', '', s or '')
    return re.sub(r'[-‐‑‒–—―]', '-', s)


def _similarity(a: str, b: str, *, cap: int = 4000) -> float:
    """Cheap normalized-text similarity (0..1) for the changed-page table.

    Capped length keeps it fast on long pages."""
    import difflib
    na, nb = normalize(a)[:cap], normalize(b)[:cap]
    if not na and not nb:
        return 1.0
    return round(difflib.SequenceMatcher(None, na, nb).ratio(), 3)


def diff_pages(baseline: list[dict], new: list[dict], *, snippet_len: int = 140) -> list[dict]:
    """Compare two page lists by normalized_text_hash + rhwp_page_index alignment.

    Returns one row per page in max(baseline, new). change_type:
      same / modified / added / deleted / moved

    ``moved`` is detected when a 'modified' page's new content hash matches a
    *different* baseline page (content relocated). ``similarity`` is filled for
    modified/moved rows so the review UI can sort by magnitude of change.
    """
    bn = len(baseline); nn = len(new)
    # baseline hash → list of baseline page indices (for moved detection)
    base_hash_to_pages: dict[str, list[int]] = {}
    for b in baseline:
        h = b.get('normalized_text_hash')
        if h:
            base_hash_to_pages.setdefault(h, []).append(b['rhwp_page_index'])

    out: list[dict] = []
    common = min(bn, nn)
    for i in range(common):
        b = baseline[i]; n = new[i]
        bh = b.get('normalized_text_hash'); nh = n.get('normalized_text_hash')
        if bh == nh:
            out.append({
                'manual_label': n.get('manual_label') or b.get('manual_label'),
                'baseline_page': b['rhwp_page_index'], 'new_page': n['rhwp_page_index'],
                'change_type': 'same', 'similarity': 1.0,
                'baseline_snippet': '', 'new_snippet': '',
                'keywords': sorted(set((b.get('keywords') or []) + (n.get('keywords') or []))),
            })
            continue
        # not same on this index — modified, or moved if new content lives
        # at a different baseline page
        ch = 'modified'; moved_from = None
        same_pages = base_hash_to_pages.get(nh or '', [])
        other = [p for p in same_pages if p != b['rhwp_page_index']]
        if other:
            ch = 'moved'; moved_from = other[0]
        out.append({
            'manual_label': n.get('manual_label') or b.get('manual_label'),
            'baseline_page': b['rhwp_page_index'],
            'new_page': n['rhwp_page_index'],
            'change_type': ch,
            'moved_from': moved_from,
            'similarity': _similarity(b.get('text', ''), n.get('text', '')),
            'baseline_snippet': b.get('text', '')[:snippet_len].replace('\n', ' '),
            'new_snippet':      n.get('text', '')[:snippet_len].replace('\n', ' '),
            'keywords': sorted(set((b.get('keywords') or []) + (n.get('keywords') or []))),
        })
    for i in range(common, nn):
        n = new[i]
        out.append({
            'manual_label': n.get('manual_label'),
            'baseline_page': None, 'new_page': n['rhwp_page_index'],
            'change_type': 'added', 'similarity': None,
            'baseline_snippet': '', 'new_snippet': n.get('text', '')[:snippet_len].replace('\n', ' '),
            'keywords': n.get('keywords') or [],
        })
    for i in range(common, bn):
        b = baseline[i]
        out.append({
            'manual_label': b.get('manual_label'),
            'baseline_page': b['rhwp_page_index'], 'new_page': None,
            'change_type': 'deleted', 'similarity': None,
            'baseline_snippet': b.get('text', '')[:snippet_len].replace('\n', ' '), 'new_snippet': '',
            'keywords': b.get('keywords') or [],
        })
    return out


# ── 사람-검토 결정 보존 (version-independent, row_id 기준) ─────────────────────
# 자동 staging 재생성 시 사람이 내린 decision/note/manual_page 입력이 사라지지 않도록
# manual_review_decisions.json 을 row_id 기준으로 left-join 병합한다. 새 추천 결과가
# 기존 결정을 덮어쓰지 않는다. 후보 페이지가 바뀌면 결정은 보존하되 재검토 플래그만 단다.
_PRESERVED_DECISION_FIELDS = (
    'decision', 'decision_note', 'reviewed', 'reviewed_candidate_page',
    'manual_page_from', 'manual_page_to', 'applied', 'updated_at', 'orphaned',
)


def load_decisions() -> dict:
    """manual_review_decisions.json 의 decisions(dict, row_id 키)를 로드.

    파일이 없거나 손상돼도 절대 raise 하지 않고 {} 반환 (파이프라인 보호)."""
    if not DECISIONS_PATH.exists():
        return {}
    try:
        data = json.loads(DECISIONS_PATH.read_text(encoding='utf-8'))
    except Exception as e:
        print(f'[warn] manual_review_decisions.json 읽기 실패 — 빈 결정으로 진행: {e}',
              file=sys.stderr)
        return {}
    d = data.get('decisions') if isinstance(data, dict) else None
    return d if isinstance(d, dict) else {}


def save_decisions(decisions: dict) -> None:
    """decisions(dict, row_id 키)를 메타와 함께 저장. immigration DB 는 건드리지 않는다."""
    payload = {
        '_schema': 1,
        '_updated_at': datetime.datetime.now().isoformat(timespec='seconds'),
        '_note': ('version-independent manual_ref review decisions keyed by row_id. '
                  'Never auto-applied to immigration_guidelines_db_v2.json — '
                  'reference/preservation only.'),
        'decisions': decisions,
    }
    DECISIONS_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def _flag_candidate_recheck(c: dict, d: dict) -> None:
    """검토 당시 후보(reviewed_candidate_page)와 새 후보(candidate_page_from)가 다르면
    결정은 유지한 채 candidate_changed/needs_recheck/recheck_reason 만 표시."""
    prev = d.get('reviewed_candidate_page')
    new = c.get('candidate_page_from')
    reasons: list[str] = []
    if prev is not None and new and int(prev) != int(new):
        c['candidate_changed'] = True
        reasons.append(f'candidate page {prev}→{new}')
        if d.get('applied'):
            reasons.append('applied page may be stale')
    if reasons:
        c['needs_recheck'] = True
        c['recheck_reason'] = '; '.join(reasons)


def merge_decisions_into_candidates(candidates: list[dict], decisions: dict) -> list[dict]:
    """저장된 결정(row_id 키)을 새로 생성된 candidates 에 left-join 병합.

    사람 입력(decision/note/manual_page/reviewed/applied)은 보존되며 새 추천이
    덮어쓰지 않는다. 결정이 없는 후보는 기본값만 채워진다."""
    for c in candidates:
        c.setdefault('candidate_changed', False)
        c.setdefault('needs_recheck', False)
        c.setdefault('recheck_reason', '')
        c.setdefault('orphaned', False)
        c.setdefault('reviewed', False)
        c.setdefault('applied', False)
        c.setdefault('reviewed_candidate_page', None)
        c.setdefault('manual_page_from', None)
        c.setdefault('manual_page_to', None)
        c.setdefault('decision', '')
        c.setdefault('decision_note', '')
        d = decisions.get(c.get('row_id'))
        if not d:
            continue
        # 사람 결정 보존 (새 추천으로 덮어쓰지 않음)
        c['decision'] = d.get('decision', '')
        c['decision_note'] = d.get('decision_note', '')
        # 기존 산출물 컬럼(user_decision/user_note)도 동기화하여 하위 호환 유지
        c['user_decision'] = d.get('decision', '')
        c['user_note'] = d.get('decision_note', '')
        c['reviewed'] = bool(d.get('reviewed'))
        c['applied'] = bool(d.get('applied'))
        c['reviewed_candidate_page'] = d.get('reviewed_candidate_page')
        if d.get('manual_page_from') is not None:
            c['manual_page_from'] = d.get('manual_page_from')
            c['manual_page_to'] = d.get('manual_page_to')
        _flag_candidate_recheck(c, d)
    return candidates


def reconcile_decision_store(decisions: dict, candidates: list[dict]) -> dict:
    """현재 후보 집합에 없는 row_id 의 결정을 orphaned=True 로 표시(보존, 삭제 안 함).
    다시 등장한 결정은 orphaned=False 로 복구. decision/note 값은 절대 변경하지 않는다."""
    present = {c.get('row_id') for c in candidates}
    for rid, d in decisions.items():
        if not isinstance(d, dict):
            continue
        if rid in present:
            if d.get('orphaned'):
                d['orphaned'] = False
        else:
            d['orphaned'] = True
    return decisions


def make_ref_candidates(changed: list[dict], new_pages_by_label: dict[str, list[dict]]) -> list[dict]:
    """Generate manual_ref update candidates ONLY for refs whose page range
    overlaps a changed page or whose match_text appears on a changed page.
    Unaffected refs are NOT included. (immigration_guidelines_db_v2.json is read
    only — never modified here.)"""
    db = json.loads(DB_PATH.read_text(encoding='utf-8'))
    master = db.get('master_rows', [])

    # Set of (manual_kr, page_no) that are non-same
    affected: dict[str, set[int]] = {}
    new_page_lookup: dict[str, dict[int, dict]] = {}
    for label, rows in new_pages_by_label.items():
        new_page_lookup[label] = {r['rhwp_page_index']: r for r in rows}
    label_to_kr = {LABEL_KR[k]: k for k in LABEL_KR}

    for c in changed:
        if c['change_type'] == 'same': continue
        label = c['manual_label']
        kr = LABEL_KR.get(label, label)
        affected.setdefault(kr, set())
        if c.get('baseline_page'):
            affected[kr].add(c['baseline_page'])
        if c.get('new_page'):
            affected[kr].add(c['new_page'])

    candidates: list[dict] = []
    for row in master:
        for i, ref in enumerate(row.get('manual_ref') or []):
            manual_kr = ref.get('manual')
            if manual_kr not in affected: continue
            pf = int(ref.get('page_from') or 0); pt = int(ref.get('page_to') or 0)
            mt = (ref.get('match_text') or '').strip()
            label = label_to_kr.get(manual_kr)
            if not label: continue
            overlap = any(p for p in range(pf, pt + 1) if p in affected[manual_kr])
            text_hit_pages: list[int] = []
            if not overlap:
                # text-hit fallback: only specific enough mt qualifies (>= 8
                # normalized chars filters generic Korean substrings).
                norm_mt = normalize(mt) if mt else ''
                if len(norm_mt) >= 8:
                    for pn in affected[manual_kr]:
                        page = new_page_lookup.get(label, {}).get(pn)
                        if page and norm_mt in normalize(page.get('text', '')):
                            text_hit_pages.append(pn)
                if not text_hit_pages:
                    continue
            # affected → build candidate
            new_snip = ''
            cand_pf = pf; cand_pt = pt
            new_page = new_page_lookup.get(label, {}).get(pf or 1)
            if new_page:
                new_snip = new_page.get('text', '')[:140].replace('\n', ' ')
            change_types = sorted({c['change_type'] for c in changed
                                   if c['manual_label'] == label
                                   and (c.get('baseline_page') in range(pf, pt+1)
                                        or c.get('new_page') in range(pf, pt+1))})
            confidence = 'high' if change_types == ['modified'] else 'review'
            action = 'review' if confidence == 'review' else 'remap_candidate'
            if text_hit_pages and not overlap:
                action = 'review'  # text matched elsewhere — needs human
                cand_pf = cand_pt = text_hit_pages[0]
                confidence = 'medium'
            candidates.append({
                'row_id': row.get('row_id'),
                'item_index': i,
                'manual_label': label,
                'old_page_from': pf, 'old_page_to': pt,
                'candidate_page_from': cand_pf, 'candidate_page_to': cand_pt,
                'reason': f'baseline pages {pf}-{pt} overlap changed pages '
                          f'{sorted(affected[manual_kr] & set(range(pf,pt+1))) or text_hit_pages}',
                'change_type': '+'.join(change_types) or 'unknown',
                'confidence': confidence,
                'action': action,
                'baseline_snippet': '',
                'new_snippet': new_snip,
                'match_text': mt,
                'detailed_code': row.get('detailed_code'),
                'business_name': row.get('business_name'),
                'major_action_std': row.get('major_action_std'),
                'user_decision': '',
                'user_note': '',
            })
    return candidates


def write_candidates(candidates: list[dict], out_dir: pathlib.Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / 'manual_ref_update_candidates.json').write_text(
        json.dumps(candidates, ensure_ascii=False, indent=2), encoding='utf-8')

    lines = ['# manual_ref_update_candidates.md\n']
    lines.append(f'count: **{len(candidates)}** affected refs '
                 f'(unaffected refs omitted by design — no full 365-entry review)\n')
    if candidates:
        lines.append('| row_id | code | manual | old | candidate | confidence | action |')
        lines.append('|---|---|---|---|---|---|---|')
        for c in candidates:
            lines.append(f'| {c["row_id"]} | `{c["detailed_code"]}` | {c["manual_label"]} | '
                         f'p.{c["old_page_from"]}-{c["old_page_to"]} | '
                         f'p.{c["candidate_page_from"]}-{c["candidate_page_to"]} | '
                         f'{c["confidence"]} | {c["action"]} |')
    (out_dir / 'manual_ref_update_candidates.md').write_text('\n'.join(lines) + '\n', encoding='utf-8')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'candidates'
        head = ['row_id','item_index','manual_label','detailed_code','business_name','major_action_std',
                'old_page_from','old_page_to','candidate_page_from','candidate_page_to',
                'reason','change_type','confidence','action',
                'match_text','baseline_snippet','new_snippet','user_decision','user_note',
                'reviewed','reviewed_candidate_page','candidate_changed','needs_recheck',
                'recheck_reason','orphaned']
        ws.append(head)
        for c in ws[1]: c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')
        for c in candidates:
            ws.append([c.get(k) for k in head])
        widths = [11,7,13,11,18,16,8,8,10,10,40,16,11,12,30,40,40,16,30,
                  9,12,10,12,30,9]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i) if i <= 26 else f'A{chr(64+i-26)}'].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row: c.alignment = Alignment(vertical='top', wrap_text=True)
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions
        wb.save(out_dir / 'manual_ref_update_candidates.xlsx')
    except Exception as e:
        print(f'[warn] xlsx skipped: {e}', file=sys.stderr)


# ── 변경 페이지 ± 이웃 검토 PDF ────────────────────────────────────────────────
def collect_affected_pages(changed: list[dict], neighbor: int,
                           page_counts: dict[str, int]) -> dict[str, list[int]]:
    """변경(non-same) 페이지의 *new_page* 를 모아 ± neighbor 로 확장하고 문서 경계로
    clamp 한다. deleted 는 new doc 에 없으므로 제외 (렌더 불가)."""
    by_label: dict[str, set[int]] = {}
    for c in changed:
        if c['change_type'] == 'same':
            continue
        np = c.get('new_page')
        if not np:
            continue
        label = c['manual_label']
        by_label.setdefault(label, set())
        pc = page_counts.get(label, np)
        for p in range(np - neighbor, np + neighbor + 1):
            if 1 <= p <= pc:
                by_label[label].add(p)
    return {label: sorted(pages) for label, pages in by_label.items() if pages}


def generate_changed_page_pdfs(affected: dict[str, list[int]],
                               src_by_label: dict[str, pathlib.Path],
                               review_dir: pathlib.Path,
                               logs_dir: pathlib.Path) -> dict[str, list[int]]:
    """변경 페이지 ± 이웃에 대해서만 검토용 PDF 생성.

    출력: review_dir/{label}/p####.pdf  (--flat, 같은 페이지 재사용 위해 --skip-existing)
    rhwp_pdf/ 트리는 만들지 않는다."""
    generated: dict[str, list[int]] = {}
    for label, pages in affected.items():
        src = src_by_label.get(label)
        if not src or not pages:
            continue
        run_node([
            str(NODE_TOOLS / 'generate_pdf.mjs'),
            '--src', str(src), '--label', label,
            '--pages', ','.join(str(p) for p in pages),
            '--flat', '--skip-existing',
            '--out-dir', str(review_dir),
        ], logs_dir / f'review_pdf_{label}.log')
        present = sorted(
            int(re.search(r'p(\d+)\.pdf$', f.name).group(1))
            for f in (review_dir / label).glob('p*.pdf')
        )
        generated[label] = present
    return generated


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    ap.add_argument('--version', required=True, help='version string e.g. 260620 or 260414_test')
    ap.add_argument('--input-dir', default=str(MANUALS / 'incoming'))
    ap.add_argument('--baseline-version', default='260414')
    ap.add_argument('--dry-run', action='store_true', default=True)
    ap.add_argument('--changed-pages-pdf', action='store_true',
                    help='변경 페이지 ± 이웃 검토 PDF 강제 생성')
    ap.add_argument('--no-changed-pages-pdf', action='store_true',
                    help='기본 review 모드의 변경 페이지 PDF 생성을 억제')
    ap.add_argument('--full-pdf', action='store_true',
                    help='전체 staging PDF 생성 (무거움 — 명시할 때만)')
    ap.add_argument('--neighbor', type=int, default=1, help='변경 페이지 ± N (기본 1)')
    ap.add_argument('--label-only', default=None,
                    choices=['residence', 'visa', 'revision_history'])
    ap.add_argument('--max-pages', type=int, default=None,
                    help='테스트용 — 추출 페이지 수 제한')
    ap.add_argument('--force', action='store_true',
                    help='같은 버전 재실행 시 재추출 강제')
    args = ap.parse_args()

    incoming = pathlib.Path(args.input_dir)
    staging = MANUALS / 'staging' / args.version

    # rerun guard
    manifest_path = staging / 'manifest.json'
    if manifest_path.exists() and not args.force:
        print(f'[notice] staging/{args.version}/manifest.json already exists. '
              f'Use --force to rerun (re-extract). Proceeding with --skip-existing reuse.',
              file=sys.stderr)

    for sub in ('input', 'rhwp_text', 'diff', 'candidates', 'reports', 'logs', 'review_pdf_pages'):
        (staging / sub).mkdir(parents=True, exist_ok=True)

    baseline_manifest_p = BASELINE_DIR_PARENT / args.baseline_version / 'manifest.json'
    if not baseline_manifest_p.exists():
        print(f'[FATAL] baseline manifest not found: {baseline_manifest_p}', file=sys.stderr)
        return 2
    baseline_manifest = json.loads(baseline_manifest_p.read_text(encoding='utf-8'))

    # Discover incoming files
    if not incoming.exists():
        print(f'[FATAL] incoming dir not found: {incoming}', file=sys.stderr)
        return 2
    candidates_in: dict[str, pathlib.Path] = {}
    for p in incoming.iterdir():
        if not p.is_file(): continue
        if p.suffix.lower() not in ('.hwp', '.hwpx'): continue
        label = classify_label(p.name)
        if not label: continue
        if args.label_only and label != args.label_only: continue
        candidates_in[label] = p

    if not candidates_in:
        print(f'[FATAL] no incoming HWP/HWPX files found in {incoming}', file=sys.stderr)
        return 3

    print(f'[input] {len(candidates_in)} file(s): '
          + ', '.join(f'{L}={p.name}' for L, p in candidates_in.items()))
    print(f'[mode]  full_pdf={args.full_pdf}  changed_pages_pdf={args.changed_pages_pdf}  '
          f'neighbor=±{args.neighbor}  max_pages={args.max_pages}')

    manifest = {
        'version': args.version,
        'baseline_version': args.baseline_version,
        'created_at': datetime.datetime.now().isoformat(timespec='seconds'),
        'status': 'STAGED_FOR_REVIEW',
        'manuals': {},
        'changed_pages_summary': {},
        'changed_page_count': 0,
        'manual_ref_candidate_count': 0,
        'pdf_mode': 'none',
        'review_pdf_pages': {},
        'dry_run': True,
    }

    all_changed: list[dict] = []
    new_pages_by_label: dict[str, list[dict]] = {}
    page_counts: dict[str, int] = {}
    src_by_label: dict[str, pathlib.Path] = {}

    for label, src in candidates_in.items():
        print(f'\n=== {label} ===  {src.name}')
        dst_input = staging / 'input' / src.name
        if args.force or not dst_input.exists():
            shutil.copy2(src, dst_input)
        src_by_label[label] = dst_input
        src_sha = sha256(dst_input)

        # 1) extract text/hash (LIGHT — no PDF)
        jsonl_path = staging / 'rhwp_text' / f'{label}_pages.jsonl'
        if args.force or not jsonl_path.exists():
            extra = ['--max-pages', str(args.max_pages)] if args.max_pages else []
            run_node([
                str(NODE_TOOLS / 'extract.mjs'),
                '--src', str(dst_input), '--label', label,
                '--out-dir', str(staging / 'rhwp_text'),
            ] + extra, staging / 'logs' / f'extract_{label}.log')
        new_pages_by_label[label] = load_jsonl(jsonl_path)
        page_count = len(new_pages_by_label[label])
        page_counts[label] = page_count
        print(f'  extracted {page_count} pages → rhwp_text/{label}_pages.jsonl')

        # 2) diff vs baseline (text/hash only)
        b_meta = (baseline_manifest['manuals'] or {}).get(label, {}) or {}
        b_jsonl_rel = (b_meta.get('rhwp_jsonl') or {}).get('path')
        ct: dict[str, int] = {}
        b_jsonl_path = _resolve_baseline_path(b_jsonl_rel) if b_jsonl_rel else None
        if b_jsonl_path and b_jsonl_path.exists():
            b_pages = load_jsonl(b_jsonl_path)
            changed = diff_pages(b_pages, new_pages_by_label[label])
            (staging / 'diff' / f'{label}_changed_pages.json').write_text(
                json.dumps(changed, ensure_ascii=False, indent=2), encoding='utf-8')
            for c in changed: ct[c['change_type']] = ct.get(c['change_type'], 0) + 1
            print(f'  diff vs baseline: {ct}')
            all_changed.extend(changed)
            manifest['changed_pages_summary'][label] = ct
        else:
            print(f'  [warn] baseline jsonl missing for {label}; diff skipped')

        manifest['manuals'][label] = {
            'source_file':   src.name,
            'source_sha256': src_sha,
            'page_count':    page_count,
            'jsonl_path':    f'backend/data/manuals/staging/{args.version}/rhwp_text/{label}_pages.jsonl',
            'text_dir':      f'backend/data/manuals/staging/{args.version}/rhwp_text/page_text/{label}',
            'pdf_path':      None,
            'pdf_size':      None,
        }

    # combined changed_pages.json (non-same only — what review cares about)
    non_same = [c for c in all_changed if c['change_type'] != 'same']
    (staging / 'diff' / 'changed_pages.json').write_text(
        json.dumps(non_same, ensure_ascii=False, indent=2), encoding='utf-8')
    manifest['changed_page_count'] = len(non_same)

    # changed_pages.md
    md = [f'# changed_pages.md (version {args.version} vs baseline {args.baseline_version})\n']
    for label, ct in manifest['changed_pages_summary'].items():
        total = sum(ct.values()) if ct else 0
        md.append(f'## {label}')
        md.append(f'- total pages compared: {total}')
        for k in ('same', 'modified', 'moved', 'added', 'deleted'):
            md.append(f'- {k}: {ct.get(k, 0)}')
        md.append(f'- non-same total: {total - ct.get("same", 0)}')
        md.append('')
    (staging / 'diff' / 'changed_pages.md').write_text('\n'.join(md) + '\n', encoding='utf-8')

    # ref candidates (only affected — never touches the production DB)
    candidates_list = make_ref_candidates(all_changed, new_pages_by_label)
    # 사람-검토 결정 보존: row_id 기준 left-join 병합 (staging 재생성과 무관하게 유지)
    decisions = load_decisions()
    candidates_list = merge_decisions_into_candidates(candidates_list, decisions)
    if decisions:
        decisions = reconcile_decision_store(decisions, candidates_list)
        save_decisions(decisions)
        preserved = sum(1 for c in candidates_list if c.get('reviewed') or c.get('applied'))
        recheck = sum(1 for c in candidates_list if c.get('needs_recheck'))
        orphaned = sum(1 for d in decisions.values()
                       if isinstance(d, dict) and d.get('orphaned'))
        print(f'[decisions] preserved={preserved} needs_recheck={recheck} '
              f'orphaned={orphaned} (store: manual_review_decisions.json)')
    write_candidates(candidates_list, staging / 'candidates')
    manifest['manual_ref_candidate_count'] = len(candidates_list)
    print(f'\n[candidates] {len(candidates_list)} affected manual_ref rows '
          f'(unaffected refs intentionally omitted)')

    # ── PDF generation decision ───────────────────────────────────────────────
    # priority: full > changed-pages (explicit OR default review mode w/ changes)
    want_changed_pdf = (args.changed_pages_pdf
                        or (not args.no_changed_pages_pdf and not args.full_pdf))
    if args.full_pdf:
        manifest['pdf_mode'] = 'full'
        print('\n[pdf] FULL mode — generating per-page + merged PDF for every page...')
        for label, src in src_by_label.items():
            gen = run_node([
                str(NODE_TOOLS / 'generate_pdf.mjs'),
                '--src', str(src), '--label', label,
                '--out-dir', str(staging / 'rhwp_pdf'), '--skip-existing',
            ], staging / 'logs' / f'gen_pdf_{label}.log')
            merged = merge_per_page_pdfs(
                staging / 'rhwp_pdf' / 'pdf_pages' / label,
                staging / 'rhwp_pdf' / f'rhwp_{args.version}_{label}.pdf',
            )
            print(f'  {label}: {gen.get("done")} pages, errs={gen.get("errs")} → '
                  f'{merged.get("merged_path")} ({merged.get("merged_size")})')
            mm = manifest['manuals'].get(label, {})
            if merged.get('merged_path'):
                mm['pdf_path'] = str(pathlib.Path(merged['merged_path']).relative_to(ROOT))
                mm['pdf_size'] = merged.get('merged_size')
    elif want_changed_pdf and non_same:
        affected = collect_affected_pages(non_same, args.neighbor, page_counts)
        if affected:
            manifest['pdf_mode'] = 'changed-pages-only'
            total_pages = sum(len(v) for v in affected.values())
            print(f'\n[pdf] CHANGED-PAGES mode — rendering {total_pages} page(s) '
                  f'(changed ± {args.neighbor}) for review...')
            generated = generate_changed_page_pdfs(
                affected, src_by_label,
                staging / 'review_pdf_pages', staging / 'logs')
            manifest['review_pdf_pages'] = generated
            for label, pages in generated.items():
                print(f'  {label}: review pages {pages}')
        else:
            print('\n[pdf] no renderable changed pages (deleted-only); pdf_mode=none')
    else:
        if not non_same:
            print('\n[pdf] no changed pages — pdf_mode=none (nothing to render)')
        else:
            print('\n[pdf] changed pages exist but PDF generation suppressed — pdf_mode=none')

    # reports
    rep = [f'# Staging report — version {args.version}\n',
           f'- baseline: {args.baseline_version}',
           f'- dry_run: True',
           f'- created_at: {manifest["created_at"]}',
           f'- pdf_mode: **{manifest["pdf_mode"]}**',
           f'- changed pages (non-same): {manifest["changed_page_count"]}',
           f'- manual_ref candidates: {len(candidates_list)} (affected only)',
           '',
           '## per-manual',
           '| label | source | sha256 | pages |',
           '|---|---|---|---:|']
    for label, m in manifest['manuals'].items():
        rep.append(f'| {label} | {m["source_file"]} | {(m["source_sha256"] or "")[:12]} | {m["page_count"]} |')
    rep.append('')
    rep.append('## changes summary')
    for label, ct in manifest['changed_pages_summary'].items():
        rep.append(f'- **{label}**: {ct}')
    rep.append('')
    rep.append('## review endpoints (admin)')
    rep.append(f'- GET /api/guidelines/manual-staging/{args.version}/manifest')
    rep.append(f'- GET /api/guidelines/manual-staging/{args.version}/changed-pages')
    rep.append(f'- GET /api/guidelines/manual-staging/{args.version}/candidates')
    rep.append(f'- GET /api/guidelines/manual-staging/{args.version}/{{label}}/review-page/{{page_no}}/pdf')
    if manifest['pdf_mode'] == 'full':
        for label, m in manifest['manuals'].items():
            if m['pdf_path']:
                rep.append(f'- GET /api/guidelines/manual-pdf-staging/{args.version}/{label}/download')
    (staging / 'reports' / 'staging_report.md').write_text('\n'.join(rep) + '\n', encoding='utf-8')

    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'\n[manifest] {manifest_path}  (pdf_mode={manifest["pdf_mode"]})')
    print(f'[staging]  {staging}')
    print('\n[done] dry-run complete. No production DB/PDF changed.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
