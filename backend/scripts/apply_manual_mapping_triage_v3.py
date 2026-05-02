"""
v3 AUTO_SAFE 적용 스크립트 (DEFAULT: dry-run)

기본 동작 = dry-run.
실제 DB 쓰기는 --apply 플래그를 명시한 경우에만 수행.

대상:
  manual_mapping_triage_v3.json 의 triage_category=AUTO_SAFE +
  apply_candidate=true + comparison_status=changed 인 row만.
  manual_ref 필드 외에는 어떤 필드도 손대지 않음.

안전장치:
  - dry-run 기본
  - 검증 실패 candidate 존재 시 abort
  - --apply 시 DB를 backend/data/backups/ 로 timestamp 백업 (실패 시 abort)
  - JSON 사전·사후 검증 (실패 시 abort + 복구 안내)
  - DB master_rows 개수 보존 검증

용법:
  # dry-run (기본)
  python backend/scripts/apply_manual_mapping_triage_v3.py

  # 실제 적용 (수동 검토 후에만)
  python backend/scripts/apply_manual_mapping_triage_v3.py --apply
"""
from __future__ import annotations
import argparse, json, sys, shutil
from datetime import datetime
from pathlib import Path

ROOT      = Path(__file__).parent.parent.parent
DB_PATH   = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
TRIAGE    = ROOT / "backend" / "data" / "manuals" / "manual_mapping_triage_v3.json"
BLOCKLIST = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_blocklist_v3.json"
BACKUPS   = ROOT / "backend" / "data" / "backups"

OUT_DRYRUN_JSON = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_dryrun_v3.json"
OUT_DRYRUN_XLSX = ROOT / "backend" / "data" / "manuals" / "manual_mapping_apply_dryrun_v3.xlsx"

VALID_MANUALS = ("체류민원", "사증민원")


def load_blocklist() -> tuple[set[str], dict]:
    """blocklist JSON 로드 — 파일 없으면 빈 set, malformed면 abort."""
    if not BLOCKLIST.exists():
        return set(), {}
    try:
        data = json.loads(BLOCKLIST.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[ABORT] blocklist JSON 파싱 실패: {e}")
        sys.exit(7)
    blocked = data.get("blocked_row_ids")
    if not isinstance(blocked, list) or not all(isinstance(x, str) for x in blocked):
        print(f"[ABORT] blocklist 형식 비정상: blocked_row_ids 가 문자열 list 아님")
        sys.exit(7)
    return set(blocked), data


# ─── 한 줄 검증 함수 ─────────────────────────────────────────────
def is_apply_candidate(triage_row: dict) -> tuple[bool, str]:
    """
    8중 조건 모두 충족해야 True.
    실패 시 (False, 거절 사유) 반환.
    """
    t = triage_row
    if t.get("triage_category") != "AUTO_SAFE":
        return False, f"category={t.get('triage_category')}"
    if not t.get("apply_candidate"):
        return False, "apply_candidate=false"
    if t.get("comparison_status") != "changed":
        return False, f"status={t.get('comparison_status')}"
    if t.get("confidence") not in ("exact", "high"):
        return False, f"confidence={t.get('confidence')}"
    if t.get("routing_warning"):
        return False, "routing_warning 있음"
    if t.get("comparison_reason") == "both_manuals_plausible":
        return False, "both_manuals_plausible"
    if t.get("action_type") == "APPLICATION_CLAIM":
        return False, "APPLICATION_CLAIM"
    if not t.get("detailed_code"):
        return False, "empty detailed_code"

    proposed = t.get("proposed_manual_ref") or []
    if not proposed:
        return False, "no proposed_manual_ref"
    p0 = proposed[0]
    manual = p0.get("manual")
    if manual not in VALID_MANUALS:
        return False, f"invalid manual={manual!r}"
    pf = p0.get("page_from")
    pt = p0.get("page_to")
    if not isinstance(pf, int) or pf <= 0:
        return False, f"invalid page_from={pf!r}"
    if not isinstance(pt, int) or pt <= 0:
        return False, f"invalid page_to={pt!r}"
    return True, ""


# ─── 메인 ───────────────────────────────────────────────────────
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    parser = argparse.ArgumentParser(description="AUTO_SAFE manual_ref 적용 (dry-run by default)")
    parser.add_argument("--apply", action="store_true",
                        help="실제 DB 쓰기 (기본은 dry-run, DB 절대 미수정)")
    args = parser.parse_args()
    mode = "apply" if args.apply else "dry-run"

    # ── abort 조건: 입력 파일 존재 ──
    if not TRIAGE.exists():
        print(f"[ABORT] triage 파일 없음: {TRIAGE}")
        sys.exit(1)
    if not DB_PATH.exists():
        print(f"[ABORT] DB 파일 없음: {DB_PATH}")
        sys.exit(1)

    print(f"[apply-v3] mode={mode}")
    print(f"[apply-v3] triage v3 로드...")
    triage = json.loads(TRIAGE.read_text(encoding="utf-8"))
    triage_rows = triage.get("rows") or []
    print(f"  triage rows: {len(triage_rows)}")

    # blocklist 로드
    blocked_ids, blocklist_meta = load_blocklist()
    if BLOCKLIST.exists():
        print(f"[apply-v3] blocklist 로드: {BLOCKLIST.relative_to(ROOT)}")
        print(f"  blocked row_ids: {len(blocked_ids)} ({sorted(blocked_ids)})")
        if blocklist_meta.get("reason"):
            print(f"  reason: {blocklist_meta['reason']}")
    else:
        print(f"[apply-v3] blocklist 없음 (선택적)")

    print(f"[apply-v3] DB 로드 (dry-run에선 검증용 메모리 카피만)...")
    db_text_orig = DB_PATH.read_text(encoding="utf-8")
    try:
        db = json.loads(db_text_orig)
    except Exception as e:
        print(f"[ABORT] DB JSON 사전 파싱 실패: {e}")
        sys.exit(4)
    db_rows = db.get("master_rows") or []
    db_index = {r.get("row_id"): r for r in db_rows}
    print(f"  DB rows: {len(db_rows)}")

    # ── 카테고리 분리 ──
    auto_safe    = [t for t in triage_rows if t.get("triage_category") == "AUTO_SAFE"]
    auto_changed = [t for t in auto_safe if t.get("comparison_status") == "changed"]
    auto_same    = [t for t in auto_safe if t.get("comparison_status") == "same"]
    non_auto     = [t for t in triage_rows if t.get("triage_category") != "AUTO_SAFE"]

    # ── candidate 검증 + blocklist 적용 ──
    candidates: list[dict] = []
    rejected:   list[dict] = []
    blocked_in_run: list[dict] = []
    original_would_update = 0
    for t in auto_changed:
        ok, why = is_apply_candidate(t)
        if not ok:
            rejected.append({"row_id": t.get("row_id"), "title": t.get("title"), "reason": why})
            continue
        rid = t.get("row_id")
        if rid not in db_index:
            rejected.append({"row_id": rid, "title": t.get("title"),
                             "reason": f"row_id {rid} not in DB"})
            continue
        # 검증을 통과한 candidate 카운트 (blocklist 적용 전)
        original_would_update += 1
        # blocklist 필터
        if rid in blocked_ids:
            blocked_in_run.append({
                "row_id": rid,
                "title": t.get("title"),
                "detailed_code": t.get("detailed_code"),
                "action_type": t.get("action_type"),
                "reason": "blocklist 차단 — manual_mapping_apply_blocklist_v3.json",
            })
            continue
        candidates.append(t)

    # ── diff 빌드 ──
    diff: list[dict] = []
    for t in candidates:
        rid = t["row_id"]
        db_row = db_index[rid]
        old_ref = db_row.get("manual_ref") or []
        new_ref = t["proposed_manual_ref"]
        old_p0 = old_ref[0] if old_ref else {}
        new_p0 = new_ref[0]
        diff.append({
            "row_id":         rid,
            "title":          t.get("title"),
            "action_type":    t.get("action_type"),
            "detailed_code":  t.get("detailed_code"),
            "old_manual_ref": old_ref,
            "new_manual_ref": new_ref,
            "old_manual":     old_p0.get("manual", ""),
            "old_page_from":  old_p0.get("page_from", ""),
            "old_page_to":    old_p0.get("page_to", ""),
            "new_manual":     new_p0.get("manual", ""),
            "new_page_from":  new_p0.get("page_from"),
            "new_page_to":    new_p0.get("page_to"),
            "confidence":     t.get("confidence"),
            "method":         t.get("method"),
        })

    summary = {
        "auto_safe_total":           len(auto_safe),
        "auto_safe_changed":         len(auto_changed),
        "auto_safe_same_skipped":    len(auto_same),
        "non_auto_safe_skipped":     len(non_auto),
        "original_would_update":     original_would_update,
        "blocked_count":             len(blocked_in_run),
        "would_update":              len(candidates),
        "rejected_candidates":       len(rejected),
    }

    # ── 콘솔 (요약 부분) ──
    print(f"\n  AUTO_SAFE total:                {summary['auto_safe_total']}")
    print(f"  AUTO_SAFE changed:              {summary['auto_safe_changed']}")
    print(f"  AUTO_SAFE same (skipped):       {summary['auto_safe_same_skipped']}")
    print(f"  non-AUTO_SAFE (skipped):        {summary['non_auto_safe_skipped']}")
    print(f"  original would-update:          {summary['original_would_update']}  (before blocklist)")
    print(f"  blocked by blocklist:           {summary['blocked_count']}")
    print(f"  final would-update:             {summary['would_update']}")
    print(f"  rejected candidates:            {summary['rejected_candidates']}")

    if blocked_in_run:
        print(f"\n  blocked rows ({len(blocked_in_run)}):")
        for b in blocked_in_run:
            print(f"    - {b['row_id']:8s} {b['detailed_code']:8s} {b['action_type']:13s} {b['title']}")

    if rejected:
        print(f"\n  rejected sample (~5):")
        for r in rejected[:5]:
            print(f"    - {r['row_id']}: {r['reason']}")
        if args.apply:
            print(f"[ABORT] --apply 모드: 검증 실패 candidate {len(rejected)}건 존재 — 중단")
            sys.exit(2)

    # ── dry-run output 저장 (apply 모드여도 진단용으로 저장) ──
    out = {
        "mode":         mode,
        "ran_at":       datetime.now().isoformat(),
        "triage_file":  str(TRIAGE.relative_to(ROOT)),
        "db_file":      str(DB_PATH.relative_to(ROOT)),
        "blocklist":    str(BLOCKLIST.relative_to(ROOT)) if BLOCKLIST.exists() else None,
        "blocklist_meta": blocklist_meta,
        "summary":      summary,
        "blocked_in_run": blocked_in_run,
        "diff":         diff,
        "rejected":     rejected,
    }
    OUT_DRYRUN_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_DRYRUN_JSON.relative_to(ROOT)} ({OUT_DRYRUN_JSON.stat().st_size:,} bytes)")

    # XLSX
    xlsx_written = False
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("would_update")
        headers = ["row_id","title","action_type","detailed_code",
                   "old_manual","old_pf","old_pt",
                   "new_manual","new_pf","new_pt",
                   "confidence","method"]
        widths  = [10, 30, 14, 12, 10, 8, 8, 10, 8, 8, 10, 24]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center")
        for d in diff:
            ws.append([
                d["row_id"], d["title"], d["action_type"], d["detailed_code"],
                d["old_manual"], d["old_page_from"], d["old_page_to"],
                d["new_manual"], d["new_page_from"], d["new_page_to"],
                d["confidence"], d["method"],
            ])
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        ws2 = wb.create_sheet("summary")
        ws2.append(["metric", "value"])
        for c in ws2[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws2.append([k, v])
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 16

        if rejected:
            ws3 = wb.create_sheet("rejected")
            ws3.append(["row_id", "title", "reason"])
            for c in ws3[1]:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFCCCB")
            for r in rejected:
                ws3.append([r["row_id"], r["title"], r["reason"]])
            ws3.column_dimensions["A"].width = 10
            ws3.column_dimensions["B"].width = 30
            ws3.column_dimensions["C"].width = 50

        wb.save(OUT_DRYRUN_XLSX)
        xlsx_written = True
        print(f"[OUT] {OUT_DRYRUN_XLSX.relative_to(ROOT)} ({OUT_DRYRUN_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl 미설치: {e}")
    except Exception as e:
        print(f"[skip xlsx] 오류: {type(e).__name__}: {e}")

    # ── 첫 30개 변경 행 ──
    print("\n=== First 30 would-update rows ===")
    for d in diff[:30]:
        title = (d["title"] or "")[:24]
        if d["old_manual"]:
            oldp = f"{d['old_manual']} p.{d['old_page_from']}~{d['old_page_to']}"
        else:
            oldp = "—"
        newp = f"{d['new_manual']} p.{d['new_page_from']}~{d['new_page_to']}"
        print(
            f"  {d['row_id']:8s} {d['detailed_code']:8s} {d['action_type']:13s} "
            f"{title:24s}  {oldp:26s} → {newp:24s}  [{d['confidence']}/{d['method']}]"
        )

    # ── dry-run 종료 분기 ──
    if not args.apply:
        print("\n[dry-run] DB 미수정. 실제 적용은 --apply 필요.")
        print("\n실행 명령:")
        print("  dry-run (이번 실행):")
        print("    python backend/scripts/apply_manual_mapping_triage_v3.py")
        print("  apply (수동 검토 후 별도 실행 — 이번엔 실행하지 않음):")
        print("    python backend/scripts/apply_manual_mapping_triage_v3.py --apply")
        return

    # ─────────────────────────────────────────────────────────────
    #  --apply 분기 (이 스크립트는 호출됐지만 이번 단계에선 실행되지 않음)
    # ─────────────────────────────────────────────────────────────
    print("\n[apply] DB 백업 생성...")
    BACKUPS.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUPS / f"immigration_guidelines_db_v2.manual_ref_backup_{ts}.json"
    try:
        shutil.copy2(DB_PATH, backup_path)
        if not backup_path.exists() or backup_path.stat().st_size < 1000:
            raise RuntimeError(
                f"백업 검증 실패 (size={backup_path.stat().st_size if backup_path.exists() else 'missing'})"
            )
        # 백업 readback 검증
        try:
            json.loads(backup_path.read_text(encoding="utf-8"))
        except Exception as e:
            raise RuntimeError(f"백업 JSON 파싱 실패: {e}")
        print(f"  backup: {backup_path.relative_to(ROOT)} ({backup_path.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[ABORT] 백업 생성/검증 실패: {e}")
        sys.exit(3)

    # 메모리에서 manual_ref만 갱신 (다른 필드 절대 미수정)
    # 추가 안전망: blocked_ids는 candidates에 들어올 수 없지만 한 번 더 검증
    for t in candidates:
        if t.get("row_id") in blocked_ids:
            print(f"[ABORT] 안전망 위반: blocked row {t['row_id']} 가 candidates에 있음 — 중단")
            sys.exit(8)

    print(f"\n[apply] {len(candidates)}건 manual_ref 갱신 중 (blocked {len(blocked_in_run)}건 제외됨)...")
    applied = 0
    for t in candidates:
        rid = t["row_id"]
        db_row = db_index[rid]
        # **manual_ref 필드만** 교체. 다른 필드는 손대지 않음.
        db_row["manual_ref"] = t["proposed_manual_ref"]
        applied += 1

    # 직렬화 + 사전 검증
    print("[apply] DB 직렬화 + JSON 검증...")
    db_text_new = json.dumps(db, ensure_ascii=False, indent=2)
    try:
        re_parsed = json.loads(db_text_new)
        if len(re_parsed.get("master_rows") or []) != len(db_rows):
            raise RuntimeError(
                f"master_rows 개수 불일치 ({len(db_rows)} → {len(re_parsed.get('master_rows') or [])})"
            )
    except Exception as e:
        print(f"[ABORT] 직렬화 후 검증 실패: {e}")
        print(f"[복구 안내] cp \"{backup_path}\" \"{DB_PATH}\"")
        sys.exit(5)

    # 디스크 쓰기
    DB_PATH.write_text(db_text_new, encoding="utf-8")

    # 사후 readback 검증
    try:
        reread = json.loads(DB_PATH.read_text(encoding="utf-8"))
        if len(reread.get("master_rows") or []) != len(db_rows):
            raise RuntimeError("쓰기 후 master_rows 개수 불일치")
    except Exception as e:
        print(f"[ABORT after-write] readback 검증 실패: {e}")
        print(f"[복구 안내] cp \"{backup_path}\" \"{DB_PATH}\"")
        sys.exit(6)

    print(f"\n[OK] applied={applied}")
    print(f"  backup: {backup_path.relative_to(ROOT)}")
    print(f"  DB:     {DB_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
