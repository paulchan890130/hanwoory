"""
Task 1 — Extract WRONG rows from judge results for v5 patch research.

READ-ONLY. No DB modification. No LLM call. No patch.
Outputs:
  backend/data/manuals/v5_patch_research.json
  backend/data/manuals/v5_patch_research.xlsx
"""
import json
import re
import sys
from pathlib import Path

import fitz  # PyMuPDF

ROOT = Path(__file__).parent.parent.parent
JUDGE_DIR = ROOT / "backend" / "data" / "manuals" / "llm_judge_results"
J1 = JUDGE_DIR / "llm_judge_results_FAIL_sonnet_20260501_131101.json"
J2 = JUDGE_DIR / "llm_judge_results_FAIL_sonnet_20260501_132723.json"
BATCH = ROOT / "backend" / "data" / "manuals" / "llm_judge_batches" / "batch_fail_cluster_lt3_20260501_125555.json"
PDF_RES = ROOT / "backend" / "data" / "manuals" / "unlocked_체류민원.pdf"
PDF_VIS = ROOT / "backend" / "data" / "manuals" / "unlocked_사증민원.pdf"
OUT_JSON = ROOT / "backend" / "data" / "manuals" / "v5_patch_research.json"
OUT_XLSX = ROOT / "backend" / "data" / "manuals" / "v5_patch_research.xlsx"


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    j1 = json.loads(J1.read_text(encoding="utf-8"))
    j2 = json.loads(J2.read_text(encoding="utf-8"))

    merged = {r["row_id"]: r for r in (j1.get("results") or [])}
    for r in (j2.get("results") or []):
        merged[r["row_id"]] = r

    wrong_rows = [r for r in merged.values() if r.get("decision") == "WRONG"]
    print(f"WRONG rows: {len(wrong_rows)}")

    batch = json.loads(BATCH.read_text(encoding="utf-8"))
    batch_idx = {p["row_id"]: p for p in (batch.get("packets") or [])}

    docs = {
        "체류민원": fitz.open(str(PDF_RES)),
        "사증민원": fitz.open(str(PDF_VIS)),
    }

    def page_heading(manual: str, page_no: int) -> str:
        if manual not in docs:
            return ""
        doc = docs[manual]
        if page_no < 1 or page_no > doc.page_count:
            return ""
        txt = doc.load_page(page_no - 1).get_text() or ""
        lines = [l.strip() for l in txt.splitlines() if l.strip()]
        top = []
        for l in lines:
            if re.match(r"^- \d+ -$", l) or l == "목차":
                continue
            top.append(l)
            if len(top) >= 3:
                break
        return " / ".join(top)

    report = []
    for r in wrong_rows:
        rid = r["row_id"]
        p = batch_idx.get(rid, {})
        cur_page = p.get("current_page_from")
        rec_page = r.get("recommended_page")

        ready = False
        rec_page_int = None
        if rec_page is not None and rec_page != "":
            s = str(rec_page).strip()
            m = re.search(r"\d+", s)
            if m:
                rec_page_int = int(m.group(0))
                if rec_page_int != cur_page:
                    ready = True

        feasibility = "READY" if ready else "NEEDS_MANUAL_CHECK"
        snippet = ""
        if ready and rec_page_int is not None:
            snippet = page_heading(p.get("current_manual", ""), rec_page_int)

        report.append({
            "row_id":               rid,
            "detailed_code":        p.get("detailed_code"),
            "action_type":          p.get("action_type"),
            "title":                p.get("title"),
            "current_manual":       p.get("current_manual"),
            "current_page_from":    cur_page,
            "current_page_to":      p.get("current_page_to"),
            "recommended_page_raw": rec_page,
            "recommended_page_int": rec_page_int,
            "patch_feasibility":    feasibility,
            "heading_snippet":      snippet,
            "positive_evidence":    (r.get("positive_evidence") or "")[:600],
            "negative_evidence":    (r.get("negative_evidence") or "")[:600],
            "confidence":           r.get("confidence"),
            "page_role":            r.get("page_role"),
        })

    ready_count = sum(1 for r in report if r["patch_feasibility"] == "READY")
    manual_count = sum(1 for r in report if r["patch_feasibility"] == "NEEDS_MANUAL_CHECK")

    out_doc = {
        "meta": {
            "sources":          [str(J1.relative_to(ROOT)), str(J2.relative_to(ROOT))],
            "total_wrong_rows": len(wrong_rows),
            "ready":            ready_count,
            "needs_manual":     manual_count,
        },
        "rows": report,
    }
    OUT_JSON.write_text(json.dumps(out_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        ws_s.append(["total_wrong_rows", len(wrong_rows)])
        ws_s.append(["ready (recommended_page available, ≠ current)", ready_count])
        ws_s.append(["needs_manual_check", manual_count])
        ws_s.column_dimensions["A"].width = 38
        ws_s.column_dimensions["B"].width = 14

        ws = wb.create_sheet("wrong_rows")
        headers = [
            "row_id", "code", "action_type", "title",
            "current_page", "recommended_page", "patch_feasibility",
            "heading_snippet", "negative_evidence", "positive_evidence",
        ]
        widths = [10, 12, 14, 32, 14, 16, 22, 60, 80, 60]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for x in report:
            cur = f"{x['current_manual']} p.{x['current_page_from']}"
            rec = (f"{x['recommended_page_int']}" if x["recommended_page_int"]
                   else (str(x["recommended_page_raw"]) if x["recommended_page_raw"] else ""))
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                cur, rec, x["patch_feasibility"],
                x["heading_snippet"], x["negative_evidence"], x["positive_evidence"],
            ])
            color = "C6F6D5" if x["patch_feasibility"] == "READY" else "FFCC99"
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[xlsx skip] {type(e).__name__}: {e}")

    print()
    print("=== Summary ===")
    print(f"Total WRONG: {len(wrong_rows)}")
    print(f"READY (recommended_page available): {ready_count}")
    print(f"NEEDS_MANUAL_CHECK: {manual_count}")
    print()
    print("=== Per-row ===")
    for r in report:
        rec_str = (f"p.{r['recommended_page_int']}" if r["recommended_page_int"]
                   else (str(r["recommended_page_raw"]) if r["recommended_page_raw"] else "null"))
        print(f"  {r['row_id']}  {r['detailed_code']:<10} {r['action_type']:<14} "
              f"{r['current_manual']} p.{r['current_page_from']} → rec={rec_str:<10} [{r['patch_feasibility']}]")
        if r["heading_snippet"]:
            print(f"     heading: {r['heading_snippet'][:120]}")


if __name__ == "__main__":
    main()
