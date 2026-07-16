"""고객 일괄추출(Excel) — 현재 tenant 고객 전체를 고객카드 기준 컬럼으로 내보낸다.

설계 원칙:
  - 하이코리아/소시넷 등 외부 사이트 계정(ID/PW)은 절대 포함하지 않는다.
  - 전화번호·등록번호 등 선행 0/문자 손실 위험이 있는 열은 Excel Text(@) 서식으로 고정한다.
  - 정렬은 고객목록과 동일하게 고객ID 내림차순(list_customers/customer_pg_service 위임).
  - 읽기 전용 — 이 모듈은 아무것도 저장/변경하지 않는다.
"""
from __future__ import annotations

import io
from typing import Iterable

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# (엑셀 헤더, 표준 키, 열 너비). 표준 키는 customer_pg_service.PG_TO_SHEET 출력과 동일한
# 한글 키를 쓴다("_phone"/"_visa"는 이 모듈에서만 쓰는 합성 키).
_EXPORT_COLUMNS: tuple[tuple[str, str, int], ...] = (
    ("고객ID", "고객ID", 12),
    ("한글명", "한글", 14),
    ("영문 성", "성", 16),
    ("영문 이름", "명", 20),
    ("성별", "성별", 8),
    ("국적", "국적", 14),
    ("전화번호", "_phone", 16),
    ("등록증 앞자리", "등록증", 18),
    ("등록번호 뒷자리", "번호", 18),
    ("등록증 발급일", "발급일", 14),
    ("등록증 만기일", "만기일", 14),
    ("여권번호", "여권", 16),
    ("여권 발급일", "발급", 14),
    ("여권 만기일", "만기", 14),
    ("체류자격", "_visa", 12),
    ("주소", "주소", 36),
    ("비고", "비고", 30),
    ("위임내역", "위임내역", 32),
    ("폴더", "폴더", 28),
)
# 선행 0/문자 손실 방지가 필요한 열 — Text(@) 서식.
_TEXT_EXPORT_KEYS = {"고객ID", "_phone", "등록증", "번호", "여권"}


def _phone(record: dict) -> str:
    parts = [str(record.get(key, "") or "").strip() for key in ("연", "락", "처")]
    return "-".join(part for part in parts if part)


def _value(record: dict, key: str) -> str:
    if key == "_phone":
        return _phone(record)
    if key == "_visa":
        return str(record.get("체류자격") or "").strip()
    return str(record.get(key, "") or "")


def build_export_bytes_from_records(records: Iterable[dict]) -> tuple[bytes, int]:
    """이미 tenant-scoped 된 레코드로 고객 일괄추출 workbook(bytes)을 만든다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    safe_records = list(records)
    wb = Workbook()
    ws = wb.active
    ws.title = "고객목록"

    headers = [column[0] for column in _EXPORT_COLUMNS]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="F0E6C8")
    for col_idx, (_header, _key, width) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, record in enumerate(safe_records, start=2):
        ws.append([_value(record, key) for _header, key, _width in _EXPORT_COLUMNS])
        for col_idx, (_header, key, _width) in enumerate(_EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if key in _TEXT_EXPORT_KEYS:
                cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=key in {"주소", "비고", "위임내역"})

    ws.freeze_panes = "A2"
    last_col = ws.cell(row=1, column=len(_EXPORT_COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(safe_records) + 1)}"
    ws.row_dimensions[1].height = 32

    info = wb.create_sheet("안내")
    notes = (
        "[고객 일괄추출 안내]",
        "",
        f"추출 고객 수: {len(safe_records)}명",
        "이 파일에는 외국인등록번호·여권번호·주소 등 개인정보가 포함될 수 있습니다.",
        "하이코리아·소시넷 아이디와 비밀번호는 추출 대상에서 제외했습니다.",
        "업무 목적이 끝난 파일은 안전하게 폐기하세요.",
    )
    for idx, note in enumerate(notes, start=1):
        cell = info.cell(row=idx, column=1, value=note)
        if idx == 1:
            cell.font = Font(bold=True, size=12)
    info.column_dimensions["A"].width = 88

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(safe_records)


def build_tenant_export_bytes(tenant_id: str) -> tuple[bytes, int]:
    """현재 tenant의 비삭제 고객 전체(고객ID 내림차순, 기존 고객목록과 동일)를 추출."""
    from backend.services.customer_pg_service import list_customers

    records = list_customers(tenant_id, reveal=True)
    return build_export_bytes_from_records(records)
