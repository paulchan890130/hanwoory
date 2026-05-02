"""
manual_ref triage v3 — v2 audit 결과를 6개 카테고리로 분류 (READ-ONLY)

입력:
  backend/data/manuals/manual_mapping_audit_v2.json

출력:
  backend/data/manuals/manual_mapping_triage_v3.json
  backend/data/manuals/manual_mapping_triage_v3.xlsx (openpyxl 있으면)

분류 카테고리 (apply_candidate=True 는 AUTO_SAFE에만):
  1. AUTO_SAFE              — 자동 적용 안전
  2. ROUTING_SAFE_REVIEW    — 현재 DB가 비선호 매뉴얼 → preferred에서 exact/high
  3. MANUAL_REVIEW          — 두 매뉴얼 다 그럴듯 / missing_current / fallback
  4. LOW_CONFIDENCE         — medium/low 또는 section_only/prefix_fallback
  5. NO_CANDIDATE           — 후보 페이지 없음
  6. APPLICATION_CLAIM_REVIEW — APPLICATION_CLAIM 또는 빈 detailed_code

쓰기 보장:
  위 출력 2개 외 어떤 파일도 변경하지 않음.
  immigration_guidelines_db_v2.json, manual_mapping_audit*.json 모두 무수정.
"""
from __future__ import annotations
import json, sys
from pathlib import Path
from collections import Counter, defaultdict

ROOT     = Path(__file__).parent.parent.parent
MANUALS  = ROOT / "backend" / "data" / "manuals"

V2_AUDIT = MANUALS / "manual_mapping_audit_v2.json"
OUT_JSON = MANUALS / "manual_mapping_triage_v3.json"
OUT_XLSX = MANUALS / "manual_mapping_triage_v3.xlsx"

VALID_MANUALS = {"체류민원", "사증민원"}


# ─── 분류 ─────────────────────────────────────────────────────
def classify(audit_row: dict) -> tuple[str, bool, str, str, str]:
    """
    Returns:
      (triage_category, apply_candidate, apply_blocked_reason, human_review_reason, notes)
    """
    at        = (audit_row.get("action_type") or "").strip()
    code      = (audit_row.get("detailed_code") or "").strip()
    routing   = audit_row.get("routing") or {}
    preferred = (routing.get("preferred_manual") or "unknown").strip()
    primary   = audit_row.get("primary_candidate")
    secondary = audit_row.get("secondary_candidate")
    comp      = audit_row.get("comparison") or {}
    rwarn     = (audit_row.get("routing_warning") or "").strip()

    p_mref = (primary or {}).get("manual_ref") or []
    p0     = p_mref[0] if p_mref else {}
    p_conf = (primary or {}).get("confidence", "none")
    p_method = (primary or {}).get("method", "none")
    p_manual = p0.get("manual")
    p_pf     = p0.get("page_from")
    p_pt     = p0.get("page_to")

    s_conf = (secondary or {}).get("confidence", "none") if secondary else "none"

    status = comp.get("status", "")
    creason = comp.get("reason", "")  # 'both_manuals_plausible' / 'current_in_wrong_manual'

    # ── 1. APPLICATION_CLAIM_REVIEW (가장 먼저) ──
    if at == "APPLICATION_CLAIM" or not code or preferred == "unknown":
        return (
            "APPLICATION_CLAIM_REVIEW", False,
            "APPLICATION_CLAIM 계열 또는 빈 detailed_code — 자동 라우팅 불가",
            "신청·증명 발급 업무: 매뉴얼 자격 섹션이 아닌 별도 절차 — 사람 검토 필요",
            f"action_type={at!r}, detailed_code={code!r}",
        )

    # ── 2. NO_CANDIDATE ──
    if not primary or not p_mref:
        return (
            "NO_CANDIDATE", False,
            "후보 페이지 없음",
            "어떤 매핑 방법으로도 페이지를 못 찾음",
            "primary_candidate is null",
        )
    if not isinstance(p_pf, int) or not isinstance(p_pt, int) or p_pf <= 0 or p_pt <= 0:
        return (
            "NO_CANDIDATE", False,
            "page_from/page_to 누락 또는 비정상",
            "유효한 페이지 번호 없음",
            f"page_from={p_pf}, page_to={p_pt}",
        )
    if p_conf == "none":
        return (
            "NO_CANDIDATE", False,
            "confidence=none",
            "매핑 신뢰도 없음",
            f"method={p_method}",
        )

    # ── 3. MANUAL_REVIEW (both_manuals_plausible) ──
    if status == "conflict" and creason == "both_manuals_plausible":
        return (
            "MANUAL_REVIEW", False,
            "both_manuals_plausible — 두 매뉴얼 모두 exact/high 후보",
            (
                f"체류·사증 양쪽에 그럴듯한 헤더가 있음 (primary {p_conf} in {p_manual}, "
                f"secondary {s_conf}) — 의미론적 사람 판단 필요"
            ),
            f"primary={p_manual}, secondary={(secondary or {}).get('manual_ref',[{}])[0].get('manual','—') if secondary else '—'}",
        )

    # ── 4. ROUTING_SAFE_REVIEW (current_in_wrong_manual) ──
    if status == "conflict" and creason == "current_in_wrong_manual":
        if p_conf in ("exact", "high") and p_manual == preferred:
            return (
                "ROUTING_SAFE_REVIEW", False,
                "현재 DB가 비선호 매뉴얼 — preferred에 exact/high 후보 있음",
                f"action_type={at} 의 권장 매뉴얼은 {preferred}. 현재 DB는 다른 매뉴얼을 가리키고 있음",
                f"current={(audit_row.get('current_manual_ref') or [{}])[0].get('manual','—')} → proposed={p_manual} ({p_conf})",
            )
        # exact/high 아니면 LOW_CONFIDENCE 쪽으로 떨어트림 (defensive)

    # ── 5. missing_current 처리 — DB가 비어있던 row ──
    if status == "missing_current":
        # 신뢰도 + preferred 일치하면 ROUTING_SAFE_REVIEW와 비슷한 위상
        if p_conf in ("exact", "high") and p_manual == preferred and not rwarn:
            return (
                "MANUAL_REVIEW", False,
                "DB에 manual_ref 없음 — 신규 매핑 검토 필요",
                f"DB에 매핑 없는 row, preferred 매뉴얼에 {p_conf} 후보 발견 — 신규 추가 가능 여부 사람 확인",
                f"new mapping candidate: {p_manual} p.{p_pf}~{p_pt}",
            )
        return (
            "LOW_CONFIDENCE", False,
            "DB에 manual_ref 없음 + 후보 신뢰도 낮음/하향됨",
            f"신규 매핑 후보가 신뢰도 부족 (conf={p_conf}, method={p_method})",
            f"routing_warning={rwarn!r}",
        )

    # ── 6. LOW_CONFIDENCE ──
    if p_conf in ("medium", "low"):
        return (
            "LOW_CONFIDENCE", False,
            f"confidence={p_conf}",
            f"자격 섹션 단위 또는 prefix 폴백 — 정확 페이지 미확정",
            f"method={p_method}",
        )
    if p_method in ("section_only", "prefix_fallback"):
        return (
            "LOW_CONFIDENCE", False,
            f"method={p_method}",
            "자격 섹션 전체만 잡힘 — 액션 페이지 미확정",
            f"confidence={p_conf}",
        )
    if rwarn:
        # routing_warning 있으면 강제 하향이 있었으므로 신뢰 못 함
        return (
            "LOW_CONFIDENCE", False,
            "routing_warning 있음 — 비선호 매뉴얼 강제 하향",
            rwarn,
            "",
        )

    # ── 7. AUTO_SAFE 마지막 검증 ──
    fail_reasons = []
    if p_manual != preferred:        fail_reasons.append(f"primary가 비선호 매뉴얼({p_manual} != {preferred})")
    if p_conf not in ("exact", "high"): fail_reasons.append(f"confidence={p_conf}")
    if status not in ("same", "changed"): fail_reasons.append(f"status={status}")
    if status == "conflict":          fail_reasons.append("conflict")
    if rwarn:                         fail_reasons.append("routing_warning")
    if p_manual not in VALID_MANUALS: fail_reasons.append(f"manual={p_manual!r}")
    if at == "APPLICATION_CLAIM":     fail_reasons.append("APPLICATION_CLAIM")
    if not code:                      fail_reasons.append("empty detailed_code")

    if fail_reasons:
        # 어떤 카테고리에도 안 들어맞은 잔여 — 보수적으로 MANUAL_REVIEW
        return (
            "MANUAL_REVIEW", False,
            "AUTO_SAFE 조건 일부 미충족",
            " / ".join(fail_reasons),
            "fallback → MANUAL_REVIEW",
        )

    return ("AUTO_SAFE", True, "", "", f"{p_manual} p.{p_pf}~{p_pt} ({p_conf}/{p_method})")


# ─── 메인 ─────────────────────────────────────────────────────
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not V2_AUDIT.exists():
        print(f"[ERR] v2 audit 파일 없음: {V2_AUDIT}")
        sys.exit(1)

    print("[triage-v3] v2 audit 로드...")
    v2 = json.loads(V2_AUDIT.read_text(encoding="utf-8"))
    rows = v2.get("rows") or []
    print(f"  rows: {len(rows)}")

    print("[triage-v3] 분류 진행...")
    triaged: list[dict] = []
    for r in rows:
        category, apply_cand, blocked, review, notes = classify(r)
        primary = r.get("primary_candidate") or {}
        p_mref = primary.get("manual_ref") or []
        p0 = p_mref[0] if p_mref else {}
        triaged.append({
            "row_id":               r.get("row_id"),
            "title":                r.get("title"),
            "action_type":          r.get("action_type"),
            "detailed_code":        r.get("detailed_code"),
            "preferred_manual":     (r.get("routing") or {}).get("preferred_manual"),
            "current_manual_ref":   r.get("current_manual_ref") or [],
            "proposed_manual_ref":  p_mref,
            "confidence":           primary.get("confidence", "none"),
            "method":               primary.get("method", "none"),
            "comparison_status":    (r.get("comparison") or {}).get("status", ""),
            "comparison_reason":    (r.get("comparison") or {}).get("reason", ""),
            "routing_warning":      r.get("routing_warning") or "",
            "triage_category":      category,
            "apply_candidate":      apply_cand,
            "apply_blocked_reason": blocked,
            "human_review_reason":  review,
            "notes":                notes,
            # 보조 필드 — xlsx 표시용
            "_proposed_manual":     p0.get("manual", ""),
            "_proposed_pf":         p0.get("page_from", ""),
            "_proposed_pt":         p0.get("page_to", ""),
            "_current_manual":      (r.get("current_manual_ref") or [{}])[0].get("manual", "") if r.get("current_manual_ref") else "",
            "_current_pf":          (r.get("current_manual_ref") or [{}])[0].get("page_from", "") if r.get("current_manual_ref") else "",
            "_current_pt":          (r.get("current_manual_ref") or [{}])[0].get("page_to", "") if r.get("current_manual_ref") else "",
        })

    # ── 통계 ──
    cat_dist = Counter(t["triage_category"] for t in triaged)
    apply_count = sum(1 for t in triaged if t["apply_candidate"])
    conflict_count = sum(1 for t in triaged if t["comparison_status"] == "conflict")
    both_plausible = sum(1 for t in triaged if t["comparison_reason"] == "both_manuals_plausible")
    routing_warns = sum(1 for t in triaged if t["routing_warning"])
    auto_changed = sum(1 for t in triaged if t["triage_category"] == "AUTO_SAFE" and t["comparison_status"] == "changed")
    current_wrong = sum(1 for t in triaged if t["comparison_reason"] == "current_in_wrong_manual")

    def by_filter(predicate):
        sub = [t for t in triaged if predicate(t)]
        return sub, dict(Counter(t["triage_category"] for t in sub))

    e9_rows, e9_by_cat       = by_filter(lambda t: (t["detailed_code"] or "").startswith("E-9"))
    f1_rows, f1_by_cat       = by_filter(lambda t: (t["detailed_code"] or "").startswith("F-1"))
    visa_rows, visa_by_cat   = by_filter(lambda t: t["action_type"] == "VISA_CONFIRM")
    appcl_rows, appcl_by_cat = by_filter(lambda t: t["action_type"] == "APPLICATION_CLAIM")

    summary = {
        "total":                       len(triaged),
        "category_count":              dict(cat_dist),
        "apply_candidate_true":        apply_count,
        "conflict_count":              conflict_count,
        "both_manuals_plausible":      both_plausible,
        "routing_warning_count":       routing_warns,
        "auto_safe_changed_count":     auto_changed,
        "current_in_wrong_manual":     current_wrong,
        "e9_rows_total":               len(e9_rows),
        "e9_by_category":              e9_by_cat,
        "f1_rows_total":               len(f1_rows),
        "f1_by_category":              f1_by_cat,
        "visa_confirm_rows_total":     len(visa_rows),
        "visa_confirm_by_category":    visa_by_cat,
        "application_claim_rows_total": len(appcl_rows),
        "application_claim_by_category": appcl_by_cat,
    }

    out = {
        "triage_meta": {
            "version":   "v3",
            "v2_audit":  str(V2_AUDIT.relative_to(ROOT)),
            "rules":     "AUTO_SAFE only when proposed=preferred & conf∈{exact,high} & status∈{same,changed} & no conflict & no routing_warning & valid pages",
        },
        "summary": summary,
        "rows":    triaged,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX (8 sheets) ──
    xlsx_written = False
    skip_reason = ""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        # 기본 sheet 제거
        wb.remove(wb.active)

        # summary 시트
        ws_sum = wb.create_sheet("summary")
        ws_sum.append(["metric", "value"])
        for c in ws_sum[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            if isinstance(v, dict):
                ws_sum.append([k, json.dumps(v, ensure_ascii=False)])
            else:
                ws_sum.append([k, v])
        ws_sum.column_dimensions["A"].width = 32
        ws_sum.column_dimensions["B"].width = 90

        # row 시트들 (공통 컬럼)
        headers = [
            "row_id", "title", "action_type", "detailed_code",
            "preferred_manual", "triage_category", "apply_candidate",
            "confidence", "method", "comparison_status", "comparison_reason",
            "current_manual", "current_pf", "current_pt",
            "proposed_manual", "proposed_pf", "proposed_pt",
            "apply_blocked_reason", "human_review_reason",
            "routing_warning", "notes",
        ]
        widths = [10, 28, 14, 12, 12, 24, 8, 10, 22, 18, 22,
                  10, 8, 8, 10, 8, 8, 36, 40, 32, 36]

        def write_sheet(ws_name: str, items: list[dict]):
            ws = wb.create_sheet(ws_name)
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
                c.alignment = Alignment(horizontal="center")
            for t in items:
                ws.append([
                    t["row_id"], t["title"], t["action_type"], t["detailed_code"],
                    t["preferred_manual"], t["triage_category"], t["apply_candidate"],
                    t["confidence"], t["method"], t["comparison_status"], t["comparison_reason"],
                    t["_current_manual"], t["_current_pf"], t["_current_pt"],
                    t["_proposed_manual"], t["_proposed_pf"], t["_proposed_pt"],
                    t["apply_blocked_reason"], t["human_review_reason"],
                    t["routing_warning"], t["notes"],
                ])
            for col_idx, w in enumerate(widths, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = w
            ws.freeze_panes = "A2"

        # 카테고리별 시트
        cats_sheets = {
            "auto_safe":                 [t for t in triaged if t["triage_category"] == "AUTO_SAFE"],
            "routing_safe_review":       [t for t in triaged if t["triage_category"] == "ROUTING_SAFE_REVIEW"],
            "manual_review":             [t for t in triaged if t["triage_category"] == "MANUAL_REVIEW"],
            "low_confidence":            [t for t in triaged if t["triage_category"] == "LOW_CONFIDENCE"],
            "no_candidate":              [t for t in triaged if t["triage_category"] == "NO_CANDIDATE"],
            "application_claim_review":  [t for t in triaged if t["triage_category"] == "APPLICATION_CLAIM_REVIEW"],
        }
        for name, items in cats_sheets.items():
            write_sheet(name, items)

        # 전체
        write_sheet("all_rows", triaged)

        wb.save(OUT_XLSX)
        xlsx_written = True
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        skip_reason = f"openpyxl import failed: {e}"
        print(f"[skip xlsx] {skip_reason}")
    except Exception as e:
        skip_reason = f"xlsx 생성 오류: {type(e).__name__}: {e}"
        print(f"[skip xlsx] {skip_reason}")

    # ── 콘솔 요약 ──
    print("\n" + "=" * 78)
    print("TRIAGE SUMMARY v3")
    print("=" * 78)
    print(f"  total rows:                        {len(triaged)}")
    for cat in ["AUTO_SAFE","ROUTING_SAFE_REVIEW","MANUAL_REVIEW","LOW_CONFIDENCE","NO_CANDIDATE","APPLICATION_CLAIM_REVIEW"]:
        print(f"  {cat:30s} {cat_dist.get(cat, 0)}")
    print(f"  apply_candidate=true:              {apply_count}")
    print(f"  conflict (overall):                {conflict_count}")
    print(f"  both_manuals_plausible:            {both_plausible}")
    print(f"  routing_warning_count:             {routing_warns}")
    print(f"  AUTO_SAFE.changed (page diff):     {auto_changed}")
    print(f"  current_in_wrong_manual:           {current_wrong}")
    print(f"\n  E-9 by category:               {e9_by_cat}")
    print(f"  F-1 by category:               {f1_by_cat}")
    print(f"  VISA_CONFIRM by category:      {visa_by_cat}")
    print(f"  APPLICATION_CLAIM by category: {appcl_by_cat}")

    def show_top(name, items, n=30):
        print(f"\n=== Top {n} {name} ===")
        for t in items[:n]:
            title = (t["title"] or "")[:24]
            print(
                f"  {t['row_id']:8s} {t['detailed_code']:8s} {t['action_type']:13s} "
                f"pref={t['preferred_manual']:6s} conf={t['confidence']:6s} "
                f"{t['_proposed_manual']:6s} p.{str(t['_proposed_pf']):>4s}~{str(t['_proposed_pt']):>4s}  "
                f"{title}"
            )

    show_top("AUTO_SAFE changed",    [t for t in triaged if t["triage_category"]=="AUTO_SAFE" and t["comparison_status"]=="changed"])
    show_top("ROUTING_SAFE_REVIEW",  [t for t in triaged if t["triage_category"]=="ROUTING_SAFE_REVIEW"])
    show_top("MANUAL_REVIEW",        [t for t in triaged if t["triage_category"]=="MANUAL_REVIEW"])

    print("\n=== Files created (only) ===")
    print(f"  - {OUT_JSON.relative_to(ROOT)}")
    if xlsx_written:
        print(f"  - {OUT_XLSX.relative_to(ROOT)}")
    else:
        print(f"  - (xlsx skipped: {skip_reason})")

    print("\n=== Files NOT modified ===")
    print(f"  - immigration_guidelines_db_v2.json")
    print(f"  - manual_mapping_audit.json (v1)")
    print(f"  - manual_mapping_audit_v2.json (v2)")
    print(f"  - manual_index_v6 / structure / CSV / PDFs / indexer / watcher / frontend")


if __name__ == "__main__":
    main()
