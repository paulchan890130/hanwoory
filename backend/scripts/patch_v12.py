"""
Patch v12 — 1 row: M1-0199 레거시 entry 제거.

DEFAULT: dry-run. `--apply` required for DB write.

Target:
  M1-0199  E-10  VISA_CONFIRM
    REPLACE ALL → 사증민원 p.294-296만 유지
    체류민원 p.332 (오매핑 레거시) 제거
    사증민원 p.294-296 (v11에서 추가한 정확한 entry) 유지
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

ROOT       = Path(__file__).parent.parent.parent
DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"
DRY_JSON   = ROOT / "backend" / "data" / "manuals" / "v12_patch_dryrun.json"
DRY_XLSX   = ROOT / "backend" / "data" / "manuals" / "v12_patch_dryrun.xlsx"

PATCHES: dict[str, dict] = {
    "M1-0199": {
        "detailed_code": "E-10",
        "action_type":   "VISA_CONFIRM",
        "title":         "선원취업 사증발급인정서",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.294-296: '선원취업(E-10)' 챕터 — "
            "p.294 활동범위·해당자(E-10-1/E-10-2/E-10-3), "
            "p.295 '선원취업(E-10) 사증은 출입국·외국인청(사무소·출장소)장이 발급한 "
            "사증발급인정서에 의해서만 발급' + 접수처리 절차, "
            "p.296 발급대상 첨부서류(공통 4종 + E-10-2 어선원 전용 4종). "
            "직접 PDF 열람 검증(2026-05-02). "
            "체류민원 p.332는 E-7 직종표 페이지 — VISA_CONFIRM 오매핑. 제거."
        ),
        "replace_all": True,
        "add_entries": [
            {
                "manual":     "사증민원",
                "page_from":  294,
                "page_to":    296,
                "match_type": "manual_override",
                "match_text": "E-10 선원취업 VISA_CONFIRM 사증발급인정서 (사증민원 p.294-296) — 수동 검증",
            },
        ],
    },
}


# ── Safety primitives ────────────────────────────────────────────
def _abort(msg: str) -> None:
    print(f"\n[ABORT] {msg}", file=sys.stderr)
    sys.exit(1)


def make_backup(db_bytes: bytes) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"immigration_guidelines_db_v2.v12_patch_backup_{ts}.json"
    bk.write_bytes(db_bytes)
    if not bk.exists() or bk.stat().st_size != len(db_bytes):
        _abort(f"백업 작성 검증 실패: {bk}")
    try:
        json.loads(bk.read_text(encoding="utf-8"))
    except Exception as e:
        _abort(f"백업 JSON readback 실패: {e}")
    return bk


def diff_rows(before: list, after: list) -> list:
    bi = {r.get("row_id"): r for r in before}
    ai = {r.get("row_id"): r for r in after}
    out = []
    for rid in set(bi) | set(ai):
        b = bi.get(rid); a = ai.get(rid)
        if b is None or a is None:
            out.append({"row_id": rid, "added_or_removed": True}); continue
        diff = [k for k in set(b) | set(a) if b.get(k) != a.get(k)]
        if diff:
            out.append({"row_id": rid, "fields": sorted(diff)})
    return out


def load_db() -> tuple:
    raw = DB_PATH.read_bytes()
    return json.loads(raw.decode("utf-8")), raw


def build_patched_db(db: dict) -> tuple:
    new_db = json.loads(json.dumps(db, ensure_ascii=False))
    rows = new_db.get("master_rows") or []
    found = {r.get("row_id"): r for r in rows if r.get("row_id") in PATCHES}
    missing = [rid for rid in PATCHES if rid not in found]
    if missing:
        _abort(f"target row_ids 누락: {missing}")
    patch_report = []
    for rid, spec in PATCHES.items():
        row = found[rid]
        if row.get("detailed_code") != spec["detailed_code"]:
            _abort(f"{rid} detailed_code 불일치")
        if row.get("action_type") != spec["action_type"]:
            _abort(f"{rid} action_type 불일치")
        old_ref = row.get("manual_ref") or []
        if spec.get("replace_all"):
            new_ref = list(spec["add_entries"])
        else:
            existing_sigs = {(e.get("manual"), e.get("page_from"), e.get("page_to")) for e in old_ref}
            new_ref = list(old_ref)
            for ne in spec.get("add_entries") or []:
                sig = (ne["manual"], ne["page_from"], ne["page_to"])
                if sig not in existing_sigs:
                    new_ref.append(ne)
        row["manual_ref"] = new_ref
        patch_report.append({
            "row_id":           rid,
            "detailed_code":    spec["detailed_code"],
            "action_type":      spec["action_type"],
            "title":            spec["title"],
            "preferred_manual": spec["preferred_manual"],
            "old_manual_ref":   old_ref,
            "new_manual_ref":   new_ref,
            "added_entries":    spec.get("add_entries") or [],
            "replace_all":      spec.get("replace_all", False),
            "page_evidence":    spec["page_evidence"],
        })
    return new_db, patch_report


def verify_patch_safety(before_db: dict, after_db: dict) -> list:
    errors = []
    bm = before_db.get("master_rows") or []; am = after_db.get("master_rows") or []
    if len(bm) != 369: errors.append(f"master_rows 개수 이상: {len(bm)}")
    if len(am) != len(bm): errors.append(f"패치 후 개수 불일치: {len(am)}")
    bi = {r.get("row_id"): r for r in bm}; ai = {r.get("row_id"): r for r in am}
    if set(bi) != set(ai): errors.append("row_id 집합 변경됨")
    for rid in bi:
        b = bi[rid]; a = ai.get(rid, {})
        if rid in PATCHES:
            for k in set(b) | set(a):
                if k == "manual_ref": continue
                if b.get(k) != a.get(k):
                    errors.append(f"{rid} non-manual_ref 필드 변경됨: {k}")
        else:
            for k in set(b) | set(a):
                if b.get(k) != a.get(k):
                    errors.append(f"비-target row {rid} 필드 변경됨: {k}")
    for k in set(before_db) | set(after_db):
        if k == "master_rows": continue
        if before_db.get(k) != after_db.get(k):
            errors.append(f"top-level 키 변경됨: {k}")
    return errors


def write_dryrun_reports(patch_report: list, summary: dict) -> None:
    DRY_JSON.write_text(json.dumps({
        "patch_meta": {"version": "v12", "mode": "DRY_RUN",
                       "scope": "1 row: E-10 VISA_CONFIRM 레거시 체류민원 entry 제거"},
        "summary": summary, "rows": patch_report,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {DRY_JSON.relative_to(ROOT)} ({DRY_JSON.stat().st_size:,} bytes)")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("patches")
        headers = ["row_id","code","action","title","preferred","replace_all",
                   "old_refs","added_entries","evidence"]
        widths  = [10,12,14,36,14,10,80,100,80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for x in patch_report:
            ws.append([x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                       x["preferred_manual"], x["replace_all"],
                       json.dumps(x["old_manual_ref"], ensure_ascii=False)[:600],
                       json.dumps(x["added_entries"], ensure_ascii=False)[:800],
                       x["page_evidence"][:400]])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCB")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(DRY_XLSX)
        print(f"[OUT] {DRY_XLSX.relative_to(ROOT)} ({DRY_XLSX.stat().st_size:,} bytes)")
    except Exception as e:
        print(f"[skip xlsx] {e}")


def main() -> int:
    try: sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    is_apply = args.apply
    if not DB_PATH.exists(): _abort(f"DB 없음: {DB_PATH}")
    db, raw = load_db()
    bytes_before = len(raw)
    bm = db.get("master_rows") or []
    if len(bm) != 369: _abort("master_rows 개수 비정상")
    print(f"[patch] mode={'APPLY' if is_apply else 'DRY-RUN'}  DB={bytes_before:,}  master_rows={len(bm)}")
    print(f"[patch] targets ({len(PATCHES)}): {sorted(PATCHES.keys())}")
    new_db, patch_report = build_patched_db(db)
    errors = verify_patch_safety(db, new_db)
    if errors:
        for e in errors: print(f"  [error] {e}")
        _abort("safety check 실패")
    diffs = diff_rows(bm, new_db.get("master_rows") or [])
    diff_ids = {d["row_id"] for d in diffs}
    if diff_ids != set(PATCHES.keys()):
        _abort(f"diff 불일치: diff={sorted(diff_ids)} expected={sorted(PATCHES.keys())}")
    for d in diffs:
        if d.get("fields") != ["manual_ref"]:
            _abort(f"{d['row_id']} manual_ref 외 필드 변경: {d.get('fields')}")
    summary = {
        "mode": "APPLY" if is_apply else "DRY_RUN",
        "production_db_modified": False,
        "patch_candidate_count": len(patch_report),
        "diff_row_ids": sorted(diff_ids),
        "diff_count": len(diff_ids),
        "non_target_row_diffs": 0,
        "non_manual_ref_field_diffs": 0,
        "db_byte_size_before": bytes_before,
    }
    write_dryrun_reports(patch_report, summary)
    if not is_apply:
        print(f"\n{'='*70}\nPATCH v12 — DRY-RUN ({len(PATCHES)} rows)\n{'='*70}")
        print(f"  targets: {sorted(diff_ids)}")
        print(f"  non-target diffs: 0  /  non-manual_ref field diffs: 0")
        print(f"  DB: {bytes_before:,} → {bytes_before:,} (unchanged)")
        print()
        for x in patch_report:
            mode = "REPLACE" if x["replace_all"] else "ADD"
            print(f"  {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}  [{mode}]")
            print(f"    old: {len(x['old_manual_ref'])} entries  →  new: {len(x['new_manual_ref'])} entries")
            print(f"    old_entries:")
            for e in x["old_manual_ref"]:
                print(f"      - {e.get('manual')} p.{e.get('page_from')}-{e.get('page_to')}  [{e.get('match_type','')}]")
            print(f"    new_entries:")
            for e in x["new_manual_ref"]:
                print(f"      + {e.get('manual')} p.{e.get('page_from')}-{e.get('page_to')}  [{e.get('match_text','')[:80]}]")
        print()
        print("  Apply command (NOT EXECUTED):\n    python backend/scripts/patch_v12.py --apply")
        return 0
    # APPLY
    print("\n[apply] 백업 생성...")
    bk = make_backup(raw)
    print(f"  backup: {bk.relative_to(ROOT)}")
    DB_PATH.write_text(json.dumps(new_db, ensure_ascii=False, indent=2), encoding="utf-8")
    readback = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rb_rows = readback.get("master_rows") or []
    if len(rb_rows) != 369:
        _abort(f"readback 이상: {len(rb_rows)}\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    errs2 = verify_patch_safety(json.loads(bk.read_text(encoding="utf-8")), readback)
    if errs2:
        for e in errs2: print(f"  [error] {e}")
        _abort(f"사후 verify 실패\n복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    bytes_after = DB_PATH.stat().st_size
    print(f"\n{'='*70}\nPATCH v12 — APPLIED\n{'='*70}")
    print(f"  backup: {bk.relative_to(ROOT)}")
    print(f"  DB: {bytes_before:,} → {bytes_after:,}  rows updated: {len(patch_report)}")
    print(f"  rollback: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
