"""
v3 AUTO_SAFE 변경 행 spot-check 패키지 빌더 (READ-ONLY)

입력:
  backend/data/manuals/manual_mapping_apply_dryrun_v3.json
  backend/data/manuals/unlocked_체류민원.pdf
  backend/data/manuals/unlocked_사증민원.pdf

출력:
  backend/data/manuals/manual_mapping_spotcheck_v3.json
  backend/data/manuals/manual_mapping_spotcheck_v3.xlsx
  backend/data/manuals/spotcheck_pages_v3/<row_id>_<manual>_p<new_pf>.png

규칙:
  - DB 미수정.
  - PDF 원본 텍스트만 사용 (AI 요약 금지).
  - 페이지당 미리보기 최대 1500 한글자.
  - 선택 기준 A~E 합집합 (중복 제거).
"""
from __future__ import annotations
import json, sys
from pathlib import Path
from collections import defaultdict, OrderedDict

import fitz

ROOT     = Path(__file__).parent.parent.parent
MANUALS  = ROOT / "backend" / "data" / "manuals"
DRYRUN   = MANUALS / "manual_mapping_apply_dryrun_v3.json"

OUT_JSON = MANUALS / "manual_mapping_spotcheck_v3.json"
OUT_XLSX = MANUALS / "manual_mapping_spotcheck_v3.xlsx"
IMG_DIR  = MANUALS / "spotcheck_pages_v3"

PDF_PATHS = {
    "체류민원": MANUALS / "unlocked_체류민원.pdf",
    "사증민원": MANUALS / "unlocked_사증민원.pdf",
}

PREVIEW_MAX  = 1500
RENDER_DPI   = 120

D2_EXTRA_CODES = {"D-2-1","D-2-2","D-2-3","D-2-4","D-2-5","D-2-6","D-2-7","D-2-8"}


# ─── 선택 규칙 ─────────────────────────────────────────────────
def is_high_jump(d: dict) -> tuple[bool, int | None]:
    """
    True if:
      |old_pf - new_pf| >= 20, OR
      old range >= 5 페이지 AND new가 단일 페이지
    """
    o_pf = d.get("old_page_from")
    o_pt = d.get("old_page_to")
    n_pf = d.get("new_page_from")
    n_pt = d.get("new_page_to")

    o_pf_i = o_pf if isinstance(o_pf, int) else None
    o_pt_i = o_pt if isinstance(o_pt, int) else None

    if o_pf_i is not None and isinstance(n_pf, int):
        if abs(o_pf_i - n_pf) >= 20:
            return True, abs(o_pf_i - n_pf)
        if o_pt_i is not None and (o_pt_i - o_pf_i) >= 5 and n_pt == n_pf:
            return True, abs(o_pf_i - n_pf)
    return False, None


def is_d2_extra(d: dict) -> bool:
    return d.get("detailed_code") in D2_EXTRA_CODES and d.get("action_type") == "EXTRA_WORK"


# ─── 메인 ──────────────────────────────────────────────────────
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not DRYRUN.exists():
        print(f"[ABORT] dry-run 파일 없음: {DRYRUN}")
        sys.exit(1)

    print("[spotcheck] dry-run 로드...")
    dry = json.loads(DRYRUN.read_text(encoding="utf-8"))
    diff = dry.get("diff") or []
    print(f"  diff rows: {len(diff)}")

    # ── 선택 (집합으로 사유 누적) ──
    selected: "OrderedDict[str, dict]" = OrderedDict()

    def add(d: dict, reason: str, focus: str):
        rid = d["row_id"]
        if rid in selected:
            selected[rid]["_reasons"].add(reason)
            if focus and not selected[rid]["review_focus"]:
                selected[rid]["review_focus"] = focus
        else:
            selected[rid] = {**d, "_reasons": {reason}, "review_focus": focus}

    # A. high-jump
    high_jump_count_total = 0
    for d in diff:
        hj, jump = is_high_jump(d)
        if hj:
            high_jump_count_total += 1
            focus = f"페이지 점프 {jump}p — 새 페이지가 정답인지 PDF 직접 확인" if jump else "페이지 범위 축소 — 새 페이지가 정답인지 확인"
            add(d, "high_jump", focus)

    # B. D-2 EXTRA_WORK
    d2_rows = [d for d in diff if is_d2_extra(d)]
    for d in d2_rows:
        add(d, "d2_extra_work",
            "p.35 시간제취업 표가 D-2 자격에 적용되는지 확인")

    # C. First 30 dry-run rows
    for d in diff[:30]:
        add(d, "first30", "콘솔 첫 30개 — 표본 검사")

    # D. by action_type — 각 3건
    by_at: dict[str, list] = defaultdict(list)
    for d in diff:
        by_at[d.get("action_type", "")].append(d)
    for at in ["CHANGE","EXTEND","REGISTRATION","REENTRY","EXTRA_WORK","WORKPLACE","VISA_CONFIRM"]:
        for d in by_at.get(at, [])[:3]:
            add(d, f"sample_{at}", f"{at} 액션 표본")

    # E. by manual — 25 체류 / 7 사증
    by_m: dict[str, list] = defaultdict(list)
    for d in diff:
        by_m[d.get("new_manual", "")].append(d)
    for d in by_m.get("체류민원", [])[:25]:
        add(d, "sample_체류민원", "체류민원 표본")
    for d in by_m.get("사증민원", [])[:7]:
        add(d, "sample_사증민원", "사증민원 표본")

    print(f"  selected (unique): {len(selected)}")

    # ── PDF 열기 ──
    print("[spotcheck] PDF 로드...")
    docs: dict[str, fitz.Document] = {}
    for label, p in PDF_PATHS.items():
        if not p.exists():
            print(f"  [skip] {label}: 파일 없음"); continue
        docs[label] = fitz.open(p)
        print(f"  {label}: {len(docs[label])} pages")

    IMG_DIR.mkdir(parents=True, exist_ok=True)

    extraction_errors: list[dict] = []
    extracted_n = 0

    enriched: list[dict] = []
    for rid, row in selected.items():
        new_manual = row.get("new_manual")
        new_pf = row.get("new_page_from")
        new_pt = row.get("new_page_to")
        old_pf = row.get("old_page_from")
        old_pt = row.get("old_page_to")

        # 텍스트 미리보기 — new
        new_preview = ""
        try:
            doc = docs.get(new_manual)
            if doc and isinstance(new_pf, int) and 1 <= new_pf <= len(doc):
                new_preview = doc[new_pf - 1].get_text()[:PREVIEW_MAX]
            else:
                extraction_errors.append({
                    "row_id": rid, "stage": "new_text",
                    "error": f"page out of range: {new_manual} p.{new_pf}",
                })
        except Exception as e:
            extraction_errors.append({"row_id": rid, "stage": "new_text", "error": str(e)})

        # 텍스트 미리보기 — old (있으면)
        old_preview = ""
        old_mref = row.get("old_manual_ref") or []
        old_manual = old_mref[0].get("manual") if old_mref else None
        try:
            if isinstance(old_pf, int) and old_manual and old_manual in docs:
                doc_o = docs[old_manual]
                if 1 <= old_pf <= len(doc_o):
                    old_preview = doc_o[old_pf - 1].get_text()[:PREVIEW_MAX]
        except Exception as e:
            extraction_errors.append({"row_id": rid, "stage": "old_text", "error": str(e)})

        # PNG 렌더 — new 페이지 1장
        img_rel = ""
        try:
            doc = docs.get(new_manual)
            if doc and isinstance(new_pf, int) and 1 <= new_pf <= len(doc):
                page = doc[new_pf - 1]
                pix = page.get_pixmap(dpi=RENDER_DPI)
                fname = f"{rid}_{new_manual}_p{new_pf}.png"
                out = IMG_DIR / fname
                pix.save(str(out))
                img_rel = str(out.relative_to(ROOT))
                extracted_n += 1
        except Exception as e:
            extraction_errors.append({"row_id": rid, "stage": "image", "error": str(e)})

        page_delta = abs(old_pf - new_pf) if (isinstance(old_pf, int) and isinstance(new_pf, int)) else None

        enriched.append({
            "row_id":            rid,
            "title":             row.get("title"),
            "action_type":       row.get("action_type"),
            "detailed_code":     row.get("detailed_code"),
            "manual":            new_manual,
            "old_page_from":     old_pf,
            "old_page_to":       old_pt,
            "new_page_from":     new_pf,
            "new_page_to":       new_pt,
            "page_delta":        page_delta,
            "old_ref":           row.get("old_manual_ref") or [],
            "new_ref":           row.get("new_manual_ref") or [],
            "confidence":        row.get("confidence"),
            "method":            row.get("method"),
            "new_page_text_preview": new_preview,
            "old_page_text_preview": old_preview,
            "review_focus":      row["review_focus"],
            "selection_reasons": sorted(row["_reasons"]),
            "image_path":        img_rel,
            "human_decision":    "",
            "human_note":        "",
        })

    # PDF 닫기
    for d in docs.values():
        d.close()

    not_extracted_n = len(enriched) - extracted_n

    summary = {
        "total_dryrun_candidates":   len(diff),
        "selected_spotcheck_rows":   len(enriched),
        "high_jump_rows_total":      high_jump_count_total,
        "high_jump_rows_in_pkg":     sum(1 for r in enriched if "high_jump" in r["selection_reasons"]),
        "d2_extra_work_rows":        len(d2_rows),
        "rows_with_image":           extracted_n,
        "rows_without_image":        not_extracted_n,
        "extraction_errors":         len(extraction_errors),
        "error_details_first_5":     extraction_errors[:5],
    }

    out = {
        "spotcheck_meta": {
            "version":           "v3",
            "dryrun_file":       str(DRYRUN.relative_to(ROOT)),
            "preview_max_chars": PREVIEW_MAX,
            "render_dpi":        RENDER_DPI,
            "image_dir":         str(IMG_DIR.relative_to(ROOT)),
        },
        "summary": summary,
        "rows":    enriched,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX (6 sheets) ──
    xlsx_written = False
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1. summary
        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            if isinstance(v, (list, dict)):
                ws_s.append([k, json.dumps(v, ensure_ascii=False)[:32000]])
            else:
                ws_s.append([k, v])
        ws_s.column_dimensions["A"].width = 32
        ws_s.column_dimensions["B"].width = 100

        common_headers = [
            "row_id","title","action_type","detailed_code","manual",
            "old_pf","old_pt","new_pf","new_pt","page_delta",
            "confidence","method","selection_reasons","review_focus",
            "image_path","new_page_preview(~600)","old_page_preview(~600)",
            "human_decision","human_note",
        ]
        common_widths = [10,28,14,12,10,8,8,8,8,9,10,22,30,42,32,60,60,18,30]

        def write_rows_sheet(name: str, items: list[dict]):
            ws = wb.create_sheet(name)
            ws.append(common_headers)
            for c in ws[1]:
                c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for r in items:
                ws.append([
                    r["row_id"], r["title"], r["action_type"], r["detailed_code"], r["manual"],
                    r["old_page_from"], r["old_page_to"], r["new_page_from"], r["new_page_to"], r["page_delta"],
                    r["confidence"], r["method"],
                    ", ".join(r["selection_reasons"]), r["review_focus"],
                    r["image_path"],
                    (r["new_page_text_preview"] or "")[:600],
                    (r["old_page_text_preview"] or "")[:600],
                    "", "",
                ])
            for i, w in enumerate(common_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 26

        # 2. spotcheck_rows (전체)
        write_rows_sheet("spotcheck_rows", enriched)

        # 3. high_jump
        hj_items = [r for r in enriched if "high_jump" in r["selection_reasons"]]
        write_rows_sheet("high_jump", hj_items)

        # 4. d2_extra_work
        d2_items = [r for r in enriched if "d2_extra_work" in r["selection_reasons"]]
        write_rows_sheet("d2_extra_work", d2_items)

        # 5. by_action_type
        ordered_at = []
        for at in ["CHANGE","EXTEND","REGISTRATION","REENTRY","EXTRA_WORK","WORKPLACE","VISA_CONFIRM"]:
            ordered_at.extend([r for r in enriched if r["action_type"] == at])
        write_rows_sheet("by_action_type", ordered_at)

        # 6. all_151_summary
        ws_all = wb.create_sheet("all_151_summary")
        ws_all.append(["row_id","title","action_type","detailed_code",
                       "old_manual","old_pf","old_pt","new_manual","new_pf","new_pt",
                       "page_delta","confidence","method","in_spotcheck"])
        for c in ws_all[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        spot_ids = {r["row_id"] for r in enriched}
        for d in diff:
            opf = d.get("old_page_from")
            npf = d.get("new_page_from")
            delta = abs(opf - npf) if isinstance(opf, int) and isinstance(npf, int) else ""
            ws_all.append([
                d["row_id"], d.get("title",""), d.get("action_type",""), d.get("detailed_code",""),
                d.get("old_manual",""), d.get("old_page_from",""), d.get("old_page_to",""),
                d.get("new_manual",""), d.get("new_page_from",""), d.get("new_page_to",""),
                delta, d.get("confidence",""), d.get("method",""),
                d["row_id"] in spot_ids,
            ])
        widths_all = [10,28,14,12,10,8,8,10,8,8,10,10,24,12]
        for i, w in enumerate(widths_all, start=1):
            ws_all.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws_all.freeze_panes = "A2"

        wb.save(OUT_XLSX)
        xlsx_written = True
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # ── 콘솔 요약 ──
    print("\n" + "=" * 72)
    print("SPOT-CHECK PACKAGE SUMMARY")
    print("=" * 72)
    print(f"  total dry-run candidates:        {summary['total_dryrun_candidates']}")
    print(f"  selected (unique):               {summary['selected_spotcheck_rows']}")
    print(f"  high-jump rows (전체 diff 안):   {summary['high_jump_rows_total']}")
    print(f"  high-jump rows (이 패키지에):    {summary['high_jump_rows_in_pkg']}")
    print(f"  D-2 EXTRA_WORK rows:             {summary['d2_extra_work_rows']}")
    print(f"  PNG extracted:                   {summary['rows_with_image']}")
    print(f"  PNG not extracted:               {summary['rows_without_image']}")
    print(f"  extraction errors:               {summary['extraction_errors']}")

    # action_type 분포 (이 패키지 안)
    at_dist = defaultdict(int)
    for r in enriched:
        at_dist[r["action_type"]] += 1
    print(f"\n  action_type dist (in pkg):  {dict(at_dist)}")

    m_dist = defaultdict(int)
    for r in enriched:
        m_dist[r["manual"]] += 1
    print(f"  manual dist (in pkg):       {dict(m_dist)}")

    print("\n=== Files created ===")
    print(f"  - {OUT_JSON.relative_to(ROOT)}")
    if xlsx_written:
        print(f"  - {OUT_XLSX.relative_to(ROOT)}")
    print(f"  - {IMG_DIR.relative_to(ROOT)}/  ({extracted_n} PNG)")

    print("\n=== Files NOT modified ===")
    print(f"  - immigration_guidelines_db_v2.json  (DB 무수정)")
    print(f"  - manual_mapping_apply_dryrun_v3.json/.xlsx  (입력만)")
    print(f"  - PDFs / structure / indexer / watcher / frontend")


if __name__ == "__main__":
    main()
