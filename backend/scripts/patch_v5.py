"""
Patch v5 — combines Task 1 LLM-judge READY rows (5) + Task 3 D-10 manual research (2).

DEFAULT: dry-run (DB 무수정).
실제 적용은 `python backend/scripts/patch_v5.py --apply` 명시 필요.

Patch targets (7 rows):
  Task 1 (LLM judge + verifier or LLM judge READY recommendation):
    M1-0158 E-4   REGISTRATION   → 체류 p.196
    M1-0166 E-6   VISA_CONFIRM   → 사증 p.165
    M1-0284 F-3-3R CHANGE        → 체류 p.614
    M1-0298 A-2   CHANGE         → 체류 p.20
    M1-0366 F-5-6 CHANGE         → 체류 p.447
  Task 3 (manual research, D-10 NEEDS_HUMAN):
    M1-0130 D-10-1 CHANGE → 체류 p.142-143 + 체류 p.146-148 (non-contig)
    M1-0097 D-10-T CHANGE → 체류 p.144-145 + 체류 p.150 (non-contig)

EXCLUDED:
  M1-0202 E-10 REENTRY (rec p.334 heading is EXTEND, not REENTRY — needs manual re-check)

쓰기 보장:
  manual_ref 외 모든 필드 무수정.
  7 target row_id 외 다른 362 row 무수정.
  master_rows 개수(369) 보존.
  실패 시 abort + 백업 복구 명령 출력.

Same v3/v4 patch pattern (patch_f1_manual_ref_v3.py, patch_f1_residual_v4.py).
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

ROOT       = Path(__file__).parent.parent.parent
DB_PATH    = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"
DRY_JSON   = ROOT / "backend" / "data" / "manuals" / "v5_patch_dryrun.json"
DRY_XLSX   = ROOT / "backend" / "data" / "manuals" / "v5_patch_dryrun.xlsx"

PATCHES: dict[str, dict] = {
    # ── Task 1 READY (LLM judge confirmed + recommended page) ──
    "M1-0158": {
        "detailed_code": "E-4",
        "action_type":   "REGISTRATION",
        "title":         "기술지도",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.196 '다. 제출서류' 헤딩이 E-4 기술지도 외국인등록 신청 제출서류 직접 명시"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  196,
                "page_to":    196,
                "match_type": "manual_override",
                "match_text": "E-4 기술지도 REGISTRATION 제출서류 (체류민원 p.196) — LLM judge 검증",
            },
        ],
    },
    "M1-0166": {
        "detailed_code": "E-6",
        "action_type":   "VISA_CONFIRM",
        "title":         "예술흥행",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.165 '사증발급인정서 / 발급대상' 헤딩 — E-6 예술흥행 사증발급 절차"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "사증민원",
                "page_from":  165,
                "page_to":    165,
                "match_type": "manual_override",
                "match_text": "E-6 예술흥행 VISA_CONFIRM 사증발급인정서 (사증민원 p.165) — LLM judge 검증",
            },
        ],
    },
    "M1-0284": {
        "detailed_code": "F-3-3R",
        "action_type":   "CHANGE",
        "title":         "지역특화숙련기능인력가족",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.614 '5년 이내 지역특화 숙련기능인력 및 동반가족 체류자격으로 / "
            "체류한 적이 있는 자' 헤딩 — F-3-3R 자격변경 요건"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  614,
                "page_to":    614,
                "match_type": "manual_override",
                "match_text": "F-3-3R 지역특화숙련기능인력 CHANGE (체류민원 p.614) — LLM judge 검증",
            },
        ],
    },
    "M1-0298": {
        "detailed_code": "A-2",
        "action_type":   "CHANGE",
        "title":         "공무",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.20 '변경허가 / 필수서류' 헤딩 — A-2 공무 체류자격 변경허가 절차"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  20,
                "page_to":    20,
                "match_type": "manual_override",
                "match_text": "A-2 공무 CHANGE 변경허가 (체류민원 p.20) — LLM judge 검증",
            },
        ],
    },
    "M1-0366": {
        "detailed_code": "F-5-6",
        "action_type":   "CHANGE",
        "title":         "결혼이민자영주",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.447 '1) 허가 요건 / 결혼이민(F-6) 자격으로 2년 이상 국내 계속 체류' "
            "헤딩 — F-5-6 영주(결혼이민) CHANGE 요건"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  447,
                "page_to":    447,
                "match_type": "manual_override",
                "match_text": "F-5-6 영주 CHANGE (체류민원 p.447) — LLM judge 검증",
            },
        ],
    },
    # ── Task 3 D-10 manual research ──
    "M1-0130": {
        "detailed_code": "D-10-1",
        "action_type":   "CHANGE",
        "title":         "구직",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.142-143 '가. 허용대상 1-1) 일반구직(D-10-1) 점수제 / 1-2) 점수제 면제'; "
            "체류민원 p.146-148 '다. 제출서류 1-1) 점수제 적용 + 1-2) 면제 cases ⅰ-ⅶ'"
        ),
        "non_contiguous": True,  # p.142-143 + p.146-148
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  142,
                "page_to":    143,
                "match_type": "manual_override",
                "match_text": "D-10-1 구직 CHANGE 신청요건 (체류민원 p.142-143) — 수동 검증",
            },
            {
                "manual":     "체류민원",
                "page_from":  146,
                "page_to":    148,
                "match_type": "manual_override",
                "match_text": "D-10-1 구직 CHANGE 제출서류 (체류민원 p.146-148) — 수동 검증",
            },
        ],
    },
    "M1-0097": {
        "detailed_code": "D-10-T",
        "action_type":   "CHANGE",
        "title":         "최우수인재 구직",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.144 '4) 최우수인재(D-10-T)' 신청요건 + p.145 '5) 최우수인재(D-10-T) 제한대상'; "
            "체류민원 p.150 '4) 최우수인재(D-10-T)' 제출서류"
        ),
        "non_contiguous": True,  # p.144-145 + p.150
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  144,
                "page_to":    145,
                "match_type": "manual_override",
                "match_text": "D-10-T 최우수인재 CHANGE 신청요건 (체류민원 p.144-145) — 수동 검증",
            },
            {
                "manual":     "체류민원",
                "page_from":  150,
                "page_to":    150,
                "match_type": "manual_override",
                "match_text": "D-10-T 최우수인재 CHANGE 제출서류 (체류민원 p.150) — 수동 검증",
            },
        ],
    },
}


# ──────────────────────────────────────────────────────────────────
# Safety primitives (v3/v4 와 동일)
# ──────────────────────────────────────────────────────────────────
def _abort(msg: str) -> None:
    print(f"\n[ABORT] {msg}", file=sys.stderr)
    sys.exit(1)


def make_backup(db_bytes: bytes) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"immigration_guidelines_db_v2.v5_patch_backup_{ts}.json"
    bk.write_bytes(db_bytes)
    if not bk.exists() or bk.stat().st_size != len(db_bytes):
        _abort(f"백업 작성 검증 실패: {bk}")
    try:
        json.loads(bk.read_text(encoding="utf-8"))
    except Exception as e:
        _abort(f"백업 JSON readback 실패: {e}")
    return bk


def diff_rows(before: list[dict], after: list[dict]) -> list[dict]:
    bi = {r.get("row_id"): r for r in before}
    ai = {r.get("row_id"): r for r in after}
    out = []
    for rid in set(bi.keys()) | set(ai.keys()):
        b = bi.get(rid)
        a = ai.get(rid)
        if b is None or a is None:
            out.append({"row_id": rid, "added_or_removed": True})
            continue
        diff_fields = []
        for k in set(b.keys()) | set(a.keys()):
            if b.get(k) != a.get(k):
                diff_fields.append(k)
        if diff_fields:
            out.append({"row_id": rid, "fields": sorted(diff_fields)})
    return out


def load_db() -> tuple[dict, bytes]:
    raw = DB_PATH.read_bytes()
    return json.loads(raw.decode("utf-8")), raw


def build_patched_db(db: dict) -> tuple[dict, list[dict]]:
    new_db = json.loads(json.dumps(db, ensure_ascii=False))
    rows = new_db.get("master_rows") or []
    found: dict[str, dict] = {}
    for r in rows:
        rid = r.get("row_id")
        if rid in PATCHES:
            found[rid] = r

    missing = [rid for rid in PATCHES if rid not in found]
    if missing:
        _abort(f"target row_ids 누락: {missing}")

    patch_report = []
    for rid, spec in PATCHES.items():
        row = found[rid]
        if row.get("detailed_code") != spec["detailed_code"]:
            _abort(f"{rid} detailed_code 불일치: {row.get('detailed_code')} != {spec['detailed_code']}")
        if row.get("action_type") != spec["action_type"]:
            _abort(f"{rid} action_type 불일치: {row.get('action_type')} != {spec['action_type']}")

        old_ref = row.get("manual_ref") or []
        new_ref = spec["new_manual_ref"]
        row["manual_ref"] = new_ref

        patch_report.append({
            "row_id":           rid,
            "detailed_code":    spec["detailed_code"],
            "action_type":      spec["action_type"],
            "title":            spec["title"],
            "preferred_manual": spec["preferred_manual"],
            "old_manual_ref":   old_ref,
            "new_manual_ref":   new_ref,
            "page_evidence":    spec["page_evidence"],
            "non_contiguous":   spec["non_contiguous"],
            "page_count":       len(new_ref),
        })

    return new_db, patch_report


def verify_patch_safety(before_db: dict, after_db: dict) -> list[str]:
    errors = []
    bm = before_db.get("master_rows") or []
    am = after_db.get("master_rows") or []
    if len(bm) != 369:
        errors.append(f"백업 master_rows 개수가 369 가 아님: {len(bm)}")
    if len(am) != len(bm):
        errors.append(f"패치 후 master_rows 개수 불일치: {len(am)} != {len(bm)}")

    bi = {r.get("row_id"): r for r in bm}
    ai = {r.get("row_id"): r for r in am}
    if set(bi.keys()) != set(ai.keys()):
        errors.append("row_id 집합 변경됨")

    for rid in bi:
        b = bi[rid]
        a = ai.get(rid, {})
        if rid in PATCHES:
            for k in set(b.keys()) | set(a.keys()):
                if k == "manual_ref":
                    continue
                if b.get(k) != a.get(k):
                    errors.append(f"{rid} non-manual_ref 필드 변경됨: {k}")
        else:
            for k in set(b.keys()) | set(a.keys()):
                if b.get(k) != a.get(k):
                    errors.append(f"비-target row {rid} 의 필드 변경됨: {k}")

    for k in set(before_db.keys()) | set(after_db.keys()):
        if k == "master_rows":
            continue
        if before_db.get(k) != after_db.get(k):
            errors.append(f"top-level 키 변경됨: {k}")

    return errors


def write_dryrun_reports(patch_report: list[dict], summary: dict) -> None:
    out_doc = {
        "patch_meta": {
            "version":     "v5",
            "mode":        "DRY_RUN",
            "scope":       "Task 1 (LLM judge READY 5) + Task 3 (D-10 manual 2) = 7 rows",
            "db_path":     str(DB_PATH.relative_to(ROOT)),
            "schema":      "minimal: manual, page_from, page_to, match_type=manual_override, match_text",
        },
        "summary": summary,
        "rows":    patch_report,
    }
    DRY_JSON.write_text(json.dumps(out_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {DRY_JSON.relative_to(ROOT)} ({DRY_JSON.stat().st_size:,} bytes)")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 80

        ws = wb.create_sheet("patches")
        headers = [
            "row_id", "detailed_code", "action_type", "title",
            "preferred_manual", "page_count", "non_contiguous",
            "old_manual_ref", "new_manual_ref", "page_evidence",
        ]
        widths = [10, 12, 14, 32, 14, 10, 14, 80, 100, 80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for x in patch_report:
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                x["preferred_manual"], x["page_count"], x["non_contiguous"],
                json.dumps(x["old_manual_ref"], ensure_ascii=False)[:600],
                json.dumps(x["new_manual_ref"], ensure_ascii=False)[:800],
                x["page_evidence"][:500],
            ])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCB")
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        wb.save(DRY_XLSX)
        print(f"[OUT] {DRY_XLSX.relative_to(ROOT)} ({DRY_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="DB 실제 갱신. 미지정 시 dry-run only.")
    args = ap.parse_args()
    is_apply = args.apply

    if not DB_PATH.exists():
        _abort(f"DB 없음: {DB_PATH}")

    print(f"[patch] mode = {'APPLY' if is_apply else 'DRY-RUN'}")
    print(f"[patch] DB 로드: {DB_PATH.relative_to(ROOT)}")
    db, raw = load_db()
    bytes_before = len(raw)
    print(f"[patch] DB byte size before: {bytes_before:,}")

    bm = db.get("master_rows") or []
    if len(bm) != 369:
        _abort(f"DB master_rows 개수 비정상: {len(bm)} (예상 369)")

    print(f"[patch] target rows: {sorted(PATCHES.keys())}")
    new_db, patch_report = build_patched_db(db)

    print(f"[patch] 변경 사항 검증...")
    errors = verify_patch_safety(db, new_db)
    if errors:
        for e in errors:
            print(f"  [error] {e}")
        _abort("safety check 실패 — abort.")

    diffs = diff_rows(db.get("master_rows") or [], new_db.get("master_rows") or [])
    diff_ids = {d["row_id"] for d in diffs}
    expected_ids = set(PATCHES.keys())
    if diff_ids != expected_ids:
        _abort(f"diff row_id 집합이 target 과 다름: diff={sorted(diff_ids)} expected={sorted(expected_ids)}")
    for d in diffs:
        if d.get("fields") != ["manual_ref"]:
            _abort(f"row {d['row_id']} 가 manual_ref 외 필드 변경: {d.get('fields')}")

    summary = {
        "mode":                       "APPLY" if is_apply else "DRY_RUN",
        "production_db_modified":     False,
        "patch_candidate_count":      len(patch_report),
        "diff_row_ids":               sorted(diff_ids),
        "diff_count":                 len(diff_ids),
        "expected_target_count":      len(PATCHES),
        "non_target_row_diffs":       0,
        "non_manual_ref_field_diffs": 0,
        "non_contiguous_rows":        [x["row_id"] for x in patch_report if x["non_contiguous"]],
        "db_byte_size_before":        bytes_before,
    }

    write_dryrun_reports(patch_report, summary)

    if not is_apply:
        print("\n" + "=" * 78)
        print("PATCH v5 — DRY-RUN (Task 1 LLM 5 + Task 3 D-10 2 = 7 rows)")
        print("=" * 78)
        print("\n1. Files created:")
        print(f"   - {DRY_JSON.relative_to(ROOT)}")
        print(f"   - {DRY_XLSX.relative_to(ROOT)}")

        print("\n2. Production DB modified or not:")
        print("   No — DB / backup / blocklist / triage / verify / 매뉴얼 모든 파일 무수정.")

        print(f"\n3. Patch candidate count: {len(patch_report)}")
        print(f"   target row_ids: {sorted(diff_ids)}")
        print(f"   non-target row diffs: {summary['non_target_row_diffs']}")
        print(f"   non-manual_ref field diffs: {summary['non_manual_ref_field_diffs']}")
        print(f"   non-contiguous rows: {summary['non_contiguous_rows']}")

        print(f"\n4. DB byte size before: {bytes_before:,}")
        print(f"   DB byte size after  : {bytes_before:,}  (unchanged — dry-run)")

        print("\n5. Old vs new manual_ref:")
        for x in patch_report:
            print(f"\n   {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}  {x['title']}")
            print(f"     preferred: {x['preferred_manual']}  non_contig={x['non_contiguous']}  "
                  f"page_count={x['page_count']}")
            print(f"     OLD: " + json.dumps(x["old_manual_ref"], ensure_ascii=False)[:200])
            print(f"     NEW: " + json.dumps(x["new_manual_ref"], ensure_ascii=False)[:300])
            print(f"     evidence: {x['page_evidence'][:200]}")

        print("\n6. Apply command (NOT EXECUTED — 사용자 확인 후 실행):")
        print("   python backend/scripts/patch_v5.py --apply")
        return 0

    # ── APPLY ──
    print("\n[apply] 백업 생성...")
    bk = make_backup(raw)
    print(f"  backup: {bk.relative_to(ROOT)}")

    print("[apply] 신규 JSON 직렬화 + 파일 쓰기...")
    new_json = json.dumps(new_db, ensure_ascii=False, indent=2)
    DB_PATH.write_text(new_json, encoding="utf-8")

    print("[apply] readback 검증...")
    readback = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rb_rows = readback.get("master_rows") or []
    if len(rb_rows) != 369:
        _abort(f"readback master_rows 개수 비정상: {len(rb_rows)}\n"
               f"복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    rb_idx = {r.get("row_id"): r for r in rb_rows}
    for rid, spec in PATCHES.items():
        if rb_idx.get(rid, {}).get("manual_ref") != spec["new_manual_ref"]:
            _abort(f"{rid} readback manual_ref 불일치\n"
                   f"복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")

    errs2 = verify_patch_safety(json.loads(bk.read_text(encoding="utf-8")), readback)
    if errs2:
        for e in errs2:
            print(f"  [error] {e}")
        _abort(f"사후 verify 실패\n복구: cp {bk.relative_to(ROOT)} "
               f"backend/data/immigration_guidelines_db_v2.json")

    bytes_after = DB_PATH.stat().st_size
    summary["production_db_modified"] = True
    summary["backup_path"] = str(bk.relative_to(ROOT))
    summary["db_byte_size_after"] = bytes_after

    print("\n" + "=" * 78)
    print("PATCH v5 — APPLIED (Task 1 LLM 5 + Task 3 D-10 2 = 7 rows)")
    print("=" * 78)
    print(f"  backup: {bk.relative_to(ROOT)}")
    print(f"  DB: {DB_PATH.relative_to(ROOT)} updated ({len(patch_report)} rows)")
    print(f"  byte size before: {bytes_before:,}")
    print(f"  byte size after : {bytes_after:,}")
    print(f"  rollback: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
