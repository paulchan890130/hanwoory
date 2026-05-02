"""
v3 비-EXTRA_WORK 최종 spot-check (READ-ONLY)

목적:
  EXTRA_WORK에서 발견된 "본문 키워드만 보고 자격 설명 페이지를 잘못 매핑" 패턴이
  CHANGE/EXTEND/REGISTRATION/REENTRY/GRANT 에도 존재하는지 확인.

선택:
  현재 dry-run의 139건 (blocklist 제외 후) 중에서:
    1. action_type ≠ EXTRA_WORK
    2. 절대 page_delta 상위 15
    3. CHANGE 최소 3
    4. EXTEND 최소 3
    5. REGISTRATION 최소 3
    6. REENTRY 최소 2
    7. GRANT 임의의 1+
  row_id 기준 dedupe.

판정:
  PASS         — heading 영역에 해당 action_type의 헤더 키워드 등장
  BLOCK        — 자격 설명 마커만 있고 action 키워드는 페이지 어디에도 없음
  NEEDS_REVIEW — 신호 혼합 또는 법조문 인용만 존재

출력:
  backend/data/manuals/manual_mapping_final_spotcheck_v3.json
  backend/data/manuals/manual_mapping_final_spotcheck_v3.xlsx
"""
from __future__ import annotations
import json, re, sys
from pathlib import Path
from collections import OrderedDict, Counter

import fitz

ROOT     = Path(__file__).parent.parent.parent
MANUALS  = ROOT / "backend" / "data" / "manuals"
DRYRUN   = MANUALS / "manual_mapping_apply_dryrun_v3.json"
SPOT_DIR = MANUALS / "spotcheck_pages_v3"

OUT_JSON = MANUALS / "manual_mapping_final_spotcheck_v3.json"
OUT_XLSX = MANUALS / "manual_mapping_final_spotcheck_v3.xlsx"

PDF_PATHS = {
    "체류민원": MANUALS / "unlocked_체류민원.pdf",
    "사증민원": MANUALS / "unlocked_사증민원.pdf",
}

# ─── action_type 헤더 키워드 ───────────────────────────────────
ACTION_HEADERS = {
    "CHANGE":       ["체류자격 변경", "체류자격변경", "변경허가", "자격변경", "체류자격 변경허가"],
    "EXTEND":       ["체류기간 연장", "체류기간연장", "연장허가", "기간연장", "체류기간 연장허가"],
    "REGISTRATION": ["외국인등록", "외국인 등록", "등록 신청", "등록사항"],
    "REENTRY":      ["재입국허가", "재입국 허가", "재입국 신청"],
    "GRANT":        ["체류자격 부여", "체류자격부여", "자격부여"],
}

QUALIF_MARKERS = [
    "자격해당자", "자격 해당자",
    "1회에부여할수", "1회에 부여할 수",
    "체류기간상한", "체류기간 상한",
    "별도허가없이", "별도 허가 없이",
    "유학활동이가능한", "유학활동이 가능한",
]

# 법조문 인용 패턴 — 본문 인용으로 분류
LAW_CITATION_PATTERNS = {
    "CHANGE":       re.compile(r"법\s*제24조|시행령\s*제30조"),
    "EXTEND":       re.compile(r"법\s*제25조|시행령\s*제31조"),
    "REGISTRATION": re.compile(r"법\s*제31조|시행령\s*제40조"),
    "REENTRY":      re.compile(r"법\s*제30조|시행령\s*제39조"),
    "GRANT":        re.compile(r"법\s*제23조|시행령\s*제29조"),
}


def heading_text(text: str, n: int = 600) -> str:
    return text[:n]


def find_action_in_head(page_text: str, action_type: str) -> tuple[bool, str]:
    head = heading_text(page_text, 600)
    for kw in ACTION_HEADERS.get(action_type, []):
        if kw in head:
            return True, kw
    return False, ""


def find_action_anywhere(page_text: str, action_type: str) -> tuple[bool, str]:
    body = page_text[:2500]
    for kw in ACTION_HEADERS.get(action_type, []):
        if kw in body:
            return True, kw
    return False, ""


def has_qualif_in_head(page_text: str) -> bool:
    head = heading_text(page_text, 600)
    return any(m in head for m in QUALIF_MARKERS)


def is_only_law_citation(page_text: str, action_type: str) -> bool:
    pat = LAW_CITATION_PATTERNS.get(action_type)
    if not pat: return False
    return bool(pat.search(page_text[:2000]))


def classify(page_text: str, action_type: str) -> tuple[str, str]:
    if not page_text:
        return "NEEDS_REVIEW", "page_text 없음 (PDF 추출 실패 가능)"

    in_head, head_kw = find_action_in_head(page_text, action_type)
    anywhere, any_kw = find_action_anywhere(page_text, action_type)
    qualif = has_qualif_in_head(page_text)

    # PASS: heading 영역에 action header 등장
    if in_head:
        if qualif:
            return "PASS", (
                f"heading 영역에 '{head_kw}' 키워드 + qualification carry-over (단일 페이지에 자격 카드+섹션 공존)"
            )
        return "PASS", f"heading 영역에 '{head_kw}' 키워드 등장"

    # BLOCK: 자격 마커만 있고 키워드는 페이지 어디에도 없음
    if qualif and not anywhere:
        return "BLOCK", "qualification 페이지로 보임 — head에 자격 마커만 있고 action 키워드 페이지 전체에 없음"

    # NEEDS_REVIEW: 키워드는 본문에만, heading은 자격
    if qualif and anywhere:
        # 법조문 인용만인지 확인
        only_law = is_only_law_citation(page_text, action_type) and not in_head
        if only_law:
            return "BLOCK", f"action 키워드 '{any_kw}' 가 법조문 인용 형태로만 등장 — 실제 섹션 헤더 아님"
        return "NEEDS_REVIEW", (
            f"qualification head + body에 '{any_kw}' 키워드 등장하나 heading 형태 아님 — 사람 검토"
        )

    # 키워드 어디에도 없음
    if not anywhere:
        return "BLOCK", f"페이지에 '{action_type}' 헤더 키워드 자체가 없음"

    # 키워드는 있지만 head 영역 밖 (qualif 마커 없음)
    return "NEEDS_REVIEW", f"action 키워드 '{any_kw}' 가 페이지 본문에는 있으나 heading 영역에 없음"


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not DRYRUN.exists():
        print(f"[ABORT] dry-run 파일 없음: {DRYRUN}")
        sys.exit(1)

    print("[final-spotcheck] dry-run 로드...")
    dry = json.loads(DRYRUN.read_text(encoding="utf-8"))
    diff = dry.get("diff") or []
    print(f"  current dry-run rows (post-blocklist): {len(diff)}")

    # ── 비-EXTRA_WORK ──
    non_ew = [d for d in diff if d.get("action_type") != "EXTRA_WORK"]
    print(f"  non-EXTRA_WORK rows: {len(non_ew)}")

    # 각 행에 page_delta 계산 추가
    for d in non_ew:
        opf = d.get("old_page_from")
        npf = d.get("new_page_from")
        d["_page_delta"] = abs(opf - npf) if isinstance(opf, int) and isinstance(npf, int) else 0

    # ── 선택 ──
    selected: "OrderedDict[str, dict]" = OrderedDict()

    def add(d, reason):
        rid = d["row_id"]
        if rid in selected:
            selected[rid]["_reasons"].add(reason)
        else:
            selected[rid] = {**d, "_reasons": {reason}}

    # 1. top 15 by abs(page_delta), 비-EXTRA_WORK
    by_delta = sorted(non_ew, key=lambda x: -x["_page_delta"])
    for d in by_delta[:15]:
        add(d, "top15_by_delta")

    # 2~6. action_type별 최소 표본 — 절대 page_delta 큰 순으로 채움
    by_at = {}
    for d in non_ew:
        by_at.setdefault(d["action_type"], []).append(d)
    for at in by_at:
        by_at[at].sort(key=lambda x: -x["_page_delta"])

    targets_per_at = {"CHANGE": 3, "EXTEND": 3, "REGISTRATION": 3, "REENTRY": 2, "GRANT": 1}
    for at, n_min in targets_per_at.items():
        in_pkg = sum(1 for r in selected.values() if r["action_type"] == at)
        if in_pkg < n_min:
            need = n_min - in_pkg
            pool = by_at.get(at, [])
            added = 0
            for d in pool:
                if d["row_id"] in selected:
                    continue
                add(d, f"min_{at}")
                added += 1
                if added >= need:
                    break

    print(f"  selected (unique): {len(selected)}")

    # ── PDF 로드 ──
    print("[final-spotcheck] PDF 로드...")
    docs: dict[str, fitz.Document] = {}
    for label, p in PDF_PATHS.items():
        if p.exists():
            docs[label] = fitz.open(p)
            print(f"  {label}: {len(docs[label])} pages")

    # ── 분류 ──
    results = []
    for rid, d in selected.items():
        manual = d.get("new_manual")
        pf = d.get("new_page_from")
        pt = d.get("new_page_to")
        at = d.get("action_type")

        page_text = ""
        if manual in docs and isinstance(pf, int) and 1 <= pf <= len(docs[manual]):
            page_text = docs[manual][pf - 1].get_text()

        judgment, reason = classify(page_text, at)

        # heading detection summary
        head = heading_text(page_text, 600)
        head_kw_match = ""
        for kw in ACTION_HEADERS.get(at, []):
            if kw in head:
                head_kw_match = kw
                break

        # spotcheck PNG path (이미 만들어진 것 재사용)
        img_rel = ""
        candidate_img = SPOT_DIR / f"{rid}_{manual}_p{pf}.png"
        if candidate_img.exists():
            img_rel = str(candidate_img.relative_to(ROOT))

        results.append({
            "row_id":               rid,
            "title":                d.get("title"),
            "action_type":          at,
            "detailed_code":        d.get("detailed_code"),
            "old_manual":           d.get("old_manual"),
            "old_page_from":        d.get("old_page_from"),
            "old_page_to":          d.get("old_page_to"),
            "new_manual":           manual,
            "new_page_from":        pf,
            "new_page_to":          pt,
            "page_delta":           d["_page_delta"],
            "page_heading_top":     head[:250],
            "head_kw_match":        head_kw_match,
            "new_page_text_preview": page_text[:1500],
            "image_path":           img_rel,
            "selection_reasons":    sorted(d["_reasons"]),
            "judgment":             judgment,
            "reason":               reason,
        })

    for d in docs.values():
        d.close()

    # ── 통계 ──
    j_dist = Counter(r["judgment"] for r in results)
    by_at_dist = Counter(r["action_type"] for r in results)

    summary = {
        "selected_total":           len(results),
        "PASS_count":               j_dist.get("PASS", 0),
        "BLOCK_count":              j_dist.get("BLOCK", 0),
        "NEEDS_REVIEW_count":       j_dist.get("NEEDS_REVIEW", 0),
        "by_action_type":           dict(by_at_dist),
    }

    out = {
        "final_spotcheck_meta": {
            "version":      "v3",
            "dryrun_file":  str(DRYRUN.relative_to(ROOT)),
            "scope":        "non-EXTRA_WORK action types (CHANGE/EXTEND/REGISTRATION/REENTRY/GRANT)",
            "rule":         "PASS = heading area contains action header keyword; BLOCK = qualification only with no action keyword; NEEDS_REVIEW otherwise",
        },
        "summary": summary,
        "rows":    results,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OUT] {OUT_JSON.relative_to(ROOT)} ({OUT_JSON.stat().st_size:,} bytes)")

    # ── XLSX ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            if isinstance(v, dict):
                ws_s.append([k, json.dumps(v, ensure_ascii=False)])
            else:
                ws_s.append([k, v])
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 80

        ws = wb.create_sheet("final_spotcheck")
        headers = [
            "row_id","detailed_code","action_type","title",
            "old_manual","old_pf","old_pt","new_manual","new_pf","new_pt","page_delta",
            "judgment","reason","head_kw_match","selection_reasons",
            "image_path","page_heading_top(~250)","new_page_text_preview(~700)",
        ]
        widths  = [10,12,14,28,10,8,8,10,8,8,10,14,60,14,30,30,50,80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 정렬: BLOCK > NEEDS_REVIEW > PASS, then by page_delta desc
        ord_j = {"BLOCK": 0, "NEEDS_REVIEW": 1, "PASS": 2}
        results_sorted = sorted(results, key=lambda r: (ord_j.get(r["judgment"], 9), -r["page_delta"]))
        for r in results_sorted:
            ws.append([
                r["row_id"], r["detailed_code"], r["action_type"], r["title"],
                r["old_manual"], r["old_page_from"], r["old_page_to"],
                r["new_manual"], r["new_page_from"], r["new_page_to"], r["page_delta"],
                r["judgment"], r["reason"], r["head_kw_match"],
                ", ".join(r["selection_reasons"]),
                r["image_path"],
                r["page_heading_top"][:250],
                (r["new_page_text_preview"] or "")[:700],
            ])
            color = {"BLOCK":"FFCCCB","NEEDS_REVIEW":"FFE599","PASS":"C6F6D5"}.get(r["judgment"])
            if color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=color)

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        wb.save(OUT_XLSX)
        print(f"[OUT] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")

    # ── 콘솔 ──
    print("\n" + "=" * 76)
    print("FINAL NON-EXTRA_WORK SPOTCHECK SUMMARY")
    print("=" * 76)
    print(f"  selected total:   {summary['selected_total']}")
    print(f"  PASS:             {summary['PASS_count']}")
    print(f"  BLOCK:            {summary['BLOCK_count']}")
    print(f"  NEEDS_REVIEW:     {summary['NEEDS_REVIEW_count']}")
    print(f"  by action_type:   {summary['by_action_type']}")

    if j_dist.get("BLOCK", 0):
        print("\n=== BLOCK (clearly false positive — recommend blocklist) ===")
        for r in [x for x in results if x["judgment"] == "BLOCK"]:
            print(f"  {r['row_id']:8s} {r['detailed_code']:8s} {r['action_type']:13s} "
                  f"{r['new_manual']} p.{r['new_page_from']} delta={r['page_delta']:>3}  "
                  f"{(r['title'] or '')[:24]}  :: {r['reason']}")

    if j_dist.get("NEEDS_REVIEW", 0):
        print("\n=== NEEDS_REVIEW (사람 시각 검토) ===")
        for r in [x for x in results if x["judgment"] == "NEEDS_REVIEW"]:
            print(f"  {r['row_id']:8s} {r['detailed_code']:8s} {r['action_type']:13s} "
                  f"{r['new_manual']} p.{r['new_page_from']} delta={r['page_delta']:>3}  "
                  f"{(r['title'] or '')[:24]}  :: {r['reason']}")

    if j_dist.get("PASS", 0):
        print("\n=== PASS samples (heading 키워드 검출, 자동 통과) — 첫 10개 ===")
        for r in [x for x in results if x["judgment"] == "PASS"][:10]:
            print(f"  {r['row_id']:8s} {r['detailed_code']:8s} {r['action_type']:13s} "
                  f"{r['new_manual']} p.{r['new_page_from']} delta={r['page_delta']:>3}  "
                  f"head_kw='{r['head_kw_match']}'  {(r['title'] or '')[:20]}")


if __name__ == "__main__":
    main()
