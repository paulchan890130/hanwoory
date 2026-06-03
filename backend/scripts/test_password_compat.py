"""Tests for backward-compatible password verification + jpup hash migration.

Run (repo root, venv active): python backend/scripts/test_password_compat.py

NO raw password is printed or hardcoded as a real secret. The jpup fixture is
asserted only for *format* and *safe handling* — never its actual password.
"""
from __future__ import annotations

import base64
import hashlib
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
except Exception:
    pass

from backend.services.accounts_service import hash_password, verify_password

_ok = {"v": True}


def check(name, cond):
    print(("PASS" if cond else "FAIL"), name)
    _ok["v"] = _ok["v"] and bool(cond)


def _legacy_hash(password: str, salt: bytes | None = None) -> str:
    """Reproduce the legacy Streamlit Accounts hash format exactly."""
    salt = salt or os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100_000)
    return base64.b64encode(salt + dk).decode("ascii")


def main() -> int:
    # 1) legacy PBKDF2 hash verifies the correct password
    h = _legacy_hash("Corr3ct-Horse")
    check("1 legacy hash verifies correct password", verify_password("Corr3ct-Horse", h) is True)

    # 2) legacy PBKDF2 hash rejects the wrong password
    check("2 legacy hash rejects wrong password", verify_password("wrong-password", h) is False)

    # current hash_password output is the same legacy format → round-trips
    h2 = hash_password("Another_PW!")
    check("2b hash_password round-trips", verify_password("Another_PW!", h2) is True
          and verify_password("nope", h2) is False)

    # 3) "new format" (bcrypt $2…): verified if bcrypt present, else safe False — never throws
    bcrypt_hash = "$2b$12$" + "A" * 53  # well-formed-looking bcrypt hash
    try:
        res = verify_password("whatever", bcrypt_hash)
        check("3 bcrypt-format hash handled without throwing", res in (True, False))
        try:
            import bcrypt  # noqa: F401
            real = bcrypt.hashpw(b"bc-pass", bcrypt.gensalt()).decode("ascii")
            check("3b bcrypt verify works when bcrypt installed",
                  verify_password("bc-pass", real) is True and verify_password("x", real) is False)
        except ImportError:
            check("3b bcrypt absent → bcrypt-format safely rejected", verify_password("whatever", bcrypt_hash) is False)
    except Exception as e:
        check(f"3 bcrypt-format must not throw (got {type(e).__name__})", False)

    # 4) malformed hashes return False and never throw
    malformed = ["", "   ", "not-base64-!!!", "短い", "$2", "QQ==", "x" * 3, None]  # incl too-short b64
    safe = True
    for m in malformed:
        try:
            if verify_password("p", m) is not False:  # type: ignore[arg-type]
                safe = False
        except Exception:
            safe = False
    check("4 malformed hashes → False, no throw", safe)

    # 5) jpup migrated hash: assert FORMAT + SAFE HANDLING only (never the password)
    try:
        from openpyxl import load_workbook
        from pathlib import Path
        root = Path(__file__).resolve().parent.parent.parent
        src = None
        for p in sorted((root / "migration_input").glob("*.xlsx")):
            wb = load_workbook(p, read_only=True, data_only=True)
            if "Accounts" in wb.sheetnames:
                src = (p, wb); break
        if src is None:
            print("INFO 5 jpup fixture skipped (no snapshot Accounts tab in this checkout)")
        else:
            _, wb = src
            ws = wb["Accounts"]
            rows = list(ws.iter_rows(values_only=True))
            header = [str(h or "").strip() for h in rows[0]]
            li, pi = header.index("login_id"), header.index("password_hash")
            jpup = next((r for r in rows[1:] if str(r[li]).strip().lower() == "jpup"), None)
            check("5 jpup row present in snapshot", jpup is not None)
            if jpup is not None:
                ph = str(jpup[pi]).strip()
                raw = base64.b64decode(ph.encode("ascii"))
                check("5 jpup hash is legacy PBKDF2 format (48 bytes: 16 salt + 32 dk)", len(raw) == 48)
                # processed as legacy, returns False for an obviously-wrong password, no crash
                check("5 jpup hash safely rejects a wrong password", verify_password("definitely-not-it", ph) is False)
    except Exception as e:
        check(f"5 jpup fixture check must not throw (got {type(e).__name__})", False)

    # 6) migration script source reader finds jpup's legacy hash (non-empty)
    try:
        from backend.scripts.migrate_account_password_hashes import _discover_source, _read_legacy_hashes
        s = _discover_source()
        if s is None:
            print("INFO 6 migration source skipped (no snapshot in this checkout)")
        else:
            legacy = _read_legacy_hashes(s)
            check("6 migration reader returns non-empty jpup hash", bool(legacy.get("jpup")))
    except Exception as e:
        check(f"6 migration reader must not throw (got {type(e).__name__})", False)

    print("\nRESULT:", "ALL PASS" if _ok["v"] else "SOME FAILED")
    return 0 if _ok["v"] else 1


if __name__ == "__main__":
    sys.exit(main())
