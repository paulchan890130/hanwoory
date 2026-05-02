"""
F-1 manual_ref surgical patch (DEFAULT: dry-run)

Patch 대상 (4 row만, USER_VALIDATED):
  M1-0209  F-1-21  VISA_CONFIRM  주한 외국공관원 가사보조인
  M1-0211  F-1-15  EXTEND        우수인재·투자자·유학생 부모 방문동거
  M1-0227  F-1-22  EXTEND        고액투자가 가사보조인 방문동거
  M1-0230  F-1-24  EXTEND        해외우수인재 가사보조인 방문동거

DEFAULT: dry-run (DB 무수정).
실제 적용은 `python backend/scripts/patch_f1_manual_ref_v3.py --apply` 명시 필요.

쓰기 보장:
  manual_ref 외 모든 필드 무수정.
  4 target row_id 외 다른 365 row 무수정.
  master_rows 개수(369) 보존.
  실패 시 abort + 백업 복구 명령 출력.
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path

ROOT      = Path(__file__).parent.parent.parent
DB_PATH   = ROOT / "backend" / "data" / "immigration_guidelines_db_v2.json"
BACKUP_DIR = ROOT / "backend" / "data" / "backups"
DRY_JSON  = ROOT / "backend" / "data" / "manuals" / "f1_manual_ref_patch_dryrun_v3.json"
DRY_XLSX  = ROOT / "backend" / "data" / "manuals" / "f1_manual_ref_patch_dryrun_v3.xlsx"

# ──────────────────────────────────────────────────────────────────
# Patch spec — 4 row만, validate_f1_manual_ref_v3 결과 기반
# 매뉴얼 페이지는 PDF 본문 직접 확인 (체류 p.347-349, p.668-669; 사증 p.298, p.305-307)
# ──────────────────────────────────────────────────────────────────
PATCHES: dict[str, dict] = {
    "M1-0209": {
        "detailed_code": "F-1-21",
        "action_type":   "VISA_CONFIRM",
        "title":         "주한 외국공관원 가사보조인",
        "preferred_manual": "사증민원",
        "page_evidence": (
            "사증민원 p.298 #2: '주한 외국공관원의 가사보조인(F-1-21) 단수사증' / "
            "사증민원 p.306-307 #3: 사증발급인정서 첨부서류 (외국공관 요청공문 + 고용계약서)"
        ),
        "non_contiguous": True,  # p.298 + p.306-307
        "new_manual_ref": [
            {
                "manual":     "사증민원",
                "page_from":  298,
                "page_to":    298,
                "match_type": "manual_override",
                "match_text": "F-1-21 주한 외국공관원 가사보조인 단수사증 (사증민원 p.298 #2) — 수동 검증",
            },
            {
                "manual":     "사증민원",
                "page_from":  306,
                "page_to":    307,
                "match_type": "manual_override",
                "match_text": "F-1-21 사증발급인정서 첨부서류 (사증민원 p.306-307 #3) — 수동 검증",
            },
        ],
    },
    "M1-0211": {
        "detailed_code": "F-1-15",
        "action_type":   "EXTEND",
        "title":         "우수인재·투자자·유학생 부모 방문동거",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.349 #8: '우수인재, 투자자 및 유학생 부모' 체류기간 연장허가 제출서류 "
            "(신청서, 신원보증서, 가족관계 입증서류, 체류지 입증서류)"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  349,
                "page_to":    349,
                "match_type": "manual_override",
                "match_text": "F-1-15 우수인재·투자자·유학생 부모 체류기간 연장 (체류민원 p.349 #8) — 수동 검증",
            },
        ],
    },
    "M1-0227": {
        "detailed_code": "F-1-22",
        "action_type":   "EXTEND",
        "title":         "고액투자가 가사보조인 방문동거",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.348: 외국인투자자 / 우수전문인력의 가사보조인 체류기간 연장 절차 "
            "(고용주 재직증명서, 외국인투자신고서, 체류지 입증서류 등)"
        ),
        "non_contiguous": False,
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  348,
                "page_to":    348,
                "match_type": "manual_override",
                "match_text": "F-1-22 고액투자가 가사보조인 체류기간 연장 (체류민원 p.348) — 수동 검증",
            },
        ],
    },
    "M1-0230": {
        "detailed_code": "F-1-24",
        "action_type":   "EXTEND",
        "title":         "해외우수인재 가사보조인 방문동거",
        "preferred_manual": "체류민원",
        "page_evidence": (
            "체류민원 p.348: 외국인투자자 / 우수전문인력의 가사보조인 체류기간 연장 절차 / "
            "체류민원 p.668-669: 최우수인재 부모·가사보조인 신청서류 (1명 초청 가능, "
            "고용계약서, 주 자격자 소득요건 입증서류)"
        ),
        "non_contiguous": True,  # p.348 + p.668-669
        "new_manual_ref": [
            {
                "manual":     "체류민원",
                "page_from":  348,
                "page_to":    348,
                "match_type": "manual_override",
                "match_text": "F-1-24 해외우수인재 가사보조인 체류기간 연장 (체류민원 p.348) — 수동 검증",
            },
            {
                "manual":     "체류민원",
                "page_from":  668,
                "page_to":    669,
                "match_type": "manual_override",
                "match_text": "F-1-24 최우수인재 부모·가사보조인 신청서류 (체류민원 p.668-669) — 수동 검증",
            },
        ],
    },
}


# ──────────────────────────────────────────────────────────────────
# Safety primitives
# ──────────────────────────────────────────────────────────────────
def _abort(msg: str) -> None:
    print(f"\n[ABORT] {msg}", file=sys.stderr)
    sys.exit(1)


def make_backup(db_bytes: bytes) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUP_DIR / f"immigration_guidelines_db_v2.f1_manual_ref_patch_backup_{ts}.json"
    bk.write_bytes(db_bytes)
    if not bk.exists() or bk.stat().st_size != len(db_bytes):
        _abort(f"백업 작성 검증 실패: {bk}")
    # readback
    try:
        json.loads(bk.read_text(encoding="utf-8"))
    except Exception as e:
        _abort(f"백업 JSON readback 실패: {e}")
    return bk


def diff_rows(before: list[dict], after: list[dict]) -> list[dict]:
    """row-level diff. row_id 기준 매칭, 변경된 row만 반환."""
    bi = {r.get("row_id"): r for r in before}
    ai = {r.get("row_id"): r for r in after}
    out = []
    for rid in set(bi.keys()) | set(ai.keys()):
        b = bi.get(rid)
        a = ai.get(rid)
        if b is None or a is None:
            out.append({"row_id": rid, "added_or_removed": True})
            continue
        diff_fields = []
        all_keys = set(b.keys()) | set(a.keys())
        for k in all_keys:
            if b.get(k) != a.get(k):
                diff_fields.append(k)
        if diff_fields:
            out.append({"row_id": rid, "fields": diff_fields})
    return out


# ──────────────────────────────────────────────────────────────────
# Build / verify
# ──────────────────────────────────────────────────────────────────
def load_db() -> tuple[dict, bytes]:
    raw = DB_PATH.read_bytes()
    return json.loads(raw.decode("utf-8")), raw


def build_patched_db(db: dict) -> tuple[dict, list[dict]]:
    """deep-copy db, target row 4건의 manual_ref 만 새 값으로 치환. 변경 사항 리포트 생성."""
    new_db = json.loads(json.dumps(db, ensure_ascii=False))
    rows = new_db.get("master_rows") or []
    found: dict[str, dict] = {}

    for r in rows:
        rid = r.get("row_id")
        if rid in PATCHES:
            found[rid] = r

    missing = [rid for rid in PATCHES if rid not in found]
    if missing:
        _abort(f"target row_ids 누락: {missing}")

    patch_report = []
    for rid, spec in PATCHES.items():
        row = found[rid]
        # sanity: detailed_code / action_type 일치
        if row.get("detailed_code") != spec["detailed_code"]:
            _abort(f"{rid} detailed_code 불일치: {row.get('detailed_code')} != {spec['detailed_code']}")
        if row.get("action_type") != spec["action_type"]:
            _abort(f"{rid} action_type 불일치: {row.get('action_type')} != {spec['action_type']}")

        old_ref = row.get("manual_ref") or []
        new_ref = spec["new_manual_ref"]
        row["manual_ref"] = new_ref

        patch_report.append({
            "row_id":           rid,
            "detailed_code":    spec["detailed_code"],
            "action_type":      spec["action_type"],
            "title":            spec["title"],
            "preferred_manual": spec["preferred_manual"],
            "old_manual_ref":   old_ref,
            "new_manual_ref":   new_ref,
            "reason": (
                f"action_type={spec['action_type']} → preferred={spec['preferred_manual']}. "
                f"기존 체류 p.543 + 사증 p.404 본문은 외국국적동포가족 방문동거 절차로, "
                f"본 row({spec['title']}) 와 무관. PDF 본문 직접 확인으로 정답 페이지 교체."
            ),
            "page_evidence":   spec["page_evidence"],
            "non_contiguous":  spec["non_contiguous"],
            "page_count":      len(new_ref),
        })

    return new_db, patch_report


def verify_patch_safety(before_db: dict, after_db: dict) -> list[str]:
    """패치 후 모든 안전 보장을 검증. 위반 사항 리스트 반환 (빈 리스트면 OK)."""
    errors = []
    bm = before_db.get("master_rows") or []
    am = after_db.get("master_rows") or []

    if len(bm) != 369:
        errors.append(f"백업 master_rows 개수가 369 가 아님: {len(bm)}")
    if len(am) != len(bm):
        errors.append(f"패치 후 master_rows 개수 불일치: {len(am)} != {len(bm)}")

    bi = {r.get("row_id"): r for r in bm}
    ai = {r.get("row_id"): r for r in am}

    if set(bi.keys()) != set(ai.keys()):
        errors.append("row_id 집합 변경됨")

    for rid in bi:
        b = bi[rid]
        a = ai.get(rid, {})
        if rid in PATCHES:
            # manual_ref 만 변경되어야 함, 다른 모든 필드는 그대로
            for k in set(b.keys()) | set(a.keys()):
                if k == "manual_ref":
                    continue
                if b.get(k) != a.get(k):
                    errors.append(f"{rid} non-manual_ref 필드 변경됨: {k}")
        else:
            # target 외 row 는 어떤 필드도 변경되어선 안 됨
            for k in set(b.keys()) | set(a.keys()):
                if b.get(k) != a.get(k):
                    errors.append(f"비-target row {rid} 의 필드 변경됨: {k}")

    # master_rows 외 다른 top-level 키도 보존
    for k in set(before_db.keys()) | set(after_db.keys()):
        if k == "master_rows":
            continue
        if before_db.get(k) != after_db.get(k):
            errors.append(f"top-level 키 변경됨: {k}")

    return errors


# ──────────────────────────────────────────────────────────────────
# Reports
# ──────────────────────────────────────────────────────────────────
def write_dryrun_reports(patch_report: list[dict], summary: dict) -> None:
    out_doc = {
        "patch_meta": {
            "version":     "v3",
            "mode":        "DRY_RUN",
            "db_path":     str(DB_PATH.relative_to(ROOT)),
            "scope":       "F-1 USER_VALIDATED 4 rows",
            "schema":      "minimal: manual, page_from, page_to, match_type=manual_override, match_text",
        },
        "summary": summary,
        "rows":    patch_report,
    }
    DRY_JSON.write_text(json.dumps(out_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OUT] {DRY_JSON.relative_to(ROOT)} ({DRY_JSON.stat().st_size:,} bytes)")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_s = wb.create_sheet("summary")
        ws_s.append(["metric", "value"])
        for c in ws_s[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
        for k, v in summary.items():
            ws_s.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 80

        ws = wb.create_sheet("patches")
        headers = [
            "row_id", "detailed_code", "action_type", "title",
            "preferred_manual", "page_count", "non_contiguous",
            "old_manual_ref", "new_manual_ref",
            "page_evidence", "reason",
        ]
        widths = [10, 12, 14, 32, 14, 10, 14, 80, 80, 80, 80]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFE599")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for x in patch_report:
            ws.append([
                x["row_id"], x["detailed_code"], x["action_type"], x["title"],
                x["preferred_manual"], x["page_count"], x["non_contiguous"],
                json.dumps(x["old_manual_ref"], ensure_ascii=False)[:600],
                json.dumps(x["new_manual_ref"], ensure_ascii=False)[:600],
                x["page_evidence"][:500],
                x["reason"][:600],
            ])
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFCCCB")
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        wb.save(DRY_XLSX)
        print(f"[OUT] {DRY_XLSX.relative_to(ROOT)} ({DRY_XLSX.stat().st_size:,} bytes)")
    except ImportError as e:
        print(f"[skip xlsx] openpyxl: {e}")
    except Exception as e:
        print(f"[skip xlsx] {type(e).__name__}: {e}")


# ──────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="DB 실제 갱신. 미지정 시 dry-run only.")
    args = ap.parse_args()
    is_apply = args.apply

    if not DB_PATH.exists():
        _abort(f"DB 없음: {DB_PATH}")

    print(f"[patch] mode = {'APPLY' if is_apply else 'DRY-RUN'}")
    print(f"[patch] DB 로드: {DB_PATH.relative_to(ROOT)}")
    db, raw = load_db()

    bm = db.get("master_rows") or []
    if len(bm) != 369:
        _abort(f"DB master_rows 개수 비정상: {len(bm)} (예상 369)")

    print(f"[patch] target rows: {sorted(PATCHES.keys())}")
    new_db, patch_report = build_patched_db(db)

    print(f"[patch] 변경 사항 검증...")
    errors = verify_patch_safety(db, new_db)
    if errors:
        for e in errors:
            print(f"  [error] {e}")
        _abort("safety check 실패 — abort.")

    # diff 집계
    diffs = diff_rows(db.get("master_rows") or [], new_db.get("master_rows") or [])
    diff_ids = {d["row_id"] for d in diffs}
    expected_ids = set(PATCHES.keys())
    if diff_ids != expected_ids:
        _abort(f"diff row_id 집합이 target 과 다름: diff={sorted(diff_ids)} expected={sorted(expected_ids)}")
    for d in diffs:
        if d.get("fields") != ["manual_ref"]:
            _abort(f"row {d['row_id']} 가 manual_ref 외 필드 변경: {d.get('fields')}")

    summary = {
        "mode":                       "APPLY" if is_apply else "DRY_RUN",
        "production_db_modified":     False,  # apply 성공 후에만 True 로 갱신
        "patch_candidate_count":      len(patch_report),
        "diff_row_ids":               sorted(diff_ids),
        "diff_count":                 len(diff_ids),
        "expected_target_count":      len(PATCHES),
        "schema_compatibility_issue": "",
        "non_contiguous_rows":        [x["row_id"] for x in patch_report if x["non_contiguous"]],
    }

    write_dryrun_reports(patch_report, summary)

    if not is_apply:
        # ── dry-run 리포트 ──
        print("\n" + "=" * 78)
        print("F-1 MANUAL_REF SURGICAL PATCH — DRY-RUN")
        print("=" * 78)
        print("\n1. Files created:")
        print(f"   - {DRY_JSON.relative_to(ROOT)}")
        print(f"   - {DRY_XLSX.relative_to(ROOT)}")
        print(f"   (no DB or backup file written in dry-run)")

        print("\n2. Production DB modified or not:")
        print("   No — DB / backup / blocklist / triage / verify / 매뉴얼 모든 파일 무수정 (dry-run).")

        print(f"\n3. Patch candidate count: {len(patch_report)}")

        print("\n4. Old vs new manual_ref:")
        for x in patch_report:
            print(f"\n   {x['row_id']}  {x['detailed_code']:8s}  {x['action_type']:13s}  {x['title']}")
            print(f"     preferred_manual: {x['preferred_manual']}  "
                  f"non_contiguous={x['non_contiguous']}  page_count={x['page_count']}")
            print(f"     OLD: " + json.dumps(x["old_manual_ref"], ensure_ascii=False))
            print(f"     NEW: " + json.dumps(x["new_manual_ref"], ensure_ascii=False))

        print("\n5. Schema compatibility issue:")
        print("   None. minimal manual_override schema (manual, page_from, page_to, "
              "match_type, match_text) — frontend ManualPdfViewer 는 page_from/page_to만 사용.")
        print("   기존 schema 의 score / section_pf / section_pt / kind / page_count 는 manual_override")
        print("   엔트리에서 생략 (auto-derived 메타이므로 사람 검증된 entry 에 부적합).")

        print("\n6. Exact dry-run command (currently executed):")
        print("   python backend/scripts/patch_f1_manual_ref_v3.py")

        print("\n7. Apply command (NOT EXECUTED):")
        print("   python backend/scripts/patch_f1_manual_ref_v3.py --apply")
        print("   ※ apply 실행 시: 자동 백업 → 사전·사후 검증 → 실패 시 abort + 복구 명령 출력")
        return

    # ── APPLY ──
    print("\n[apply] 백업 생성...")
    bk = make_backup(raw)
    print(f"  backup: {bk.relative_to(ROOT)}")

    print("[apply] 신규 JSON 직렬화 + 파일 쓰기...")
    new_json = json.dumps(new_db, ensure_ascii=False, indent=2)
    DB_PATH.write_text(new_json, encoding="utf-8")

    # readback 검증
    print("[apply] readback 검증...")
    readback = json.loads(DB_PATH.read_text(encoding="utf-8"))
    rb_rows = readback.get("master_rows") or []
    if len(rb_rows) != 369:
        _abort(f"readback master_rows 개수 비정상: {len(rb_rows)}\n"
               f"복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    rb_idx = {r.get("row_id"): r for r in rb_rows}
    for rid, spec in PATCHES.items():
        if rb_idx.get(rid, {}).get("manual_ref") != spec["new_manual_ref"]:
            _abort(f"{rid} readback manual_ref 불일치\n"
                   f"복구: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")

    # 사후 verify (back-against-current)
    errs2 = verify_patch_safety(json.loads(bk.read_text(encoding="utf-8")), readback)
    if errs2:
        for e in errs2:
            print(f"  [error] {e}")
        _abort(f"사후 verify 실패\n복구: cp {bk.relative_to(ROOT)} "
               f"backend/data/immigration_guidelines_db_v2.json")

    summary["production_db_modified"] = True
    summary["backup_path"] = str(bk.relative_to(ROOT))

    print("\n" + "=" * 78)
    print("F-1 MANUAL_REF SURGICAL PATCH — APPLIED")
    print("=" * 78)
    print(f"  backup: {bk.relative_to(ROOT)}")
    print(f"  DB: {DB_PATH.relative_to(ROOT)} updated ({len(patch_report)} rows)")
    print(f"  rollback: cp {bk.relative_to(ROOT)} backend/data/immigration_guidelines_db_v2.json")
    print(f"  next: python backend/scripts/audit_post_apply_manual_ref_anomalies_v3.py")


if __name__ == "__main__":
    main()
